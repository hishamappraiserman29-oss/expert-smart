import sys, io, traceback, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
sys.path.insert(0, r'C:\Users\Lenovo\Desktop\expert_smart - Copy\core_engine')

import bridge_api as api

data = {
    'location': 'القاهرة الجديدة',
    'area': 200, 'price_per_meter': 35000,
    'rent_per_sqm': 380, 'cap_rate': 0.08,
    'building_age': 5, 'floor': 3, 'year_built': 2019,
    'property_type': 'شقة',
    'report_id': 'TEST-003',
    'report_date': '18/04/2026'
}
out = r'C:\Users\Lenovo\Desktop\expert_smart - Copy\core_engine\outputs\test_full3.xlsm'
try:
    api.write_to_excel_template(data, out)
    print('Done - no exception')
except Exception as e:
    print('ERROR:')
    traceback.print_exc()

# Verify sheet count
import openpyxl
wb = openpyxl.load_workbook(out, keep_vba=True)
print(f'Sheet count: {len(wb.sheetnames)}')
for i,n in enumerate(wb.sheetnames, 1):
    print(f'  {i:2}. {n}')
wb.close()
