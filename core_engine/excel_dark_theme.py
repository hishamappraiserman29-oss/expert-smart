"""
excel_dark_theme.py
===================
بوابة تصميم فخامة التقارير (Dark Mode v26 - Black & Gold)
تُستدعى قبل حفظ أي تقرير Excel بصيغة XLSM لضمان المعايير السيادية والطباعة الخالية من إهدار الحبر.
"""

import logging
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side
except ImportError:
    Workbook = None  # Mocking if not installed locally for some reason

def apply_sovereign_dark_theme(wb: Workbook) -> Workbook:
    """
    يطبق النمط الاستشاري الليلي على الشيتات الحساسة
    ويضبط إعدادات الطباعة السيادية (أسود وأبيض) لمنع طباعة الخلفيات الداكنة.
    """
    logger = logging.getLogger("ExcelSovereignTheme")
    
    if not wb:
        return wb

    # 1. كود الألوان (Palettes)
    dark_carbon_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    
    # 2. الخطوط (Fonts)
    gold_font    = Font(color="D4AF37", bold=True, name="Tajawal", size=12) # عناوين ذهبية
    silver_font  = Font(color="E0E0E0", name="Tajawal", size=11)            # بيانات فضية/رمادية فاتحة
    
    success_font = Font(color="10B981", bold=True, name="Tajawal")          # ختم الجودة الإيجابي (زمردي)
    error_font   = Font(color="EF4444", bold=True, name="Tajawal")          # ختم الجودة السلبي (أحمر)

    # 3. خطوط وعمق ثلاثي الأبعاد (Borders 3D Effect)
    gold_side    = Side(border_style="medium", color="D4AF37")
    dim_side     = Side(border_style="thin", color="333333")
    gold_border  = Border(top=gold_side, bottom=gold_side, left=dim_side, right=dim_side)

    # المرور على جميع الشيتات
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # ⚠️ المعيار السيادي للطباعة: (Print Settings)
        # هذا الأمر يجعل الإكسيل يتجاهل الخلفية السوداء ويعكس الألوان تلقائياً للورق!
        ws.page_setup.blackAndWhite = True  
        
        # تطبيق التنسيق فقط على Dashboard، الملخص، واتجاهات السوق (لتسريع الأداء)
        if any(keyword in sheet_name.lower() for keyword in ["dash", "ملخص", "رئيسية", "trends"]):
            logger.info(f"Injecting V26 Dark Mode into Sheet: {sheet_name}")

            for row in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=15):
                for cell in row:
                    # أ) الخلفية الكربونية
                    cell.fill = dark_carbon_fill
                    
                    # ب) التعامل مع الخطوط والنصوص
                    if cell.value:
                        text_val = str(cell.value)
                        
                        # علامات/أختام الجودة (Validation Stamp I1)
                        if "✅" in text_val:
                            cell.font = success_font
                        elif "❌" in text_val or "⚠" in text_val:
                            cell.font = error_font
                            
                        # القيم المالية والعناوين المركزية
                        elif any(k in text_val for k in ["قيمة", "إجمالي", "سعر", "متر", "Market Trends"]):
                            cell.font = gold_font
                            cell.border = gold_border # إضفاء عمق للأرقام الرئيسية
                        
                        # البيانات العادية
                        else:
                            cell.font = silver_font

    return wb
