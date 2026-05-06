"""
core_engine/mass_appraisal_excel.py
=====================================
Mass Appraisal Phase 1.3 — Professional XLSX Export

Exposes:
    build_mass_appraisal_workbook(result: dict) -> bytes

Rules:
- No disk I/O.  Returns raw bytes only.
- Does NOT call write_to_excel_template(), write_word_summary(),
  write_purpose_specific_report(), or any per-row report writer.
- Does NOT create one file per row.
- Uses openpyxl only (already in requirements.txt).
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Palette / shared styles ───────────────────────────────────────────────────
_FILL_HDR   = PatternFill("solid", fgColor="1F2937")   # dark header
_FILL_SEC   = PatternFill("solid", fgColor="374151")   # section sub-header
_FILL_OK    = PatternFill("solid", fgColor="D1FAE5")   # success row
_FILL_ERR   = PatternFill("solid", fgColor="FEE2E2")   # error row
_FILL_SKIP  = PatternFill("solid", fgColor="FEF3C7")   # skipped row

_FNT_HDR    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FNT_TITLE  = Font(name="Calibri", bold=True, color="FFD700", size=14)
_FNT_SEC    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FNT_LABEL  = Font(name="Calibri", bold=True)
_FNT_NOTE   = Font(name="Calibri", italic=True, color="6B7280", size=9)

_THIN_SIDE  = Side(style="thin")
_BORDER     = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                     top=_THIN_SIDE,  bottom=_THIN_SIDE)

_EGP_FMT    = '#,##0 "EGP"'
_PCT_FMT    = '0.00%'


# ── Helpers ───────────────────────────────────────────────────────────────────

def _hdr(ws, row: int, col: int, value: str) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _FNT_HDR
    c.fill      = _FILL_HDR
    c.border    = _BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _kv(ws, row: int, label: str, value: Any,
        num_fmt: Optional[str] = None) -> None:
    lc = ws.cell(row=row, column=1, value=label)
    lc.font   = _FNT_LABEL
    lc.border = _BORDER
    vc = ws.cell(row=row, column=2, value=value)
    vc.border = _BORDER
    if num_fmt and isinstance(value, (int, float)):
        vc.number_format = num_fmt


def _section_title(ws, row: int, n_cols: int, text: str) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    c = ws[f"A{row}"]
    c.value     = text
    c.font      = _FNT_SEC
    c.fill      = _FILL_SEC
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _auto_width(ws, min_w: int = 8, max_w: int = 42) -> None:
    for col in ws.columns:
        best = min_w
        for cell in col:
            try:
                v = str(cell.value or "")
                if len(v) > best:
                    best = len(v)
            except Exception:
                pass
        ws.column_dimensions[
            get_column_letter(col[0].column)
        ].width = min(best + 2, max_w)


# ── Sheet 1: Executive Summary ─────────────────────────────────────────────────

def _sheet_executive_summary(wb: Workbook, result: dict,
                              reviewed_summary: Optional[dict] = None,
                              audit: Optional[dict] = None) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.rightToLeft = True

    summary       = result.get("summary", {}) or {}
    rows_data     = result.get("rows", []) or []
    total         = summary.get("total_rows", len(rows_data))
    successful    = summary.get("successful_rows", 0)
    failed        = summary.get("failed_rows", 0)
    skipped       = summary.get("skipped_rows", 0)
    total_mv      = summary.get("total_market_value", 0.0)
    avg_mv        = summary.get("average_market_value", 0.0)
    med_mv        = summary.get("median_market_value", 0.0)
    avm_count     = summary.get("avm_applied_count", 0)
    purpose_cnts  = summary.get("purpose_counts", {}) or {}
    elapsed_ms    = summary.get("elapsed_ms")

    # Title banner
    ws.merge_cells("A1:D1")
    tc = ws["A1"]
    tc.value     = "تقرير التقييم الجماعي للعقارات"
    tc.font      = _FNT_TITLE
    tc.fill      = _FILL_HDR
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border    = _BORDER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    sc = ws["A2"]
    sc.value     = (f"Expert_Smart — Mass Appraisal  |  "
                    f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    sc.font      = Font(name="Calibri", italic=True, color="9CA3AF", size=10)
    sc.alignment = Alignment(horizontal="center")

    # Portfolio summary KVs
    _section_title(ws, 4, 4, "ملخص المحفظة العقارية")
    r = 5
    _kv(ws, r, "إجمالي العقارات",        total);                   r += 1
    _kv(ws, r, "نجاح",                    successful);              r += 1
    _kv(ws, r, "فشل",                     failed);                  r += 1
    _kv(ws, r, "مستبعد",                  skipped);                 r += 1
    _kv(ws, r, "إجمالي القيمة السوقية",   total_mv,    _EGP_FMT);  r += 1
    _kv(ws, r, "متوسط القيمة السوقية",    avg_mv,      _EGP_FMT);  r += 1
    _kv(ws, r, "الوسيط",                  med_mv,      _EGP_FMT);  r += 1
    _kv(ws, r, "صفوف AVM مُطبَّق",         avm_count);               r += 1
    if elapsed_ms is not None:
        _kv(ws, r, "زمن المعالجة (ms)", elapsed_ms);                r += 1

    # Purpose breakdown
    r += 1
    _section_title(ws, r, 4, "توزيع الأغراض")
    r += 1
    for purpose, count in purpose_cnts.items():
        _kv(ws, r, purpose, count)
        r += 1

    # Note
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    nc = ws[f"A{r}"]
    nc.value     = ("ملاحظة: هذا التقرير يمثّل مخرجات التقييم الجماعي الإجمالي. "
                    "لا يشمل تقارير Excel/Word فردية لكل عقار. "
                    "يُستخدم للتحليل الإجمالي والتدقيق الضريبي والمصرفي وفق معايير IVSC/IVS 2022.")
    nc.font      = _FNT_NOTE
    nc.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 40

    # ── Phase 1.8: Reviewed Portfolio Summary ────────────────────────────────
    if reviewed_summary and isinstance(reviewed_summary, dict):
        r += 2
        _section_title(ws, r, 4, "الملخص النهائي بعد المراجعة")
        r += 1
        _kv(ws, r, "إجمالي الصفوف الخام",                   reviewed_summary.get("raw_total_rows"));        r += 1
        _kv(ws, r, "صفوف ناجحة خام",                         reviewed_summary.get("raw_successful_rows"));   r += 1
        _kv(ws, r, "مُدرجة في المحفظة النهائية",             reviewed_summary.get("reviewed_included_rows")); r += 1
        _kv(ws, r, "مستبعدة من المحفظة النهائية",            reviewed_summary.get("reviewed_excluded_rows")); r += 1
        _kv(ws, r, "إجمالي القيمة السوقية (مراجَع)",         reviewed_summary.get("reviewed_total_market_value"),    _EGP_FMT); r += 1
        _kv(ws, r, "متوسط القيمة السوقية (مراجَع)",          reviewed_summary.get("reviewed_average_market_value"),  _EGP_FMT); r += 1
        _kv(ws, r, "وسيط القيمة السوقية (مراجَع)",           reviewed_summary.get("reviewed_median_market_value"),   _EGP_FMT); r += 1
        _kv(ws, r, "إجمالي المساحة (مراجَع م²)",             reviewed_summary.get("reviewed_total_area"));          r += 1
        _kv(ws, r, "متوسط القيمة/م² (مراجَع)",               reviewed_summary.get("reviewed_average_value_per_m2"), _EGP_FMT); r += 1
        r += 1
        _section_title(ws, r, 4, "توزيع قرارات المراجعة")
        r += 1
        _kv(ws, r, "معتمد",                                   reviewed_summary.get("approved_count"));              r += 1
        _kv(ws, r, "قيد المراجعة",                            reviewed_summary.get("pending_review_count"));        r += 1
        _kv(ws, r, "يحتاج تصحيح بيانات",                     reviewed_summary.get("needs_data_correction_count")); r += 1
        _kv(ws, r, "يحتاج تحقق سوقي",                        reviewed_summary.get("needs_market_verification_count")); r += 1
        _kv(ws, r, "مستبعد (قرار محلل)",                     reviewed_summary.get("excluded_count"));             r += 1
        _kv(ws, r, "مستبعد من الملخص النهائي",               reviewed_summary.get("excluded_from_final_summary_count")); r += 1
        final_ready = reviewed_summary.get("final_ready")
        _kv(ws, r, "جاهز للتقرير النهائي",                   "نعم" if final_ready else "لا");                    r += 1
        blocking = reviewed_summary.get("blocking_issues")
        if blocking:
            r += 1
            ws.merge_cells(f"A{r}:D{r}")
            bc = ws[f"A{r}"]
            bc.value     = "تحذيرات تمنع الإغلاق: " + str(blocking)
            bc.font      = Font(name="Calibri", bold=True, color="DC2626", size=10)
            bc.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 28

    # ── Phase 1.9: Audit info block ───────────────────────────────────────
    if audit and isinstance(audit, dict):
        r += 2
        _section_title(ws, r, 4, "معلومات التدقيق والحوكمة — Audit & Governance")
        r += 1
        _kv(ws, r, "Batch ID",          audit.get("batch_id"));           r += 1
        _kv(ws, r, "اسم الدفعة",        audit.get("batch_label"));        r += 1
        _kv(ws, r, "المراجع / Analyst", audit.get("analyst_name"));       r += 1
        _kv(ws, r, "وقت التشغيل",       audit.get("run_timestamp"));      r += 1
        _kv(ws, r, "وقت المراجعة",      audit.get("review_timestamp"));   r += 1
        _kv(ws, r, "وقت التصدير",       audit.get("export_timestamp"));   r += 1
        _kv(ws, r, "الوحدة",            audit.get("app_module"));          r += 1
        _kv(ws, r, "المرحلة",           audit.get("app_phase"));           r += 1
        mn = audit.get("methodology_note")
        if mn:
            r += 1
            ws.merge_cells(f"A{r}:D{r}")
            mc = ws[f"A{r}"]
            mc.value     = "المنهجية: " + str(mn)
            mc.font      = _FNT_NOTE
            mc.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 36

    # ── Phase 3.1: Zone & Property-Class Summary ─────────────────────────────
    zone_sum   = summary.get("zone_summary", {})   or {}
    pclass_sum = summary.get("property_class_summary", {}) or {}

    if zone_sum:
        r += 2
        _section_title(ws, r, 4, "ملخص المناطق الجغرافية (Zone Summary)")
        r += 1
        for zone_id, zd in sorted(zone_sum.items()):
            ws.merge_cells(f"A{r}:D{r}")
            zh = ws[f"A{r}"]
            zh.value     = f"Zone: {zone_id}"
            zh.font      = Font(name="Calibri", bold=True, size=10)
            zh.fill      = _FILL_SKIP
            zh.border    = _BORDER
            r += 1
            nbhd = zd.get("neighborhood")
            smkt = zd.get("submarket")
            if nbhd:
                _kv(ws, r, "  الحي",    nbhd); r += 1
            if smkt:
                _kv(ws, r, "  السوق الفرعي", smkt); r += 1
            _kv(ws, r, "  إجمالي الصفوف",       zd.get("row_count", 0)); r += 1
            _kv(ws, r, "  ناجحة",                zd.get("successful_rows", 0)); r += 1
            _kv(ws, r, "  إجمالي القيمة السوقية", zd.get("total_market_value", 0.0), _EGP_FMT); r += 1
            _kv(ws, r, "  متوسط القيمة السوقية",  zd.get("average_market_value", 0.0), _EGP_FMT); r += 1
            _kv(ws, r, "  متوسط القيمة/م²",       zd.get("average_value_per_m2", 0.0), _EGP_FMT); r += 1

    if pclass_sum:
        r += 2
        _section_title(ws, r, 4, "ملخص فئات العقار (Property Class Summary)")
        r += 1
        for pc, pd in sorted(pclass_sum.items()):
            ws.merge_cells(f"A{r}:D{r}")
            ph = ws[f"A{r}"]
            ph.value     = f"Class: {pc}"
            ph.font      = Font(name="Calibri", bold=True, size=10)
            ph.fill      = _FILL_SKIP
            ph.border    = _BORDER
            r += 1
            _kv(ws, r, "  إجمالي الصفوف",       pd.get("row_count", 0)); r += 1
            _kv(ws, r, "  ناجحة",                pd.get("successful_rows", 0)); r += 1
            _kv(ws, r, "  إجمالي القيمة السوقية", pd.get("total_market_value", 0.0), _EGP_FMT); r += 1
            _kv(ws, r, "  متوسط القيمة السوقية",  pd.get("average_market_value", 0.0), _EGP_FMT); r += 1
            _kv(ws, r, "  متوسط القيمة/م²",       pd.get("average_value_per_m2", 0.0), _EGP_FMT); r += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A3"


# ── Sheet 2: Portfolio Results ─────────────────────────────────────────────────

def _sheet_portfolio(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("Portfolio Results")

    # Phase 1.4 + 1.5: location / property_type / area, ppm traceability, DQ columns
    headers = [
        "row_index", "row_id", "status", "valuation_purpose",
        "location", "property_type", "area",
        "market_value", "currency",
        "price_per_meter_input", "price_per_meter_effective", "calculation_source",
        "avm_applied", "avm_confidence", "avm_n_records", "avm_ppm",
        "tax_annual_tax", "tax_taxable_amount",
        "usufruct_years", "usufruct_discount_rate", "usufruct_pv_factor",
        "uncertainty_low", "uncertainty_best", "uncertainty_high",
        "uncertainty_spread_pct", "error", "skip_reason", "warnings",
        # Phase 1.5
        "data_quality_score", "data_quality_level", "data_quality_flags",
        # Phase 1.6
        "value_per_m2", "review_required", "outlier_score", "outlier_level", "review_reasons",
        # Phase 1.8
        "included_in_final_summary", "review_status", "analyst_note", "exclude_from_final_summary",
        # Phase 3.1
        "zone_id", "neighborhood", "submarket", "property_class",
    ]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    # col 8 = market_value, 17/18 = tax, 22/23/24 = uncertainty
    MONEY_COLS = {8, 17, 18, 22, 23, 24}
    N_COLS     = len(headers)

    for r_idx, row in enumerate(rows_data, 2):
        avm = row.get("avm")             or {}
        tax = row.get("tax_assessment")  or {}
        usu = row.get("usufruct")        or {}
        unc = row.get("uncertainty_range") or {}

        values = [
            row.get("row_index"),
            row.get("row_id"),
            row.get("status"),
            row.get("valuation_purpose"),
            row.get("location"),
            row.get("property_type"),
            row.get("area"),
            row.get("market_value"),
            row.get("currency"),
            row.get("price_per_meter_input"),
            row.get("price_per_meter_effective"),
            row.get("calculation_source"),
            avm.get("applied"),
            avm.get("confidence"),
            avm.get("n_records"),
            avm.get("avm_ppm"),
            tax.get("annual_tax"),
            tax.get("taxable_amount"),
            usu.get("usufruct_years"),
            usu.get("discount_rate"),
            usu.get("pv_factor"),
            unc.get("low"),
            unc.get("best"),
            unc.get("high"),
            unc.get("spread_pct"),
            row.get("error"),
            row.get("skip_reason"),
            "; ".join(row.get("warnings") or []) or None,
            # Phase 1.5
            row.get("data_quality_score"),
            row.get("data_quality_level"),
            "\n".join(
                "[%s] %s: %s" % (f.get("severity", "?"), f.get("code", "?"),
                                  f.get("message", ""))
                for f in (row.get("data_quality_flags") or [])
            ) or None,
            # Phase 1.6
            row.get("value_per_m2"),
            row.get("review_required"),
            row.get("outlier_score"),
            row.get("outlier_level"),
            "\n".join(
                "[%s] %s: %s" % (f.get("severity", "?"), f.get("code", "?"),
                                  f.get("message", ""))
                for f in (row.get("review_reasons") or [])
            ) or None,
            # Phase 1.8
            row.get("included_in_final_summary"),
            row.get("review_status"),
            row.get("analyst_note"),
            row.get("exclude_from_final_summary"),
            # Phase 3.1
            row.get("zone_id"),
            row.get("neighborhood"),
            row.get("submarket"),
            row.get("property_class"),
        ]

        status = row.get("status", "")
        fill   = (_FILL_OK   if status == "success"
                  else _FILL_ERR  if status == "error"
                  else _FILL_SKIP)

        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = fill
            cell.border = _BORDER
            if c_idx in (28, N_COLS):   # warnings and data_quality_flags
                cell.alignment = Alignment(wrap_text=True)
            if c_idx in MONEY_COLS and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT

    ws.auto_filter.ref = f"A1:{get_column_letter(N_COLS)}1"
    ws.freeze_panes    = "A2"
    _auto_width(ws)


# ── Sheet 3: Purpose Analysis ──────────────────────────────────────────────────

def _sheet_purpose_analysis(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("Purpose Analysis")

    agg: Dict[str, Dict[str, Any]] = {}
    for row in rows_data:
        p = row.get("valuation_purpose") or "unknown"
        if p not in agg:
            agg[p] = {"row_count": 0, "successful_count": 0, "failed_count": 0,
                      "skipped_count": 0, "total_mv": 0.0, "mvs": [],
                      "avm_count": 0}
        d = agg[p]
        d["row_count"] += 1
        status = row.get("status", "")
        if status == "success":
            d["successful_count"] += 1
            mv = row.get("market_value") or 0.0
            d["total_mv"] += mv
            d["mvs"].append(mv)
        elif status == "error":
            d["failed_count"] += 1
        elif status == "skipped":
            d["skipped_count"] += 1
        if (row.get("avm") or {}).get("applied"):
            d["avm_count"] += 1

    headers = ["purpose", "row_count", "successful_count", "failed_count",
               "skipped_count", "total_market_value", "average_market_value",
               "avm_applied_count"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    purposes = list(agg.keys())
    for r_idx, p in enumerate(purposes, 2):
        d = agg[p]
        mvs = d["mvs"]
        avg = (d["total_mv"] / len(mvs)) if mvs else 0.0
        vals = [p, d["row_count"], d["successful_count"], d["failed_count"],
                d["skipped_count"], d["total_mv"], avg, d["avm_count"]]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDER
            if c_idx in (6, 7) and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT

    n = len(purposes)
    if n > 0:
        chart_row = n + 4

        bar = BarChart()
        bar.type  = "col"
        bar.title = "Total Market Value by Purpose"
        bar.y_axis.title = "EGP"
        bar.x_axis.title = "Purpose"
        bar.add_data(Reference(ws, min_col=6, max_col=6, min_row=1, max_row=n + 1),
                     titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
        bar.shape = 4
        ws.add_chart(bar, f"A{chart_row}")

        pie = PieChart()
        pie.title = "Row Count by Purpose"
        pie.add_data(Reference(ws, min_col=2, max_col=2, min_row=1, max_row=n + 1),
                     titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
        ws.add_chart(pie, f"J{chart_row}")

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Sheet 4: Location Analysis ─────────────────────────────────────────────────

def _sheet_location_analysis(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("Location Analysis")

    # Phase 1.4: location is now echoed in every row result
    has_location = any(row.get("location") for row in rows_data)
    if not has_location:
        ws.merge_cells("A1:F1")
        nc = ws["A1"]
        nc.value     = "Location data not available in this batch result."
        nc.font      = _FNT_NOTE
        nc.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[1].height = 40
        ws.column_dimensions["A"].width = 60
        return

    agg: Dict[str, Dict[str, Any]] = {}
    for row in rows_data:
        loc = row.get("location") or "unknown"
        if loc not in agg:
            agg[loc] = {"row_count": 0, "mvs": [], "total_mv": 0.0}
        agg[loc]["row_count"] += 1
        if row.get("status") == "success":
            mv = row.get("market_value") or 0.0
            agg[loc]["total_mv"] += mv
            agg[loc]["mvs"].append(mv)

    headers = ["location", "row_count", "total_market_value",
               "average_market_value", "max_market_value", "min_market_value"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    locs = list(agg.keys())
    for r_idx, loc in enumerate(locs, 2):
        d   = agg[loc]
        mvs = d["mvs"]
        avg = (d["total_mv"] / len(mvs)) if mvs else 0.0
        vals = [loc, d["row_count"], d["total_mv"], avg,
                max(mvs) if mvs else None, min(mvs) if mvs else None]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDER
            if c_idx in (3, 4, 5, 6) and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT

    n = len(locs)
    if n > 0:
        chart_row = n + 4
        bar = BarChart()
        bar.type  = "col"
        bar.title = "Total Market Value by Location"
        bar.y_axis.title = "EGP"
        bar.add_data(Reference(ws, min_col=3, max_col=3, min_row=1, max_row=n + 1),
                     titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
        ws.add_chart(bar, f"A{chart_row}")

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Sheet 4b: Asset Type Analysis ─────────────────────────────────────────────

def _sheet_asset_type_analysis(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("Asset Type Analysis")

    agg: Dict[str, Dict[str, Any]] = {}
    for row in rows_data:
        pt = row.get("property_type") or "unknown"
        if pt not in agg:
            agg[pt] = {"row_count": 0, "mvs": [], "total_mv": 0.0,
                       "total_area": 0.0}
        agg[pt]["row_count"] += 1
        if row.get("status") == "success":
            mv = row.get("market_value") or 0.0
            agg[pt]["total_mv"] += mv
            agg[pt]["mvs"].append(mv)
            area = row.get("area")
            if isinstance(area, (int, float)) and area > 0:
                agg[pt]["total_area"] += area

    if not agg:
        ws["A1"]      = "No property type data available."
        ws["A1"].font = _FNT_NOTE
        return

    headers = ["property_type", "row_count", "total_market_value",
               "average_market_value", "average_value_per_m2"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    pts = list(agg.keys())
    for r_idx, pt in enumerate(pts, 2):
        d   = agg[pt]
        mvs = d["mvs"]
        avg_mv   = (d["total_mv"] / len(mvs))         if mvs else 0.0
        avg_pm2  = (d["total_mv"] / d["total_area"])  if d["total_area"] > 0 else 0.0
        vals = [pt, d["row_count"], d["total_mv"], avg_mv, avg_pm2]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDER
            if c_idx in (3, 4, 5) and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT

    n = len(pts)
    if n > 0:
        chart_row = n + 4
        bar = BarChart()
        bar.type  = "col"
        bar.title = "Total Market Value by Asset Type"
        bar.y_axis.title = "EGP"
        bar.add_data(Reference(ws, min_col=3, max_col=3, min_row=1, max_row=n + 1),
                     titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
        ws.add_chart(bar, f"A{chart_row}")

        pie = PieChart()
        pie.title = "Row Count by Asset Type"
        pie.add_data(Reference(ws, min_col=2, max_col=2, min_row=1, max_row=n + 1),
                     titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
        ws.add_chart(pie, f"J{chart_row}")

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Sheet 5: AVM Analysis ──────────────────────────────────────────────────────

def _sheet_avm_analysis(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("AVM Analysis")

    headers = ["row_id", "valuation_purpose", "market_value",
               "avm_applied", "avm_confidence", "avm_n_records", "avm_ppm"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    applied_cnt     = 0
    not_applied_cnt = 0
    conf_dist: Dict[str, int] = {}

    for r_idx, row in enumerate(rows_data, 2):
        avm     = row.get("avm") or {}
        applied = avm.get("applied")
        conf    = avm.get("confidence")
        if applied:
            applied_cnt += 1
        elif avm:
            not_applied_cnt += 1
        if conf:
            conf_dist[conf] = conf_dist.get(conf, 0) + 1
        vals = [row.get("row_id"), row.get("valuation_purpose"),
                row.get("market_value"), applied,
                conf, avm.get("n_records"), avm.get("avm_ppm")]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDER
            if c_idx == 3 and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT

    n_data  = len(rows_data)
    s_start = n_data + 3

    ws.cell(row=s_start,     column=1, value="AVM Applied").font     = _FNT_LABEL
    ws.cell(row=s_start,     column=2, value=applied_cnt)
    ws.cell(row=s_start + 1, column=1, value="AVM Not Applied").font = _FNT_LABEL
    ws.cell(row=s_start + 1, column=2, value=not_applied_cnt)

    # Pie: applied vs not applied
    ws.cell(row=s_start,     column=4, value="AVM Applied")
    ws.cell(row=s_start,     column=5, value=applied_cnt)
    ws.cell(row=s_start + 1, column=4, value="AVM Not Applied")
    ws.cell(row=s_start + 1, column=5, value=not_applied_cnt)

    if applied_cnt + not_applied_cnt > 0:
        pie = PieChart()
        pie.title = "AVM Applied vs Not Applied"
        pie.add_data(Reference(ws, min_col=5, max_col=5,
                               min_row=s_start, max_row=s_start + 1))
        pie.set_categories(Reference(ws, min_col=4,
                                     min_row=s_start, max_row=s_start + 1))
        ws.add_chart(pie, f"A{s_start + 5}")

    # Bar: count by confidence
    if conf_dist:
        cb = s_start + 5
        ws.cell(row=cb, column=8, value="Confidence").font = _FNT_LABEL
        ws.cell(row=cb, column=9, value="Count").font      = _FNT_LABEL
        for i, (conf, cnt) in enumerate(conf_dist.items(), 1):
            ws.cell(row=cb + i, column=8, value=conf)
            ws.cell(row=cb + i, column=9, value=cnt)
        nc = len(conf_dist)
        bar = BarChart()
        bar.type  = "col"
        bar.title = "AVM Count by Confidence Level"
        bar.add_data(Reference(ws, min_col=9, max_col=9,
                               min_row=cb, max_row=cb + nc), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=8,
                                     min_row=cb + 1, max_row=cb + nc))
        ws.add_chart(bar, f"J{s_start + 5}")

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Sheet 6: Tax Assessment ────────────────────────────────────────────────────

def _sheet_tax_assessment(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("Tax Assessment")

    tax_rows = [r for r in rows_data if r.get("tax_assessment")]
    if not tax_rows:
        ws["A1"]      = "لا توجد صفوف ضريبة عقارية في هذه الدفعة."
        ws["A1"].font = _FNT_NOTE
        return

    headers = ["row_id", "market_value", "annual_rental_value",
               "tax_base", "taxable_amount", "tax_rate", "annual_tax",
               "governorate_factor", "construction_factor", "location_factor",
               "exemption_threshold", "legal_basis",
               "assessed_value", "exemption_amount", "taxable_value",
               "effective_tax_rate", "tax_due",
               "policy_profile", "property_class",
               "appeal_strength", "operator_recommendation"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    MONEY_COLS_TAX = {2, 3, 4, 5, 7, 11, 13, 14, 15, 17}
    PCT_COLS_TAX   = {6, 16}
    total_tax = 0.0

    for r_idx, row in enumerate(tax_rows, 2):
        tax = row.get("tax_assessment") or {}
        at  = tax.get("annual_tax") or 0.0
        total_tax += at
        vals = [row.get("row_id"),
                tax.get("market_value"), tax.get("annual_rental_value"),
                tax.get("tax_base"),     tax.get("taxable_amount"),
                tax.get("tax_rate"),     at,
                tax.get("governorate_factor"), tax.get("construction_factor"),
                tax.get("location_factor"),    tax.get("exemption_threshold"),
                tax.get("legal_basis"),
                tax.get("assessed_value"),   tax.get("exemption_amount"),
                tax.get("taxable_value"),    tax.get("effective_tax_rate"),
                tax.get("tax_due"),
                tax.get("policy_profile"),   tax.get("property_class"),
                (tax.get("tax_appeal_package") or {}).get("appeal_strength"),
                (tax.get("tax_appeal_package") or {}).get("operator_recommendation")]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDER
            if c_idx in MONEY_COLS_TAX and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT
            if c_idx in PCT_COLS_TAX and isinstance(val, (int, float)):
                cell.number_format = _PCT_FMT

    sr = len(tax_rows) + 3
    tax_vals_list = [
        (r.get("tax_assessment") or {}).get("annual_tax") or 0.0
        for r in tax_rows
    ]
    _kv(ws, sr,     "إجمالي الضريبة السنوية", total_tax,
        _EGP_FMT)
    _kv(ws, sr + 1, "متوسط الضريبة السنوية",
        total_tax / len(tax_rows) if tax_rows else 0, _EGP_FMT)
    _kv(ws, sr + 2, "الحد الأقصى للضريبة السنوية",
        max(tax_vals_list), _EGP_FMT)

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Sheet 7: Usufruct & Uncertainty ───────────────────────────────────────────

def _sheet_usufruct_uncertainty(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("Usufruct & Uncertainty")

    usu_rows = [r for r in rows_data if r.get("usufruct")]
    unc_rows = [r for r in rows_data if r.get("uncertainty_range")]
    ptr = 1

    # Usufruct section
    _section_title(ws, ptr, 6, "حق الانتفاع (Usufruct)")
    ptr += 1
    usu_hdrs = ["row_id", "market_value", "usufruct_years",
                "discount_rate", "pv_factor", "note"]
    for c, h in enumerate(usu_hdrs, 1):
        _hdr(ws, ptr, c, h)
    ptr += 1
    if usu_rows:
        for row in usu_rows:
            usu  = row.get("usufruct") or {}
            vals = [row.get("row_id"), row.get("market_value"),
                    usu.get("usufruct_years"), usu.get("discount_rate"),
                    usu.get("pv_factor"),      usu.get("note")]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=ptr, column=c_idx, value=val)
                cell.border = _BORDER
                if c_idx == 2 and isinstance(val, (int, float)):
                    cell.number_format = _EGP_FMT
            ptr += 1
    else:
        ws.cell(row=ptr, column=1, value="لا توجد صفوف حق انتفاع.").font = _FNT_NOTE
        ptr += 1

    ptr += 2

    # Uncertainty section
    _section_title(ws, ptr, 6, "نطاق عدم اليقين (IFRS 13 §93)")
    ptr += 1
    unc_hdrs = ["row_id", "market_value", "low", "best", "high", "spread_pct"]
    for c, h in enumerate(unc_hdrs, 1):
        _hdr(ws, ptr, c, h)
    ptr += 1
    if unc_rows:
        for row in unc_rows:
            unc  = row.get("uncertainty_range") or {}
            vals = [row.get("row_id"), row.get("market_value"),
                    unc.get("low"), unc.get("best"),
                    unc.get("high"), unc.get("spread_pct")]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=ptr, column=c_idx, value=val)
                cell.border = _BORDER
                if c_idx in (2, 3, 4, 5) and isinstance(val, (int, float)):
                    cell.number_format = _EGP_FMT
            ptr += 1
    else:
        ws.cell(row=ptr, column=1, value="لا توجد صفوف عدم يقين.").font = _FNT_NOTE

    _auto_width(ws)


# ── Sheet 8: Data Quality ─────────────────────────────────────────────────────

_FILL_DQ_HIGH     = PatternFill("solid", fgColor="D1FAE5")   # green  (high)
_FILL_DQ_MEDIUM   = PatternFill("solid", fgColor="FEF3C7")   # yellow (medium)
_FILL_DQ_LOW      = PatternFill("solid", fgColor="FED7AA")   # orange (low)
_FILL_DQ_CRITICAL = PatternFill("solid", fgColor="FEE2E2")   # red    (critical)

_DQ_FILL_MAP = {
    "high":     _FILL_DQ_HIGH,
    "medium":   _FILL_DQ_MEDIUM,
    "low":      _FILL_DQ_LOW,
    "critical": _FILL_DQ_CRITICAL,
}


def _sheet_data_quality(wb: Workbook, rows_data: List[dict], summary: dict) -> None:
    ws = wb.create_sheet("Data Quality")

    dq_counts  = summary.get("data_quality_counts") or {}
    avg_score  = summary.get("average_data_quality_score")
    crit_count = summary.get("rows_with_critical_quality")
    err_count  = summary.get("rows_with_errors")
    warn_count = summary.get("rows_with_warnings")

    _section_title(ws, 1, 5, "ملخص جودة البيانات")
    r = 2
    _kv(ws, r, "متوسط درجة الجودة",    avg_score);                      r += 1
    _kv(ws, r, "صفوف عالية الجودة",    dq_counts.get("high", 0));       r += 1
    _kv(ws, r, "صفوف جودة متوسطة",     dq_counts.get("medium", 0));     r += 1
    _kv(ws, r, "صفوف جودة منخفضة",     dq_counts.get("low", 0));        r += 1
    _kv(ws, r, "صفوف جودة حرجة",       dq_counts.get("critical", 0));   r += 1
    _kv(ws, r, "صفوف بأخطاء جودة",     err_count);                      r += 1
    _kv(ws, r, "صفوف بتحذيرات جودة",   warn_count);                     r += 1
    r += 1

    hdrs = ["row_index", "row_id", "status", "location", "property_type",
            "data_quality_score", "data_quality_level", "data_quality_flags"]
    for c, h in enumerate(hdrs, 1):
        _hdr(ws, r, c, h)
    hdr_row = r
    r += 1

    for row in rows_data:
        dql      = row.get("data_quality_level") or "critical"
        flags    = row.get("data_quality_flags") or []
        row_fill = _DQ_FILL_MAP.get(dql, _FILL_DQ_CRITICAL)
        flags_str = "\n".join(
            "[%s] %s: %s" % (f.get("severity", "?"), f.get("code", "?"),
                              f.get("message", ""))
            for f in flags
        ) or None
        vals = [
            row.get("row_index"),
            row.get("row_id"),
            row.get("status"),
            row.get("location"),
            row.get("property_type"),
            row.get("data_quality_score"),
            dql,
            flags_str,
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.fill   = row_fill
            cell.border = _BORDER
            if c_idx == 8:
                cell.alignment = Alignment(wrap_text=True)
        r += 1

    # Auto-filter on header row
    ws.auto_filter.ref = f"A{hdr_row}:H{hdr_row}"

    # Distribution bar chart (score distribution)
    chart_row = r + 2
    dist_data = [("high", dq_counts.get("high", 0)),
                 ("medium", dq_counts.get("medium", 0)),
                 ("low", dq_counts.get("low", 0)),
                 ("critical", dq_counts.get("critical", 0))]
    for i, (level, count) in enumerate(dist_data):
        ws.cell(row=chart_row + i, column=10, value=level)
        ws.cell(row=chart_row + i, column=11, value=count)

    if any(c > 0 for _, c in dist_data):
        bar = BarChart()
        bar.type  = "col"
        bar.title = "Data Quality Distribution"
        bar.y_axis.title = "Row Count"
        bar.add_data(Reference(ws, min_col=11, max_col=11,
                               min_row=chart_row, max_row=chart_row + 3))
        bar.set_categories(Reference(ws, min_col=10,
                                     min_row=chart_row, max_row=chart_row + 3))
        ws.add_chart(bar, f"A{chart_row + 6}")

    ws.freeze_panes = f"A{hdr_row + 1}"
    _auto_width(ws)


# ── Sheet 9: Errors & Skipped Rows ────────────────────────────────────────────

def _sheet_errors_skipped(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("Errors & Skipped Rows")

    bad = [r for r in rows_data if r.get("status") in ("error", "skipped")]
    if not bad:
        ws["A1"]      = "لا توجد أخطاء أو صفوف مستبعدة."
        ws["A1"].font = _FNT_NOTE
        return

    headers = ["row_index", "row_id", "status", "valuation_purpose",
               "error", "skip_reason", "warnings"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    for r_idx, row in enumerate(bad, 2):
        status = row.get("status", "")
        fill   = _FILL_ERR if status == "error" else _FILL_SKIP
        vals   = [row.get("row_index"), row.get("row_id"), status,
                  row.get("valuation_purpose"), row.get("error"),
                  row.get("skip_reason"),
                  "; ".join(row.get("warnings") or []) or None]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = fill
            cell.border = _BORDER
            if c_idx == 7:
                cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Sheet 11 (new): Review Queue ─────────────────────────────────────────────

_OL_FILL_MAP = {
    "critical": _FILL_DQ_CRITICAL,
    "review":   _FILL_DQ_LOW,       # orange
    "watch":    _FILL_DQ_MEDIUM,    # yellow
    "normal":   _FILL_DQ_HIGH,      # green
}


def _sheet_review_queue(wb: Workbook, rows_data: List[dict], summary: dict) -> None:
    ws = wb.create_sheet("Review Queue")

    ol_summary = summary.get("outlier_summary") or {}
    rrc        = summary.get("review_required_count", 0)
    top_codes  = summary.get("top_review_reasons") or {}
    pm         = summary.get("portfolio_medians") or {}

    _section_title(ws, 1, 6, "قائمة المراجعة — Outlier Review Queue")
    r = 2
    _kv(ws, r, "صفوف تتطلب مراجعة",     rrc);                            r += 1
    _kv(ws, r, "normal",                  ol_summary.get("normal",   0));  r += 1
    _kv(ws, r, "watch",                   ol_summary.get("watch",    0));  r += 1
    _kv(ws, r, "review",                  ol_summary.get("review",   0));  r += 1
    _kv(ws, r, "critical",                ol_summary.get("critical", 0));  r += 1
    _kv(ws, r, "وسيط القيمة السوقية",    pm.get("median_market_value"), _EGP_FMT); r += 1
    _kv(ws, r, "وسيط المساحة (م²)",      pm.get("median_area"));           r += 1
    _kv(ws, r, "وسيط القيمة/م²",         pm.get("median_value_per_m2"), _EGP_FMT); r += 1
    r += 1

    hdrs = [
        "row_index", "row_id", "status", "location", "property_type",
        "area", "market_value", "value_per_m2", "valuation_purpose",
        "data_quality_score", "data_quality_level",
        "outlier_score", "outlier_level",
        "reason_severity", "reason_code", "reason_message",
    ]
    for c, h in enumerate(hdrs, 1):
        _hdr(ws, r, c, h)
    hdr_row = r
    r += 1

    MONEY_COLS_RQ = {7, 8}   # market_value, value_per_m2

    # Include rows that are review_required OR error/skipped OR review/critical level
    review_rows = [
        row for row in rows_data
        if (row.get("review_required") or
            row.get("status") in ("error", "skipped") or
            row.get("outlier_level") in ("review", "critical"))
    ]

    for row in review_rows:
        reasons = list(row.get("review_reasons") or [])
        if not reasons:
            reasons = [{"severity": "info", "code": "REVIEW_REQUIRED",
                        "message": "Row flagged for review."}]

        ol  = row.get("outlier_level") or "normal"
        row_fill = _OL_FILL_MAP.get(ol, _FILL_DQ_HIGH)

        for reason in reasons:
            vals = [
                row.get("row_index"),
                row.get("row_id"),
                row.get("status"),
                row.get("location"),
                row.get("property_type"),
                row.get("area"),
                row.get("market_value"),
                row.get("value_per_m2"),
                row.get("valuation_purpose"),
                row.get("data_quality_score"),
                row.get("data_quality_level"),
                row.get("outlier_score"),
                row.get("outlier_level"),
                reason.get("severity"),
                reason.get("code"),
                reason.get("message"),
            ]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.fill   = row_fill
                cell.border = _BORDER
                if c_idx == 16:
                    cell.alignment = Alignment(wrap_text=True)
                if c_idx in MONEY_COLS_RQ and isinstance(val, (int, float)):
                    cell.number_format = _EGP_FMT
            r += 1

    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(hdrs))}{hdr_row}"
    ws.freeze_panes    = f"A{hdr_row + 1}"

    # ── Summary by reason code ─────────────────────────────────────────────────
    sr = r + 2
    _section_title(ws, sr, 4, "Top Review Reason Codes")
    sr += 1
    _hdr(ws, sr, 1, "reason_code")
    _hdr(ws, sr, 2, "count")
    sr += 1
    sorted_codes = sorted(top_codes.items(), key=lambda x: -x[1])
    code_start   = sr
    for code, count in sorted_codes[:15]:
        ws.cell(row=sr, column=1, value=code).border   = _BORDER
        ws.cell(row=sr, column=2, value=count).border  = _BORDER
        sr += 1

    # Charts
    n_codes = min(len(sorted_codes), 15)
    if n_codes > 0:
        bar = BarChart()
        bar.type  = "col"
        bar.title = "Top Review Reason Codes"
        bar.y_axis.title = "Count"
        bar.add_data(Reference(ws, min_col=2, max_col=2,
                               min_row=code_start - 1, max_row=code_start + n_codes - 1),
                     titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1,
                                     min_row=code_start, max_row=code_start + n_codes - 1))
        ws.add_chart(bar, f"D{code_start - 1}")

    ol_dist_data = [("normal",   ol_summary.get("normal",   0)),
                    ("watch",    ol_summary.get("watch",    0)),
                    ("review",   ol_summary.get("review",   0)),
                    ("critical", ol_summary.get("critical", 0))]
    if any(c > 0 for _, c in ol_dist_data):
        pie_start = sr + 2
        for i, (level, count) in enumerate(ol_dist_data):
            ws.cell(row=pie_start + i, column=10, value=level)
            ws.cell(row=pie_start + i, column=11, value=count)
        pie = PieChart()
        pie.title = "Outlier Level Distribution"
        pie.add_data(Reference(ws, min_col=11, max_col=11,
                               min_row=pie_start, max_row=pie_start + 3))
        pie.set_categories(Reference(ws, min_col=10,
                                     min_row=pie_start, max_row=pie_start + 3))
        ws.add_chart(pie, f"D{sr + 4}")

    _auto_width(ws)


# ── Sheet 13: Final Reviewed Portfolio ────────────────────────────────────────

def _sheet_final_reviewed_portfolio(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("Final Reviewed Portfolio")

    included = [
        row for row in rows_data
        if (row.get("status") == "success" and
            row.get("review_status") == "approved" and
            not row.get("exclude_from_final_summary"))
    ]

    if not included:
        ws["A1"]      = "لا توجد صفوف معتمدة ومُدرجة في المحفظة النهائية."
        ws["A1"].font = _FNT_NOTE
        return

    headers = [
        "row_index", "row_id", "location", "property_type", "area",
        "valuation_purpose", "market_value", "value_per_m2",
        "data_quality_level", "outlier_level", "review_status", "analyst_note",
    ]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    MONEY_COLS_FRP = {7, 8}

    for r_idx, row in enumerate(included, 2):
        vals = [
            row.get("row_index"),
            row.get("row_id"),
            row.get("location"),
            row.get("property_type"),
            row.get("area"),
            row.get("valuation_purpose"),
            row.get("market_value"),
            row.get("value_per_m2"),
            row.get("data_quality_level"),
            row.get("outlier_level"),
            row.get("review_status"),
            row.get("analyst_note"),
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = _FILL_OK
            cell.border = _BORDER
            if c_idx in MONEY_COLS_FRP and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes    = "A2"
    _auto_width(ws)


# ── Sheet 14: Final Exclusions ─────────────────────────────────────────────────

def _sheet_final_exclusions(wb: Workbook, rows_data: List[dict]) -> None:
    ws = wb.create_sheet("Final Exclusions")

    def _exclusion_reason(row: dict) -> str:
        if row.get("status") in ("error", "skipped"):
            return row.get("error") or row.get("skip_reason") or row.get("status", "")
        if row.get("exclude_from_final_summary"):
            return "excluded_by_analyst"
        rs = row.get("review_status", "")
        if rs and rs != "approved":
            return f"review_status={rs}"
        return "not_approved"

    excluded = [
        row for row in rows_data
        if not (row.get("status") == "success" and
                row.get("review_status") == "approved" and
                not row.get("exclude_from_final_summary"))
    ]

    if not excluded:
        ws["A1"]      = "لا توجد صفوف مستبعدة من المحفظة النهائية."
        ws["A1"].font = _FNT_NOTE
        return

    headers = [
        "row_index", "row_id", "status", "location", "property_type", "area",
        "valuation_purpose", "market_value", "review_status",
        "exclude_from_final_summary", "exclusion_reason", "analyst_note",
    ]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    for r_idx, row in enumerate(excluded, 2):
        status = row.get("status", "")
        fill   = (_FILL_ERR  if status == "error"
                  else _FILL_SKIP if status == "skipped"
                  else PatternFill("solid", fgColor="F3F4F6"))
        vals = [
            row.get("row_index"),
            row.get("row_id"),
            status,
            row.get("location"),
            row.get("property_type"),
            row.get("area"),
            row.get("valuation_purpose"),
            row.get("market_value"),
            row.get("review_status"),
            row.get("exclude_from_final_summary"),
            _exclusion_reason(row),
            row.get("analyst_note"),
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = fill
            cell.border = _BORDER
            if c_idx == 8 and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes    = "A2"
    _auto_width(ws)


# ── Sheet 15: Audit & Governance ─────────────────────────────────────────────

def _sheet_audit_governance(wb: Workbook, audit: Optional[dict],
                             summary: dict,
                             reviewed_summary: Optional[dict] = None) -> None:
    ws = wb.create_sheet("Audit & Governance")

    _section_title(ws, 1, 3, "Audit & Governance — Phase 1.9")

    r = 2
    # Audit identity
    _kv(ws, r, "Batch ID",                        (audit or {}).get("batch_id", ""));          r += 1
    _kv(ws, r, "Batch Label",                     (audit or {}).get("batch_label", ""));       r += 1
    _kv(ws, r, "Analyst Name",                    (audit or {}).get("analyst_name", ""));      r += 1
    _kv(ws, r, "Run Timestamp",                   (audit or {}).get("run_timestamp", ""));     r += 1
    _kv(ws, r, "Review Timestamp",                (audit or {}).get("review_timestamp", ""));  r += 1
    _kv(ws, r, "Export Timestamp",                (audit or {}).get("export_timestamp", ""));  r += 1
    _kv(ws, r, "App Module",                      (audit or {}).get("app_module", ""));        r += 1
    _kv(ws, r, "App Phase",                       (audit or {}).get("app_phase", ""));         r += 1
    _kv(ws, r, "Source",                          (audit or {}).get("source", ""));            r += 1
    r += 1

    # Portfolio counts
    _section_title(ws, r, 3, "Portfolio Counts")
    r += 1
    _kv(ws, r, "Total Rows",       summary.get("total_rows",      0));                         r += 1
    _kv(ws, r, "Successful Rows",  summary.get("successful_rows", 0));                         r += 1
    _kv(ws, r, "Failed Rows",      summary.get("failed_rows",     0));                         r += 1
    _kv(ws, r, "Skipped Rows",     summary.get("skipped_rows",    0));                         r += 1
    if reviewed_summary and isinstance(reviewed_summary, dict):
        _kv(ws, r, "Reviewed Included", reviewed_summary.get("reviewed_included_rows"));       r += 1
        _kv(ws, r, "Reviewed Excluded", reviewed_summary.get("reviewed_excluded_rows"));       r += 1
        final_ready = reviewed_summary.get("final_ready")
        _kv(ws, r, "Final Ready",       "Yes" if final_ready else "No");                       r += 1
    r += 1

    # Narrative notes (wide cells)
    _section_title(ws, r, 3, "Methodology & Assumptions")
    r += 1
    for label, key in [
        ("Methodology Note",  "methodology_note"),
        ("Assumptions Note",  "assumptions_note"),
        ("Limitations Note",  "limitations_note"),
    ]:
        text = (audit or {}).get(key, "")
        lc = ws.cell(row=r, column=1, value=label)
        lc.font   = _FNT_LABEL
        lc.border = _BORDER
        vc = ws.cell(row=r, column=2, value=text)
        vc.border = _BORDER
        vc.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"B{r}:C{r}")
        ws.row_dimensions[r].height = max(30, min(80, len(str(text)) // 2))
        r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20
    ws.freeze_panes = "A2"


# ── Sheet 12: Assumptions ─────────────────────────────────────────────────────

def _sheet_assumptions(wb: Workbook, import_validation: Optional[dict] = None) -> None:
    ws = wb.create_sheet("Assumptions")

    _section_title(ws, 1, 3, "Assumptions & Notes")

    points = [
        "This XLSX is generated from Mass Appraisal Run results.",
        "It is an aggregate portfolio-level workbook — not a per-property report.",
        "It does NOT call write_to_excel_template(), write_word_summary(), or any per-row report writer.",
        "AVM confidence and n_records are carried from the valuation engine.",
        "Tax, usufruct, and uncertainty blocks are included only when present in row results.",
        "CSV export remains client-side (Download CSV button).",
        "This XLSX export is generated server-side in memory — no disk writes.",
        "Phase 1.4: location, property_type, area are now echoed in every row result.",
        "Phase 1.4: success rows include price_per_meter_input, price_per_meter_effective, calculation_source.",
        "Phase 1.4: summary includes location_counts, property_type_counts, average_value_per_m2.",
        "Phase 1.5: every row carries data_quality_score (0-100), data_quality_level, data_quality_flags.",
        "Phase 1.5: new Data Quality sheet with per-row DQ coloring and portfolio DQ summary.",
        "Phase 1.6: portfolio-level outlier detection; every row carries value_per_m2, "
        "review_required, review_reasons, outlier_score, outlier_level.",
        "Phase 1.6: new Review Queue sheet expanding review_reasons into one row per reason.",
        "Standards: IVSC, IVS 2022, Basel III/IV, IFRS 13.",
        f"Export timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ]
    if import_validation:
        iv_sum = import_validation.get("summary") or {}
        mr     = iv_sum.get("matching_readiness") or {}
        points.append(
            f"Phase 3.14: Import source: Excel Upload | "
            f"Properties: {iv_sum.get('total_rows', '')} | Sales: {iv_sum.get('total_sales', '')} | "
            f"Warnings: {iv_sum.get('warnings_count', '')} | Errors: {iv_sum.get('errors_count', '')}"
        )
        points.append(
            f"Phase 3.14: Matching ready: {mr.get('matching_ready', '')} | "
            f"Matched: {mr.get('matched_count', '')} | "
            f"Unmatched properties: {mr.get('unmatched_properties', '')} | "
            f"Unmatched sales: {mr.get('unmatched_sales', '')}"
        )
    for i, point in enumerate(points, 2):
        c = ws.cell(row=i, column=1, value=f"• {point}")
        c.alignment = Alignment(wrap_text=True)
        c.border    = _BORDER
        ws.row_dimensions[i].height = 28

    ws.column_dimensions["A"].width = 90
    ws.row_dimensions[1].height = 24


# ── Phase 3.1: Zone Analysis Sheet ───────────────────────────────────────────

def _sheet_zone_analysis(wb: Workbook, rows_data: List[dict], summary: dict) -> None:
    ws = wb.create_sheet("Zone Analysis")
    ws.sheet_view.rightToLeft = True

    zone_sum = (summary.get("zone_summary") or {})
    if not zone_sum:
        ws.cell(row=1, column=1, value="No zone data — rows did not supply zone_id.").font = _FNT_NOTE
        return

    # Header row
    headers = [
        "zone_id", "neighborhood", "submarket",
        "row_count", "successful_rows", "failed_rows", "skipped_rows",
        "total_market_value", "average_market_value",
        "total_area", "average_value_per_m2",
    ]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    MONEY_COLS = {8, 9, 11}
    for r_idx, (zone_id, zd) in enumerate(sorted(zone_sum.items()), 2):
        values = [
            zone_id,
            zd.get("neighborhood"),
            zd.get("submarket"),
            zd.get("row_count", 0),
            zd.get("successful_rows", 0),
            zd.get("failed_rows", 0),
            zd.get("skipped_rows", 0),
            zd.get("total_market_value", 0.0),
            zd.get("average_market_value", 0.0),
            zd.get("total_area", 0.0),
            zd.get("average_value_per_m2", 0.0),
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDER
            if c_idx in MONEY_COLS and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes    = "A2"
    _auto_width(ws)


# ── Phase 3.1: Property Class Analysis Sheet ──────────────────────────────────

def _sheet_property_class_analysis(wb: Workbook, rows_data: List[dict], summary: dict) -> None:
    ws = wb.create_sheet("Property Class Analysis")
    ws.sheet_view.rightToLeft = True

    pclass_sum = (summary.get("property_class_summary") or {})
    if not pclass_sum:
        ws.cell(row=1, column=1,
                value="No property class data — rows did not supply property_class.").font = _FNT_NOTE
        return

    headers = [
        "property_class",
        "row_count", "successful_rows", "failed_rows", "skipped_rows",
        "total_market_value", "average_market_value",
        "total_area", "average_value_per_m2",
    ]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)

    MONEY_COLS = {6, 7, 9}
    for r_idx, (pc, pd) in enumerate(sorted(pclass_sum.items()), 2):
        values = [
            pc,
            pd.get("row_count", 0),
            pd.get("successful_rows", 0),
            pd.get("failed_rows", 0),
            pd.get("skipped_rows", 0),
            pd.get("total_market_value", 0.0),
            pd.get("average_market_value", 0.0),
            pd.get("total_area", 0.0),
            pd.get("average_value_per_m2", 0.0),
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDER
            if c_idx in MONEY_COLS and isinstance(val, (int, float)):
                cell.number_format = _EGP_FMT

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes    = "A2"
    _auto_width(ws)


# ── Phase 3.3: Ratio Study Sheet ─────────────────────────────────────────────

def _sheet_ratio_study(wb: Workbook, ratio_study: dict) -> None:
    ws = wb.create_sheet("Ratio Study")
    ws.sheet_view.rightToLeft = True

    summary  = ratio_study.get("summary", {}) or {}
    pm       = summary.get("portfolio_metrics", {}) or {}

    # Title banner
    ws.merge_cells("A1:D1")
    tc = ws["A1"]
    tc.value     = "دراسة النسب — Ratio Study (Phase 3.3)"
    tc.font      = _FNT_TITLE
    tc.fill      = _FILL_HDR
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border    = _BORDER
    ws.row_dimensions[1].height = 28

    r = 3
    _section_title(ws, r, 4, "ملخص دراسة النسب — Study Summary"); r += 1
    _kv(ws, r, "Subject Count",          summary.get("subject_count"));          r += 1
    _kv(ws, r, "Sales Count",            summary.get("sales_count"));            r += 1
    _kv(ws, r, "Usable Sales",           summary.get("usable_sales_count"));     r += 1
    _kv(ws, r, "Matched Pairs",          summary.get("matched_pair_count"));     r += 1
    _kv(ws, r, "Unmatched Subjects",     summary.get("unmatched_subject_count")); r += 1
    _kv(ws, r, "Unmatched Sales",        summary.get("unmatched_sales_count"));  r += 1

    r += 1
    _section_title(ws, r, 4, "Portfolio Metrics"); r += 1
    _kv(ws, r, "Sample Size",            pm.get("sample_size"));          r += 1
    _kv(ws, r, "Mean Ratio",             pm.get("mean_ratio"));           r += 1
    _kv(ws, r, "Median Ratio",           pm.get("median_ratio"));         r += 1
    _kv(ws, r, "Weighted Mean Ratio",    pm.get("weighted_mean_ratio"));  r += 1
    _kv(ws, r, "COD (%)",               pm.get("cod"));                   r += 1
    _kv(ws, r, "PRD",                   pm.get("prd"));                   r += 1
    _kv(ws, r, "Min Ratio",             pm.get("min_ratio"));             r += 1
    _kv(ws, r, "Max Ratio",             pm.get("max_ratio"));             r += 1

    # ── Zone Metrics ──────────────────────────────────────────────────────────
    zone_m = ratio_study.get("zone_metrics", {}) or {}
    if zone_m:
        r += 1
        _section_title(ws, r, 4, "Zone Metrics"); r += 1
        for zone_id, zm in sorted(zone_m.items()):
            ws.merge_cells(f"A{r}:D{r}")
            zh = ws[f"A{r}"]
            zh.value  = f"Zone: {zone_id}"
            zh.font   = Font(name="Calibri", bold=True, size=10)
            zh.fill   = _FILL_SKIP
            zh.border = _BORDER
            r += 1
            _kv(ws, r, "  Sample Size",  zm.get("sample_size"));  r += 1
            _kv(ws, r, "  Mean Ratio",   zm.get("mean_ratio"));    r += 1
            _kv(ws, r, "  Median Ratio", zm.get("median_ratio"));  r += 1
            _kv(ws, r, "  COD (%)",      zm.get("cod"));           r += 1
            _kv(ws, r, "  PRD",          zm.get("prd"));           r += 1

    # ── Property Class Metrics ────────────────────────────────────────────────
    pclass_m = ratio_study.get("property_class_metrics", {}) or {}
    if pclass_m:
        r += 1
        _section_title(ws, r, 4, "Property Class Metrics"); r += 1
        for pc, pm2 in sorted(pclass_m.items()):
            ws.merge_cells(f"A{r}:D{r}")
            ph = ws[f"A{r}"]
            ph.value  = f"Class: {pc}"
            ph.font   = Font(name="Calibri", bold=True, size=10)
            ph.fill   = _FILL_SKIP
            ph.border = _BORDER
            r += 1
            _kv(ws, r, "  Sample Size",  pm2.get("sample_size"));  r += 1
            _kv(ws, r, "  Mean Ratio",   pm2.get("mean_ratio"));   r += 1
            _kv(ws, r, "  Median Ratio", pm2.get("median_ratio")); r += 1
            _kv(ws, r, "  COD (%)",      pm2.get("cod"));          r += 1
            _kv(ws, r, "  PRD",          pm2.get("prd"));          r += 1

    # ── Matched Pairs table ───────────────────────────────────────────────────
    matched = ratio_study.get("matched_pairs", []) or []
    if matched:
        r += 2
        pair_hdrs = [
            "subject_row_id", "sale_id", "zone_id", "property_class",
            "appraised_value", "sale_price", "ratio", "ratio_pct",
            "sale_price_per_m2", "appraised_value_per_m2",
            "abs_deviation_from_median", "pct_deviation_from_median",
            "match_basis",
        ]
        _section_title(ws, r, len(pair_hdrs), "Matched Pairs"); r += 1
        for c, h in enumerate(pair_hdrs, 1):
            _hdr(ws, r, c, h)
        r += 1
        MONEY_COLS = {5, 6, 9, 10}
        for p in matched:
            vals = [
                p.get("subject_row_id"), p.get("sale_id"),
                p.get("zone_id"), p.get("property_class"),
                p.get("appraised_value"), p.get("sale_price"),
                p.get("ratio"), p.get("ratio_pct"),
                p.get("sale_price_per_m2"), p.get("appraised_value_per_m2"),
                p.get("absolute_deviation_from_median"),
                p.get("percent_deviation_from_median"),
                p.get("match_basis"),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.border = _BORDER
                if ci in MONEY_COLS and isinstance(v, (int, float)):
                    cell.number_format = _EGP_FMT
            r += 1

    # ── Unmatched Subjects ────────────────────────────────────────────────────
    unmatched_s = ratio_study.get("unmatched_subjects", []) or []
    if unmatched_s:
        r += 1
        _section_title(ws, r, 5, "Unmatched Subjects"); r += 1
        us_hdrs = ["row_id", "zone_id", "property_class", "location", "area"]
        for c, h in enumerate(us_hdrs, 1):
            _hdr(ws, r, c, h)
        r += 1
        for s in unmatched_s:
            vals = [s.get("row_id"), s.get("zone_id"), s.get("property_class"),
                    s.get("location"), s.get("area")]
            for ci, v in enumerate(vals, 1):
                ws.cell(row=r, column=ci, value=v).border = _BORDER
            r += 1

    # ── Unmatched Sales ───────────────────────────────────────────────────────
    unmatched_sl = ratio_study.get("unmatched_sales", []) or []
    if unmatched_sl:
        r += 1
        _section_title(ws, r, 5, "Unmatched Sales"); r += 1
        sl_hdrs = ["sale_id", "zone_id", "property_class", "location", "sale_price"]
        for c, h in enumerate(sl_hdrs, 1):
            _hdr(ws, r, c, h)
        r += 1
        for s in unmatched_sl:
            vals = [s.get("sale_id"), s.get("zone_id"), s.get("property_class"),
                    s.get("location"), s.get("sale_price")]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.border = _BORDER
                if ci == 5 and isinstance(v, (int, float)):
                    cell.number_format = _EGP_FMT
            r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A3"


# ── Phase 3.6: Calibration Preview sheet ─────────────────────────────────────

def _sheet_calibration_preview(wb: Workbook, calibration: dict) -> None:
    """Add a Calibration Preview sheet (Phase 3.6 — preview only)."""
    ws = wb.create_sheet("Calibration Preview")
    ws.sheet_view.rightToLeft = True

    summary = calibration.get("summary", {}) or {}
    port    = calibration.get("portfolio_calibration", {}) or {}
    zone_c  = calibration.get("zone_calibration", {}) or {}
    pclass_c = calibration.get("property_class_calibration", {}) or {}
    combo_c  = calibration.get("zone_property_class_calibration", {}) or {}
    notes    = calibration.get("calibration_notes", []) or []

    # Title banner
    ws.merge_cells("A1:E1")
    tc = ws["A1"]
    tc.value     = "معاينة المعايرة — Calibration Preview (Phase 3.6 — Preview Only)"
    tc.font      = _FNT_TITLE
    tc.fill      = _FILL_HDR
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border    = _BORDER
    ws.row_dimensions[1].height = 28

    r = 3
    # ── Summary ───────────────────────────────────────────────────────────────
    _section_title(ws, r, 5, "ملخص — Summary"); r += 1
    _kv(ws, r, "Subject Count",                  summary.get("subject_count"));               r += 1
    _kv(ws, r, "Sales Count",                    summary.get("sales_count"));                 r += 1
    _kv(ws, r, "Matched Pair Count",             summary.get("matched_pair_count"));          r += 1
    _kv(ws, r, "Pairs Source",                   summary.get("pairs_source"));                r += 1
    _kv(ws, r, "Target Median Ratio",            summary.get("target_median_ratio"));         r += 1
    _kv(ws, r, "Portfolio Suggested Factor",     summary.get("portfolio_suggested_factor"));  r += 1
    _kv(ws, r, "Portfolio Recommendation",       summary.get("portfolio_recommendation"));    r += 1
    _kv(ws, r, "Groups With Low Sample",         summary.get("groups_with_low_sample"));      r += 1
    _kv(ws, r, "Groups Requiring Major Review",  summary.get("groups_requiring_major_review")); r += 1

    # ── Portfolio Calibration ─────────────────────────────────────────────────
    r += 1
    _section_title(ws, r, 5, "معايرة المحفظة — Portfolio Calibration"); r += 1
    for label, key in [
        ("Sample Size",         "sample_size"),
        ("Median Ratio",        "median_ratio"),
        ("Mean Ratio",          "mean_ratio"),
        ("Weighted Mean Ratio", "weighted_mean_ratio"),
        ("COD (%)",             "cod"),
        ("PRD",                 "prd"),
        ("Suggested Factor",    "suggested_factor"),
        ("Recommendation",      "recommendation"),
    ]:
        _kv(ws, r, label, port.get(key)); r += 1

    # ── Zone Calibration table ────────────────────────────────────────────────
    r += 1
    _section_title(ws, r, 5, "معايرة المناطق — Zone Calibration"); r += 1
    hdrs = ["Zone", "Sample Size", "Median Ratio", "Suggested Factor", "Recommendation"]
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font   = _FNT_HDR
        cell.fill   = _FILL_SEC
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")
    r += 1
    for zone_id, m in sorted(zone_c.items()):
        ws.cell(row=r, column=1, value=zone_id).border  = _BORDER
        ws.cell(row=r, column=2, value=m.get("sample_size")).border = _BORDER
        ws.cell(row=r, column=3, value=m.get("median_ratio")).border = _BORDER
        ws.cell(row=r, column=4, value=m.get("suggested_factor")).border = _BORDER
        ws.cell(row=r, column=5, value=m.get("recommendation")).border  = _BORDER
        r += 1

    # ── Property Class Calibration table ──────────────────────────────────────
    r += 1
    _section_title(ws, r, 5, "معايرة فئات العقارات — Property Class Calibration"); r += 1
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=r, column=ci, value=h.replace("Zone", "Property Class"))
        cell.font   = _FNT_HDR
        cell.fill   = _FILL_SEC
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")
    r += 1
    for pc, m in sorted(pclass_c.items()):
        ws.cell(row=r, column=1, value=pc).border           = _BORDER
        ws.cell(row=r, column=2, value=m.get("sample_size")).border  = _BORDER
        ws.cell(row=r, column=3, value=m.get("median_ratio")).border = _BORDER
        ws.cell(row=r, column=4, value=m.get("suggested_factor")).border = _BORDER
        ws.cell(row=r, column=5, value=m.get("recommendation")).border  = _BORDER
        r += 1

    # ── Zone+Class Combination table ──────────────────────────────────────────
    if combo_c:
        r += 1
        _section_title(ws, r, 5, "Zone + Property Class Combination"); r += 1
        combo_hdrs = ["Zone | Class", "Sample Size", "Median Ratio", "Suggested Factor", "Recommendation"]
        for ci, h in enumerate(combo_hdrs, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font   = _FNT_HDR
            cell.fill   = _FILL_SEC
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
        r += 1
        for combo_key, m in sorted(combo_c.items()):
            ws.cell(row=r, column=1, value=combo_key).border        = _BORDER
            ws.cell(row=r, column=2, value=m.get("sample_size")).border  = _BORDER
            ws.cell(row=r, column=3, value=m.get("median_ratio")).border = _BORDER
            ws.cell(row=r, column=4, value=m.get("suggested_factor")).border = _BORDER
            ws.cell(row=r, column=5, value=m.get("recommendation")).border  = _BORDER
            r += 1

    # ── Warnings ──────────────────────────────────────────────────────────────
    top_warns = summary.get("warnings") or []
    if top_warns:
        r += 1
        _section_title(ws, r, 5, "تحذيرات — Warnings"); r += 1
        for w in top_warns:
            ws.merge_cells(f"A{r}:E{r}")
            cell = ws.cell(row=r, column=1,
                           value=f"[{w.get('code','?')}] {w.get('message','')}")
            cell.font   = Font(name="Calibri", color="991B1B", italic=True, size=9)
            cell.border = _BORDER
            r += 1

    # ── Notes ─────────────────────────────────────────────────────────────────
    r += 1
    _section_title(ws, r, 5, "ملاحظات — Notes"); r += 1
    for note in notes:
        ws.merge_cells(f"A{r}:E{r}")
        cell = ws.cell(row=r, column=1, value=note)
        cell.font   = _FNT_NOTE
        cell.border = _BORDER
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 32
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["E"].width = 22
    ws.freeze_panes = "A3"


# ── Phase 3.7: Calibration Sandbox sheet ─────────────────────────────────────

def _sheet_calibration_sandbox(wb: Workbook, sandbox: dict) -> None:
    """Add a Calibration Sandbox sheet (Phase 3.7 — sandbox only)."""
    ws = wb.create_sheet("Calibration Sandbox")
    ws.sheet_view.rightToLeft = True

    summary = sandbox.get("summary", {}) or {}
    rows    = sandbox.get("rows", []) or []

    # Title banner
    ws.merge_cells("A1:I1")
    tc = ws["A1"]
    tc.value     = "المعايرة التجريبية — Calibration Sandbox (Phase 3.7 — Sandbox Only)"
    tc.font      = _FNT_TITLE
    tc.fill      = _FILL_HDR
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border    = _BORDER
    ws.row_dimensions[1].height = 28

    # Sandbox disclaimer
    ws.merge_cells("A2:I2")
    disc = ws["A2"]
    disc.value     = ("⚠  هذه معايرة تجريبية لا تغيّر القيم الأصلية ولا يتم استخدامها كاعتماد نهائي."
                      " — SANDBOX: original values unchanged, not for official use.")
    disc.font      = Font(name="Calibri", bold=True, color="92400E", size=9)
    disc.fill      = PatternFill("solid", fgColor="FEF3C7")
    disc.alignment = Alignment(horizontal="center")
    disc.border    = _BORDER
    ws.row_dimensions[2].height = 18

    r = 4
    # ── Summary ───────────────────────────────────────────────────────────────
    _section_title(ws, r, 9, "ملخص — Sandbox Summary"); r += 1
    _kv(ws, r, "Total Rows",                   summary.get("total_rows"));               r += 1
    _kv(ws, r, "Calibrated Rows",              summary.get("calibrated_rows"));          r += 1
    _kv(ws, r, "Unchanged Rows",               summary.get("unchanged_rows"));           r += 1
    _kv(ws, r, "Original Total Market Value",  summary.get("original_total_market_value")); r += 1
    _kv(ws, r, "Sandbox Total Market Value",   summary.get("sandbox_total_market_value")); r += 1
    _kv(ws, r, "Total Value Delta",            summary.get("total_value_delta"));        r += 1
    _kv(ws, r, "Total Value Delta %",          summary.get("total_value_delta_pct"));    r += 1
    _kv(ws, r, "Average Calibration Factor",   summary.get("average_calibration_factor")); r += 1

    # Factor source counts
    source_counts = summary.get("factor_source_counts") or {}
    if source_counts:
        r += 1
        _section_title(ws, r, 9, "مصادر المعامل — Factor Source Counts"); r += 1
        for src, cnt in sorted(source_counts.items()):
            _kv(ws, r, src, cnt); r += 1

    # ── Calibrated Rows table ─────────────────────────────────────────────────
    if rows:
        r += 1
        _section_title(ws, r, 9, "الصفوف المعايرة — Calibrated Rows"); r += 1
        hdrs = [
            "row_id", "zone_id", "property_class",
            "original_market_value", "factor_applied", "factor_source",
            "sandbox_calibrated_value", "value_delta", "warnings",
        ]
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font   = _FNT_HDR
            cell.fill   = _FILL_SEC
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
        r += 1
        for row_d in rows:
            sb_val    = row_d.get("sandbox_calibrated_value")
            orig_val  = row_d.get("original_market_value")
            fill      = _FILL_OK if sb_val is not None else PatternFill("solid", fgColor="F3F4F6")
            warns_str = "; ".join(row_d.get("calibration_sandbox_warnings") or []) or ""

            ws.cell(row=r, column=1, value=row_d.get("row_id")).border        = _BORDER
            ws.cell(row=r, column=2, value=row_d.get("zone_id")).border        = _BORDER
            ws.cell(row=r, column=3, value=row_d.get("property_class")).border = _BORDER
            ws.cell(row=r, column=4, value=orig_val).border                    = _BORDER
            ws.cell(row=r, column=5, value=row_d.get("calibration_factor_applied")).border = _BORDER
            ws.cell(row=r, column=6, value=row_d.get("calibration_factor_source")).border  = _BORDER
            ws.cell(row=r, column=7, value=sb_val).border                      = _BORDER
            ws.cell(row=r, column=8, value=row_d.get("sandbox_value_delta")).border        = _BORDER
            ws.cell(row=r, column=9, value=warns_str).border                   = _BORDER

            for ci in range(1, 10):
                ws.cell(row=r, column=ci).fill = fill
            r += 1

    # Column widths
    col_widths = [14, 14, 16, 24, 16, 20, 26, 14, 40]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"


# ── Phase 3.8: Governance Approval sheet ─────────────────────────────────────

def _sheet_governance(wb: Workbook, governance: dict) -> None:
    ws = wb.create_sheet("Governance Approval")

    DARK   = PatternFill("solid", fgColor="1F2937")
    GOLD   = Font(bold=True, color="FBBF24", size=12)
    HDR_F  = Font(bold=True, color="FFFFFF")
    HDR_BG = PatternFill("solid", fgColor="374151")
    APPR_F = Font(bold=True, color="166534")
    APPR_B = PatternFill("solid", fgColor="DCFCE7")
    REJ_F  = Font(bold=True, color="991B1B")
    REJ_B  = PatternFill("solid", fgColor="FEE2E2")
    REV_F  = Font(bold=True, color="1E40AF")
    REV_B  = PatternFill("solid", fgColor="DBEAFE")
    DRAFT_F= Font(bold=True, color="374151")
    DRAFT_B= PatternFill("solid", fgColor="F3F4F6")
    WARN_B = PatternFill("solid", fgColor="FEF3C7")
    thin   = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    _status_styles = {
        "approved":     (APPR_F,  APPR_B),
        "rejected":     (REJ_F,   REJ_B),
        "under_review": (REV_F,   REV_B),
    }

    # Title row
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value       = "Governance Approval — Phase 3.8"
    t.font        = GOLD
    t.fill        = DARK
    t.alignment   = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Disclaimer
    ws.merge_cells("A2:D2")
    d = ws["A2"]
    d.value     = "هذه وثيقة حوكمة داخلية للمراجعة فقط. لا تُعدّ اعتمادًا رسميًا بموجب أي إطار قانوني أو تنظيمي."
    d.font      = Font(italic=True, color="7C3AED", size=9)
    d.fill      = PatternFill("solid", fgColor="F5F3FF")
    d.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Metadata KVs
    meta_fields = [
        ("governance_id",      governance.get("governance_id", "")),
        ("status",             governance.get("status", "")),
        ("reviewer_name",      governance.get("reviewer_name", "")),
        ("reviewer_role",      governance.get("reviewer_role", "")),
        ("decision_note",      governance.get("decision_note", "")),
        ("approved_outputs",   ", ".join(governance.get("approved_outputs") or [])),
        ("created_at",         governance.get("created_at", "")),
        ("last_updated",       governance.get("last_updated", "")),
        ("approval_timestamp", governance.get("approval_timestamp") or ""),
    ]

    row = 3
    # Header for KVs
    for col, hdr in enumerate(["Field", "Value"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font   = HDR_F
        c.fill   = HDR_BG
        c.border = border
    row += 1

    status_val = governance.get("status", "")
    for field, value in meta_fields:
        kc = ws.cell(row=row, column=1, value=field)
        vc = ws.cell(row=row, column=2, value=str(value) if value is not None else "")
        kc.border = border
        vc.border = border
        if field == "status" and status_val in _status_styles:
            sf, sb = _status_styles[status_val]
            kc.font = sf; kc.fill = sb
            vc.font = sf; vc.fill = sb
        elif field == "decision_note":
            kc.fill = WARN_B; vc.fill = WARN_B
        row += 1

    row += 1  # blank separator

    # Governance history table
    history = governance.get("governance_history") or []
    if history:
        ws.merge_cells(f"A{row}:D{row}")
        hdr_c = ws.cell(row=row, column=1, value="Governance History")
        hdr_c.font  = HDR_F
        hdr_c.fill  = HDR_BG
        hdr_c.alignment = Alignment(horizontal="center")
        row += 1
        for col, hdr in enumerate(["Timestamp", "From Status", "To Status", "Actor / Note"], 1):
            c = ws.cell(row=row, column=col, value=hdr)
            c.font   = HDR_F
            c.fill   = HDR_BG
            c.border = border
        row += 1
        for entry in history:
            ts   = str(entry.get("timestamp", "")).replace("T", " ")[:16]
            frm  = str(entry.get("from_status", ""))
            to   = str(entry.get("to_status", ""))
            act  = str(entry.get("actor", ""))
            note = str(entry.get("note", ""))
            actor_note = act + ((" — " + note[:60]) if note else "")
            for col, val in enumerate([ts, frm, to, actor_note], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = border
                sf, sb = _status_styles.get(to, (DRAFT_F, DRAFT_B))
                if col in (2, 3):
                    c.font = sf; c.fill = sb
            row += 1

    # Column widths
    for ci, w in enumerate([22, 20, 20, 60], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ── Phase 3.9: Model & Revaluation Cycle sheet ───────────────────────────────

def _sheet_model_cycle(wb: Workbook, mc: dict) -> None:
    ws = wb.create_sheet("Model & Revaluation Cycle")

    NAVY   = PatternFill("solid", fgColor="0C4A6E")
    BLUE_F = Font(bold=True, color="7DD3FC", size=12)
    HDR_F  = Font(bold=True, color="FFFFFF")
    HDR_BG = PatternFill("solid", fgColor="075985")
    INFO_B = PatternFill("solid", fgColor="E0F2FE")
    INFO_F = Font(color="0369A1")
    thin   = Side(style="thin", color="BAE6FD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    _status_color = {
        "draft":        "6B7280",
        "under_review": "2563EB",
        "approved":     "16A34A",
        "retired":      "92400E",
        "superseded":   "7C3AED",
    }

    # Title
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = "Model & Revaluation Cycle — Phase 3.9"
    t.font      = BLUE_F
    t.fill      = NAVY
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Disclaimer
    ws.merge_cells("A2:C2")
    d = ws["A2"]
    d.value     = ("هذه البيانات للتتبع والمراجعة فقط — لا تُطبَّق معايرة تلقائية "
                   "ولا تُعدَّل قيم التقييم.")
    d.font      = INFO_F
    d.fill      = INFO_B
    d.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Column headers
    row = 3
    for col, hdr in enumerate(["Field", "Value", "Notes"], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.font   = HDR_F
        c.fill   = HDR_BG
        c.border = border
    row += 1

    status_val = mc.get("model_status", "draft")
    sc_hex     = _status_color.get(status_val, "6B7280")

    fields = [
        ("cycle_id",              mc.get("cycle_id", ""),              "Unique cycle identifier"),
        ("cycle_name",            mc.get("cycle_name", ""),            "Human-readable cycle name"),
        ("revaluation_year",      mc.get("revaluation_year", ""),      "Target revaluation year"),
        ("valuation_date",        mc.get("valuation_date", ""),        "As-of valuation date"),
        ("effective_date",        mc.get("effective_date", ""),        "Date values take effect"),
        ("model_version",         mc.get("model_version", ""),         "Model version tag"),
        ("model_family",          mc.get("model_family", ""),          "Model family / methodology"),
        ("model_status",          status_val,                          "Lifecycle status"),
        ("calibration_reference", mc.get("calibration_reference", ""), "Phase 3.6/3.7 reference ID"),
        ("ratio_study_reference", mc.get("ratio_study_reference", ""), "Phase 3.3 reference ID"),
        ("notes",                 mc.get("notes", ""),                 "Additional notes"),
        ("created_at",            mc.get("created_at", ""),            "Record creation timestamp"),
        ("last_updated",          mc.get("last_updated", ""),          "Last modification timestamp"),
    ]

    for field, value, note in fields:
        kc = ws.cell(row=row, column=1, value=field)
        vc = ws.cell(row=row, column=2, value=str(value) if value is not None else "")
        nc = ws.cell(row=row, column=3, value=note)
        kc.border = border
        vc.border = border
        nc.border = border
        nc.font   = Font(color="6B7280", italic=True, size=9)
        if field == "model_status":
            status_fill = PatternFill("solid", fgColor=sc_hex)
            status_font = Font(bold=True, color="FFFFFF")
            vc.fill = status_fill
            vc.font = status_font
        elif field in ("cycle_id", "model_version"):
            kc.font = Font(bold=True)
        elif field in ("created_at", "last_updated"):
            kc.fill = PatternFill("solid", fgColor="F8FAFC")
            vc.fill = PatternFill("solid", fgColor="F8FAFC")
            kc.font = Font(color="64748B")
            vc.font = Font(color="64748B")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 40


# ── Phase 3.14: Export Metadata ───────────────────────────────────────────────

def _sheet_export_metadata(wb: Workbook, result: dict, import_validation: Optional[dict]) -> None:
    ws = wb.create_sheet("Export_Metadata")
    _section_title(ws, 1, 2, "Export Metadata")
    summary = result.get("summary") or {}
    iv_sum  = (import_validation or {}).get("summary") or {}
    mr      = iv_sum.get("matching_readiness") or {}
    fields: List[tuple] = [
        ("Export Timestamp",   datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Run Status",         result.get("status", "unknown")),
        ("Total Properties",   summary.get("total_rows", len(result.get("rows") or []))),
        ("Successful Rows",    summary.get("successful_rows", "")),
        ("Failed Rows",        summary.get("failed_rows", "")),
        ("Total Market Value", summary.get("total_market_value", "")),
        ("Import Source",      "Excel Upload" if import_validation else "Manual / API"),
        ("Import Properties",  iv_sum.get("total_rows", "")),
        ("Import Sales",       iv_sum.get("total_sales", "")),
        ("Warnings Count",     iv_sum.get("warnings_count", "")),
        ("Errors Count",       iv_sum.get("errors_count", "")),
        ("Matching Ready",     str(mr.get("matching_ready", "")) if mr else ""),
    ]
    for r, (key, val) in enumerate(fields, 2):
        kc = ws.cell(row=r, column=1, value=key)
        vc = ws.cell(row=r, column=2, value=str(val) if val not in (None, "") else "")
        kc.border = _BORDER
        vc.border = _BORDER
        kc.font   = Font(bold=True)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40


# ── Phase 3.14: Import Validation Detail ──────────────────────────────────────

def _sheet_import_validation(wb: Workbook, import_validation: dict) -> None:
    ws = wb.create_sheet("Import_Validation")
    ws.sheet_view.rightToLeft = True
    _section_title(ws, 1, 3, "Import Validation — Warnings & Errors")
    warnings = import_validation.get("warnings") or []
    if not warnings:
        ws.cell(row=2, column=1, value="لا توجد تحذيرات أو أخطاء في عملية الاستيراد.").font = _FNT_NOTE
        return
    row = 2
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    for col, txt in enumerate(["الصنف", "الخطورة", "الرسالة"], 1):
        hc = ws.cell(row=row, column=col, value=txt)
        hc.fill = hdr_fill
        hc.font = Font(bold=True, color="FFFFFF")
        hc.border = _BORDER
    row += 1
    for w in warnings:
        if isinstance(w, str):
            cat, sev, msg = "general", "warning", w
        else:
            cat = w.get("category", w.get("type", "general"))
            sev = w.get("severity", "warning")
            msg = w.get("message", str(w))
        row_fill = PatternFill("solid", fgColor="FEE2E2" if sev == "error" else "FEF9C3")
        for col, val in enumerate([cat, sev, msg], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = row_fill
            c.border    = _BORDER
            c.alignment = Alignment(wrap_text=True)
        row += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A3"


# ── Phase 3.14: Readiness Summary ─────────────────────────────────────────────

def _sheet_readiness(wb: Workbook, import_validation: dict) -> None:
    ws = wb.create_sheet("Readiness")
    ws.sheet_view.rightToLeft = True
    _section_title(ws, 1, 2, "Import Readiness Summary")
    summary  = import_validation.get("summary") or {}
    matching = summary.get("matching_readiness") or {}
    norma    = summary.get("normalization_readiness") or {}

    def _kv(r: int, label: str, value: object, fill_hex: Optional[str] = None) -> None:
        kc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=str(value) if value is not None else "")
        kc.border = _BORDER
        vc.border = _BORDER
        kc.font   = Font(bold=True)
        if fill_hex:
            vc.fill = PatternFill("solid", fgColor=fill_hex)

    row = 2
    _section_title(ws, row, 2, "Matching Readiness"); row += 1
    mr_ok = bool(matching.get("matching_ready"))
    _kv(row, "Matching Ready",       "نعم ✓" if mr_ok else "لا ✗",
        "D1FAE5" if mr_ok else "FEE2E2"); row += 1
    _kv(row, "Properties Imported",  matching.get("properties_imported", "")); row += 1
    _kv(row, "Sales Imported",       matching.get("sales_imported", "")); row += 1
    _kv(row, "Matched (subject_id)", matching.get("matched_count", "")); row += 1
    _kv(row, "Unmatched Properties", matching.get("unmatched_properties", "")); row += 1
    _kv(row, "Unmatched Sales",      matching.get("unmatched_sales", "")); row += 1

    row += 1
    _section_title(ws, row, 2, "Normalization Readiness"); row += 1
    nr_ok = bool(norma.get("normalization_ok"))
    _kv(row, "Normalization OK",        "نعم ✓" if nr_ok else "تحذير",
        "D1FAE5" if nr_ok else "FEF9C3"); row += 1
    _kv(row, "Zone Variants Detected",  norma.get("zone_variants_detected", "")); row += 1
    _kv(row, "Class Variants Detected", norma.get("class_variants_detected", "")); row += 1
    _kv(row, "Coverage Gaps",           norma.get("coverage_gaps", "")); row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24


# ── Phase 3.14: Sales Verification ───────────────────────────────────────────

def _sheet_sales_verification(wb: Workbook, sales_verification: dict) -> None:
    ws = wb.create_sheet("Sales_Verification")
    ws.sheet_view.rightToLeft = True
    _section_title(ws, 1, 7, "Sales Verification")
    records = sales_verification.get("records") or sales_verification.get("results") or []
    if not records:
        ws.cell(row=2, column=1, value="لا توجد بيانات مبيعات للتحقق.").font = _FNT_NOTE
        return
    row = 2
    headers = ["sale_id", "verified", "arms_length", "usable_for_ratio_study",
               "market_value", "sale_price", "flags"]
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    for col, h in enumerate(headers, 1):
        hc = ws.cell(row=row, column=col, value=h)
        hc.fill = hdr_fill
        hc.font = Font(bold=True, color="FFFFFF")
        hc.border = _BORDER
    row += 1
    for rec in records:
        usable   = rec.get("usable_for_ratio_study", rec.get("usable", False))
        row_fill = PatternFill("solid", fgColor="D1FAE5" if usable else "FEE2E2")
        for col, key in enumerate(["sale_id", "verified", "arms_length", "usable_for_ratio_study",
                                    "market_value", "sale_price"], 1):
            c        = ws.cell(row=row, column=col, value=rec.get(key, ""))
            c.fill   = row_fill
            c.border = _BORDER
        fc           = ws.cell(row=row, column=7, value="; ".join(rec.get("flags", []) or []))
        fc.fill      = row_fill
        fc.border    = _BORDER
        fc.alignment = Alignment(wrap_text=True)
        row += 1
    for i, w in enumerate([14, 12, 14, 22, 18, 14, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


# ── Phase 3.14: Time Adjustment ───────────────────────────────────────────────

def _sheet_time_adjustment(wb: Workbook, time_adjustment: dict) -> None:
    ws = wb.create_sheet("Time_Adjustment")
    ws.sheet_view.rightToLeft = True
    _section_title(ws, 1, 6, "Sales Time Adjustment")
    summary  = time_adjustment.get("summary") or {}
    adjusted = time_adjustment.get("adjusted_sales") or time_adjustment.get("records") or []
    row = 2
    if summary:
        for label, val in [
            ("Status",           time_adjustment.get("status", "")),
            ("Monthly Rate (%)", summary.get("monthly_rate_pct", summary.get("monthly_rate", ""))),
            ("Reference Date",   summary.get("reference_date", "")),
            ("Adjusted Records", summary.get("adjusted_count", len(adjusted))),
        ]:
            kc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=2, value=str(val) if val is not None else "")
            kc.border = _BORDER
            vc.border = _BORDER
            kc.font   = Font(bold=True)
            row += 1
        row += 1
    if not adjusted:
        ws.cell(row=row, column=1, value="لا توجد بيانات تعديل زمني.").font = _FNT_NOTE
        return
    hdr_row = row
    headers = ["sale_id", "sale_date", "original_price", "adjusted_price",
               "months_elapsed", "adjustment_factor"]
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    for col, h in enumerate(headers, 1):
        hc = ws.cell(row=row, column=col, value=h)
        hc.fill = hdr_fill
        hc.font = Font(bold=True, color="FFFFFF")
        hc.border = _BORDER
    row += 1
    for rec in adjusted:
        for col, key in enumerate(headers, 1):
            c        = ws.cell(row=row, column=col, value=rec.get(key, ""))
            c.border = _BORDER
        row += 1
    for i, w in enumerate([14, 14, 18, 18, 16, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hdr_row + 1}"


# ── Phase 3.14: Sales Adjustments ────────────────────────────────────────────

def _sheet_sales_adjustments(wb: Workbook, sales_adjustments: dict) -> None:
    ws = wb.create_sheet("Sales_Adjustments")
    ws.sheet_view.rightToLeft = True
    _section_title(ws, 1, 6, "Sales Adjustments")
    records = sales_adjustments.get("adjusted_sales") or sales_adjustments.get("records") or []
    summary = sales_adjustments.get("summary") or {}
    row = 2
    if summary:
        for k, v in summary.items():
            kc = ws.cell(row=row, column=1, value=str(k))
            vc = ws.cell(row=row, column=2, value=str(v) if v is not None else "")
            kc.border = _BORDER
            vc.border = _BORDER
            kc.font   = Font(bold=True)
            row += 1
        row += 1
    if not records:
        ws.cell(row=row, column=1, value="لا توجد بيانات تعديلات مبيعات.").font = _FNT_NOTE
        return
    all_keys = list(records[0].keys())
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_row  = row
    for col, h in enumerate(all_keys, 1):
        hc = ws.cell(row=row, column=col, value=h)
        hc.fill = hdr_fill
        hc.font = Font(bold=True, color="FFFFFF")
        hc.border = _BORDER
    row += 1
    for rec in records:
        for col, key in enumerate(all_keys, 1):
            c        = ws.cell(row=row, column=col, value=rec.get(key, ""))
            c.border = _BORDER
        row += 1
    for i in range(len(all_keys)):
        ws.column_dimensions[get_column_letter(i + 1)].width = 18
    ws.freeze_panes = f"A{hdr_row + 1}"


# ── Public API ────────────────────────────────────────────────────────────────

def build_mass_appraisal_workbook(result: dict,
                                   reviewed_summary: Optional[dict] = None,
                                   audit: Optional[dict] = None,
                                   ratio_study: Optional[dict] = None,
                                   calibration_preview: Optional[dict] = None,
                                   calibration_sandbox: Optional[dict] = None,
                                   governance: Optional[dict] = None,
                                   model_cycle: Optional[dict] = None,
                                   import_validation: Optional[dict] = None,
                                   sales_verification: Optional[dict] = None,
                                   time_adjustment: Optional[dict] = None,
                                   sales_adjustments: Optional[dict] = None) -> bytes:
    """
    Build a professional XLSX workbook for the given Mass Appraisal Run result.
    Pass reviewed_summary (Phase 1.8) to include the Reviewed Portfolio section.
    Pass audit (Phase 1.9) to include Audit & Governance sheet and metadata.
    Pass ratio_study (Phase 3.3) to include the Ratio Study sheet.
    Pass calibration_preview (Phase 3.6) to include the Calibration Preview sheet.
    Pass calibration_sandbox (Phase 3.7) to include the Calibration Sandbox sheet.
    Pass governance (Phase 3.8) to include the Governance Approval sheet.
    Pass model_cycle (Phase 3.9) to include the Model & Revaluation Cycle sheet.
    No disk I/O — returns raw bytes only.
    """
    # Support both old schema (rows/summary) and new schema (units/n_units/total_portfolio_value)
    raw_units: List[dict] = result.get("units") or result.get("rows") or []
    rows_data = []
    for _i, _u in enumerate(raw_units):
        _r = dict(_u)
        if "row_id" not in _r:
            _r["row_id"] = _r.get("id") or f"UNIT-{_i+1:04d}"
        if "row_index" not in _r:
            _r["row_index"] = _i + 1
        if "market_value" not in _r:
            _r["market_value"] = _r.get("unit_value") or 0.0
        if "status" not in _r:
            _r["status"] = "success"
        if "price_per_meter_effective" not in _r:
            _r["price_per_meter_effective"] = _r.get("final_ppm") or _r.get("avm_ppm") or 0.0
        if "value_per_m2" not in _r:
            _r["value_per_m2"] = _r.get("final_ppm") or _r.get("avm_ppm") or 0.0
        rows_data.append(_r)

    summary: dict = result.get("summary") or {}
    if not summary:
        _port_s = result.get("portfolio_summary") or {}
        summary = {
            "total_rows":           result.get("n_units") or len(rows_data),
            "successful_rows":      result.get("n_units") or len(rows_data),
            "failed_rows":          0,
            "skipped_rows":         0,
            "total_market_value":   result.get("total_portfolio_value") or 0.0,
            "average_market_value": _port_s.get("median_val") or 0.0,
            "median_market_value":  _port_s.get("median_val") or 0.0,
            "avm_applied_count":    result.get("n_units") or len(rows_data),
            "purpose_counts":       {result.get("purpose", "fair_market"): result.get("n_units") or len(rows_data)},
        }

    _result = dict(result)
    _result["rows"]    = rows_data
    _result["summary"] = summary

    wb = Workbook()

    _sheet_executive_summary(wb, _result, reviewed_summary, audit)  # Phase 1.8+1.9
    _sheet_portfolio(wb, rows_data)
    _sheet_purpose_analysis(wb, rows_data)
    _sheet_location_analysis(wb, rows_data)
    _sheet_asset_type_analysis(wb, rows_data)
    _sheet_avm_analysis(wb, rows_data)
    _sheet_tax_assessment(wb, rows_data)
    _sheet_usufruct_uncertainty(wb, rows_data)
    _sheet_data_quality(wb, rows_data, summary)                    # Phase 1.5
    _sheet_review_queue(wb, rows_data, summary)                    # Phase 1.6
    _sheet_errors_skipped(wb, rows_data)
    _sheet_final_reviewed_portfolio(wb, rows_data)                 # Phase 1.8
    _sheet_final_exclusions(wb, rows_data)                         # Phase 1.8
    _sheet_audit_governance(wb, audit, summary, reviewed_summary)  # Phase 1.9
    _sheet_zone_analysis(wb, rows_data, summary)                    # Phase 3.1
    _sheet_property_class_analysis(wb, rows_data, summary)          # Phase 3.1
    if isinstance(ratio_study, dict) and ratio_study.get("status") == "success":
        _sheet_ratio_study(wb, ratio_study)                         # Phase 3.3
    if isinstance(calibration_preview, dict) and calibration_preview.get("status") == "success":
        _sheet_calibration_preview(wb, calibration_preview)         # Phase 3.6
    if isinstance(calibration_sandbox, dict) and calibration_sandbox.get("status") == "success":
        _sheet_calibration_sandbox(wb, calibration_sandbox)         # Phase 3.7
    if isinstance(governance, dict) and governance.get("governance_id"):
        _sheet_governance(wb, governance)                           # Phase 3.8
    if isinstance(model_cycle, dict) and model_cycle.get("cycle_id"):
        _sheet_model_cycle(wb, model_cycle)                        # Phase 3.9
    _sheet_export_metadata(wb, _result, import_validation)          # Phase 3.14
    if isinstance(import_validation, dict):
        _sheet_import_validation(wb, import_validation)             # Phase 3.14
        _sheet_readiness(wb, import_validation)                     # Phase 3.14
    if isinstance(sales_verification, dict) and sales_verification.get("status") == "success":
        _sheet_sales_verification(wb, sales_verification)           # Phase 3.14
    if isinstance(time_adjustment, dict):
        _sheet_time_adjustment(wb, time_adjustment)                 # Phase 3.14
    if isinstance(sales_adjustments, dict):
        _sheet_sales_adjustments(wb, sales_adjustments)             # Phase 3.14
    _sheet_assumptions(wb, import_validation=import_validation)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
