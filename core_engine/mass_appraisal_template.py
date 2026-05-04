"""
core_engine/mass_appraisal_template.py
=======================================
Phase 3.12 Prompt 1 — Excel Template Builder

Exposes:
    build_mass_appraisal_template_workbook() -> bytes

Rules:
- No disk I/O. Returns raw bytes only.
- No valuation engine imports.
- No Mass Appraisal logic.
- No pandas dependency.
- Uses openpyxl only (already in requirements.txt).
"""
from __future__ import annotations

from io import BytesIO
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
_FILL_HDR      = PatternFill("solid", fgColor="1F2937")   # dark blue-grey header
_FILL_REQ      = PatternFill("solid", fgColor="312E81")   # indigo — required col
_FILL_INSTR    = PatternFill("solid", fgColor="0F172A")   # instructions bg
_FILL_SECTION  = PatternFill("solid", fgColor="374151")   # section sub-header
_FILL_LISTS    = PatternFill("solid", fgColor="1E293B")   # lists sheet header
_FILL_SAMPLE   = PatternFill("solid", fgColor="1A2035")   # sample data row tint

_FONT_HDR   = Font(name="Calibri", bold=True, color="F8FAFC", size=11)
_FONT_REQ   = Font(name="Calibri", bold=True, color="A5B4FC", size=11)
_FONT_TITLE = Font(name="Calibri", bold=True, color="C7D2FE", size=13)
_FONT_BODY  = Font(name="Calibri", color="CBD5E1", size=10)
_FONT_MUTED = Font(name="Calibri", color="6B7280", size=9)
_FONT_WARN  = Font(name="Calibri", bold=True, color="FCD34D", size=10)

_ALIGN_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_RTL  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

_THIN = Side(style="thin", color="374151")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _w(ws, row: int, col: int, value, font=None, fill=None,
       alignment=None, border=None, number_format=None) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:          cell.font          = font
    if fill:          cell.fill          = fill
    if alignment:     cell.alignment     = alignment
    if border:        cell.border        = border
    if number_format: cell.number_format = number_format


def _header_row(ws, row: int, columns: List[str],
                required: set, start_col: int = 1) -> None:
    for ci, col_name in enumerate(columns, start=start_col):
        is_req = col_name in required
        _w(ws, row, ci, col_name,
           font=_FONT_REQ if is_req else _FONT_HDR,
           fill=_FILL_REQ if is_req else _FILL_HDR,
           alignment=_ALIGN_CTR,
           border=_BORDER)


def _data_row(ws, row: int, values: List, start_col: int = 1) -> None:
    fill = _FILL_SAMPLE if row % 2 == 0 else None
    for ci, val in enumerate(values, start=start_col):
        _w(ws, row, ci, val,
           font=_FONT_BODY,
           fill=fill,
           alignment=_ALIGN_LFT,
           border=_BORDER)


def _col_widths(ws, widths: List[int], start_col: int = 1) -> None:
    for ci, w in enumerate(widths, start=start_col):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _freeze(ws, cell: str = "A2") -> None:
    ws.freeze_panes = cell


def _set_rtl(ws) -> None:
    ws.sheet_view.rightToLeft = True


# ── Sheet 1: Instructions ─────────────────────────────────────────────────────

def _sheet_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.rightToLeft = False

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72

    # Title
    ws.merge_cells("A1:B1")
    _w(ws, 1, 1,
       "Mass Appraisal Excel Template — Expert Smart Platform",
       font=_FONT_TITLE,
       fill=_FILL_INSTR,
       alignment=_ALIGN_CTR)
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.merge_cells("A2:B2")
    _w(ws, 2, 1,
       "قالب Excel للتقييم الجماعي — منصة Expert Smart",
       font=Font(name="Calibri", bold=True, color="818CF8", size=12),
       fill=_FILL_INSTR,
       alignment=_ALIGN_CTR)
    ws.row_dimensions[2].height = 22

    instructions: List[Tuple[str, str]] = [
        ("", ""),
        ("Sheet",         "Purpose / الغرض"),
        ("Properties",    "أدخل وحدات العقارات المراد تقييمها — Enter the property units for Mass Appraisal."),
        ("Sales",         "أدخل بيانات صفقات البيع — Enter sale transaction evidence."),
        ("Assumptions",   "معاملات التقييم الاختيارية — Optional valuation parameters and overrides."),
        ("Governance",    "بيانات الحوكمة والمراجع — Reviewer / approval metadata."),
        ("Model_Cycle",   "بيانات دورة النموذج — Model version and revaluation cycle info."),
        ("Lists",         "قوائم القيم المقبولة — Lookup lists for reference only."),
        ("", ""),
        ("Rules / قواعد", ""),
        ("Required",      "الحقول المطلوبة موضحة بلون indigo (أزرق بنفسجي) في صف العناوين."),
        ("Required",      "Required fields are highlighted in indigo in the header row."),
        ("Dates",         "استخدم صيغة YYYY-MM-DD للتواريخ — Use YYYY-MM-DD format for all dates."),
        ("Booleans",      "استخدم TRUE أو FALSE (بالأحرف الكبيرة) — Use TRUE or FALSE for boolean fields."),
        ("Sheet names",   "لا تغيّر أسماء الأوراق — Do NOT rename any sheet."),
        ("Column names",  "لا تغيّر أسماء الأعمدة — Do NOT rename column headers."),
        ("Empty rows",    "لا تترك صفوفاً فارغة في منتصف البيانات."),
        ("Empty rows",    "Do not leave blank rows in the middle of your data."),
        ("", ""),
        ("Next steps",    ""),
        ("Step 1",        "Fill the Properties sheet with your portfolio units."),
        ("Step 2",        "Fill the Sales sheet with sales evidence (optional but recommended for Ratio Study)."),
        ("Step 3",        "Review the Assumptions sheet and update values if needed."),
        ("Step 4",        "Fill Governance and Model_Cycle if required by your workflow."),
        ("Step 5",        "Upload this file using the Import Excel button in the Mass Appraisal workspace."),
        ("Step 6",        "Preview parsed data, then press Run to execute the appraisal."),
        ("", ""),
        ("Note",          "يحتوي هذا القالب على بيانات تجريبية للتوضيح فقط."),
        ("Note",          "Delete sample rows before uploading your real data."),
        ("Note",          "This template is generated by Phase 3.12 of the Expert Smart platform."),
    ]

    for ri, (label, text) in enumerate(instructions, start=3):
        ws.row_dimensions[ri].height = 18
        if label in ("Sheet", "Rules / قواعد", "Next steps"):
            _w(ws, ri, 1, label,
               font=_FONT_WARN,
               fill=_FILL_SECTION,
               alignment=_ALIGN_LFT)
            _w(ws, ri, 2, text,
               font=_FONT_WARN,
               fill=_FILL_SECTION,
               alignment=_ALIGN_LFT)
        elif label == "":
            ws.cell(row=ri, column=1).fill = _FILL_INSTR
            ws.cell(row=ri, column=2).fill = _FILL_INSTR
        else:
            _w(ws, ri, 1, label,
               font=_FONT_MUTED,
               fill=_FILL_INSTR,
               alignment=_ALIGN_LFT)
            _w(ws, ri, 2, text,
               font=_FONT_BODY,
               fill=_FILL_INSTR,
               alignment=_ALIGN_LFT)


# ── Sheet 2: Properties ───────────────────────────────────────────────────────

def _sheet_properties(wb: Workbook) -> None:
    ws = wb.create_sheet("Properties")
    _set_rtl(ws)

    columns  = ["row_id", "location", "zone_id", "property_class",
                "property_type", "area", "floor", "year_built",
                "condition", "valuation_purpose", "notes"]
    required = {"row_id", "location", "property_type", "area"}

    _header_row(ws, 1, columns, required)
    _freeze(ws)

    samples = [
        ["U-001", "التجمع الخامس", "Z-NC-01", "A",
         "شقة سكنية", 150, 3, 2020, "excellent",  "fair_market_value", "sample"],
        ["U-002", "مدينة نصر",    "Z-NS-02", "B",
         "شقة سكنية", 120, 8, 2012, "good",       "fair_market_value", "sample"],
        ["U-003", "المعادي",      "Z-MA-03", "B",
         "محل تجاري", 90,  0, 2015, "very_good",  "fair_market_value", "sample"],
    ]
    for ri, row in enumerate(samples, start=2):
        _data_row(ws, ri, row)

    _col_widths(ws, [12, 22, 14, 14, 18, 10, 8, 12, 14, 22, 24])


# ── Sheet 3: Sales ────────────────────────────────────────────────────────────

def _sheet_sales(wb: Workbook) -> None:
    ws = wb.create_sheet("Sales")
    _set_rtl(ws)

    columns  = ["sale_id", "subject_id", "location", "zone_id",
                "property_class", "property_type", "area", "sale_price",
                "sale_date", "source", "verified", "arms_length",
                "buyer_seller_related", "usable_for_ratio_study", "notes"]
    required = {"sale_id", "location", "property_type", "area",
                "sale_price", "sale_date"}

    _header_row(ws, 1, columns, required)
    _freeze(ws)

    samples = [
        ["S-001", "U-001", "التجمع الخامس", "Z-NC-01", "A",
         "شقة سكنية",  150, 7900000,  "2026-01-15", "broker_verified",
         True,  True,  False, True,  "sample"],
        ["S-002", "U-002", "مدينة نصر",    "Z-NS-02", "B",
         "شقة سكنية",  120, 4200000,  "2025-12-20", "registered_sale",
         True,  True,  False, True,  "sample"],
        ["S-003", "U-003", "المعادي",      "Z-MA-03", "B",
         "محل تجاري",  90,  8250000,  "2026-03-01", "registered_sale",
         True,  True,  False, True,  "sample"],
    ]
    for ri, row in enumerate(samples, start=2):
        _data_row(ws, ri, row)

    _col_widths(ws, [10, 10, 22, 14, 14, 18, 10, 14,
                     14, 18, 10, 12, 18, 20, 20])


# ── Sheet 4: Assumptions ──────────────────────────────────────────────────────

def _sheet_assumptions(wb: Workbook) -> None:
    ws = wb.create_sheet("Assumptions")

    columns  = ["key", "value", "notes"]
    required = {"key", "value"}

    _header_row(ws, 1, columns, required)
    _freeze(ws)

    rows = [
        ["valuation_date",           "2026-05-03", "تاريخ التقييم"],
        ["monthly_growth_rate",      0.0075,       "معدل نمو شهري للتعديل الزمني"],
        ["base_market_ppm",          50000,        "سعر متر أساسي اختياري"],
        ["target_ratio",             1.0,          "هدف دراسة النسب (IAAO)"],
        ["min_factor",               0.85,         "أقل عامل معايرة مسموح به"],
        ["max_factor",               1.15,         "أعلى عامل معايرة مسموح به"],
        ["manual_calibration_factor", 1.05,        "عامل يدوي لاختبار Sandbox (اختياري)"],
    ]
    for ri, row in enumerate(rows, start=2):
        _data_row(ws, ri, row)

    _col_widths(ws, [28, 18, 42])


# ── Sheet 5: Governance ───────────────────────────────────────────────────────

def _sheet_governance(wb: Workbook) -> None:
    ws = wb.create_sheet("Governance")

    columns  = ["key", "value", "notes"]
    required = {"key"}

    _header_row(ws, 1, columns, required)
    _freeze(ws)

    rows = [
        ["governance_status", "pending",           "pending / approved_for_review / rejected"],
        ["decision",          "pending",           "قرار المراجعة"],
        ["approved_by",       "م. هشام المهدي",    "اسم المراجع"],
        ["approval_date",     "2026-05-03",        "تاريخ القرار (YYYY-MM-DD)"],
        ["review_level",      "internal_review",   "مستوى المراجعة"],
        ["notes",             "تم إنشاء القالب للاختبار", "ملاحظات الحوكمة"],
    ]
    for ri, row in enumerate(rows, start=2):
        _data_row(ws, ri, row)

    _col_widths(ws, [24, 30, 42])


# ── Sheet 6: Model_Cycle ──────────────────────────────────────────────────────

def _sheet_model_cycle(wb: Workbook) -> None:
    ws = wb.create_sheet("Model_Cycle")

    columns  = ["key", "value", "notes"]
    required = {"key"}

    _header_row(ws, 1, columns, required)
    _freeze(ws)

    rows = [
        ["model_version",    "MA-2026-Q2-v1",
         "إصدار النموذج"],
        ["cycle_id",         "Mass Appraisal Test Cycle - May 2026",
         "معرف الدورة"],
        ["cycle_start_date", "2026-04-30",
         "بداية دورة التقييم (YYYY-MM-DD)"],
        ["cycle_end_date",   "2026-05-03",
         "نهاية دورة التقييم (YYYY-MM-DD)"],
        ["model_status",     "draft",
         "draft / active / archived"],
        ["notes",            "دورة اختبار قبل الاستيراد الكامل",
         "ملاحظات"],
    ]
    for ri, row in enumerate(rows, start=2):
        _data_row(ws, ri, row)

    _col_widths(ws, [22, 42, 36])


# ── Sheet 7: Lists ────────────────────────────────────────────────────────────

def _sheet_lists(wb: Workbook) -> None:
    ws = wb.create_sheet("Lists")

    _w(ws, 1, 1, "valuation_purpose",
       font=_FONT_HDR, fill=_FILL_LISTS, alignment=_ALIGN_CTR, border=_BORDER)
    _w(ws, 1, 2, "condition",
       font=_FONT_HDR, fill=_FILL_LISTS, alignment=_ALIGN_CTR, border=_BORDER)
    _w(ws, 1, 3, "property_class",
       font=_FONT_HDR, fill=_FILL_LISTS, alignment=_ALIGN_CTR, border=_BORDER)
    _w(ws, 1, 4, "boolean_values",
       font=_FONT_HDR, fill=_FILL_LISTS, alignment=_ALIGN_CTR, border=_BORDER)
    _freeze(ws)

    purposes = [
        "fair_market_value", "bank_financing", "insurance",
        "financial_reporting", "tax_assessment", "rental_arbitration",
        "acquisition", "judicial_liquidation", "usufruct",
        "uncertainty_valuation",
    ]
    conditions    = ["excellent", "very_good", "good", "average", "poor"]
    prop_classes  = ["A", "B", "C"]
    booleans      = ["TRUE", "FALSE"]

    max_len = max(len(purposes), len(conditions), len(prop_classes), len(booleans))
    for ri in range(max_len):
        row_n = ri + 2
        _w(ws, row_n, 1,
           purposes[ri]    if ri < len(purposes)     else "",
           font=_FONT_BODY, alignment=_ALIGN_LFT, border=_BORDER)
        _w(ws, row_n, 2,
           conditions[ri]  if ri < len(conditions)   else "",
           font=_FONT_BODY, alignment=_ALIGN_LFT, border=_BORDER)
        _w(ws, row_n, 3,
           prop_classes[ri] if ri < len(prop_classes) else "",
           font=_FONT_BODY, alignment=_ALIGN_CTR, border=_BORDER)
        _w(ws, row_n, 4,
           booleans[ri]    if ri < len(booleans)     else "",
           font=_FONT_BODY, alignment=_ALIGN_CTR, border=_BORDER)

    _col_widths(ws, [26, 16, 16, 16])


# ── Public API ────────────────────────────────────────────────────────────────

def build_mass_appraisal_template_workbook() -> bytes:
    """
    Build a Mass Appraisal import template workbook in memory.
    Returns raw bytes (xlsx). No disk I/O.
    """
    wb = Workbook()

    _sheet_instructions(wb)   # uses wb.active
    _sheet_properties(wb)
    _sheet_sales(wb)
    _sheet_assumptions(wb)
    _sheet_governance(wb)
    _sheet_model_cycle(wb)
    _sheet_lists(wb)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Parser helpers ────────────────────────────────────────────────────────────

def _cell_str(cell) -> str:
    """Return stripped string from cell value, or empty string."""
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def _cell_float(cell) -> Optional[float]:
    """Return float from cell value, or None if missing/invalid."""
    v = cell.value
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _cell_bool(cell) -> Optional[bool]:
    """Parse boolean cell: TRUE/FALSE, true/false, yes/no, نعم/لا, 1/0."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "yes", "نعم", "1"):
        return True
    if s in ("false", "no", "لا", "0"):
        return False
    return None


def _cell_date_str(cell) -> Optional[str]:
    """Return ISO YYYY-MM-DD string from a date cell or date-like string."""
    import datetime as _dt
    v = cell.value
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return _dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # return as-is if unparseable, validation will catch it


def _parse_sheet_headers(ws) -> Tuple[List[str], int]:
    """
    Find the first non-empty row in ws and return (header_list, row_index).
    Returns ([], 0) if worksheet is empty.
    """
    for ri, row in enumerate(ws.iter_rows(), start=1):
        headers = [_cell_str(c) for c in row]
        if any(h for h in headers):
            return headers, ri
    return [], 0


def _is_row_empty(row) -> bool:
    return all(c.value is None or str(c.value).strip() == "" for c in row)


def _normalize_kv_value(cell):
    """Normalize key-value sheet value cells."""
    import datetime as _dt
    v = cell.value
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s.lower() == "true":
        return True
    if s.lower() == "false":
        return False
    return s


# ── Sheet parsers ─────────────────────────────────────────────────────────────

_PROP_REQUIRED = {"row_id", "location", "property_type", "area"}
_SALE_REQUIRED = {"sale_id", "location", "property_type", "area",
                  "sale_price", "sale_date"}


def _parse_properties(ws, errors: list, warnings: list) -> list:
    headers, hdr_row = _parse_sheet_headers(ws)
    if not headers:
        warnings.append({"sheet": "Properties", "row": None, "field": None,
                         "message": "Properties sheet appears empty."})
        return []

    col_map = {h: i for i, h in enumerate(headers) if h}
    rows_out = []

    for row in ws.iter_rows(min_row=hdr_row + 1):
        if _is_row_empty(row):
            continue
        ri = row[0].row

        def _get(field: str):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        row_id       = _cell_str(_get("row_id"))   if _get("row_id")   else ""
        location     = _cell_str(_get("location")) if _get("location") else ""
        prop_type    = _cell_str(_get("property_type")) if _get("property_type") else ""
        area_cell    = _get("area")
        area         = _cell_float(area_cell) if area_cell is not None else None

        # Required field validation
        row_errors = []
        if not row_id:
            row_errors.append({"sheet": "Properties", "row": ri, "field": "row_id",
                                "message": f"Row {ri}: row_id is required."})
        if not location:
            row_errors.append({"sheet": "Properties", "row": ri, "field": "location",
                                "message": f"Row {ri}: location is required."})
        if not prop_type:
            row_errors.append({"sheet": "Properties", "row": ri, "field": "property_type",
                                "message": f"Row {ri}: property_type is required."})
        if area is None or area <= 0:
            row_errors.append({"sheet": "Properties", "row": ri, "field": "area",
                                "message": f"Row {ri}: area must be a positive number."})

        errors.extend(row_errors)
        if row_errors:
            continue  # skip invalid row but keep collecting errors

        # Optional numeric fields
        def _opt_float(field: str) -> Optional[float]:
            c = _get(field)
            return _cell_float(c) if c is not None else None

        rows_out.append({
            "id":               row_id,
            "row_id":           row_id,
            "location":         location,
            "zone_id":          _cell_str(_get("zone_id"))          if _get("zone_id")          else None,
            "property_class":   _cell_str(_get("property_class"))   if _get("property_class")   else None,
            "property_type":    prop_type,
            "area":             area,
            "floor":            _opt_float("floor"),
            "year_built":       _opt_float("year_built"),
            "condition":        _cell_str(_get("condition"))        if _get("condition")        else None,
            "valuation_purpose": _cell_str(_get("valuation_purpose")) if _get("valuation_purpose") else None,
            "notes":            _cell_str(_get("notes"))            if _get("notes")            else None,
        })

    return rows_out


def _parse_sales(ws, errors: list, warnings: list) -> list:
    headers, hdr_row = _parse_sheet_headers(ws)
    if not headers:
        warnings.append({"sheet": "Sales", "row": None, "field": None,
                         "message": "Sales sheet appears empty."})
        return []

    col_map = {h: i for i, h in enumerate(headers) if h}
    rows_out = []

    for row in ws.iter_rows(min_row=hdr_row + 1):
        if _is_row_empty(row):
            continue
        ri = row[0].row

        def _get(field: str):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        sale_id   = _cell_str(_get("sale_id"))   if _get("sale_id")   else ""
        location  = _cell_str(_get("location"))  if _get("location")  else ""
        prop_type = _cell_str(_get("property_type")) if _get("property_type") else ""
        area_cell = _get("area")
        area      = _cell_float(area_cell) if area_cell is not None else None
        price_cell = _get("sale_price")
        sale_price = _cell_float(price_cell) if price_cell is not None else None
        date_cell  = _get("sale_date")
        sale_date  = _cell_date_str(date_cell) if date_cell is not None else None

        row_errors = []
        if not sale_id:
            row_errors.append({"sheet": "Sales", "row": ri, "field": "sale_id",
                                "message": f"Row {ri}: sale_id is required."})
        if not location:
            row_errors.append({"sheet": "Sales", "row": ri, "field": "location",
                                "message": f"Row {ri}: location is required."})
        if not prop_type:
            row_errors.append({"sheet": "Sales", "row": ri, "field": "property_type",
                                "message": f"Row {ri}: property_type is required."})
        if area is None or area <= 0:
            row_errors.append({"sheet": "Sales", "row": ri, "field": "area",
                                "message": f"Row {ri}: area must be a positive number."})
        if sale_price is None or sale_price <= 0:
            row_errors.append({"sheet": "Sales", "row": ri, "field": "sale_price",
                                "message": f"Row {ri}: sale_price must be a positive number."})
        if not sale_date:
            row_errors.append({"sheet": "Sales", "row": ri, "field": "sale_date",
                                "message": f"Row {ri}: sale_date is required."})

        errors.extend(row_errors)
        if row_errors:
            continue

        def _opt_float(field: str) -> Optional[float]:
            c = _get(field)
            return _cell_float(c) if c is not None else None

        def _opt_bool(field: str) -> Optional[bool]:
            c = _get(field)
            return _cell_bool(c) if c is not None else None

        usable = _opt_bool("usable_for_ratio_study")
        usability = "usable" if usable is True else ("excluded" if usable is False else "unknown")

        rows_out.append({
            "sale_id":               sale_id,
            "subject_id":            _cell_str(_get("subject_id"))      if _get("subject_id")     else None,
            "location":              location,
            "zone_id":               _cell_str(_get("zone_id"))         if _get("zone_id")        else None,
            "property_class":        _cell_str(_get("property_class"))  if _get("property_class") else None,
            "property_type":         prop_type,
            "area":                  area,
            "sale_price":            sale_price,
            "sale_date":             sale_date,
            "source":                _cell_str(_get("source"))          if _get("source")         else None,
            "verified":              _opt_bool("verified"),
            "arms_length":           _opt_bool("arms_length"),
            "buyer_seller_related":  _opt_bool("buyer_seller_related"),
            "usable_for_ratio_study": usable,
            "usability_status":      usability,
            "notes":                 _cell_str(_get("notes"))           if _get("notes")          else None,
        })

    return rows_out


def _parse_kv_sheet(ws, sheet_name: str, errors: list, warnings: list) -> dict:
    """Parse key-value sheets (Assumptions, Governance, Model_Cycle)."""
    headers, hdr_row = _parse_sheet_headers(ws)
    if not headers:
        warnings.append({"sheet": sheet_name, "row": None, "field": None,
                         "message": f"{sheet_name} sheet appears empty."})
        return {}

    col_map = {h: i for i, h in enumerate(headers) if h}
    key_idx = col_map.get("key")
    val_idx = col_map.get("value")

    if key_idx is None:
        warnings.append({"sheet": sheet_name, "row": None, "field": "key",
                         "message": f"{sheet_name}: 'key' column not found."})
        return {}

    result: dict = {}
    for row in ws.iter_rows(min_row=hdr_row + 1):
        if _is_row_empty(row):
            continue
        key = _cell_str(row[key_idx]) if key_idx < len(row) else ""
        if not key:
            continue
        val = _normalize_kv_value(row[val_idx]) if (val_idx is not None and val_idx < len(row)) else None
        if key in result:
            warnings.append({"sheet": sheet_name, "row": row[0].row, "field": key,
                             "message": f"{sheet_name}: duplicate key '{key}' — using latest value."})
        result[key] = val

    return result


# ── Public parse API ──────────────────────────────────────────────────────────

def parse_mass_appraisal_template_workbook(file_bytes: bytes) -> dict:
    """
    Parse a Mass Appraisal import template workbook.

    Parameters
    ----------
    file_bytes : bytes
        Raw bytes of an .xlsx workbook.

    Returns
    -------
    dict with keys:
        status         "success" | "validation_error"
        summary        counts dict
        data           parsed data dict
        validation     {"errors": [...], "warnings": [...]}
    """
    from io import BytesIO as _BytesIO
    from openpyxl import load_workbook as _load_wb

    try:
        wb = _load_wb(_BytesIO(file_bytes), read_only=False, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read workbook: {exc}") from exc

    errors: list = []
    warnings: list = []

    # ── Check required sheets ─────────────────────────────────────────────────
    sheet_names = wb.sheetnames
    required_sheets = ["Properties", "Sales", "Assumptions", "Governance", "Model_Cycle"]
    for sname in required_sheets:
        if sname not in sheet_names:
            errors.append({"sheet": sname, "row": None, "field": None,
                           "message": f"Missing required sheet: {sname}"})

    # ── Parse sheets (even if some are missing, parse what exists) ────────────
    properties = _parse_properties(wb["Properties"], errors, warnings) \
        if "Properties" in sheet_names else []

    sales = _parse_sales(wb["Sales"], errors, warnings) \
        if "Sales" in sheet_names else []

    assumptions = _parse_kv_sheet(wb["Assumptions"], "Assumptions", errors, warnings) \
        if "Assumptions" in sheet_names else {}

    governance = _parse_kv_sheet(wb["Governance"], "Governance", errors, warnings) \
        if "Governance" in sheet_names else {}

    model_cycle = _parse_kv_sheet(wb["Model_Cycle"], "Model_Cycle", errors, warnings) \
        if "Model_Cycle" in sheet_names else {}

    # ── Build response ────────────────────────────────────────────────────────
    status = "validation_error" if errors else "success"

    return {
        "status": status,
        "summary": {
            "properties_count":  len(properties),
            "sales_count":       len(sales),
            "assumptions_count": len(assumptions),
            "governance_count":  len(governance),
            "model_cycle_count": len(model_cycle),
            "warnings_count":    len(warnings),
            "errors_count":      len(errors),
        },
        "data": {
            "properties":  properties,
            "sales":       sales,
            "assumptions": assumptions,
            "governance":  governance,
            "model_cycle": model_cycle,
        },
        "validation": {
            "errors":   errors,
            "warnings": warnings,
        },
    }


__all__ = ["build_mass_appraisal_template_workbook",
           "parse_mass_appraisal_template_workbook"]
