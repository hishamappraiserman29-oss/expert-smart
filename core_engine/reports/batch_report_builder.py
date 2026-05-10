"""
batch_report_builder.py — Batch Valuation Excel Report (Phase 12.2)

Generates a 3-sheet workbook from a BatchProcessor completion report dict.
Standalone builder — does not modify ExcelReportBuilder.

Sheets:
    1. Batch Summary          — headline metrics block
    2. Completed Properties   — one data row per successful property
    3. Failed & Skipped       — error messages and skip reasons

Color scheme: amber 843C0C / light FCE4D6 — distinct from all Phase 9-11 fills.
"""
from __future__ import annotations

from datetime import datetime
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Style constants ────────────────────────────────────────────────────────────

_FILL_BATCH     = PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid")
_FILL_BATCH_COL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_FILL_FAIL      = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
_FILL_WARN      = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

_FONT_TITLE   = Font(bold=True, color="FFFFFF", size=14)
_FONT_SECTION = Font(bold=True, color="FFFFFF", size=12)
_FONT_BOLD    = Font(bold=True)
_FONT_MUTED   = Font(italic=True, color="888888")
_FONT_ERR     = Font(bold=True, color="FFFFFF")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_WRAP   = Alignment(wrap_text=True, vertical="top")
_ALIGN_RIGHT  = Alignment(horizontal="right")

_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

_FMT_CURRENCY = "#,##0"
_FMT_PCT      = "0.00%"


class BatchReportBuilder:
    """Generate a 3-sheet Excel workbook from a batch completion report."""

    def __init__(self, completion_report: Dict) -> None:
        self.report      = completion_report
        self.workbook    = Workbook()
        self.workbook.remove(self.workbook.active)
        self.report_date = datetime.now().strftime("%Y-%m-%d")

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _section_row(self, ws, row: int, title: str, ncols: int) -> None:
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c           = ws.cell(row=row, column=1)
        c.value     = title
        c.fill      = _FILL_BATCH
        c.font      = _FONT_SECTION
        c.alignment = _ALIGN_WRAP

    def _kv(self, ws, row: int, label: str, value) -> None:
        lc           = ws.cell(row=row, column=1)
        lc.value     = label
        lc.font      = _FONT_BOLD
        lc.alignment = _ALIGN_WRAP
        vc           = ws.cell(row=row, column=2)
        vc.value     = value if value is not None else "—"
        vc.alignment = _ALIGN_WRAP

    def _col_header_row(self, ws, row: int, headers: list) -> None:
        for col_i, hdr in enumerate(headers, 1):
            c           = ws.cell(row=row, column=col_i)
            c.value     = hdr
            c.font      = _FONT_BOLD
            c.fill      = _FILL_BATCH_COL
            c.alignment = _ALIGN_WRAP
            c.border    = _BORDER

    def _data_row(self, ws, row: int, values: list) -> None:
        for col_i, v in enumerate(values, 1):
            c           = ws.cell(row=row, column=col_i)
            c.value     = v
            c.alignment = _ALIGN_WRAP
            c.border    = _BORDER

    # ── Sheet 1: Batch Summary ────────────────────────────────────────────────

    def sheet_batch_summary(self) -> None:
        ws = self.workbook.create_sheet("Batch Summary", 0)
        for col, w in zip("ABCD", (28, 24, 16, 20)):
            ws.column_dimensions[col].width = w

        summary = self.report.get("summary", {})
        total   = summary.get("total_submitted", 0)
        done    = summary.get("completed", 0)
        failed  = summary.get("failed", 0)
        skipped = summary.get("skipped", 0)
        success_rate = (done / (total - skipped)) if (total - skipped) > 0 else 0.0

        # Title
        ws.merge_cells("A1:D1")
        tc           = ws["A1"]
        tc.value     = "BATCH VALUATION REPORT"
        tc.fill      = _FILL_BATCH
        tc.font      = _FONT_TITLE
        tc.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:D2")
        ws["A2"].value     = f"Report Date: {self.report_date}"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = _ALIGN_RIGHT

        r = 4
        self._section_row(ws, r, "BATCH OVERVIEW", 4);               r += 1
        self._kv(ws, r, "Batch ID",   self.report.get("batch_id", "—")); r += 1
        self._kv(ws, r, "Status",     self.report.get("status", "—"));    r += 1
        self._kv(ws, r, "Completed At",
                 self.report.get("completed_at", "—"));                r += 2

        self._section_row(ws, r, "PROCESSING RESULTS", 4);           r += 1
        self._kv(ws, r, "Total Submitted",   total);                  r += 1
        self._kv(ws, r, "Completed",         done);                   r += 1
        self._kv(ws, r, "Failed",            failed);                 r += 1
        self._kv(ws, r, "Skipped (Invalid)", skipped);                r += 1
        self._kv(ws, r, "Success Rate",
                 f"{success_rate * 100:.1f}%");                        r += 2

        self._section_row(ws, r, "VALUATION SUMMARY", 4);            r += 1
        tv = summary.get("total_valuation_value", 0)
        av = summary.get("average_valuation",     0)
        self._kv(ws, r, "Total Valuation (EGP)",
                 f"EGP {tv:,.0f}" if tv else "—");                    r += 1
        self._kv(ws, r, "Average Valuation (EGP)",
                 f"EGP {av:,.0f}" if av else "—");                    r += 1

        # Min / max from completed properties
        completed = self.report.get("completed_properties", [])
        if completed:
            vals = [p.get("valuation_value") or 0 for p in completed]
            self._kv(ws, r, "Min Valuation (EGP)",
                     f"EGP {min(vals):,.0f}");                         r += 1
            self._kv(ws, r, "Max Valuation (EGP)",
                     f"EGP {max(vals):,.0f}");                         r += 1

    # ── Sheet 2: Completed Properties ────────────────────────────────────────

    def sheet_completed_properties(self) -> None:
        ws = self.workbook.create_sheet("Completed Properties", 1)
        for col, w in zip("ABCDEFG", (14, 28, 16, 10, 20, 18, 22)):
            ws.column_dimensions[col].width = w

        # Title
        ws.merge_cells("A1:G1")
        tc           = ws["A1"]
        tc.value     = "COMPLETED PROPERTIES"
        tc.fill      = _FILL_BATCH
        tc.font      = _FONT_TITLE
        tc.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 24

        r = 3
        headers = [
            "Property ID", "Property Name", "Type",
            "Area (sqm)", "Valuation Value (EGP)",
            "Purpose", "Processed At",
        ]
        self._col_header_row(ws, r, headers); r += 1

        completed = self.report.get("completed_properties", [])
        if completed:
            for prop in completed:
                self._data_row(ws, r, [
                    prop.get("property_id",     ""),
                    prop.get("property_name",   ""),
                    prop.get("property_type",   ""),
                    prop.get("area_sqm",        ""),
                    f"EGP {prop.get('valuation_value') or 0:,.0f}",
                    prop.get("primary_purpose", "market_value"),
                    prop.get("processed_at",    ""),
                ])
                r += 1
        else:
            ws.merge_cells(f"A{r}:G{r}")
            c           = ws.cell(row=r, column=1)
            c.value     = "No completed properties"
            c.font      = _FONT_MUTED
            c.alignment = _ALIGN_CENTER

    # ── Sheet 3: Failed & Skipped ─────────────────────────────────────────────

    def sheet_failed_skipped(self) -> None:
        ws = self.workbook.create_sheet("Failed & Skipped", 2)
        for col, w in zip("ABCD", (14, 28, 16, 50)):
            ws.column_dimensions[col].width = w

        ws.merge_cells("A1:D1")
        tc           = ws["A1"]
        tc.value     = "FAILED & SKIPPED PROPERTIES"
        tc.fill      = _FILL_BATCH
        tc.font      = _FONT_TITLE
        tc.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 24

        r = 3

        # Failed section
        failed = self.report.get("failed_properties", [])
        self._section_row(ws, r, f"FAILED ({len(failed)})", 4); r += 1
        if failed:
            self._col_header_row(ws, r, ["ID", "Name", "Type", "Error Message"]); r += 1
            for p in failed:
                self._data_row(ws, r, [
                    p.get("id",    ""),
                    p.get("name",  ""),
                    "",
                    p.get("error", ""),
                ])
                ws.cell(row=r, column=4).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
                r += 1
        else:
            ws.merge_cells(f"A{r}:D{r}")
            c           = ws.cell(row=r, column=1)
            c.value     = "No failed properties"
            c.font      = _FONT_MUTED
            c.alignment = _ALIGN_CENTER
            r += 1

        r += 1

        # Skipped section
        skipped = self.report.get("skipped_properties", [])
        self._section_row(ws, r, f"SKIPPED / INVALID ({len(skipped)})", 4); r += 1
        if skipped:
            self._col_header_row(ws, r, ["ID", "Name", "Reason", ""]); r += 1
            for p in skipped:
                self._data_row(ws, r, [
                    p.get("id",   ""),
                    p.get("name", ""),
                    "Failed validation (missing field or invalid area)",
                    "",
                ])
                r += 1
        else:
            ws.merge_cells(f"A{r}:D{r}")
            c           = ws.cell(row=r, column=1)
            c.value     = "No skipped properties"
            c.font      = _FONT_MUTED
            c.alignment = _ALIGN_CENTER

    # ── Public entry point ────────────────────────────────────────────────────

    def build(self, filename: str) -> str:
        """Build all 3 sheets and save to filename. Returns filename."""
        self.sheet_batch_summary()
        self.sheet_completed_properties()
        self.sheet_failed_skipped()
        self.workbook.save(filename)
        return filename
