from datetime import datetime
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from adapters.asset import AssetValuationResult
from reports.report_theme import BuilderPalette as _BP, NumFormat as _NF, Palette as _Palette, get_fill as _gf


# ── Style constants ────────────────────────────────────────────────────────────

_FILL_HEADER   = _gf(_BP.HEADER)
_FILL_SUBHEAD  = _gf(_BP.SUBHEAD)
_FILL_SECTION  = _gf(_BP.SECTION_MID)
_FILL_CB         = _gf(_BP.CB_GREEN)
_FILL_PORTFOLIO  = _gf(_BP.PORTFOLIO)
_FILL_PORT_COL   = _gf(_BP.PORT_COL)
_FILL_ERROR      = _gf(_BP.ERROR)
_FILL_WARNING  = _gf(_BP.WARNING)

_FONT_HEADER   = Font(bold=True, color=_Palette.WHITE, size=12)
_FONT_TITLE    = Font(bold=True, size=14)
_FONT_SECTION  = Font(bold=True, size=12)
_FONT_BOLD     = Font(bold=True)
_FONT_ERR_LBL  = Font(color=_Palette.WHITE, bold=True)
_FONT_MUTED    = Font(italic=True, color=_BP.MUTED)

_ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_WRAP    = Alignment(wrap_text=True, vertical="top")
_BORDER_THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

_FMT_CURRENCY  = _NF.CURRENCY_2DP
_FMT_PCT       = _NF.PERCENT_DETAILED

# ── Legacy Arabic-sheet style constants ───────────────────────────────────────
_FILL_INPUT_SECT  = _gf(_BP.SECTION_DARK)
_FILL_INPUT_CELL  = _gf(_BP.INPUT_CELL)
_FILL_CALC_CELL   = _gf(_BP.CALC_CELL)
_FILL_FINAL_VALUE = _gf(_BP.SUCCESS_LIGHT)
_FILL_ROW_BAND    = _gf(_BP.ROW_BAND)
_FONT_FINAL_VALUE = Font(bold=True, size=12, color=_BP.SUCCESS_DARK)
_BORDER_MEDIUM    = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)

# ── Sales comparison adjustment fills (promoted from method scope) ─────────────
_F_ADJ_NEG  = _gf(_BP.ADJ_NEG)
_F_ADJ_ZERO = _gf(_BP.ADJ_ZERO)
_F_GOLD     = _gf(_BP.ADJ_GOLD)
_F_EMERALD  = _gf(_BP.ADJ_EMERALD)
_FONT_GOLD    = Font(bold=True, size=10, color=_BP.GOLD_DARK)
_FONT_EMERALD = Font(bold=True, size=10, color=_BP.SUCCESS_DARK)
_FONT_FINAL   = Font(bold=True, size=13, color=_BP.AMBER_DARK)
_FONT_COLHDR  = Font(bold=True, color=_Palette.WHITE, size=9)

# ── Advanced-analytics sheets excluded from the legacy export ─────────────────
# Sourced from the central report profiles registry (report_profiles.py).
from reports.report_profiles import get_legacy_excluded_sheets as _get_legacy_excluded_sheets
_LEGACY_EXCLUDED_SHEETS: frozenset = _get_legacy_excluded_sheets()

# EGVS / IFRS reference descriptions
_DISCLOSURE_DESCRIPTIONS: dict[str, str] = {
    "EGVS_1.0":  "Definition of Market Value",
    "EGVS_1.1":  "Basis of Valuation",
    "EGVS_2.0":  "Assumptions and Limiting Conditions",
    "EGVS_2.1":  "Market Conditions and Valuation Date",
    "EGVS_2.2":  "Special Assumptions",
    "EGVS_2.3":  "Valuation Date and Purpose",
    "EGVS_2.4":  "Insurance Valuation Purpose",
    "EGVS_2.5":  "Underwriting Assumptions",
    "EGVS_3.0":  "Three Approaches to Value",
    "EGVS_3.1":  "Sales Comparison Approach",
    "EGVS_3.2":  "Cost Approach",
    "EGVS_3.3":  "Income Capitalisation Approach",
    "EGVS_4.0":  "Replacement Cost Basis",
    "IFRS_13-1":  "Scope and Objective of Fair Value",
    "IFRS_13-54": "Definition of Fair Value",
    "IFRS_13-55": "Market Participant Assumption",
    "IFRS_13-72": "Fair Value Hierarchy (Introduction)",
    "IFRS_13-73": "Level 1 — Quoted Prices",
    "IFRS_13-74": "Level 2 — Observable Inputs",
    "IFRS_13-75": "Level 3 — Unobservable Inputs",
    "IFRS_13-89": "Non-Performance Risk",
    "IFRS_13-90": "Valuation Adjustments",
    "CBE_Circular_Mortgage_Rules": "Central Bank of Egypt Mortgage Rules",
    "Basel_III_LGD":               "Basel III Loss Given Default",
    "Egyptian_Lender_Standards":   "Egyptian Lender Standards",
    "Insurance_Industry_Standards":"Egyptian Insurance Industry Standards",
    "AssetAdapter_RESIDENTIAL":    "Residential Asset Adapter (80/15/5 weighting)",
    "AssetAdapter_COMMERCIAL":     "Commercial Asset Adapter (income-focused weighting)",
}


class ExcelReportBuilder:
    """
    Generate an EGVS-compliant Excel workbook from an AssetValuationResult.

    Sheet order
    -----------
    0  Summary            — executive overview
    1  Three Approaches   — Phase 4 raw values
    2  Weights Analysis   — weighted reconciliation table
    3  Audit Trail        — full step-by-step log
    4  Property Details   — all metadata fields
    5  Disclosures        — EGVS + IFRS compliance refs
    6  Issues & Warnings  — validation issues
    7  Certification      — appraiser signature block
    8  IVSC Compliance    — optional; added when ivsc_disclosure is provided
    """

    def __init__(self, result: Optional[AssetValuationResult] = None) -> None:
        self.result      = result
        self.workbook    = Workbook()
        self.workbook.remove(self.workbook.active)   # remove default blank sheet
        self.report_date = datetime.now().strftime("%Y-%m-%d")

    # ──────────────────────────────────────────────────────────────────────────
    # Shared styling helpers
    # ──────────────────────────────────────────────────────────────────────────

    def _apply_header_style(self, ws, row: int, cols: int) -> None:
        """Dark-blue fill, white bold text, thin border on every cell in the row."""
        for col in range(1, cols + 1):
            cell             = ws.cell(row=row, column=col)
            cell.fill        = _FILL_HEADER
            cell.font        = _FONT_HEADER
            cell.alignment   = _ALIGN_CENTER
            cell.border      = _BORDER_THIN

    def _apply_currency_format(self, ws, start_row: int, end_row: int, col: int) -> None:
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=col).number_format = _FMT_CURRENCY

    def _apply_percentage_format(self, ws, start_row: int, end_row: int, col: int) -> None:
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=col).number_format = _FMT_PCT

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet builders
    # ──────────────────────────────────────────────────────────────────────────

    def sheet_summary(self) -> None:
        ws = self.workbook.create_sheet("Summary", 0)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 22

        ws["A1"]      = "VALUATION REPORT SUMMARY"
        ws["A1"].font = _FONT_TITLE

        row = 3
        def _row(label, value, *, currency=False, pct=False):
            nonlocal row
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = value
            if currency:
                self._apply_currency_format(ws, row, row, 2)
            if pct:
                self._apply_percentage_format(ws, row, row, 2)
            row += 1

        _row("Report Date",    self.report_date)
        _row("Asset Type",     self.result.asset_type)
        _row("Primary Purpose",self.result.primary_purpose)
        _row("Final Value (EGP)",
             float(self.result.primary_value) if self.result.primary_value else 0,
             currency=True)
        _row("Confidence",     self.result.confidence)

        row += 1
        ws.cell(row=row, column=1).value = "COMPARABLE ANALYSIS"
        ws.cell(row=row, column=1).font  = _FONT_SECTION
        row += 1

        md = self.result.metadata
        _row("Comparable Value", md.get("comparable", 0), currency=True)
        _row("Cost Value",       md.get("cost",       0), currency=True)
        _row("Income Value",     md.get("income",     0), currency=True)

        row += 1
        ws.cell(row=row, column=1).value = "WEIGHTS APPLIED"
        ws.cell(row=row, column=1).font  = _FONT_SECTION
        row += 1

        w = self.result.weights_applied
        _row("Comparable", w.get("comparable", 0), pct=True)
        _row("Cost",        w.get("cost",       0), pct=True)
        _row("Income",      w.get("income",     0), pct=True)

        if self.result.alternative_values:
            row += 1
            ws.cell(row=row, column=1).value = "ALTERNATIVE VALUES"
            ws.cell(row=row, column=1).font  = _FONT_SECTION
            row += 1
            for purpose, value in self.result.alternative_values.items():
                _row(purpose.replace("_", " ").title(),
                     float(value) if value else 0, currency=True)

    def sheet_three_approaches(self) -> None:
        ws = self.workbook.create_sheet("Three Approaches", 1)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 16

        headers = ["Approach", "Value (EGP)", "% of Total"]
        self._apply_header_style(ws, 1, len(headers))
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = h

        md  = self.result.metadata
        comp = md.get("comparable", 0)
        cost = md.get("cost",       0)
        inc  = md.get("income",     0)
        total = comp + cost + inc

        rows = [
            ("Comparable Sales", comp),
            ("Cost Approach",    cost),
            ("Income Approach",  inc),
        ]
        for r, (label, val) in enumerate(rows, 2):
            ws.cell(row=r, column=1).value = label
            ws.cell(row=r, column=2).value = val
            self._apply_currency_format(ws, r, r, 2)
            if total > 0:
                ws.cell(row=r, column=3).value = val / total
                self._apply_percentage_format(ws, r, r, 3)

        # Total row
        tr = len(rows) + 2
        ws.cell(row=tr, column=1).value = "TOTAL"
        ws.cell(row=tr, column=1).font  = _FONT_BOLD
        ws.cell(row=tr, column=2).value = total
        ws.cell(row=tr, column=2).font  = _FONT_BOLD
        self._apply_currency_format(ws, tr, tr, 2)

    def sheet_weights_analysis(self) -> None:
        ws = self.workbook.create_sheet("Weights Analysis", 2)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 22

        headers = ["Approach", "Weight %", "Weighted Value (EGP)"]
        self._apply_header_style(ws, 1, len(headers))
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = h

        md      = self.result.metadata
        weights = self.result.weights_applied
        entries = [
            ("Comparable", md.get("comparable", 0), weights.get("comparable", 0)),
            ("Cost",        md.get("cost",       0), weights.get("cost",       0)),
            ("Income",      md.get("income",     0), weights.get("income",     0)),
        ]
        for r, (label, val, wt) in enumerate(entries, 2):
            ws.cell(row=r, column=1).value = label
            ws.cell(row=r, column=2).value = wt
            self._apply_percentage_format(ws, r, r, 2)
            ws.cell(row=r, column=3).value = val * wt
            self._apply_currency_format(ws, r, r, 3)

        # Reconciled value row
        tr = len(entries) + 2
        ws.cell(row=tr, column=1).value = "RECONCILED VALUE"
        ws.cell(row=tr, column=1).font  = _FONT_BOLD
        ws.cell(row=tr, column=3).value = float(self.result.primary_value) if self.result.primary_value else 0
        ws.cell(row=tr, column=3).font  = _FONT_BOLD
        self._apply_currency_format(ws, tr, tr, 3)

    def sheet_audit_trail(self) -> None:
        ws = self.workbook.create_sheet("Audit Trail", 3)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 22

        headers = ["Step Name", "Inputs", "Outputs", "Formula", "References"]
        self._apply_header_style(ws, 1, len(headers))
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = h

        for r, entry in enumerate(self.result.audit_trail, 2):
            ws.cell(row=r, column=1).value = entry.step_name
            ws.cell(row=r, column=2).value = str(entry.inputs)
            ws.cell(row=r, column=3).value = str(entry.outputs)
            ws.cell(row=r, column=4).value = entry.formula
            ws.cell(row=r, column=5).value = ", ".join(entry.references)
            for col in range(1, 6):
                ws.cell(row=r, column=col).alignment = _ALIGN_WRAP

    def sheet_property_details(self) -> None:
        ws = self.workbook.create_sheet("Property Details", 4)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

        ws["A1"]      = "SUBJECT PROPERTY DETAILS"
        ws["A1"].font = _FONT_SECTION

        row = 3
        for key, value in self.result.metadata.items():
            ws.cell(row=row, column=1).value = key.replace("_", " ").title()
            is_value_field = "value" in key.lower() and isinstance(value, (int, float))
            is_weight_dict = "weight" in key.lower() and isinstance(value, dict)

            if is_value_field:
                ws.cell(row=row, column=2).value = value
                self._apply_currency_format(ws, row, row, 2)
            elif is_weight_dict:
                ws.cell(row=row, column=2).value = str(value)
            else:
                ws.cell(row=row, column=2).value = str(value)
            row += 1

    def sheet_disclosures(self) -> None:
        ws = self.workbook.create_sheet("Disclosures", 5)
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 60

        ws["A1"]      = "STANDARDS & COMPLIANCE REFERENCES"
        ws["A1"].font = _FONT_SECTION

        row = 3
        for disc in self.result.disclosures:
            ws.cell(row=row, column=1).value = disc
            ws.cell(row=row, column=1).font  = _FONT_BOLD
            ws.cell(row=row, column=2).value = _DISCLOSURE_DESCRIPTIONS.get(disc, "See standards documentation")
            row += 1

    def sheet_issues_warnings(self) -> None:
        ws = self.workbook.create_sheet("Issues & Warnings", 6)
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 55

        headers = ["Severity", "Code", "Message"]
        self._apply_header_style(ws, 1, len(headers))
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = h

        if not self.result.issues:
            cell = ws.cell(row=2, column=1)
            cell.value = "No issues"
            cell.font  = _FONT_MUTED
            return

        for r, issue in enumerate(self.result.issues, 2):
            sev_cell = ws.cell(row=r, column=1)
            sev_cell.value = issue.severity
            if issue.severity == "error":
                sev_cell.fill = _FILL_ERROR
                sev_cell.font = _FONT_ERR_LBL
            elif issue.severity == "warning":
                sev_cell.fill = _FILL_WARNING

            ws.cell(row=r, column=2).value = issue.code
            msg_cell = ws.cell(row=r, column=3)
            msg_cell.value     = issue.message
            msg_cell.alignment = _ALIGN_WRAP

    def sheet_certification(self) -> None:
        ws = self.workbook.create_sheet("Certification", 7)
        ws.column_dimensions["A"].width = 80

        ws["A1"]      = "APPRAISER CERTIFICATION"
        ws["A1"].font = _FONT_TITLE

        cert_text = (
            "I certify that, to the best of my knowledge and belief, the statements of fact "
            "contained in this appraisal are true and correct, and the reported analyses, "
            "conclusions, and recommendations are limited only by the reported assumptions "
            "and limiting conditions and are my personal, impartial, and unbiased professional "
            "analysis, opinions, and conclusions.\n\n"
            "I have no present or prospective interest in the property that is the subject of "
            "this report, and I have performed no services, as an appraiser or in any other "
            "capacity, regarding the subject property that would create a conflict of interest.\n\n"
            "The value opinion in this appraisal was developed independently and in accordance "
            "with the Egyptian Valuation Standard (EgVS) and International Financial Reporting "
            "Standard (IFRS) 13 Fair Value Measurement."
        )

        row = 3
        cert_cell           = ws.cell(row=row, column=1)
        cert_cell.value     = cert_text
        cert_cell.alignment = _ALIGN_WRAP
        ws.row_dimensions[row].height = 130

        row += 10
        for line in (
            "Appraiser Name: ___________________________",
            "License No.:    ___________________________",
            f"Date:           {self.report_date}",
            "Signature:      ___________________________",
        ):
            ws.cell(row=row, column=1).value = line
            row += 2

    def sheet_ivsc_compliance(self, ivsc_disclosure) -> None:
        ws = self.workbook.create_sheet("IVSC Compliance", 8)
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 70

        def _section(row: int, title: str) -> None:
            cell = ws.cell(row=row, column=1)
            cell.value     = title
            cell.fill      = _FILL_SECTION
            cell.font      = Font(bold=True, color=_Palette.WHITE, size=11)
            cell.alignment = _ALIGN_WRAP
            ws.cell(row=row, column=2).fill = _FILL_SECTION

        def _row(row: int, label: str, value) -> None:
            lc = ws.cell(row=row, column=1)
            lc.value     = label
            lc.font      = _FONT_BOLD
            lc.alignment = _ALIGN_WRAP
            vc = ws.cell(row=row, column=2)
            vc.value     = value if value is not None else "—"
            vc.alignment = _ALIGN_WRAP

        # ── Title row ─────────────────────────────────────────────────────────
        ws.merge_cells("A1:B1")
        title_cell           = ws["A1"]
        title_cell.value     = "IVSC COMPLIANCE DISCLOSURE"
        title_cell.fill      = _FILL_HEADER
        title_cell.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        title_cell.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:B2")
        ws["A2"].value     = f"Report Date: {self.report_date}"
        ws["A2"].alignment = Alignment(horizontal="right")
        ws["A2"].font      = _FONT_MUTED

        r = 4
        # ── Scope of Work ─────────────────────────────────────────────────────
        _section(r, "SCOPE OF WORK (IVS 101)"); r += 1
        _row(r, "Scope of Work",        ivsc_disclosure.scope_of_work);        r += 1
        _row(r, "Purpose of Valuation", ivsc_disclosure.purpose_of_valuation); r += 1
        _row(r, "Effective Date",
             ivsc_disclosure.effective_date.isoformat()
             if ivsc_disclosure.effective_date else "—");                       r += 2

        # ── Basis of Valuation ────────────────────────────────────────────────
        _section(r, "BASIS OF VALUATION (IVS 102)"); r += 1
        _row(r, "Basis",      ivsc_disclosure.basis_of_valuation); r += 1
        _row(r, "Definition", ivsc_disclosure.definition);         r += 2

        # ── Valuation Approaches ──────────────────────────────────────────────
        _section(r, "VALUATION APPROACHES (IVS 103-105)"); r += 1
        _row(r, "Approaches Used",
             ", ".join(ivsc_disclosure.approaches_used)
             if ivsc_disclosure.approaches_used else "—");         r += 1
        _row(r, "Methodology Summary", ivsc_disclosure.methodology_summary)
        ws.row_dimensions[r].height = 72;                          r += 2

        # ── Key Assumptions ───────────────────────────────────────────────────
        _section(r, "KEY ASSUMPTIONS (IVS 105)"); r += 1
        for assumption in ivsc_disclosure.key_assumptions:
            _row(r, "", f"• {assumption}"); r += 1
        r += 1

        # ── Limiting Conditions ───────────────────────────────────────────────
        _section(r, "LIMITING CONDITIONS"); r += 1
        for condition in ivsc_disclosure.limiting_conditions:
            _row(r, "", f"• {condition}"); r += 1
        r += 1

        # ── Market Conditions ─────────────────────────────────────────────────
        _section(r, "MARKET CONDITIONS (IVS 104)"); r += 1
        _row(r, "Market Conditions",  ivsc_disclosure.market_conditions_summary); r += 1
        _row(r, "Economic Conditions", ivsc_disclosure.economic_conditions);       r += 1
        _row(r, "Currency",            ivsc_disclosure.currency);                  r += 1
        if ivsc_disclosure.exchange_rate_basis:
            _row(r, "Exchange Rate Basis", ivsc_disclosure.exchange_rate_basis);   r += 1
        r += 1

        # ── Inspection ────────────────────────────────────────────────────────
        _section(r, "INSPECTION (IVS 105)"); r += 1
        _row(r, "Inspection Date",
             ivsc_disclosure.inspection_date.isoformat()
             if ivsc_disclosure.inspection_date else "—");           r += 1
        _row(r, "Inspection Notes", ivsc_disclosure.inspection_notes); r += 2

        # ── Appraiser Credentials ─────────────────────────────────────────────
        _section(r, "APPRAISER CREDENTIALS"); r += 1
        _row(r, "Appraiser Name",           ivsc_disclosure.appraiser_name);           r += 1
        _row(r, "Qualifications",           ivsc_disclosure.appraiser_qualifications); r += 1
        _row(r, "Declaration",              ivsc_disclosure.appraiser_declaration)
        ws.row_dimensions[r].height = 60;                                               r += 2

        # ── Standards Applied ─────────────────────────────────────────────────
        _section(r, "STANDARDS APPLIED"); r += 1
        ivsc_vals = [s.value for s in ivsc_disclosure.ivsc_standards_applied]
        _row(r, "IVSC Standards",
             "\n".join(f"• {v}" for v in ivsc_vals) if ivsc_vals else "—")
        ws.row_dimensions[r].height = max(15 * len(ivsc_vals), 15);  r += 1
        national = ivsc_disclosure.national_standards_applied
        _row(r, "National Standards",
             ", ".join(national) if national else "—");               r += 2

        # ── Certification ─────────────────────────────────────────────────────
        _section(r, "CERTIFICATION"); r += 1
        _row(r, "Certification Statement", ivsc_disclosure.certification_statement)
        ws.row_dimensions[r].height = 80;                             r += 1
        _row(r, "Date Certified",
             ivsc_disclosure.date_certified.isoformat()
             if ivsc_disclosure.date_certified else "—");             r += 1

    def sheet_cross_border(self, cb_disclosure) -> None:
        ws = self.workbook.create_sheet("Cross-Border", 9)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25

        _cb_section_font = Font(bold=True, color=_Palette.WHITE, size=11)

        def _section(row: int, title: str) -> None:
            ws.merge_cells(f"A{row}:D{row}")
            cell            = ws.cell(row=row, column=1)
            cell.value      = title
            cell.fill       = _FILL_CB
            cell.font       = _cb_section_font
            cell.alignment  = _ALIGN_WRAP

        def _row(row: int, label: str, value, merge_value: bool = False) -> None:
            lc            = ws.cell(row=row, column=1)
            lc.value      = label
            lc.font       = _FONT_BOLD
            lc.alignment  = _ALIGN_WRAP
            if merge_value:
                ws.merge_cells(f"B{row}:D{row}")
            vc            = ws.cell(row=row, column=2)
            vc.value      = value if value is not None else "—"
            vc.alignment  = _ALIGN_WRAP

        # ── Title ─────────────────────────────────────────────────────────────
        ws.merge_cells("A1:D1")
        title_cell           = ws["A1"]
        title_cell.value     = "CROSS-BORDER COMPLIANCE DISCLOSURE"
        title_cell.fill      = _FILL_CB
        title_cell.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        title_cell.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:D2")
        ws["A2"].value     = f"Report Date: {self.report_date}"
        ws["A2"].alignment = Alignment(horizontal="right")
        ws["A2"].font      = _FONT_MUTED

        r = 4
        # ── Currencies & Locations ────────────────────────────────────────────
        _section(r, "CURRENCIES & LOCATIONS"); r += 1
        _row(r, "Subject Property Currency",
             cb_disclosure.subject_property_currency.value);   r += 1
        _row(r, "Reporting Currency",
             cb_disclosure.reporting_currency.value);          r += 1
        _row(r, "Property Location Country",
             cb_disclosure.property_location_country);         r += 1
        _row(r, "Valuation Purpose Country",
             cb_disclosure.valuation_purpose_country);         r += 2

        # ── Exchange Rate Assumptions ─────────────────────────────────────────
        if cb_disclosure.exchange_rate_assumption:
            era = cb_disclosure.exchange_rate_assumption
            _section(r, "EXCHANGE RATE ASSUMPTIONS"); r += 1
            _row(r, "Exchange Rate",   era.format_rate());              r += 1
            _row(r, "Effective Date",  era.effective_date.isoformat()); r += 1
            _row(r, "Source",          era.source);                     r += 1
            _row(r, "Currency Risk",   era.currency_risk_disclosure,
                 merge_value=True)
            ws.row_dimensions[r].height = 48;                           r += 2

        # ── Currency Risk Statement ───────────────────────────────────────────
        _section(r, "CURRENCY RISK STATEMENT"); r += 1
        ws.merge_cells(f"A{r}:D{r}")
        risk_cell           = ws.cell(row=r, column=1)
        risk_cell.value     = cb_disclosure.currency_risk_statement
        risk_cell.alignment = _ALIGN_WRAP
        ws.row_dimensions[r].height = 60;                               r += 2

        # ── Multi-Currency Values ─────────────────────────────────────────────
        _section(r, "VALUATION IN MULTIPLE CURRENCIES"); r += 1
        _row(r, "Value (EGP)",
             f"EGP {cb_disclosure.primary_value_egp:,.0f}");   r += 1
        if cb_disclosure.primary_value_usd > 0:
            _row(r, "Value (USD)",
                 f"USD {cb_disclosure.primary_value_usd:,.0f}"); r += 1
        if cb_disclosure.primary_value_eur > 0:
            _row(r, "Value (EUR)",
                 f"EUR {cb_disclosure.primary_value_eur:,.0f}"); r += 1
        r += 1

        # ── Reporting Assumptions ─────────────────────────────────────────────
        _section(r, "REPORTING ASSUMPTIONS"); r += 1
        for idx, assumption in enumerate(cb_disclosure.reporting_assumptions, 1):
            ws.cell(row=r, column=1).value      = f"{idx}."
            ws.cell(row=r, column=1).font       = _FONT_BOLD
            ws.merge_cells(f"B{r}:D{r}")
            vc                 = ws.cell(row=r, column=2)
            vc.value           = assumption
            vc.alignment       = _ALIGN_WRAP
            ws.row_dimensions[r].height = 30;  r += 1
        r += 1

        # ── Certification ─────────────────────────────────────────────────────
        _section(r, "CERTIFICATION"); r += 1
        ws.merge_cells(f"A{r}:D{r}")
        cert_cell           = ws.cell(row=r, column=1)
        cert_cell.value     = cb_disclosure.certification_statement
        cert_cell.alignment = _ALIGN_WRAP
        ws.row_dimensions[r].height = 72

    def sheet_portfolio_summary(self, ps: Dict) -> None:
        """Portfolio Summary sheet — overview of all properties in a portfolio."""
        ws = self.workbook.create_sheet("Portfolio Summary")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 15

        _port_sect_font = Font(bold=True, color=_Palette.WHITE, size=11)

        def _section(row: int, title: str) -> None:
            ws.merge_cells(f"A{row}:D{row}")
            cell           = ws.cell(row=row, column=1)
            cell.value     = title
            cell.fill      = _FILL_PORTFOLIO
            cell.font      = _port_sect_font
            cell.alignment = _ALIGN_WRAP

        def _row(row: int, label: str, value, col_b: int = 2, merge_cd: bool = False) -> None:
            lc           = ws.cell(row=row, column=1)
            lc.value     = label
            lc.font      = _FONT_BOLD
            lc.alignment = _ALIGN_WRAP
            if merge_cd:
                ws.merge_cells(f"B{row}:D{row}")
            vc           = ws.cell(row=row, column=col_b)
            vc.value     = value if value is not None else "—"
            vc.alignment = _ALIGN_WRAP

        metrics = ps.get("metrics", {})

        # ── Title ─────────────────────────────────────────────────────────────
        ws.merge_cells("A1:D1")
        title_cell           = ws["A1"]
        title_cell.value     = "PORTFOLIO SUMMARY"
        title_cell.fill      = _FILL_PORTFOLIO
        title_cell.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        title_cell.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:D2")
        ws["A2"].value     = f"Report Date: {self.report_date}"
        ws["A2"].alignment = Alignment(horizontal="right")
        ws["A2"].font      = _FONT_MUTED

        r = 4
        # ── Portfolio Overview ────────────────────────────────────────────────
        _section(r, "PORTFOLIO OVERVIEW"); r += 1
        _row(r, "Portfolio Name",
             ps.get("portfolio_name", "Unnamed"));                          r += 1
        _row(r, "Total Portfolio Value",
             f"EGP {metrics.get('total_portfolio_value', 0):,.0f}");        r += 1
        _row(r, "Number of Properties",
             metrics.get("number_of_properties", 0));                       r += 1
        _row(r, "Portfolio Cap Rate",
             f"{metrics.get('portfolio_cap_rate', 0) * 100:.2f}%");         r += 2

        # ── Value Distribution by Type ────────────────────────────────────────
        _section(r, "VALUE DISTRIBUTION BY TYPE"); r += 1
        type_vals = metrics.get("value_by_type", {})
        type_pcts = metrics.get("type_percentages", {})
        for pt in sorted(type_vals):
            ws.cell(row=r, column=1).value = f"{pt.title()}:"
            ws.cell(row=r, column=1).font  = _FONT_BOLD
            ws.cell(row=r, column=2).value = f"EGP {type_vals[pt]:,.0f}"
            ws.cell(row=r, column=3).value = f"{type_pcts.get(pt, 0) * 100:.1f}%"
            r += 1
        r += 1

        # ── Income Analysis ───────────────────────────────────────────────────
        _section(r, "INCOME ANALYSIS"); r += 1
        _row(r, "Total Annual Gross Income",
             f"EGP {metrics.get('total_annual_gross_income', 0):,.0f}");    r += 1
        _row(r, "Total Annual NOI",
             f"EGP {metrics.get('total_annual_noi', 0):,.0f}");             r += 1
        _row(r, "Portfolio NOI Margin",
             f"{metrics.get('portfolio_noi_margin', 0) * 100:.2f}%");       r += 2

        # ── Risk & Diversification ────────────────────────────────────────────
        _section(r, "RISK & DIVERSIFICATION"); r += 1
        cr = metrics.get("concentration_ratio", 0)
        hi = metrics.get("herfindahl_index", 0)
        ds = round(1 - hi, 4)
        _row(r, "Concentration Ratio",      f"{cr * 100:.2f}%");            r += 1
        ws.cell(row=r - 1, column=3).value = "(Lower = more diversified)"
        _row(r, "Herfindahl Index",         f"{hi:.4f}");                   r += 1
        _row(r, "Diversification Score",    f"{ds:.4f}");                   r += 1
        ws.cell(row=r - 1, column=3).value = "(1.0 = perfectly diversified)"
        r += 1

        # ── Valuation Confidence ──────────────────────────────────────────────
        _section(r, "VALUATION CONFIDENCE"); r += 1
        high_val = metrics.get("high_confidence_value", 0)
        _row(r, "High Confidence",
             f"{metrics.get('high_confidence_count', 0)} properties  "
             f"(EGP {high_val:,.0f})");                                     r += 1
        _row(r, "Medium Confidence",
             f"{metrics.get('medium_confidence_count', 0)} properties");    r += 1
        _row(r, "Low Confidence",
             f"{metrics.get('low_confidence_count', 0)} properties");       r += 2

        # ── Property Listing ──────────────────────────────────────────────────
        _section(r, "PROPERTY LISTING"); r += 1
        for col, hdr in enumerate(("Property Name", "Type", "Value (EGP)", "Weight %"), 1):
            cell       = ws.cell(row=r, column=col)
            cell.value = hdr
            cell.font  = _FONT_BOLD
            cell.fill  = _FILL_PORT_COL
        r += 1
        for prop in ps.get("properties", []):
            ws.cell(row=r, column=1).value = prop.get("property_name", "")
            ws.cell(row=r, column=2).value = prop.get("property_type",  "")
            ws.cell(row=r, column=3).value = f"{prop.get('valuation_value', 0):,.0f}"
            ws.cell(row=r, column=4).value = f"{prop.get('portfolio_weight', 0) * 100:.2f}%"
            r += 1

    def sheet_portfolio_performance(self, perf: Dict) -> None:
        """Portfolio Performance sheet — scenario comparison + risk summary (Phase 11.3)."""
        ws = self.workbook.create_sheet("Portfolio Performance")

        for col, width in zip("ABCDEFGH", (24, 22, 14, 22, 14, 16, 14, 14)):
            ws.column_dimensions[col].width = width

        _port_sect_font = Font(bold=True, color=_Palette.WHITE, size=12)

        def _section(row: int, title: str) -> None:
            ws.merge_cells(f"A{row}:H{row}")
            cell           = ws.cell(row=row, column=1)
            cell.value     = title
            cell.fill      = _FILL_PORTFOLIO
            cell.font      = _port_sect_font
            cell.alignment = _ALIGN_WRAP

        def _kv(row: int, label: str, value) -> None:
            lc           = ws.cell(row=row, column=1)
            lc.value     = label
            lc.font      = _FONT_BOLD
            lc.alignment = _ALIGN_WRAP
            vc           = ws.cell(row=row, column=2)
            vc.value     = value if value is not None else "—"
            vc.alignment = _ALIGN_WRAP

        # Safe accessor — conditional keys may be absent when no scenarios ran
        def _g(key, default=None):
            return perf.get(key, default)

        # ── Title ─────────────────────────────────────────────────────────────
        ws.merge_cells("A1:H1")
        tc           = ws["A1"]
        tc.value     = "PORTFOLIO PERFORMANCE ANALYSIS"
        tc.fill      = _FILL_PORTFOLIO
        tc.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        tc.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:H2")
        ws["A2"].value     = f"Report Date: {self.report_date}"
        ws["A2"].alignment = Alignment(horizontal="right")
        ws["A2"].font      = _FONT_MUTED

        r = 4

        # ── Section 1: Baseline Metrics ───────────────────────────────────────
        _section(r, "BASELINE METRICS"); r += 1
        _kv(r, "Portfolio Name",        _g("portfolio_name", "—"));                   r += 1
        _kv(r, "Base Portfolio Value",  f"EGP {_g('base_portfolio_value', 0):,.0f}");     r += 1
        _kv(r, "Base Total NOI",        f"EGP {_g('base_total_noi', 0):,.0f}");           r += 1
        _kv(r, "Base Cap Rate",         f"{_g('base_cap_rate', 0) * 100:.2f}%");          r += 1
        _kv(r, "Diversification Score", f"{_g('diversification_score', 0):.4f}");        r += 1
        _kv(r, "Scenario Count",        _g("scenario_count", 0));                         r += 2

        # ── Section 2: Scenario Table ─────────────────────────────────────────
        _section(r, "SCENARIO ANALYSIS"); r += 1

        col_headers = [
            "Scenario", "Stressed Value (EGP)", "Value Δ%",
            "Stressed NOI (EGP)", "NOI Δ%", "Stressed Cap Rate",
            "NOI Margin", "IRR Estimate",
        ]
        for col_i, hdr in enumerate(col_headers, 1):
            cell           = ws.cell(row=r, column=col_i)
            cell.value     = hdr
            cell.font      = _FONT_BOLD
            cell.fill      = _FILL_PORT_COL
            cell.alignment = _ALIGN_WRAP
            cell.border    = _BORDER_THIN
        r += 1

        scenarios = _g("scenarios") or []
        if scenarios:
            for sc in scenarios:
                row_vals = [
                    sc.get("scenario_label", ""),
                    f"EGP {sc.get('stressed_portfolio_value', 0):,.0f}",
                    f"{sc.get('value_change_pct', 0) * 100:.2f}%",
                    f"EGP {sc.get('stressed_total_noi', 0):,.0f}",
                    f"{sc.get('noi_change_pct', 0) * 100:.2f}%",
                    f"{sc.get('stressed_cap_rate', 0) * 100:.2f}%",
                    f"{sc.get('stressed_noi_margin', 0) * 100:.2f}%",
                    f"{sc.get('irr_estimate', 0) * 100:.2f}%",
                ]
                for col_i, v in enumerate(row_vals, 1):
                    cell           = ws.cell(row=r, column=col_i)
                    cell.value     = v
                    cell.alignment = _ALIGN_WRAP
                    cell.border    = _BORDER_THIN
                r += 1
        else:
            ws.merge_cells(f"A{r}:H{r}")
            cell           = ws.cell(row=r, column=1)
            cell.value     = "No scenario data available"
            cell.font      = _FONT_MUTED
            cell.alignment = _ALIGN_CENTER
            r += 1

        r += 1  # gap before risk section

        # ── Section 3: Risk Summary ────────────────────────────────────────────
        _section(r, "RISK SUMMARY"); r += 1

        min_val = _g("min_stressed_value")
        max_val = _g("max_stressed_value")
        min_noi = _g("min_stressed_noi")
        max_noi = _g("max_stressed_noi")
        var_pct = _g("value_at_risk_pct")

        _kv(r, "Min Stressed Value",
            f"EGP {min_val:,.0f}" if min_val is not None else "—");   r += 1
        _kv(r, "Max Stressed Value",
            f"EGP {max_val:,.0f}" if max_val is not None else "—");   r += 1
        _kv(r, "Min Stressed NOI",
            f"EGP {min_noi:,.0f}" if min_noi is not None else "—");   r += 1
        _kv(r, "Max Stressed NOI",
            f"EGP {max_noi:,.0f}" if max_noi is not None else "—");   r += 1
        _kv(r, "Scenario-Implied Downside %",
            f"{var_pct * 100:.2f}%" if var_pct is not None else "—"); r += 1

    # ──────────────────────────────────────────────────────────────────────────
    # Legacy-only enhanced Arabic sheets
    # ──────────────────────────────────────────────────────────────────────────

    def sheet_assumptions_inputs(self) -> None:
        """'الافتراضات والمدخلات' — Assumptions & Inputs sheet (legacy only)."""
        ws = self.workbook.create_sheet("الافتراضات والمدخلات")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 22

        md  = self.result.metadata if self.result else {}
        fv  = float(self.result.primary_value) if self.result and self.result.primary_value else 0
        w   = self.result.weights_applied if self.result else {}

        # Banner
        ws.merge_cells("A1:D1")
        t            = ws["A1"]
        t.value      = "الافتراضات والمدخلات — Assumptions & Inputs"
        t.fill       = _FILL_HEADER
        t.font       = Font(bold=True, color=_Palette.WHITE, size=14)
        t.alignment  = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:D2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4

        def _sect(label: str, en: str = "") -> None:
            nonlocal r
            ws.merge_cells(f"A{r}:D{r}")
            c            = ws.cell(row=r, column=1)
            c.value      = f"  {label}" + (f"  — {en}" if en else "")
            c.fill       = _FILL_INPUT_SECT
            c.font       = Font(bold=True, color=_Palette.WHITE, size=11)
            c.alignment  = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

        def _inp(label: str, value, note: str = "", is_calc: bool = False,
                 fmt: str = "") -> None:
            nonlocal r
            fill = _FILL_CALC_CELL if is_calc else _FILL_INPUT_CELL
            lc            = ws.cell(row=r, column=2)
            lc.value      = label
            lc.font       = Font(bold=True, size=10)
            lc.alignment  = Alignment(horizontal="right", vertical="center", wrap_text=True)
            lc.fill       = fill
            lc.border     = _BORDER_THIN

            vc            = ws.cell(row=r, column=3)
            vc.value      = value if value is not None else "—"
            vc.alignment  = Alignment(horizontal="center", vertical="center")
            vc.fill       = fill
            vc.border     = _BORDER_THIN
            if fmt:
                vc.number_format = fmt

            if note:
                nc            = ws.cell(row=r, column=4)
                nc.value      = note
                nc.font       = Font(italic=True, color=_BP.NOTE, size=9)
                nc.alignment  = Alignment(horizontal="right", vertical="center", wrap_text=True)
                nc.fill       = fill
                nc.border     = _BORDER_THIN

            ws.row_dimensions[r].height = 18
            r += 1

        # ── Section 1: بيانات التقرير ─────────────────────────────────────────
        _sect("بيانات التقرير", "Report Data")
        _inp("رقم التقرير",    md.get("report_id") or "—")
        _inp("تاريخ التقييم", md.get("valuation_date") or self.report_date)
        _inp("تاريخ المعاينة", md.get("inspection_date") or "—")
        _inp("غرض التقييم",   self.result.primary_purpose if self.result else "—")
        _inp("أساس القيمة",   "القيمة السوقية")
        r += 1

        # ── Section 2: بيانات العميل والمقيم ────────────────────────────────
        _sect("بيانات العميل والمقيم", "Client & Appraiser")
        _inp("اسم العميل",    md.get("client_name") or md.get("borrower_name") or "—")
        _inp("اسم المقيم",   md.get("appraiser_name") or md.get("reviewer_name") or "—")
        _inp("رقم الترخيص",  md.get("license_no") or "—")
        _inp("الجهة المعينة", md.get("instructed_by") or "—")
        r += 1

        # ── Section 3: بيانات العقار ─────────────────────────────────────────
        _sect("بيانات العقار", "Property Data")
        _inp("نوع الأصل",          self.result.asset_type if self.result else "—")
        _inp("الموقع / العنوان",   md.get("location") or md.get("address") or "—")
        _inp("المساحة (م²)",       md.get("area") or md.get("floor_area_m2") or "—")
        _inp("سنة الإنشاء",       md.get("year_built") or "—")
        _inp("الحالة",             md.get("condition") or "—")
        r += 1

        # ── Section 4: بيانات السوق ──────────────────────────────────────────
        _sect("بيانات السوق", "Market Data")
        _inp("متوسط سعر السوق (EGP/م²)", md.get("market_avg_price_sqm") or "—")
        _inp("معدل الرسملة",    md.get("cap_rate") or "—")
        _inp("معدل الشاغر",    md.get("vacancy_rate") or "—")
        _inp("عدد المقارنات",  md.get("comparables_count") or
             len(md.get("comparables") or []) or "—")
        r += 1

        # ── Section 5: افتراضات التقييم ─────────────────────────────────────
        _sect("افتراضات التقييم", "Valuation Assumptions")
        comp = float(md.get("comparable") or 0)
        cost = float(md.get("cost")       or 0)
        inc  = float(md.get("income")     or 0)

        _inp("القيمة — المقارنة البيعية (EGP)", comp, is_calc=True, fmt=_FMT_CURRENCY)
        _inp("الوزن — المقارنة",               w.get("comparable", 0), fmt=_FMT_PCT)
        _inp("القيمة — طريقة التكلفة (EGP)",   cost, is_calc=True, fmt=_FMT_CURRENCY)
        _inp("الوزن — التكلفة",                w.get("cost", 0), fmt=_FMT_PCT)
        _inp("القيمة — رأسمالة الدخل (EGP)",   inc,  is_calc=True, fmt=_FMT_CURRENCY)
        _inp("الوزن — الدخل",                  w.get("income", 0), fmt=_FMT_PCT)
        r += 1

        # Final value banner
        ws.merge_cells(f"B{r}:D{r}")
        fv_cell            = ws.cell(row=r, column=2)
        fv_cell.value      = f"القيمة السوقية النهائية (EGP):  {fv:,.0f}"
        fv_cell.font       = _FONT_FINAL_VALUE
        fv_cell.fill       = _FILL_FINAL_VALUE
        fv_cell.alignment  = Alignment(horizontal="center", vertical="center")
        fv_cell.border     = _BORDER_MEDIUM
        ws.row_dimensions[r].height = 24
        r += 2

        # ── Section 6: حدود ومحددات الاستخدام ──────────────────────────────
        _sect("حدود ومحددات الاستخدام", "Limiting Conditions")
        for cond in (
            "يُعدّ هذا التقرير سارياً فقط بالغرض المُبيَّن أعلاه.",
            "لا يجوز الاستشهاد بجزء منه دون الرجوع إلى التقرير الكامل.",
            "القيمة محددة وفق أحوال السوق في تاريخ التقييم.",
            "لم يُراعَ في التقدير أي تكاليف بيع أو ضرائب.",
        ):
            ws.merge_cells(f"B{r}:D{r}")
            c            = ws.cell(row=r, column=2)
            c.value      = f"• {cond}"
            c.font       = Font(size=9)
            c.alignment  = Alignment(horizontal="right", vertical="center", wrap_text=True)
            ws.row_dimensions[r].height = 16
            r += 1
        r += 1

        # ── Section 7: إعدادات التقرير ──────────────────────────────────────
        _sect("إعدادات التقرير", "Report Settings")
        _inp("نمط التقرير",      "Legacy — أساسي")
        _inp("مستوى الثقة",     self.result.confidence if self.result else "—")
        _inp("المعايير المطبقة", "EGVS / IFRS 13")
        r += 2

        # ── Legend ────────────────────────────────────────────────────────────
        ws.cell(row=r, column=2).value     = "دليل الألوان:"
        ws.cell(row=r, column=2).font      = Font(bold=True, size=9)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
        r += 1
        for lbl, clr in (
            ("خلية إدخال يدوي",    "EBF3FB"),
            ("خلية محسوبة / ناتجة", "F5F5F5"),
            ("القيمة النهائية",    "C6EFCE"),
        ):
            ws.cell(row=r, column=2).value     = lbl
            ws.cell(row=r, column=2).font      = Font(size=9)
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
            ic       = ws.cell(row=r, column=3)
            ic.fill  = _gf(clr)
            ic.border = _BORDER_THIN
            r += 1

        ws.freeze_panes = "B3"

    def sheet_main_report(self) -> None:
        """'التقرير' — Executive main report sheet (legacy only)."""
        ws = self.workbook.create_sheet("التقرير")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16

        md   = self.result.metadata if self.result else {}
        fv   = float(self.result.primary_value) if self.result and self.result.primary_value else 0
        comp = float(md.get("comparable") or 0)
        cost = float(md.get("cost")       or 0)
        inc  = float(md.get("income")     or 0)
        w    = self.result.weights_applied if self.result else {}

        # ── Title block ───────────────────────────────────────────────────────
        ws.merge_cells("A1:E1")
        t            = ws["A1"]
        t.value      = "تقرير التقييم العقاري"
        t.fill       = _FILL_HEADER
        t.font       = Font(bold=True, color=_Palette.WHITE, size=16)
        t.alignment  = _ALIGN_CENTER
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:E2")
        ws["A2"].value     = "Real Estate Valuation Report"
        ws["A2"].fill      = _FILL_SECTION
        ws["A2"].font      = Font(bold=True, color=_Palette.WHITE, size=12, italic=True)
        ws["A2"].alignment = _ALIGN_CENTER
        ws.row_dimensions[2].height = 22

        asset_type = self.result.asset_type if self.result else "—"
        confidence = self.result.confidence if self.result else "—"
        ws.merge_cells("A3:E3")
        ws["A3"].value     = (f"تاريخ التقرير: {self.report_date}  |  "
                              f"نوع الأصل: {asset_type}  |  الثقة: {confidence}")
        ws["A3"].fill      = _FILL_INPUT_SECT
        ws["A3"].font      = Font(italic=True, color=_Palette.WHITE, size=10)
        ws["A3"].alignment = _ALIGN_CENTER
        ws.row_dimensions[3].height = 18

        # ── Final Value card ──────────────────────────────────────────────────
        ws.merge_cells("B5:E6")
        fv_cell            = ws["B5"]
        fv_cell.value      = f"القيمة السوقية النهائية\n{fv:,.0f}  جنيه مصري"
        fv_cell.fill       = _FILL_FINAL_VALUE
        fv_cell.font       = Font(bold=True, size=14, color=_BP.SUCCESS_DARK)
        fv_cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fv_cell.border     = _BORDER_MEDIUM
        ws.row_dimensions[5].height = 28
        ws.row_dimensions[6].height = 28

        r = 8

        def _section_hd(label: str) -> None:
            nonlocal r
            ws.merge_cells(f"B{r}:E{r}")
            c            = ws.cell(row=r, column=2)
            c.value      = f"  {label}"
            c.fill       = _FILL_INPUT_SECT
            c.font       = Font(bold=True, color=_Palette.WHITE, size=11)
            c.alignment  = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

        def _kv(label: str, value, currency: bool = False,
                pct: bool = False, alt: bool = False) -> None:
            nonlocal r
            fill = _FILL_ROW_BAND if alt else None
            lc            = ws.cell(row=r, column=2)
            lc.value      = label
            lc.font       = Font(bold=True, size=10)
            lc.alignment  = Alignment(horizontal="right", vertical="center")
            lc.border     = _BORDER_THIN
            if fill: lc.fill = fill

            vc            = ws.cell(row=r, column=3)
            vc.value      = value if value is not None else "—"
            vc.alignment  = Alignment(horizontal="center", vertical="center")
            vc.border     = _BORDER_THIN
            if fill: vc.fill = fill
            if currency: vc.number_format = _FMT_CURRENCY
            if pct:      vc.number_format = _FMT_PCT
            ws.row_dimensions[r].height = 18
            r += 1

        # ── Property basics ───────────────────────────────────────────────────
        _section_hd("بيانات العقار الأساسية")
        _kv("نوع الأصل",      asset_type)
        _kv("الموقع",         md.get("location") or md.get("address") or "—", alt=True)
        _kv("غرض التقييم",   self.result.primary_purpose if self.result else "—")
        _kv("تاريخ التقييم", md.get("valuation_date") or self.report_date, alt=True)
        _kv("مستوى الثقة",   confidence)
        r += 1

        # ── Three approaches table ────────────────────────────────────────────
        _section_hd("نتائج أساليب التقييم الثلاثة")

        for col, hdr in enumerate(("الأسلوب", "القيمة (EGP)", "الوزن"), 2):
            hc            = ws.cell(row=r, column=col)
            hc.value      = hdr
            hc.fill       = _FILL_SUBHEAD
            hc.font       = Font(bold=True, size=10)
            hc.alignment  = _ALIGN_CENTER
            hc.border     = _BORDER_THIN
            ws.row_dimensions[r].height = 18
        ws.cell(row=r, column=5).fill   = _FILL_SUBHEAD
        ws.cell(row=r, column=5).border = _BORDER_THIN
        r += 1

        for i, (lbl, val, wk) in enumerate((
            ("المقارنة البيعية", comp, "comparable"),
            ("طريقة التكلفة",   cost, "cost"),
            ("رأسمالة الدخل",   inc,  "income"),
        )):
            fill = _FILL_ROW_BAND if i % 2 else None
            nc = ws.cell(row=r, column=2)
            nc.value     = lbl
            nc.font      = Font(bold=True, size=10)
            nc.alignment = Alignment(horizontal="right", vertical="center")
            nc.border    = _BORDER_THIN
            if fill: nc.fill = fill

            vc = ws.cell(row=r, column=3)
            vc.value         = val
            vc.number_format = _FMT_CURRENCY
            vc.alignment     = _ALIGN_CENTER
            vc.border        = _BORDER_THIN
            if fill: vc.fill = fill

            wc = ws.cell(row=r, column=4)
            wc.value         = w.get(wk, 0)
            wc.number_format = _FMT_PCT
            wc.alignment     = _ALIGN_CENTER
            wc.border        = _BORDER_THIN
            if fill: wc.fill = fill

            ws.row_dimensions[r].height = 18
            r += 1

        # Reconciled total
        rc = ws.cell(row=r, column=2)
        rc.value     = "القيمة التوفيقية النهائية"
        rc.font      = _FONT_FINAL_VALUE
        rc.fill      = _FILL_FINAL_VALUE
        rc.alignment = Alignment(horizontal="right", vertical="center")
        rc.border    = _BORDER_MEDIUM

        vc = ws.cell(row=r, column=3)
        vc.value         = fv
        vc.number_format = _FMT_CURRENCY
        vc.font          = _FONT_FINAL_VALUE
        vc.fill          = _FILL_FINAL_VALUE
        vc.alignment     = _ALIGN_CENTER
        vc.border        = _BORDER_MEDIUM
        ws.row_dimensions[r].height = 22
        r += 2

        # ── Simple bar chart ──────────────────────────────────────────────────
        # Write chart data in hidden-ish rows (no explicit hide — keeps things simple)
        chart_data_row = r
        for col, hdr in enumerate(("الأسلوب", "القيمة"), 2):
            ws.cell(row=r, column=col).value = hdr
            ws.cell(row=r, column=col).font  = _FONT_BOLD
        r += 1
        for lbl, val in (
            ("المقارنة", comp),
            ("التكلفة",  cost),
            ("الدخل",    inc),
            ("النهائية", fv),
        ):
            ws.cell(row=r, column=2).value = lbl
            ws.cell(row=r, column=3).value = val
            ws.cell(row=r, column=3).number_format = _FMT_CURRENCY
            r += 1
        chart_last_row = r - 1

        try:
            from openpyxl.chart import BarChart, Reference
            chart              = BarChart()
            chart.type         = "col"
            chart.title        = "مقارنة أساليب التقييم"
            chart.y_axis.title = "القيمة (EGP)"
            chart.style        = 10
            chart.width        = 18
            chart.height       = 12
            data = Reference(ws, min_col=3, min_row=chart_data_row,
                             max_row=chart_last_row)
            cats = Reference(ws, min_col=2, min_row=chart_data_row + 1,
                             max_row=chart_last_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f"B{r + 1}")
        except Exception:
            pass  # chart is optional

        ws.freeze_panes = "B4"

    def sheet_sales_comparison(self) -> None:
        """'مقارنات البيوع' — USPAP/IVS Sales Comparison Adjustment Grid (legacy only)."""
        ws = self.workbook.create_sheet("مقارنات البيوع")
        ws.sheet_view.rightToLeft = True

        md = self.result.metadata if self.result else {}

        # ── Comparable data (cap at 5, pad to at least 3) ─────────────────────
        raw_comps = md.get("comparables") or md.get("comparable_sales") or []
        comps = list(raw_comps[:5])
        while len(comps) < 3:
            comps.append({})
        n = len(comps)

        # ── 9 Adjustment items ─────────────────────────────────────────────────
        _ADJ_ITEMS = [
            ("الموقع",   "adj_location",  "location"),
            ("المساحة",  "adj_area",      "area"),
            ("الدور",    "adj_floor",     "floor"),
            ("العمر",    "adj_age",       "age"),
            ("التشطيب",  "adj_condition", "condition"),
            ("الإطلالة", "adj_view",      "view"),
            ("التوقيت",  "adj_timing",    "timing"),
            ("الواجهة",  "adj_facade",    "facade"),
            ("الخدمات",  "adj_services",  "services"),
        ]
        N_ADJ = len(_ADJ_ITEMS)

        # ── Column layout: B=label, then (val, adj%) pairs per comparable ──────
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 26
        for i in range(n):
            ws.column_dimensions[get_column_letter(3 + 2 * i)].width = 17
            ws.column_dimensions[get_column_letter(4 + 2 * i)].width = 12
        last_data_col = 2 + 2 * n
        last_col_lt   = get_column_letter(last_data_col)

        # ── Row 1: Professional matrix banner ─────────────────────────────────
        ws.merge_cells(f"A1:{last_col_lt}1")
        t           = ws["A1"]
        t.value     = "مصفوفة الضبط الاحترافية"
        t.fill      = _FILL_HEADER
        t.font      = Font(bold=True, color=_Palette.WHITE, size=16)
        t.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 36

        # ── Row 2: English subtitle ────────────────────────────────────────────
        ws.merge_cells(f"A2:{last_col_lt}2")
        ws["A2"].value     = "Professional Sales Adjustment Grid — USPAP / IVS"
        ws["A2"].font      = Font(bold=True, color=_Palette.WHITE, size=11)
        ws["A2"].fill      = _FILL_INPUT_SECT
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # ── Row 3: Report info sub-banner ──────────────────────────────────────
        ws.merge_cells(f"A3:{last_col_lt}3")
        ws["A3"].value     = (
            f"تاريخ التقرير: {self.report_date}  |  "
            f"نوع الأصل: {getattr(self.result, 'asset_type', '—') if self.result else '—'}  |  "
            "المعيار: USPAP / IVS 105"
        )
        ws["A3"].font      = _FONT_MUTED
        ws["A3"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[3].height = 16

        # ── Row 4: Column headers ──────────────────────────────────────────────
        HDR_ROW = 4
        ws.row_dimensions[HDR_ROW].height = 22
        lhc           = ws.cell(row=HDR_ROW, column=2)
        lhc.value     = "بند الضبط"
        lhc.fill      = _FILL_INPUT_SECT
        lhc.font      = Font(bold=True, color=_Palette.WHITE, size=10)
        lhc.alignment = Alignment(horizontal="right", vertical="center")
        lhc.border    = _BORDER_THIN

        for i in range(n):
            comp_label = f"مقارن {i + 1}"
            for offset, suffix in ((0, " - الخاصية"), (1, " - الضبط %")):
                c           = ws.cell(row=HDR_ROW, column=3 + 2 * i + offset)
                c.value     = comp_label + suffix
                c.fill      = _FILL_INPUT_SECT
                c.font      = _FONT_COLHDR
                c.alignment = _ALIGN_CENTER
                c.border    = _BORDER_THIN

        # ── Rows 5-13: Adjustment items ────────────────────────────────────────
        ADJ_START = 5
        ADJ_END   = ADJ_START + N_ADJ - 1   # = 13

        for row_i, (ar_label, adj_key, val_key) in enumerate(_ADJ_ITEMS):
            r    = ADJ_START + row_i
            band = _FILL_ROW_BAND if row_i % 2 else _FILL_INPUT_CELL

            lc           = ws.cell(row=r, column=2)
            lc.value     = ar_label
            lc.font      = Font(bold=True, size=10)
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.fill      = band
            lc.border    = _BORDER_THIN
            ws.row_dimensions[r].height = 18

            for i, comp in enumerate(comps):
                adj_pct = 0.0
                try:
                    adj_pct = float(comp.get(adj_key, 0) or 0)
                except (TypeError, ValueError):
                    pass

                char_val = str(comp.get(val_key, "") or "")
                vcol     = 3 + 2 * i
                acol     = 4 + 2 * i

                vc           = ws.cell(row=r, column=vcol)
                vc.value     = char_val
                vc.alignment = _ALIGN_CENTER
                vc.fill      = _F_ADJ_ZERO
                vc.border    = _BORDER_THIN

                ac              = ws.cell(row=r, column=acol)
                ac.value        = adj_pct
                ac.number_format = _FMT_PCT
                ac.alignment    = _ALIGN_CENTER
                ac.border       = _BORDER_THIN
                ac.fill = (
                    _FILL_FINAL_VALUE  if adj_pct > 0 else
                    _F_ADJ_NEG  if adj_pct < 0 else
                    _F_ADJ_ZERO
                )

        # ── Row 14: thin spacer ────────────────────────────────────────────────
        ws.row_dimensions[14].height = 6

        # ── Row 15: إجمالي الضبط الصافي (SUM formula) ─────────────────────────
        NET_ROW = 15
        ws.row_dimensions[NET_ROW].height = 20
        nc           = ws.cell(row=NET_ROW, column=2)
        nc.value     = "إجمالي الضبط الصافي"
        nc.font      = _FONT_GOLD
        nc.fill      = _F_GOLD
        nc.alignment = Alignment(horizontal="right", vertical="center")
        nc.border    = _BORDER_THIN

        for i in range(n):
            vcol    = 3 + 2 * i
            acol    = 4 + 2 * i
            adj_let = get_column_letter(acol)

            vc        = ws.cell(row=NET_ROW, column=vcol)
            vc.fill   = _F_GOLD
            vc.border = _BORDER_THIN

            ac              = ws.cell(row=NET_ROW, column=acol)
            ac.value        = f"=SUM({adj_let}{ADJ_START}:{adj_let}{ADJ_END})"
            ac.number_format = _FMT_PCT
            ac.font         = _FONT_GOLD
            ac.fill         = _F_GOLD
            ac.alignment    = _ALIGN_CENTER
            ac.border       = _BORDER_THIN

        # ── Row 16: السعر الأصلي (EGP/م²) ────────────────────────────────────
        ORIG_ROW = 16
        ws.row_dimensions[ORIG_ROW].height = 18
        oc           = ws.cell(row=ORIG_ROW, column=2)
        oc.value     = "السعر الأصلي (EGP/م²)"
        oc.font      = Font(bold=True, size=10)
        oc.fill      = _FILL_SUBHEAD
        oc.alignment = Alignment(horizontal="right", vertical="center")
        oc.border    = _BORDER_THIN

        for i, comp in enumerate(comps):
            vcol      = 3 + 2 * i
            acol      = 4 + 2 * i
            price_sqm = 0.0
            try:
                price_sqm = float(
                    comp.get("price_per_sqm") or comp.get("price_sqm") or
                    comp.get("unit_price") or 0
                )
                if not price_sqm:
                    total     = float(comp.get("price") or comp.get("sale_price") or
                                      comp.get("value") or 0)
                    area      = float(comp.get("area") or comp.get("floor_area_m2") or 0)
                    price_sqm = (total / area) if area > 0 else 0.0
            except (TypeError, ValueError):
                pass

            vc              = ws.cell(row=ORIG_ROW, column=vcol)
            vc.value        = price_sqm
            vc.number_format = _FMT_CURRENCY
            vc.alignment    = _ALIGN_CENTER
            vc.fill         = _FILL_SUBHEAD
            vc.border       = _BORDER_THIN

            ac        = ws.cell(row=ORIG_ROW, column=acol)
            ac.fill   = _FILL_SUBHEAD
            ac.border = _BORDER_THIN

        # ── Row 17: السعر بعد الضبط (formula) ────────────────────────────────
        ADJ_PRICE_ROW = 17
        ws.row_dimensions[ADJ_PRICE_ROW].height = 20
        apc           = ws.cell(row=ADJ_PRICE_ROW, column=2)
        apc.value     = "السعر بعد الضبط (EGP/م²)"
        apc.font      = _FONT_EMERALD
        apc.fill      = _F_EMERALD
        apc.alignment = Alignment(horizontal="right", vertical="center")
        apc.border    = _BORDER_THIN

        for i in range(n):
            vcol    = 3 + 2 * i
            acol    = 4 + 2 * i
            val_let = get_column_letter(vcol)
            adj_let = get_column_letter(acol)

            vc              = ws.cell(row=ADJ_PRICE_ROW, column=vcol)
            vc.value        = f"={val_let}{ORIG_ROW}*(1+{adj_let}{NET_ROW})"
            vc.number_format = _FMT_CURRENCY
            vc.font         = _FONT_EMERALD
            vc.fill         = _F_EMERALD
            vc.alignment    = _ALIGN_CENTER
            vc.border       = _BORDER_THIN

            ac        = ws.cell(row=ADJ_PRICE_ROW, column=acol)
            ac.fill   = _F_EMERALD
            ac.border = _BORDER_THIN

        # ── Row 18: وزن المقارن (Match Quality) ──────────────────────────────
        WEIGHT_ROW = 18
        ws.row_dimensions[WEIGHT_ROW].height = 18
        wc           = ws.cell(row=WEIGHT_ROW, column=2)
        wc.value     = "وزن المقارن (Match Quality)"
        wc.font      = Font(bold=True, size=10)
        wc.fill      = _FILL_ROW_BAND
        wc.alignment = Alignment(horizontal="right", vertical="center")
        wc.border    = _BORDER_THIN

        default_wt = 1.0 / n if n else 1.0
        for i, comp in enumerate(comps):
            vcol = 3 + 2 * i
            acol = 4 + 2 * i
            wt   = default_wt
            try:
                raw_wt = comp.get("weight") or comp.get("match_quality")
                if raw_wt is not None:
                    wt = float(raw_wt)
            except (TypeError, ValueError):
                pass

            vc              = ws.cell(row=WEIGHT_ROW, column=vcol)
            vc.value        = wt
            vc.number_format = _FMT_PCT
            vc.alignment    = _ALIGN_CENTER
            vc.fill         = _FILL_ROW_BAND
            vc.border       = _BORDER_THIN

            ac        = ws.cell(row=WEIGHT_ROW, column=acol)
            ac.fill   = _FILL_ROW_BAND
            ac.border = _BORDER_THIN

        # ── Row 19: spacer ────────────────────────────────────────────────────
        ws.row_dimensions[19].height = 6

        # ── Row 20: السعر النهائي الموزون (weighted average formula) ──────────
        FINAL_ROW = 20
        ws.row_dimensions[FINAL_ROW].height = 26
        ws.merge_cells(f"A{FINAL_ROW}:B{FINAL_ROW}")
        fc_lbl           = ws.cell(row=FINAL_ROW, column=1)
        fc_lbl.value     = "السعر النهائي الموزون (EGP/م²)"
        fc_lbl.font      = _FONT_FINAL
        fc_lbl.fill      = _F_GOLD
        fc_lbl.alignment = Alignment(horizontal="right", vertical="center")
        fc_lbl.border    = _BORDER_MEDIUM

        # Formula: =SUMPRODUCT(adj_prices, weights) / SUM(weights)
        # Intermediate (adjustment %) columns are empty so their zero values
        # are harmless inside the contiguous range used by SUMPRODUCT / SUM.
        final_formula = (
            f"=SUMPRODUCT(C{ADJ_PRICE_ROW}:{last_col_lt}{ADJ_PRICE_ROW},"
            f"C{WEIGHT_ROW}:{last_col_lt}{WEIGHT_ROW})/"
            f"SUM(C{WEIGHT_ROW}:{last_col_lt}{WEIGHT_ROW})"
        )

        fc_val              = ws.cell(row=FINAL_ROW, column=3)
        fc_val.value        = final_formula
        fc_val.number_format = _FMT_CURRENCY
        fc_val.font         = _FONT_FINAL
        fc_val.fill         = _F_GOLD
        fc_val.alignment    = _ALIGN_CENTER
        fc_val.border       = _BORDER_MEDIUM
        for col in range(4, last_data_col + 1):
            c        = ws.cell(row=FINAL_ROW, column=col)
            c.fill   = _F_GOLD
            c.border = _BORDER_THIN

        # ── Legend (row 22+) ──────────────────────────────────────────────────
        LEGEND_ROW = 22
        ws.row_dimensions[LEGEND_ROW].height = 18
        ws.merge_cells(f"B{LEGEND_ROW}:E{LEGEND_ROW}")
        lhdr           = ws.cell(row=LEGEND_ROW, column=2)
        lhdr.value     = "دليل الألوان — Color Legend"
        lhdr.fill      = _FILL_INPUT_SECT
        lhdr.font      = Font(bold=True, color=_Palette.WHITE, size=10)
        lhdr.alignment = _ALIGN_CENTER

        legend_items = [
            (_FILL_FINAL_VALUE,  "ضبط موجب — المقارن أدنى من العقار (يُضاف للسعر)"),
            (_F_ADJ_NEG,  "ضبط سالب — المقارن أفضل من العقار (يُخصم من السعر)"),
            (_F_ADJ_ZERO, "لا يوجد ضبط — المقارن مماثل للعقار"),
            (_F_GOLD,     "صفوف الملخص (إجمالي الضبط / السعر النهائي الموزون)"),
            (_F_EMERALD,  "السعر بعد تطبيق الضبط (Adjusted Price)"),
        ]
        for j, (fill, text) in enumerate(legend_items):
            lr = LEGEND_ROW + 1 + j
            ws.row_dimensions[lr].height = 16
            ic        = ws.cell(row=lr, column=2)
            ic.fill   = fill
            ic.border = _BORDER_THIN
            ws.merge_cells(f"C{lr}:E{lr}")
            tc           = ws.cell(row=lr, column=3)
            tc.value     = text
            tc.font      = Font(size=9)
            tc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

        # ── Methodology block ─────────────────────────────────────────────────
        METH_ROW = LEGEND_ROW + len(legend_items) + 2
        ws.row_dimensions[METH_ROW].height = 20
        ws.merge_cells(f"B{METH_ROW}:{last_col_lt}{METH_ROW}")
        mhdr           = ws.cell(row=METH_ROW, column=2)
        mhdr.value     = "منهجية التقييم بأسلوب المقارنة البيعية — Methodology"
        mhdr.fill      = _FILL_INPUT_SECT
        mhdr.font      = Font(bold=True, color=_Palette.WHITE, size=11)
        mhdr.alignment = Alignment(horizontal="right", vertical="center")

        steps = [
            "١. اختيار المقارنات الأقرب للعقار محل التقييم من حيث الموقع والنوع والحجم.",
            "٢. مقارنة خصائص كل مقارن مع العقار محل التقييم للكشف عن أوجه الاختلاف.",
            "٣. تحديد فروق الموقع والمساحة والعمر والتشطيب وباقي البنود.",
            "٤. تطبيق نسب الضبط الموجبة (المقارن أدنى) أو السالبة (المقارن أفضل).",
            "٥. حساب السعر بعد الضبط = السعر الأصلي × (١ + إجمالي الضبط الصافي).",
            "٦. حساب السعر النهائي المرجح بناءً على جودة المقارنات (وزن كل مقارن).",
        ]
        for k, step in enumerate(steps):
            sr = METH_ROW + 1 + k
            ws.row_dimensions[sr].height = 16
            ws.merge_cells(f"B{sr}:{last_col_lt}{sr}")
            sc           = ws.cell(row=sr, column=2)
            sc.value     = step
            sc.font      = Font(size=9)
            sc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            if k % 2:
                sc.fill = _FILL_ROW_BAND

        ws.freeze_panes = f"C{ADJ_START}"

    # ── Remaining legacy Arabic sheets ────────────────────────────────────────

    def sheet_rental_comparables(self) -> None:
        """'المقارنات الإيجارية' — Rental Comparables (legacy only)."""
        ws = self.workbook.create_sheet("المقارنات الإيجارية")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20

        md     = self.result.metadata if self.result else {}
        r_comp = list((md.get("rental_comparables") or md.get("rent_comps") or [])[:3])
        while len(r_comp) < 3:
            r_comp.append({})

        ws.merge_cells("A1:F1")
        b           = ws["A1"]
        b.value     = "المقارنات الإيجارية — Rental Comparables"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:F2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}  |  المعيار: IVS 105"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4
        ws.row_dimensions[r].height = 22
        for ci, hdr in enumerate(("البند", "مقارن 1", "مقارن 2", "مقارن 3", "العقار"), 2):
            c           = ws.cell(row=r, column=ci)
            c.value     = hdr
            c.fill      = _FILL_INPUT_SECT
            c.font      = Font(bold=True, color=_Palette.WHITE, size=10)
            c.alignment = _ALIGN_CENTER
            c.border    = _BORDER_THIN
        r += 1

        _ROWS = [
            ("الموقع",             "location",    md.get("location") or "—"),
            ("المساحة (م²)",      "area",         md.get("area") or "—"),
            ("الطابق",             "floor",        "—"),
            ("العمر (سنة)",       "age",          md.get("year_built") or "—"),
            ("الحالة",             "condition",    "—"),
            ("الإيجار (EGP/م²)", "rent_sqm",     "—"),
            ("نسبة الشاغر",       "vacancy_rate", "—"),
            ("تاريخ الإيجار",     "lease_date",   self.report_date),
        ]
        for ri, (lbl, key, subj) in enumerate(_ROWS):
            ws.row_dimensions[r].height = 18
            fill = _FILL_ROW_BAND if ri % 2 else _FILL_INPUT_CELL
            lc           = ws.cell(row=r, column=2)
            lc.value     = lbl
            lc.font      = Font(bold=True, size=10)
            lc.fill      = fill
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border    = _BORDER_THIN
            for ci, comp in enumerate(r_comp, 3):
                v           = str(comp.get(key, "") or "—")
                c           = ws.cell(row=r, column=ci)
                c.value     = v
                c.alignment = _ALIGN_CENTER
                c.fill      = fill
                c.border    = _BORDER_THIN
            sc           = ws.cell(row=r, column=6)
            sc.value     = str(subj)
            sc.font      = Font(bold=True)
            sc.fill      = _FILL_FINAL_VALUE
            sc.alignment = _ALIGN_CENTER
            sc.border    = _BORDER_THIN
            r += 1

        ws.freeze_panes = "C5"

    def sheet_cost_approach(self) -> None:
        """'طريقة التكلفة' — Cost Approach (legacy only)."""
        ws = self.workbook.create_sheet("طريقة التكلفة")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 16

        md         = self.result.metadata if self.result else {}
        area       = float(md.get("area") or md.get("floor_area_m2") or 0)
        land_val   = float(md.get("land_value") or md.get("land_price") or 0)
        cost_sqm   = float(md.get("construction_cost_sqm") or 8000)
        const_cost = float(md.get("construction_cost") or md.get("replacement_cost") or (area * cost_sqm))
        depr_pct   = float(md.get("depreciation_rate") or md.get("depreciation") or 0)
        add_items  = float(md.get("additional_items") or 0)
        depr_amt   = const_cost * depr_pct
        indicated  = land_val + const_cost + add_items - depr_amt

        ws.merge_cells("A1:D1")
        b           = ws["A1"]
        b.value     = "طريقة التكلفة — Cost Approach"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:D2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}  |  المعيار: EGVS 3.2 / IVS 105"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4

        def _sect(label: str) -> None:
            nonlocal r
            ws.merge_cells(f"A{r}:D{r}")
            c           = ws.cell(row=r, column=1)
            c.value     = f"  {label}"
            c.fill      = _FILL_INPUT_SECT
            c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
            c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

        def _row(label: str, value, fmt: str = _FMT_CURRENCY,
                 is_calc: bool = False, alt: bool = False) -> None:
            nonlocal r
            fill         = _FILL_CALC_CELL if is_calc else (_FILL_ROW_BAND if alt else _FILL_INPUT_CELL)
            lc           = ws.cell(row=r, column=2)
            lc.value     = label
            lc.font      = Font(bold=True, size=10)
            lc.fill      = fill
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border    = _BORDER_THIN
            vc           = ws.cell(row=r, column=3)
            vc.value     = value
            vc.number_format = fmt
            vc.alignment = _ALIGN_CENTER
            vc.fill      = fill
            vc.border    = _BORDER_THIN
            ws.row_dimensions[r].height = 18
            r += 1

        _sect("قيمة الأرض")
        _row("مساحة الأرض (م²)", area, '#,##0.00 "م²"')
        _row("قيمة الأرض الإجمالية (EGP)", land_val)
        r += 1

        _sect("تكلفة الإنشاء والبناء")
        _row("المساحة المبنية (م²)", area, '#,##0.00 "م²"')
        _row("تكلفة البناء للمتر (EGP/م²)", cost_sqm, alt=True)
        _row("إجمالي تكلفة البناء (EGP)", const_cost, is_calc=True)
        _row("بنود إضافية (EGP)", add_items, alt=True)
        r += 1

        _sect("الاستهلاك والتقادم")
        _row("نسبة الاستهلاك", depr_pct, _FMT_PCT)
        _row("قيمة الاستهلاك (EGP)", depr_amt, is_calc=True)
        r += 1

        _sect("القيمة الاستدلالية بطريقة التكلفة")
        _row("المباني بعد الاستهلاك (EGP)", const_cost + add_items - depr_amt, is_calc=True)
        ws.merge_cells(f"B{r}:D{r}")
        fc           = ws.cell(row=r, column=2)
        fc.value     = f"القيمة الاستدلالية الإجمالية (EGP):  {indicated:,.0f}"
        fc.font      = _FONT_FINAL_VALUE
        fc.fill      = _FILL_FINAL_VALUE
        fc.alignment = Alignment(horizontal="center", vertical="center")
        fc.border    = _BORDER_MEDIUM
        ws.row_dimensions[r].height = 24

        ws.freeze_panes = "B3"

    def sheet_income_capitalization(self) -> None:
        """'رأسمالة الدخل' — Income Capitalization Approach (legacy only)."""
        ws = self.workbook.create_sheet("رأسمالة الدخل")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 16

        md         = self.result.metadata if self.result else {}
        area       = float(md.get("area") or md.get("floor_area_m2") or 0)
        rent_sqm   = float(md.get("rent_sqm") or md.get("rent_per_sqm") or 0)
        vacancy    = float(md.get("vacancy_rate") or 0.05)
        mgmt_pct   = float(md.get("management_rate") or 0.05)
        maint_pct  = float(md.get("maintenance_rate") or 0.02)
        tax_pct    = float(md.get("tax_rate") or 0.01)
        cap_rate   = float(md.get("cap_rate") or md.get("capitalization_rate") or 0.08)

        gross_income = float(md.get("gross_income") or (area * rent_sqm * 12 if area and rent_sqm else 0))
        vac_loss     = gross_income * vacancy
        egi          = gross_income - vac_loss
        total_exp    = egi * (mgmt_pct + maint_pct + tax_pct)
        noi          = egi - total_exp
        indicated    = noi / cap_rate if cap_rate else 0

        ws.merge_cells("A1:D1")
        b           = ws["A1"]
        b.value     = "رأسمالة الدخل — Income Capitalization"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:D2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}  |  المعيار: EGVS 3.3 / IVS 105"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4

        def _sect(label: str) -> None:
            nonlocal r
            ws.merge_cells(f"A{r}:D{r}")
            c           = ws.cell(row=r, column=1)
            c.value     = f"  {label}"
            c.fill      = _FILL_INPUT_SECT
            c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
            c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

        def _row(label: str, value, fmt: str = _FMT_CURRENCY,
                 is_calc: bool = False, alt: bool = False) -> None:
            nonlocal r
            fill         = _FILL_CALC_CELL if is_calc else (_FILL_ROW_BAND if alt else _FILL_INPUT_CELL)
            lc           = ws.cell(row=r, column=2)
            lc.value     = label
            lc.font      = Font(bold=True, size=10)
            lc.fill      = fill
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border    = _BORDER_THIN
            vc           = ws.cell(row=r, column=3)
            vc.value     = value
            vc.number_format = fmt
            vc.alignment = _ALIGN_CENTER
            vc.fill      = fill
            vc.border    = _BORDER_THIN
            ws.row_dimensions[r].height = 18
            r += 1

        _sect("بيانات الدخل الإجمالي")
        _row("المساحة القابلة للإيجار (م²)", area, '#,##0.00 "م²"')
        _row("معدل الإيجار (EGP/م²/شهر)", rent_sqm, alt=True)
        _row("الدخل الإجمالي السنوي (EGP)", gross_income, is_calc=True)
        r += 1

        _sect("الشاغر والخسائر")
        _row("نسبة الشاغر", vacancy, _FMT_PCT)
        _row("خسارة الشاغر (EGP)", vac_loss, is_calc=True, alt=True)
        _row("الدخل الفعلي الإجمالي — EGI (EGP)", egi, is_calc=True)
        r += 1

        _sect("المصروفات التشغيلية والنتيجة")
        _row("إجمالي المصروفات (EGP)", total_exp, is_calc=True)
        _row("صافي الدخل التشغيلي — NOI (EGP)", noi, is_calc=True, alt=True)
        _row("معدل الرسملة", cap_rate, _FMT_PCT)
        ws.merge_cells(f"B{r}:D{r}")
        fc           = ws.cell(row=r, column=2)
        fc.value     = f"القيمة الاستدلالية (EGP):  {indicated:,.0f}"
        fc.font      = _FONT_FINAL_VALUE
        fc.fill      = _FILL_FINAL_VALUE
        fc.alignment = Alignment(horizontal="center", vertical="center")
        fc.border    = _BORDER_MEDIUM
        ws.row_dimensions[r].height = 24

        ws.freeze_panes = "B3"

    def sheet_reconciliation(self) -> None:
        """'توفيق النتائج' — Reconciliation of Value Indications (legacy only)."""
        ws = self.workbook.create_sheet("توفيق النتائج")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 24

        md   = self.result.metadata if self.result else {}
        fv   = float(self.result.primary_value) if self.result and self.result.primary_value else 0
        comp = float(md.get("comparable") or 0)
        cost = float(md.get("cost") or 0)
        inc  = float(md.get("income") or 0)
        w    = self.result.weights_applied if self.result else {}

        ws.merge_cells("A1:E1")
        b           = ws["A1"]
        b.value     = "توفيق النتائج — Reconciliation of Value Indications"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:E2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}  |  المعيار: EGVS 3.0 / IVS 105"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4
        ws.row_dimensions[r].height = 22
        for ci, hdr in enumerate(
            ("أسلوب التقييم", "القيمة الاستدلالية (EGP)", "الوزن", "القيمة الموزونة (EGP)"), 2
        ):
            c           = ws.cell(row=r, column=ci)
            c.value     = hdr
            c.fill      = _FILL_INPUT_SECT
            c.font      = Font(bold=True, color=_Palette.WHITE, size=10)
            c.alignment = _ALIGN_CENTER
            c.border    = _BORDER_THIN
        r += 1

        for i, (lbl, val, wt_key) in enumerate((
            ("أسلوب المقارنة البيعية", comp, "comparable"),
            ("أسلوب التكلفة",          cost, "cost"),
            ("رأسمالة الدخل",          inc,  "income"),
        )):
            wt   = w.get(wt_key, 0)
            fill = _FILL_ROW_BAND if i % 2 else _FILL_INPUT_CELL
            ws.row_dimensions[r].height = 18
            lc           = ws.cell(row=r, column=2)
            lc.value     = lbl
            lc.font      = Font(bold=True, size=10)
            lc.fill      = fill
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border    = _BORDER_THIN
            for ci, (v, fmt) in enumerate(((val, _FMT_CURRENCY), (wt, _FMT_PCT), (val * wt, _FMT_CURRENCY)), 3):
                c                = ws.cell(row=r, column=ci)
                c.value          = v
                c.number_format  = fmt
                c.alignment      = _ALIGN_CENTER
                c.fill           = fill
                c.border         = _BORDER_THIN
            r += 1

        ws.row_dimensions[r].height = 26
        for ci, (v, fmt) in enumerate((
            ("القيمة التوفيقية النهائية", None),
            (fv, _FMT_CURRENCY),
            ("—", None),
            (fv, _FMT_CURRENCY),
        ), 2):
            c           = ws.cell(row=r, column=ci)
            c.value     = v
            c.font      = _FONT_FINAL_VALUE
            c.fill      = _FILL_FINAL_VALUE
            c.alignment = _ALIGN_CENTER
            c.border    = _BORDER_MEDIUM
            if fmt:
                c.number_format = fmt
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right", vertical="center")
        r += 2

        ws.merge_cells(f"B{r}:E{r}")
        nh           = ws.cell(row=r, column=2)
        nh.value     = "  ملاحظات التوفيق"
        nh.fill      = _FILL_INPUT_SECT
        nh.font      = Font(bold=True, color=_Palette.WHITE, size=11)
        nh.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1
        for note in (
            "• تم إعطاء الوزن الأكبر لأسلوب المقارنة البيعية لتوافر بيانات السوق.",
            "• تم دعم النتيجة بأسلوب التكلفة والدخل لتأكيد القيمة.",
            "• تتوافق نتائج الأساليب الثلاثة مع مؤشرات السوق الحالية.",
        ):
            ws.merge_cells(f"B{r}:E{r}")
            nc           = ws.cell(row=r, column=2)
            nc.value     = note
            nc.font      = Font(size=9)
            nc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            ws.row_dimensions[r].height = 16
            r += 1

        ws.freeze_panes = "B3"

    def sheet_valuation_parameters(self) -> None:
        """'محددات التقييم' — Valuation Parameters & Limiting Conditions (legacy only)."""
        ws = self.workbook.create_sheet("محددات التقييم")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 36

        md = self.result.metadata if self.result else {}

        ws.merge_cells("A1:C1")
        b           = ws["A1"]
        b.value     = "محددات التقييم — Valuation Parameters & Limiting Conditions"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:C2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}  |  المعيار: EGVS 2.0"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4

        def _sect(label: str) -> None:
            nonlocal r
            ws.merge_cells(f"A{r}:C{r}")
            c           = ws.cell(row=r, column=1)
            c.value     = f"  {label}"
            c.fill      = _FILL_INPUT_SECT
            c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
            c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

        def _param(label: str, value) -> None:
            nonlocal r
            fill         = _FILL_ROW_BAND if r % 2 else _FILL_INPUT_CELL
            lc           = ws.cell(row=r, column=2)
            lc.value     = label
            lc.font      = Font(bold=True, size=10)
            lc.fill      = fill
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border    = _BORDER_THIN
            vc           = ws.cell(row=r, column=3)
            vc.value     = str(value) if value is not None else "—"
            vc.fill      = fill
            vc.border    = _BORDER_THIN
            vc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            ws.row_dimensions[r].height = 18
            r += 1

        def _cond(text: str) -> None:
            nonlocal r
            fill = _FILL_ROW_BAND if r % 2 else None
            ws.merge_cells(f"B{r}:C{r}")
            c           = ws.cell(row=r, column=2)
            c.value     = f"• {text}"
            c.font      = Font(size=9)
            c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            if fill:
                c.fill = fill
            ws.row_dimensions[r].height = 16
            r += 1

        _sect("الافتراضات العامة")
        _param("تاريخ التقييم",  md.get("valuation_date") or self.report_date)
        _param("غرض التقييم",    self.result.primary_purpose if self.result else "—")
        _param("أساس القيمة",    "القيمة السوقية — Market Value")
        _param("نوع الأصل",      getattr(self.result, "asset_type", "—") if self.result else "—")
        _param("حق الانتفاع",    md.get("tenure") or "ملكية حرة — Freehold")
        _param("عملة التقرير",   "جنيه مصري — EGP")
        r += 1

        _sect("الافتراضات الخاصة")
        _param("حالة العقار",          md.get("condition") or "جيدة")
        _param("مصدر بيانات المساحة",  md.get("area_source") or "وثائق الملكية")
        _param("بيانات السوق",         md.get("market_data_source") or "مصادر السوق المحلية")
        r += 1

        _sect("محددات الاستخدام")
        for cond in (
            "يُعدّ هذا التقرير سارياً فقط في تاريخ التقييم المُبيَّن.",
            "لا يجوز الاستشهاد بجزء من التقرير دون الرجوع إلى النص الكامل.",
            "لا يُؤخذ في الاعتبار أي عبء مالي غير ظاهر في المستندات الرسمية.",
            "القيم المُحددة خاضعة لأحوال السوق في تاريخ التقييم.",
            "لم يُراعَ في التقدير أي تكاليف بيع أو ضرائب نقل ملكية.",
        ):
            _cond(cond)

        ws.freeze_panes = "B3"

    def sheet_certification_ar(self) -> None:
        """'شهادة' — Arabic Certification (legacy only)."""
        ws = self.workbook.create_sheet("شهادة")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 70

        md = self.result.metadata if self.result else {}

        ws.merge_cells("A1:B1")
        b           = ws["A1"]
        b.value     = "شهادة المقيم — Appraiser Certification"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:B2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4
        cert_ar = (
            "أشهد أنا الموقع أدناه بأن البيانات الواردة في هذا التقرير صحيحة وصادقة "
            "وفق أفضل ما لديَّ من معرفة ومعلومات، وأن الاستنتاجات المُبيَّنة فيه تمثل "
            "رأيي المهني المستقل والمحايد، وقد توصلتُ إليها وفق المعايير المهنية المعتمدة.\n\n"
            "أُقرّ بأنه لا توجد لديَّ أي مصلحة حالية أو مستقبلية في العقار موضوع التقرير.\n\n"
            "تم إعداد هذا التقرير وفقاً للمعايير المصرية للتقييم (EgVS) "
            "والمعيار الدولي للتقارير المالية (IFRS 13)."
        )
        ws.merge_cells(f"B{r}:B{r + 5}")
        cc           = ws.cell(row=r, column=2)
        cc.value     = cert_ar
        cc.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        cc.fill      = _FILL_INPUT_CELL
        cc.border    = _BORDER_THIN
        ws.row_dimensions[r].height = 120
        r += 7

        for lbl, val in (
            ("اسم المقيم:",    md.get("appraiser_name") or "________________________________"),
            ("رقم الترخيص:", md.get("license_no")      or "________________________________"),
            ("التاريخ:",       self.report_date),
            ("التوقيع:",       "________________________________"),
        ):
            ws.row_dimensions[r].height = 22
            c           = ws.cell(row=r, column=2)
            c.value     = f"{lbl}  {val}"
            c.font      = Font(bold=True, size=11)
            c.alignment = Alignment(horizontal="right", vertical="center")
            r += 2

    def sheet_data_sources(self) -> None:
        """'مصادر البيانات والمنهجية' — Data Sources & Methodology (legacy only)."""
        ws = self.workbook.create_sheet("مصادر البيانات والمنهجية")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 46

        md = self.result.metadata if self.result else {}

        ws.merge_cells("A1:C1")
        b           = ws["A1"]
        b.value     = "مصادر البيانات والمنهجية — Data Sources & Methodology"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:C2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4

        def _sect(label: str) -> None:
            nonlocal r
            ws.merge_cells(f"A{r}:C{r}")
            c           = ws.cell(row=r, column=1)
            c.value     = f"  {label}"
            c.fill      = _FILL_INPUT_SECT
            c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
            c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

        def _src(name: str, desc: str) -> None:
            nonlocal r
            fill         = _FILL_ROW_BAND if r % 2 else _FILL_INPUT_CELL
            lc           = ws.cell(row=r, column=2)
            lc.value     = name
            lc.font      = Font(bold=True, size=10)
            lc.fill      = fill
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border    = _BORDER_THIN
            vc           = ws.cell(row=r, column=3)
            vc.value     = desc
            vc.fill      = fill
            vc.border    = _BORDER_THIN
            vc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            ws.row_dimensions[r].height = 18
            r += 1

        _sect("مصادر البيانات")
        _src("بيانات السوق",          "صفقات البيع والإيجار المسجلة من مصادر محلية معتمدة")
        _src("المستندات الرسمية",     "مستندات الملكية ووثائق المساحة الرسمية")
        _src("المعاينة الميدانية",    f"تاريخ المعاينة: {md.get('inspection_date') or self.report_date}")
        _src("الصور الجوية والخرائط", "خرائط الموقع والصور الجوية المحدّثة")
        _src("قواعد البيانات",        md.get("data_source") or "قاعدة بيانات Expert Smart EgVS")
        r += 1

        _sect("المنهجية المتبعة")
        for step in (
            "تحليل السوق وتحديد العوامل المؤثرة في القيمة.",
            "اختيار المقارنات الأكثر ملاءمة وتطبيق الضبط.",
            "تطبيق الأساليب الثلاثة: المقارنة، التكلفة، الدخل.",
            "توفيق النتائج وإعطاء الأوزان المناسبة لكل أسلوب.",
            "إعداد التقرير وفق معايير EgVS / IVS.",
        ):
            fill = _FILL_ROW_BAND if r % 2 else None
            ws.merge_cells(f"B{r}:C{r}")
            c           = ws.cell(row=r, column=2)
            c.value     = f"• {step}"
            c.font      = Font(size=9)
            c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            if fill:
                c.fill = fill
            ws.row_dimensions[r].height = 16
            r += 1
        r += 1

        _sect("المعايير المطبقة")
        for name, desc in (
            ("EgVS",   "المعايير المصرية للتقييم العقاري"),
            ("IVS 2022", "معايير التقييم الدولية"),
            ("IFRS 13",  "المعيار الدولي للتقارير المالية — القيمة العادلة"),
            ("USPAP",    "مبادئ التقييم المقبولة عموماً"),
        ):
            _src(name, desc)

        ws.freeze_panes = "B3"

    def sheet_dcf(self) -> None:
        """'DCF — التدفقات النقدية' — Discounted Cash Flow (legacy only)."""
        ws = self.workbook.create_sheet("DCF — التدفقات النقدية")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 22

        md          = self.result.metadata if self.result else {}
        noi         = float(md.get("noi") or md.get("net_operating_income") or 0)
        growth_r    = float(md.get("income_growth_rate") or md.get("growth_rate") or 0.03)
        discount_r  = float(md.get("discount_rate") or 0.10)
        hold_yrs    = int(md.get("holding_period") or 10)
        terminal_cr = float(md.get("terminal_cap_rate") or md.get("exit_cap_rate") or 0.09)

        ws.merge_cells("A1:E1")
        b           = ws["A1"]
        b.value     = "DCF — التدفقات النقدية المخصومة"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:E2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}  |  المعيار: IVS 105"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4

        def _sect(label: str) -> None:
            nonlocal r
            ws.merge_cells(f"A{r}:E{r}")
            c           = ws.cell(row=r, column=1)
            c.value     = f"  {label}"
            c.fill      = _FILL_INPUT_SECT
            c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
            c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

        def _param(label: str, value, fmt: str = "") -> None:
            nonlocal r
            fill         = _FILL_ROW_BAND if r % 2 else _FILL_INPUT_CELL
            lc           = ws.cell(row=r, column=2)
            lc.value     = label
            lc.font      = Font(bold=True, size=10)
            lc.fill      = fill
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border    = _BORDER_THIN
            vc           = ws.cell(row=r, column=3)
            vc.value     = value
            if fmt:
                vc.number_format = fmt
            vc.alignment = _ALIGN_CENTER
            vc.fill      = fill
            vc.border    = _BORDER_THIN
            ws.row_dimensions[r].height = 18
            r += 1

        _sect("افتراضات التحليل")
        _param("صافي الدخل التشغيلي الأولي (NOI)", noi, _FMT_CURRENCY)
        _param("معدل نمو الدخل السنوي", growth_r, _FMT_PCT)
        _param("معدل الخصم (Discount Rate)", discount_r, _FMT_PCT)
        _param("فترة الاحتجاز (سنوات)", hold_yrs)
        _param("معدل الرسملة عند الخروج", terminal_cr, _FMT_PCT)
        r += 1

        _sect("جدول التدفقات النقدية المخصومة")
        ws.row_dimensions[r].height = 22
        for ci, hdr in enumerate(("السنة", "NOI (EGP)", "معامل الخصم", "القيمة الحالية (EGP)"), 2):
            c           = ws.cell(row=r, column=ci)
            c.value     = hdr
            c.fill      = _FILL_SUBHEAD
            c.font      = Font(bold=True, size=10)
            c.alignment = _ALIGN_CENTER
            c.border    = _BORDER_THIN
        r += 1

        total_pv = 0.0
        for yr in range(1, hold_yrs + 1):
            yr_noi = noi * ((1 + growth_r) ** yr)
            factor = 1 / ((1 + discount_r) ** yr)
            pv     = yr_noi * factor
            total_pv += pv
            fill = _FILL_ROW_BAND if yr % 2 else _FILL_INPUT_CELL
            ws.row_dimensions[r].height = 17
            for ci, (v, fmt) in enumerate((
                (yr,     ""),
                (yr_noi, _FMT_CURRENCY),
                (factor, "0.000000"),
                (pv,     _FMT_CURRENCY),
            ), 2):
                c           = ws.cell(row=r, column=ci)
                c.value     = v
                if fmt:
                    c.number_format = fmt
                c.alignment = _ALIGN_CENTER
                c.fill      = fill
                c.border    = _BORDER_THIN
            r += 1

        terminal_noi  = noi * ((1 + growth_r) ** (hold_yrs + 1))
        terminal_val  = terminal_noi / terminal_cr if terminal_cr else 0
        pv_terminal   = terminal_val / ((1 + discount_r) ** hold_yrs)
        total_pv     += pv_terminal

        ws.row_dimensions[r].height = 18
        for ci, (v, fmt) in enumerate((
            ("القيمة النهائية", ""),
            (terminal_val, _FMT_CURRENCY),
            ("—", ""),
            (pv_terminal, _FMT_CURRENCY),
        ), 2):
            c           = ws.cell(row=r, column=ci)
            c.value     = v
            if fmt:
                c.number_format = fmt
            c.font      = Font(bold=True)
            c.fill      = _FILL_SUBHEAD
            c.alignment = _ALIGN_CENTER
            c.border    = _BORDER_THIN
        r += 2

        ws.merge_cells(f"B{r}:D{r}")
        fc           = ws.cell(row=r, column=2)
        fc.value     = f"القيمة الحالية الصافية — NPV (EGP):  {total_pv:,.0f}"
        fc.font      = _FONT_FINAL_VALUE
        fc.fill      = _FILL_FINAL_VALUE
        fc.alignment = Alignment(horizontal="center", vertical="center")
        fc.border    = _BORDER_MEDIUM
        ws.row_dimensions[r].height = 24

        ws.freeze_panes = "B3"

    def sheet_rent_vs_buy(self) -> None:
        """'الإيجار مقابل الشراء' — Rent vs Buy Analysis (legacy only)."""
        ws = self.workbook.create_sheet("الإيجار مقابل الشراء")
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 24

        md           = self.result.metadata if self.result else {}
        fv           = float(self.result.primary_value) if self.result and self.result.primary_value else 0
        area         = float(md.get("area") or md.get("floor_area_m2") or 0)
        rent_monthly = float(md.get("rent_monthly") or
                             (float(md.get("rent_sqm", 0) or 0) * area) or 0)
        dp_pct       = float(md.get("down_payment_pct") or 0.20)
        mort_r       = float(md.get("mortgage_rate") or 0.12)
        years        = int(md.get("mortgage_years") or 20)
        down_pmt     = fv * dp_pct
        loan         = fv - down_pmt
        monthly_r    = mort_r / 12
        n_months     = years * 12
        mort_pmt     = (
            loan * monthly_r / (1 - (1 + monthly_r) ** (-n_months))
            if monthly_r > 0 and n_months > 0 else
            (loan / n_months if n_months else 0)
        )

        ws.merge_cells("A1:D1")
        b           = ws["A1"]
        b.value     = "الإيجار مقابل الشراء — Rent vs Buy Analysis"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:D2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4
        ws.row_dimensions[r].height = 22
        for ci, hdr in enumerate(("البند", "الشراء", "الإيجار"), 2):
            c           = ws.cell(row=r, column=ci)
            c.value     = hdr
            c.fill      = _FILL_INPUT_SECT
            c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
            c.alignment = _ALIGN_CENTER
            c.border    = _BORDER_THIN
        r += 1

        _ROWS = [
            ("سعر الشراء / الإيجار الشهري",     fv,         rent_monthly,        _FMT_CURRENCY),
            ("الدفعة المقدمة",                   down_pmt,   0,                   _FMT_CURRENCY),
            ("قيمة القرض / الصفر",               loan,       0,                   _FMT_CURRENCY),
            ("معدل الفائدة / نمو الإيجار",       mort_r,     float(md.get("rent_growth", 0.05) or 0), _FMT_PCT),
            ("القسط/الإيجار الشهري (EGP)",      mort_pmt,   rent_monthly,        _FMT_CURRENCY),
            (f"إجمالي التكلفة على {years} سنة",  mort_pmt * n_months, rent_monthly * n_months, _FMT_CURRENCY),
        ]
        for ri, (lbl, buy_v, rent_v, fmt) in enumerate(_ROWS):
            ws.row_dimensions[r].height = 18
            fill = _FILL_ROW_BAND if ri % 2 else _FILL_INPUT_CELL
            lc           = ws.cell(row=r, column=2)
            lc.value     = lbl
            lc.font      = Font(bold=True, size=10)
            lc.fill      = fill
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border    = _BORDER_THIN
            for ci, v in enumerate((buy_v, rent_v), 3):
                c                = ws.cell(row=r, column=ci)
                c.value          = v
                c.number_format  = fmt
                c.alignment      = _ALIGN_CENTER
                c.fill           = fill
                c.border         = _BORDER_THIN
            r += 1

        r += 1
        conclusion = "الشراء" if (mort_pmt < rent_monthly * 1.5 or rent_monthly == 0) else "الإيجار"
        ws.merge_cells(f"B{r}:D{r}")
        fc           = ws.cell(row=r, column=2)
        fc.value     = f"التوصية: {conclusion} هو الخيار الأمثل بناءً على الافتراضات المُبيَّنة."
        fc.font      = _FONT_FINAL_VALUE
        fc.fill      = _FILL_FINAL_VALUE
        fc.alignment = Alignment(horizontal="center", vertical="center")
        fc.border    = _BORDER_MEDIUM
        ws.row_dimensions[r].height = 24

        ws.freeze_panes = "B3"

    def sheet_habu(self) -> None:
        """'أفضل وأعلى استخدام — HABU' — Highest and Best Use (legacy only)."""
        ws = self.workbook.create_sheet(
            "أفضل وأعلى استخدام — HABU"
        )
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 52

        md = self.result.metadata if self.result else {}

        ws.merge_cells("A1:C1")
        b           = ws["A1"]
        b.value     = "أفضل وأعلى استخدام — Highest & Best Use (HABU)"
        b.fill      = _FILL_HEADER
        b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
        b.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:C2")
        ws["A2"].value     = f"تاريخ التقرير: {self.report_date}  |  المعيار: IVS 102"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="center")

        r = 4
        criteria = [
            ("١. المشروعية القانونية",
             "هل الاستخدام مسموح به بموجب التقسيم واللوائح التنظيمية؟",
             md.get("habu_legal") or "الاستخدام السكاني / التجاري مُرخَّص ومتوافق مع لوائح المنطقة."),
            ("٢. الجدوى الفيزيائية",
             "هل الموقع يتحمّل الاستخدام المقترح هيكلياً ومساحياً؟",
             md.get("habu_physical") or "المساحة وطبيعة الموقع تدعم الاستخدام المقترح."),
            ("٣. الجدوى المالية",
             "هل الاستخدام يحقق عائداً يتجاوز تكلفة الاستثمار؟",
             md.get("habu_financial") or "الاستخدام المقترح يُحقق عائداً إيجابياً بناءً على تحليل السوق."),
            ("٤. تعظيم الإنتاجية",
             "أي استخدام يحقق أعلى قيمة للعقار؟",
             md.get("habu_productive") or "الاستخدام الحالي يُعظّم قيمة العقار في السوق الراهن."),
        ]
        for i, (crit, question, answer) in enumerate(criteria):
            fill = _FILL_ROW_BAND if i % 2 else _FILL_INPUT_CELL
            ws.merge_cells(f"B{r}:C{r}")
            ch           = ws.cell(row=r, column=2)
            ch.value     = crit
            ch.fill      = _FILL_INPUT_SECT
            ch.font      = Font(bold=True, color=_Palette.WHITE, size=11)
            ch.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

            ws.row_dimensions[r].height = 18
            lc           = ws.cell(row=r, column=2)
            lc.value     = "السؤال:"
            lc.font      = Font(bold=True, italic=True, size=10)
            lc.fill      = fill
            lc.alignment = Alignment(horizontal="right", vertical="center")
            lc.border    = _BORDER_THIN
            vc           = ws.cell(row=r, column=3)
            vc.value     = question
            vc.font      = Font(italic=True, size=10)
            vc.fill      = fill
            vc.border    = _BORDER_THIN
            vc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            r += 1

            ws.row_dimensions[r].height = 30
            la           = ws.cell(row=r, column=2)
            la.value     = "الجواب:"
            la.font      = Font(bold=True, size=10)
            la.fill      = fill
            la.alignment = Alignment(horizontal="right", vertical="center")
            la.border    = _BORDER_THIN
            va           = ws.cell(row=r, column=3)
            va.value     = answer
            va.fill      = fill
            va.border    = _BORDER_THIN
            va.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            r += 2

        ws.merge_cells(f"B{r}:C{r}")
        ch           = ws.cell(row=r, column=2)
        ch.value     = "  خلاصة أفضل وأعلى استخدام"
        ch.fill      = _FILL_INPUT_SECT
        ch.font      = Font(bold=True, color=_Palette.WHITE, size=11)
        ch.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

        ws.merge_cells(f"B{r}:C{r}")
        cc           = ws.cell(row=r, column=2)
        cc.value     = (
            md.get("habu_conclusion") or
            f"أفضل وأعلى استخدام للعقار هو: "
            f"{getattr(self.result, 'asset_type', 'سكني') if self.result else 'سكني'} "
            "كما هو قائم حالياً، وهو الاستخدام الذي يُحقق أعلى قيمة سوقية."
        )
        cc.font      = Font(bold=True, size=10)
        cc.fill      = _FILL_FINAL_VALUE
        cc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cc.border    = _BORDER_MEDIUM
        ws.row_dimensions[r].height = 40

        ws.freeze_panes = "B3"

    # ──────────────────────────────────────────────────────────────────────────
    # Style helper
    # ──────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _should_include_sheet(sheet_name: str, report_style: str) -> bool:
        """Return False for advanced-analytics sheets when report_style == 'legacy'."""
        if report_style == "legacy":
            return sheet_name.strip().lower() not in _LEGACY_EXCLUDED_SHEETS
        return True

    # ──────────────────────────────────────────────────────────────────────────
    # Advanced-analytics sheets (detailed only — excluded from legacy)
    # ──────────────────────────────────────────────────────────────────────────

    def _adv_sheet(self, ar_title: str, en_title: str, description: str) -> None:
        """Create an advanced-analytics sheet with a standard header."""
        ws = self.workbook.create_sheet(ar_title)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 55
        ws.merge_cells("A1:B1")
        tc           = ws["A1"]
        tc.value     = f"{ar_title}  —  {en_title}"
        tc.font      = _FONT_TITLE
        tc.fill      = _FILL_HEADER
        tc.font      = Font(bold=True, color=_Palette.WHITE, size=13)
        tc.alignment = _ALIGN_CENTER
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:B2")
        ws["A2"].value     = f"Report Date: {self.report_date}"
        ws["A2"].font      = _FONT_MUTED
        ws["A2"].alignment = Alignment(horizontal="right")

        ws.merge_cells("A4:B4")
        desc_cell           = ws["A4"]
        desc_cell.value     = description
        desc_cell.alignment = _ALIGN_WRAP
        ws.row_dimensions[4].height = 45

        if self.result is not None:
            md = self.result.metadata or {}
            ws.cell(row=6, column=1).value = "Asset Type"
            ws.cell(row=6, column=1).font  = _FONT_BOLD
            ws.cell(row=6, column=2).value = self.result.asset_type
            ws.cell(row=7, column=1).value = "Primary Value (EGP)"
            ws.cell(row=7, column=1).font  = _FONT_BOLD
            ws.cell(row=7, column=2).value = float(self.result.primary_value) if self.result.primary_value else 0
            self._apply_currency_format(ws, 7, 7, 2)
            ws.cell(row=8, column=1).value = "Confidence"
            ws.cell(row=8, column=1).font  = _FONT_BOLD
            ws.cell(row=8, column=2).value = self.result.confidence
            ws.cell(row=9, column=1).value = "Location"
            ws.cell(row=9, column=1).font  = _FONT_BOLD
            ws.cell(row=9, column=2).value = md.get("location") or md.get("address") or "—"

    def sheet_spatial_analysis(self) -> None:
        self._adv_sheet(
            "التحليل المكاني", "Spatial Analysis",
            "يحتوي هذا القسم على التحليل المكاني للعقارات المقارنة والموقع الجغرافي "
            "للأصل. يشمل خرائط الكثافة السعرية، نصف قطر البحث، وتأثير الموقع على القيمة.",
        )

    def sheet_multiple_regression(self) -> None:
        ws_name = "الانحدار المتعدد"
        self._adv_sheet(
            ws_name, "Multiple Regression",
            "يحتوي هذا القسم على نموذج الانحدار المتعدد لتقدير القيمة. يشمل المتغيرات "
            "المستقلة (المساحة، الموقع، العمر)، معاملات الانحدار، ومعامل التحديد R².",
        )
        if self.result is not None:
            ws  = self.workbook[ws_name]
            md  = self.result.metadata or {}
            row = 11
            ws.cell(row=row, column=1).value = "Inputs (Three Approaches)"; row += 1
            for key, label in (("comparable", "Comparable"), ("cost", "Cost"), ("income", "Income")):
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=1).font  = _FONT_BOLD
                ws.cell(row=row, column=2).value = float(md.get(key) or 0)
                self._apply_currency_format(ws, row, row, 2)
                row += 1

    def sheet_real_options(self) -> None:
        self._adv_sheet(
            "الخيارات الحقيقية", "Real Options",
            "يحتوي هذا القسم على تحليل الخيارات الحقيقية للأصل. يشمل قيمة خيار التطوير، "
            "خيار التأخير، خيار التوسع، وخيار التخلي مع التقلبات الضمنية في السوق.",
        )

    def sheet_executive_dashboard(self) -> None:
        ws_name = "لوحة القيادة التنفيذية"
        self._adv_sheet(
            ws_name, "Executive Dashboard",
            "لوحة القيادة التنفيذية — ملخص المؤشرات الرئيسية للتقرير للإطلاع السريع.",
        )
        if self.result is not None:
            ws  = self.workbook[ws_name]
            md  = self.result.metadata or {}
            row = 11
            headers = ["Metric", "Value"]
            self._apply_header_style(ws, row, 2)
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col).value = h
            row += 1
            for label, val in (
                ("Asset Type",      self.result.asset_type),
                ("Primary Purpose", self.result.primary_purpose),
                ("Confidence",      self.result.confidence),
                ("Comparable (EGP)", float(md.get("comparable") or 0)),
                ("Cost (EGP)",       float(md.get("cost")       or 0)),
                ("Income (EGP)",     float(md.get("income")     or 0)),
                ("Final Value (EGP)", float(self.result.primary_value) if self.result.primary_value else 0),
            ):
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=1).font  = _FONT_BOLD
                ws.cell(row=row, column=2).value = val
                if "EGP" in label:
                    self._apply_currency_format(ws, row, row, 2)
                row += 1

    def sheet_neural_networks(self) -> None:
        self._adv_sheet(
            "الشبكات العصبية", "Neural Networks",
            "يحتوي هذا القسم على نموذج الشبكة العصبية الاصطناعية لتقدير القيمة السوقية. "
            "يشمل البنية المعمارية للنموذج، الطبقات المخفية، دالة التفعيل، ومقاييس الأداء.",
        )

    def sheet_time_series(self) -> None:
        ws_name = "السلاسل الزمنية"
        self._adv_sheet(
            ws_name, "Time Series",
            "يحتوي هذا القسم على تحليل السلاسل الزمنية لأسعار العقارات في المنطقة. "
            "يشمل اتجاهات الأسعار التاريخية، نماذج ARIMA، والتنبؤ بالأسعار المستقبلية.",
        )
        if self.result is not None:
            ws  = self.workbook[ws_name]
            row = 11
            ws.cell(row=row, column=1).value = "Valuation Date"
            ws.cell(row=row, column=1).font  = _FONT_BOLD
            ws.cell(row=row, column=2).value = self.report_date

    def sheet_market_intelligence(self) -> None:
        self._adv_sheet(
            "استخبارات السوق", "Market Intelligence",
            "يحتوي هذا القسم على تقرير استخبارات السوق العقاري. يشمل مؤشرات الطلب والعرض، "
            "متوسطات أسعار المنطقة، حجم الصفقات، وتحليل المنافسين.",
        )

    # ──────────────────────────────────────────────────────────────────────────
    # Public entry point
    # ──────────────────────────────────────────────────────────────────────────

    def build(self, filename: str, ivsc_disclosure=None,
              cross_border_disclosure=None, portfolio_summary=None,
              portfolio_performance=None,
              report_style: str = "legacy") -> str:
        """Build all sheets and save to filename. Returns the filename."""
        # ── Try professional template export (individual valuation) ───────────
        # Only attempted when report_style == "professional_template".
        # Any failure silently falls through to the existing builder below.
        if self.result is not None and report_style == "professional_template":
            try:
                from pathlib import Path as _Path
                try:
                    from reports.excel_template_renderer import (
                        INDIVIDUAL_VALUATION_TEMPLATE as _TPL,
                        build_individual_valuation_report as _build_tpl,
                    )
                except ImportError:
                    from core_engine.reports.excel_template_renderer import (  # type: ignore
                        INDIVIDUAL_VALUATION_TEMPLATE as _TPL,
                        build_individual_valuation_report as _build_tpl,
                    )
                if _TPL.is_file():
                    _r    = self.result
                    _meta = _r.metadata or {}
                    _context = {
                        "report_date":       self.report_date,
                        "valuation_date":    _meta.get("valuation_date") or self.report_date,
                        "client_name":       _meta.get("client_name") or _meta.get("borrower_name") or "N/A",
                        "property_type":     _r.asset_type,
                        "location":          _meta.get("location") or _meta.get("address") or "N/A",
                        "area":              _meta.get("area") or _meta.get("floor_area_m2") or "",
                        "valuation_purpose": _r.primary_purpose,
                        "market_value":      float(_r.primary_value) if _r.primary_value else 0,
                        "comparative_value": float(_meta.get("comparable") or 0),
                        "cost_value":        float(_meta.get("cost") or 0),
                        "income_value":      float(_meta.get("income") or 0),
                        "final_value":       float(_r.primary_value) if _r.primary_value else 0,
                        "confidence":        _r.confidence,
                        "reviewer_name":     _meta.get("reviewer_name") or _meta.get("appraiser_name") or "N/A",
                        "report_id":         _meta.get("report_id") or "N/A",
                    }
                    _tables: Dict[str, list] = {}
                    if _r.audit_trail:
                        _tables["audit_trail"] = [
                            {
                                "Step":       e.step_name,
                                "Formula":    e.formula or "",
                                "References": ", ".join(e.references or []),
                            }
                            for e in _r.audit_trail
                        ]
                    if _r.issues:
                        _tables["issues"] = [
                            {"Severity": i.severity, "Code": i.code, "Message": i.message}
                            for i in _r.issues
                        ]
                    _methods = [
                        {
                            "Approach":    label,
                            "Value (EGP)": float(_meta.get(key) or 0),
                            "Weight":      (_r.weights_applied or {}).get(key, ""),
                        }
                        for key, label in (
                            ("comparable", "Comparable Sales"),
                            ("cost",       "Cost Approach"),
                            ("income",     "Income Approach"),
                        )
                        if _meta.get(key)
                    ]
                    if _methods:
                        _tables["valuation_methods"] = _methods
                    _comps = _meta.get("comparables") or _meta.get("comparable_sales") or []
                    if _comps:
                        _tables["comparables"] = [
                            dict(c) if isinstance(c, dict) else {"Value": str(c)}
                            for c in _comps[:20]
                        ]
                    if _r.disclosures:
                        _tables["assumptions"] = [{"Reference": d} for d in _r.disclosures]

                    # Use .xlsm extension when the template is .xlsm but
                    # the requested filename was .xlsx (bridge_api default).
                    _req   = _Path(filename)
                    _out_path = (
                        _req.with_suffix(".xlsm")
                        if _TPL.suffix.lower() == ".xlsm"
                        and _req.suffix.lower() == ".xlsx"
                        else _req
                    )
                    _out = _build_tpl(
                        output_path=str(_out_path),
                        context=_context,
                        tables=_tables,
                    )
                    if _out is not None and _out_path.is_file():
                        return str(_out_path)
            except Exception:
                pass
        # ─────────────────────────────────────────────────────────────────────

        if self.result is not None:
            self.sheet_summary()
            self.sheet_three_approaches()
            self.sheet_weights_analysis()
            self.sheet_audit_trail()
            self.sheet_property_details()
            self.sheet_disclosures()
            self.sheet_issues_warnings()
            self.sheet_certification()
            if ivsc_disclosure is not None:
                self.sheet_ivsc_compliance(ivsc_disclosure)
            if cross_border_disclosure is not None:
                self.sheet_cross_border(cross_border_disclosure)
            # ── Legacy-only enhanced Arabic sheets ────────────────────────────
            if report_style == "legacy":
                self.sheet_assumptions_inputs()
                self.sheet_main_report()
                self.sheet_sales_comparison()
                self.sheet_rental_comparables()
                self.sheet_cost_approach()
                self.sheet_income_capitalization()
                self.sheet_reconciliation()
                self.sheet_valuation_parameters()
                self.sheet_certification_ar()
                self.sheet_data_sources()
                self.sheet_dcf()
                self.sheet_rent_vs_buy()
                self.sheet_habu()
            # ── Advanced-analytics sheets (detailed only) ─────────────────────
            _inc = self._should_include_sheet
            if _inc("التحليل المكاني", report_style):
                self.sheet_spatial_analysis()
            if _inc("الانحدار المتعدد", report_style):
                self.sheet_multiple_regression()
            if _inc("الخيارات الحقيقية", report_style):
                self.sheet_real_options()
            if _inc("لوحة القيادة التنفيذية", report_style):
                self.sheet_executive_dashboard()
            if _inc("الشبكات العصبية", report_style):
                self.sheet_neural_networks()
            if _inc("السلاسل الزمنية", report_style):
                self.sheet_time_series()
            if _inc("استخبارات السوق", report_style):
                self.sheet_market_intelligence()
        if portfolio_summary is not None:
            self.sheet_portfolio_summary(portfolio_summary)
        if portfolio_performance is not None:
            self.sheet_portfolio_performance(portfolio_performance)
        self.workbook.save(filename)
        return filename
