"""
core_engine/reports/excel_template_renderer.py

Template-based Excel report renderer for Expert Smart.

Supports both ``.xlsx`` and ``.xlsm`` (macro-enabled) templates.
When the template is ``.xlsm`` the workbook is opened with ``keep_vba=True``
so VBA code is preserved in the output file.

Usage — generic template
------------------------
    from core_engine.reports.excel_template_renderer import build_from_template

    build_from_template(
        template_path="templates/reports/valuation_template.xlsx",
        output_path="outputs/report_2026.xlsx",
        context={
            "report_date":           "2026-05-10",
            "valuation_purpose":     "Mortgage Lending",
            "total_portfolio_value": 125_000_000,
            "avg_price_per_meter":   18_500,
            "median_ratio":          0.982,
            "cod":                   4.5,
            "prd":                   1.02,
            "reviewer_name":         "Hisham Elmahdy",
        },
        tables={
            "properties_results": [
                {"Property ID": "P001", "Address": "Cairo",      "Value": 5_000_000},
                {"Property ID": "P002", "Address": "Alexandria", "Value": 3_200_000},
            ],
            "ratio_study": [...],
        },
    )

Usage — Mass Appraisal professional template (.xlsm)
-----------------------------------------------------
    from core_engine.reports.excel_template_renderer import build_mass_appraisal_report

    result = build_mass_appraisal_report(
        output_path="outputs/mass_appraisal_2026.xlsm",
        context={
            "report_date":           "2026-05-10",
            "valuation_date":        "2026-05-01",
            "total_portfolio_value": 125_000_000,
        },
        tables={
            "properties_results": [...],
            "ratio_study":        [...],
            "calibration":        [...],
        },
    )
    # result is Path on success, None if template missing → use fallback export

Fallback
--------
    If *template_path* is None or the file does not exist, ``build_from_template``
    returns None and callers should use their existing export logic unchanged.
"""

from __future__ import annotations

import copy
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Registered professional templates ─────────────────────────────────────────
MASS_APPRAISAL_TEMPLATE = Path(
    "templates/reports/mass_appraisal_professional_template.xlsm"
)

# ── Regex patterns ─────────────────────────────────────────────────────────────
_RE_PLACEHOLDER = re.compile(r"\{\{([A-Za-z_]\w*)\}\}")
_RE_TABLE       = re.compile(r"\{\{TABLE:([A-Za-z_]\w*)\}\}")

# ── Default table header style ─────────────────────────────────────────────────
_TBL_FILL_HDR   = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_TBL_FONT_HDR   = Font(bold=True, color="FFFFFF", size=10)
_TBL_ALIGN_HDR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TBL_BORDER     = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
_TBL_FONT_DATA  = Font(size=10)
_TBL_ALIGN_DATA = Alignment(vertical="center", wrap_text=False)

TableData = List[Dict[str, Any]]


class ExcelTemplateRenderer:
    """
    Render an Expert Smart Excel report from an openpyxl ``.xlsx`` template.

    Workflow
    --------
    1. Load the template workbook (preserving styles, merged cells, formulas).
    2. Walk every cell; replace ``{{key}}`` placeholders with context values.
    3. Walk every cell; expand ``{{TABLE:name}}`` anchors into tabular data.
    4. Save the result to *output_path*.

    Fallback
    --------
    If *template_path* is None or the file does not exist, :meth:`build`
    returns ``None``.  Callers must check for that and fall back to their
    existing export logic — this class never raises for a missing template.
    """

    def __init__(
        self,
        template_path: Optional[Union[str, Path]] = None,
        context: Optional[Dict[str, Any]] = None,
        tables: Optional[Dict[str, TableData]] = None,
    ) -> None:
        self.template_path: Optional[Path] = (
            Path(template_path) if template_path else None
        )
        self.context: Dict[str, Any] = context or {}
        self.tables: Dict[str, TableData] = tables or {}
        self._wb: Optional[Workbook] = None

    # ── Public API ──────────────────────────────────────────────────────────────

    def is_available(self) -> bool:
        """Return True only when a valid template file exists on disk."""
        return self.template_path is not None and self.template_path.is_file()

    def build(self, output_path: Union[str, Path]) -> Optional[Path]:
        """
        Render the template and save to *output_path*.

        Returns the resolved ``Path`` on success, or ``None`` when no template
        is available.  Callers should treat ``None`` as "use existing fallback."
        """
        if not self.is_available():
            return None

        keep_vba = self.template_path.suffix.lower() == ".xlsm"
        self._wb = load_workbook(str(self.template_path), keep_vba=keep_vba)
        self._replace_placeholders()
        self._expand_tables()

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(str(out))
        self._wb.close()
        return out

    # ── Internal: placeholder replacement ──────────────────────────────────────

    def _replace_placeholders(self) -> None:
        """Replace every ``{{key}}`` token in every writable cell of every sheet."""
        for ws in self._wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue  # only the master cell carries the value
                    if not isinstance(cell.value, str):
                        continue
                    if cell.value.startswith("="):
                        continue  # preserve Excel formulas untouched
                    new_val = self._substitute(cell.value)
                    if new_val != cell.value:
                        cell.value = new_val

    def _substitute(self, text: str) -> Any:
        """
        Replace ``{{key}}`` occurrences in *text* with their context values.

        When the *entire* string is a single placeholder (e.g. ``{{cod}}``)
        and the matching value is not a string, the raw value is returned so
        openpyxl can store it with the correct type and honour number formats.
        For all other cases a string substitution is performed.
        """
        sole = _RE_PLACEHOLDER.fullmatch(text.strip())
        if sole:
            key = sole.group(1)
            if key in self.context:
                return self.context[key]  # numeric / date — preserve type

        def _repl(m: re.Match) -> str:
            key = m.group(1)
            val = self.context.get(key)
            return str(val) if val is not None else m.group(0)  # leave if missing

        return _RE_PLACEHOLDER.sub(_repl, text)

    # ── Internal: table expansion ───────────────────────────────────────────────

    def _expand_tables(self) -> None:
        """Locate every ``{{TABLE:name}}`` anchor and render its table in-place."""
        for ws in self._wb.worksheets:
            anchors = self._find_table_anchors(ws)
            # Process bottom-up so row insertions below don't shift higher anchors
            for anchor_row, anchor_col, table_name in sorted(
                anchors, key=lambda x: x[0], reverse=True
            ):
                rows = self.tables.get(table_name)
                # Clear the anchor cell regardless of whether we have data
                ws.cell(row=anchor_row, column=anchor_col).value = None
                if rows:
                    self._render_table(ws, anchor_row, anchor_col, rows)

    def _find_table_anchors(self, ws) -> List[tuple]:
        """Return ``[(row, col, table_name), ...]`` for every TABLE anchor in *ws*."""
        found: List[tuple] = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if not isinstance(cell.value, str):
                    continue
                m = _RE_TABLE.fullmatch(cell.value.strip())
                if m:
                    found.append((cell.row, cell.column, m.group(1)))
        return found

    def _render_table(
        self,
        ws,
        start_row: int,
        start_col: int,
        rows: TableData,
    ) -> None:
        """
        Write a table starting at (*start_row*, *start_col*).

        The anchor row becomes the header row.  Data rows are inserted
        immediately below, pushing existing content further down.

        *rows* may be:
        - a list of ``dict`` — keys become column headers;
        - a list of ``list`` — first sub-list treated as header row.
        """
        if not rows:
            return

        if isinstance(rows[0], dict):
            headers: List[str] = list(rows[0].keys())
            data_rows: List[List[Any]] = [
                [r.get(h, "") for h in headers] for r in rows
            ]
        else:
            headers = [str(v) for v in rows[0]]
            data_rows = [list(r) for r in rows[1:]]

        num_cols = len(headers)

        # ── Header row (reuses the anchor row) ────────────────────────────────
        for col_offset, header in enumerate(headers):
            c = ws.cell(row=start_row, column=start_col + col_offset)
            c.value     = header
            c.font      = copy.copy(_TBL_FONT_HDR)
            c.fill      = copy.copy(_TBL_FILL_HDR)
            c.alignment = copy.copy(_TBL_ALIGN_HDR)
            c.border    = copy.copy(_TBL_BORDER)

        # ── Insert blank rows for data, then fill them ─────────────────────────
        if data_rows:
            ws.insert_rows(start_row + 1, amount=len(data_rows))
            for row_offset, data_row in enumerate(data_rows):
                for col_offset, value in enumerate(data_row):
                    c = ws.cell(
                        row=start_row + 1 + row_offset,
                        column=start_col + col_offset,
                    )
                    c.value     = value
                    c.font      = copy.copy(_TBL_FONT_DATA)
                    c.alignment = copy.copy(_TBL_ALIGN_DATA)
                    c.border    = copy.copy(_TBL_BORDER)

        # ── Best-effort column width ──────────────────────────────────────────
        for col_offset in range(num_cols):
            col_letter = get_column_letter(start_col + col_offset)
            all_vals = [str(headers[col_offset])] + [
                str(dr[col_offset]) for dr in data_rows if col_offset < len(dr)
            ]
            max_len = max((len(v) for v in all_vals), default=10)
            current = ws.column_dimensions[col_letter].width or 0
            ws.column_dimensions[col_letter].width = max(
                current, min(max_len + 4, 40)
            )


# ── Convenience function ────────────────────────────────────────────────────────

def build_from_template(
    template_path: Union[str, Path],
    output_path: Union[str, Path],
    context: Dict[str, Any],
    tables: Optional[Dict[str, TableData]] = None,
) -> Optional[Path]:
    """
    Render *template_path* with *context* and *tables*, saving to *output_path*.

    Returns the output ``Path`` on success, or ``None`` when the template file
    does not exist.  Callers should use their existing fallback export when this
    function returns ``None``.

    Parameters
    ----------
    template_path:
        Path to an ``.xlsx`` file used as the layout template.
    output_path:
        Destination for the rendered workbook.
    context:
        Scalar placeholders, e.g. ``{"report_date": "2026-05-10",
        "reviewer_name": "Hisham Elmahdy", "cod": 4.5, ...}``.
    tables:
        Named tabular data.  Each value is a list of dicts (preferred) or a
        list of lists.  Keys must match the TABLE anchor names inside the
        template (e.g. ``{{TABLE:properties_results}}``).
    """
    renderer = ExcelTemplateRenderer(
        template_path=template_path,
        context=context,
        tables=tables or {},
    )
    return renderer.build(output_path)


def build_mass_appraisal_report(
    output_path: Union[str, Path],
    context: Dict[str, Any],
    tables: Optional[Dict[str, TableData]] = None,
    template_path: Optional[Union[str, Path]] = None,
) -> Optional[Path]:
    """
    Render the Mass Appraisal professional template (``.xlsm``) and save to
    *output_path*.

    Uses ``MASS_APPRAISAL_TEMPLATE`` by default.  Accepts an optional
    *template_path* override (useful for testing or alternate layouts).

    Returns the output ``Path`` on success.  Returns ``None`` when the template
    file is missing or rendering fails — callers must then fall back to their
    existing export logic.

    Parameters
    ----------
    output_path:
        Destination for the rendered workbook.  Use a ``.xlsm`` extension to
        preserve VBA macros in the output file.
    context:
        Scalar placeholders matched against ``{{key}}`` tokens in the template,
        e.g. ``{"report_date": "2026-05-10", "valuation_date": "2026-05-01",
        "total_portfolio_value": 125_000_000}``.
    tables:
        Named row data matched against ``{{TABLE:name}}`` anchors in the
        template.  Supported anchors: ``properties_results``, ``ratio_study``,
        ``calibration``.
    template_path:
        Override the registered ``MASS_APPRAISAL_TEMPLATE`` path.
    """
    tpl = Path(template_path) if template_path else MASS_APPRAISAL_TEMPLATE
    renderer = ExcelTemplateRenderer(
        template_path=tpl,
        context=context,
        tables=tables or {},
    )
    if not renderer.is_available():
        return None
    try:
        return renderer.build(output_path)
    except Exception:
        return None
