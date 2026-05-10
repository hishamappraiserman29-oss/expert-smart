"""
multilingual_builder.py — Phase 23 Multilingual Report Builder
Generates Excel valuation reports in Arabic or English with RTL support.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from i18n.localization import Language, get_localization
from i18n.arabic_support import ArabicSupport


class MultilingualReportBuilder:
    """Build valuation reports in multiple languages."""

    def __init__(self, language: Language = Language.ENGLISH) -> None:
        self.language = language
        self.loc = get_localization()
        self.loc.set_language(language)

    # -- Public ---------------------------------------------------------------

    def build_valuation_report(
        self, valuation_data: Dict[str, Any], output_path: str
    ) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = self._t("report.title")[:31]  # Excel sheet name limit

        if self.language == Language.ARABIC:
            ws.sheet_view.rightToLeft = True

        # Title
        ws["A1"] = self._t("report.title")
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        self._add_section(ws, 3, "report.details")
        self._add_results_section(ws, 10, valuation_data)
        self._add_methodology_section(ws, 15, valuation_data)
        self._add_comparables_section(ws, 20, valuation_data)
        self._add_footer(ws)

        wb.save(output_path)
        return output_path

    def set_language(self, language: Language) -> None:
        self.language = language
        self.loc.set_language(language)

    # -- Private --------------------------------------------------------------

    def _t(self, key: str, **kwargs) -> str:
        return self.loc.t(key, language=self.language, **kwargs)

    def _add_section(self, ws, start_row: int, section_key: str) -> None:
        cell = ws[f"A{start_row}"]
        cell.value = self._t(section_key)
        cell.font = Font(size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        ws.merge_cells(f"A{start_row}:D{start_row}")

    def _add_results_section(self, ws, start_row: int, data: Dict) -> None:
        self._add_section(ws, start_row, "report.summary")
        row = start_row + 1

        ws[f"A{row}"] = self._t("result.primary_value")
        ws[f"B{row}"] = self.loc.format_currency(data.get("primary_value", 0), self.language)
        row += 1

        ws[f"A{row}"] = self._t("result.confidence")
        confidence = data.get("confidence", "medium")
        ws[f"B{row}"] = self._t(f"confidence.{confidence}")
        row += 1

        ws[f"A{row}"] = self._t("result.unit_value")
        ws[f"B{row}"] = self.loc.format_currency(data.get("unit_value_sqm", 0), self.language)

    def _add_methodology_section(self, ws, start_row: int, data: Dict) -> None:
        self._add_section(ws, start_row, "report.methodology")
        row = start_row + 1
        ws[f"A{row}"] = self._t("result.methodology")
        ws[f"B{row}"] = data.get("methodology", "Comparable Sales Approach")

    def _add_comparables_section(self, ws, start_row: int, data: Dict) -> None:
        self._add_section(ws, start_row, "report.comparables")
        row = start_row + 1
        for comp in data.get("comparables", [])[:5]:
            ws[f"A{row}"] = comp.get("property_id", "")
            ws[f"B{row}"] = comp.get("location", "")
            ws[f"C{row}"] = self.loc.format_currency(comp.get("price", 0), self.language)
            row += 1

    def _add_footer(self, ws) -> None:
        last = ws.max_row + 2
        ws[f"A{last}"] = self._t("info.powered_by")
        ws[f"A{last}"].font = Font(size=9, italic=True)
        ws[f"A{last + 1}"] = self.loc.format_date(datetime.now(), self.language)
        ws[f"A{last + 1}"].font = Font(size=9)
