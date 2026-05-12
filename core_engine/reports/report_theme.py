#!/usr/bin/env python3
"""
report_theme.py — Single source of truth for the EXPERT_SMART "Midnight Gold"
visual identity.

استخدام:
    from core_engine.reports.report_theme import (
        Palette, Typography, NumFormat,
        style_page_title, style_section_header, style_table_header,
        style_label, style_value, style_input_value, style_body,
        draw_banner, draw_section, apply_sheet_defaults,
        get_font, get_fill, get_alignment,
    )

تصميم:
  - كل الثوابت فى classes (Palette, Typography, NumFormat) لتجنب تلوّث الـ
    module namespace.
  - الدوال الـpublic بـ snake_case + type hints كاملة.
  - دوال factory صغيرة (get_font/get_fill/get_alignment) للاستخدام المباشر.
  - دوال style_* عالية المستوى (تطبّق font+fill+alignment+border معاً).
  - لا توجد آثار جانبية: كل دالة تُعدّل خلية أو ورقة تُمرَّر لها فقط.
"""

from __future__ import annotations

from typing import Literal

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# 1. Palette — Midnight Gold colours
# ═══════════════════════════════════════════════════════════════════════════


class Palette:
    """Centralized color palette. All hex without leading #."""

    # Navy spectrum (5 درجات)
    INK: str = "0B1437"
    NAVY_DEEP: str = "12224F"
    NAVY: str = "1B3263"
    NAVY_LIGHT: str = "294479"
    INDIGO: str = "3A5491"

    # Gold spectrum (3 درجات)
    GOLD: str = "C9A961"
    GOLD_LIGHT: str = "EBD79A"
    GOLD_PALE: str = "F8EFC9"

    # Status (positive / warning / info)
    EMERALD: str = "0E8B6E"
    EMERALD_LT: str = "C9EAD9"
    EMERALD_VLT: str = "EDF8F2"
    CORAL: str = "D85842"
    CORAL_LT: str = "FBE0D7"
    CORAL_VLT: str = "FDF0EB"
    PURPLE: str = "6B4FA1"
    PURPLE_LT: str = "E0D7EE"
    TEAL: str = "3A9CA1"

    # Neutrals
    PAPER: str = "FAF7F0"
    GREY_50: str = "F8F9FC"
    GREY_100: str = "EEF1F7"
    GREY_300: str = "C8CDDB"
    GREY_500: str = "8C93A8"
    GREY_700: str = "586079"
    WHITE: str = "FFFFFF"
    BLACK: str = "000000"

    # Input convention (Wall-Street blue for editable cells)
    INPUT_BLUE: str = "0000FF"


# ═══════════════════════════════════════════════════════════════════════════
# 2. Typography
# ═══════════════════════════════════════════════════════════════════════════


class Typography:
    """Fonts and standard sizes."""

    PRIMARY: str = "Cairo"       # Arabic-first font
    FALLBACK: str = "Calibri"
    NUMERIC: str = "Cairo"

    # Standard size scale (pt)
    SIZE_PAGE_TITLE: float = 22
    SIZE_SECTION_HEADER: float = 13
    SIZE_TABLE_HEADER: float = 10.5
    SIZE_LABEL: float = 10.5
    SIZE_VALUE: float = 10.5
    SIZE_INPUT: float = 12
    SIZE_BODY: float = 10.5
    SIZE_KPI_VALUE: float = 18
    SIZE_FOOTER: float = 9


# ═══════════════════════════════════════════════════════════════════════════
# 3. NumFormat
# ═══════════════════════════════════════════════════════════════════════════


class NumFormat:
    """Standard Excel number-format codes."""

    CURRENCY: str = '#,##0;(#,##0);"-"'
    CURRENCY_EGP: str = '#,##0" EGP"'
    CURRENCY_PER_M2: str = '#,##0" EGP/م²"'
    CURRENCY_2DP: str = '#,##0.00'
    PERCENT: str = "0.0%"
    PERCENT_DETAILED: str = "0.00%"
    PERCENT_SIGNED: str = "+0.0%;-0.0%;0.0%"
    RATIO: str = '0.00"x"'
    INTEGER: str = "#,##0"
    YEAR: str = "0"
    DATE: str = "yyyy/mm/dd"
    QUANTITY: str = "0.00"


# ═══════════════════════════════════════════════════════════════════════════
# 4. BuilderPalette — Classic Office Blue (ExcelReportBuilder legacy theme)
# ═══════════════════════════════════════════════════════════════════════════


class BuilderPalette:
    """Hex tokens for ExcelReportBuilder (legacy Classic-Office-Blue theme).

    These are SEPARATE from the Midnight Gold palette used by the sheet
    modules. They represent the pre-existing color scheme of excel_builder.py
    and are centralized here to eliminate inline hex literals.
    """

    # Header / section fills
    HEADER: str        = "1F4E78"
    SECTION_MID: str   = "4472C4"
    SECTION_DARK: str  = "2E4057"
    SUBHEAD: str       = "D6E4F7"

    # Cell-type fills
    INPUT_CELL: str    = "EBF3FB"
    CALC_CELL: str     = "F5F5F5"
    SUCCESS_LIGHT: str = "C6EFCE"
    ROW_BAND: str      = "F2F7FF"

    # Special fills
    PORT_COL: str      = "D9E1F2"
    PORTFOLIO: str     = "203864"
    CB_GREEN: str      = "70AD47"

    # Status fills
    ERROR: str         = "FF0000"
    WARNING: str       = "FFC000"
    ADJ_NEG: str       = "FFCCCC"
    ADJ_ZERO: str      = "E8E8E8"
    ADJ_GOLD: str      = "FFD966"
    ADJ_EMERALD: str   = "A9D18E"

    # Font colors
    SUCCESS_DARK: str  = "1A6B2A"
    MUTED: str         = "888888"
    NOTE: str          = "666666"
    GOLD_DARK: str     = "7F5700"
    AMBER_DARK: str    = "7F3700"


# ═══════════════════════════════════════════════════════════════════════════
# 5. Border presets
# ═══════════════════════════════════════════════════════════════════════════

_THIN_GREY = Side(style="thin", color=Palette.GREY_300)
_THIN_GOLD = Side(style="thin", color=Palette.GOLD)
_MEDIUM_NAVY = Side(style="medium", color=Palette.NAVY)
_MEDIUM_GOLD = Side(style="medium", color=Palette.GOLD)


class Borders:
    """Common border presets."""

    LIGHT: Border = Border(
        left=_THIN_GREY, right=_THIN_GREY,
        top=_THIN_GREY, bottom=_THIN_GREY,
    )
    GOLD: Border = Border(
        left=_THIN_GOLD, right=_THIN_GOLD,
        top=_THIN_GOLD, bottom=_THIN_GOLD,
    )
    HEAD: Border = Border(
        top=_MEDIUM_NAVY, bottom=_MEDIUM_NAVY,
        left=_THIN_GREY, right=_THIN_GREY,
    )
    BANNER_BOTTOM: Border = Border(bottom=_MEDIUM_GOLD)


# ═══════════════════════════════════════════════════════════════════════════
# 6. Factory helpers (low-level)
# ═══════════════════════════════════════════════════════════════════════════

Align_H = Literal["left", "center", "right", "justify"]
Align_V = Literal["top", "center", "bottom"]


def get_font(
    size: float = Typography.SIZE_BODY,
    bold: bool = False,
    italic: bool = False,
    color: str = Palette.INK,
    name: str = Typography.PRIMARY,
) -> Font:
    """Create a Font object with Midnight Gold defaults."""
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)


def get_fill(color: str) -> PatternFill:
    """Solid fill helper."""
    return PatternFill("solid", start_color=color, end_color=color)


def get_alignment(
    h: Align_H = "right",
    v: Align_V = "center",
    wrap: bool = True,
) -> Alignment:
    """RTL-aware alignment."""
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, readingOrder=2)


# ═══════════════════════════════════════════════════════════════════════════
# 7. Cell-level style applicators (high-level)
# ═══════════════════════════════════════════════════════════════════════════


def style_page_title(cell) -> None:
    """A1 of any sheet — large gold title on ink background."""
    cell.font = get_font(
        size=Typography.SIZE_PAGE_TITLE, bold=True, color=Palette.GOLD_LIGHT,
    )
    cell.fill = get_fill(Palette.INK)
    cell.alignment = get_alignment(h="center")
    cell.border = Borders.BANNER_BOTTOM


def style_section_header(cell, bg: str = Palette.NAVY) -> None:
    """Section divider — coloured background, white bold text."""
    cell.font = get_font(
        size=Typography.SIZE_SECTION_HEADER, bold=True, color=Palette.WHITE,
    )
    cell.fill = get_fill(bg)
    cell.alignment = get_alignment(h="center")
    cell.border = Borders.HEAD


def style_table_header(cell, bg: str = Palette.NAVY_DEEP) -> None:
    """Table header row — dark navy, white bold."""
    cell.font = get_font(
        size=Typography.SIZE_TABLE_HEADER, bold=True, color=Palette.WHITE,
    )
    cell.fill = get_fill(bg)
    cell.alignment = get_alignment(h="center")
    cell.border = Borders.LIGHT


def style_label(cell) -> None:
    """Left column label — bold ink on grey."""
    cell.font = get_font(size=Typography.SIZE_LABEL, bold=True, color=Palette.INK)
    cell.fill = get_fill(Palette.GREY_100)
    cell.alignment = get_alignment(h="right")
    cell.border = Borders.LIGHT


def style_value(cell) -> None:
    """Read-only value cell."""
    cell.font = get_font(
        size=Typography.SIZE_VALUE, color=Palette.INK, name=Typography.NUMERIC,
    )
    cell.fill = get_fill(Palette.WHITE)
    cell.alignment = get_alignment(h="center")
    cell.border = Borders.LIGHT


def style_input_value(cell) -> None:
    """Editable input cell — Wall Street blue convention."""
    cell.font = get_font(
        size=Typography.SIZE_INPUT, bold=True,
        color=Palette.INPUT_BLUE, name=Typography.NUMERIC,
    )
    cell.fill = get_fill(Palette.WHITE)
    cell.alignment = get_alignment(h="center")
    cell.border = Borders.GOLD


def style_body(cell) -> None:
    """Plain body text."""
    cell.font = get_font(size=Typography.SIZE_BODY, color=Palette.INK)
    cell.fill = get_fill(Palette.WHITE)
    cell.alignment = get_alignment(h="right")
    cell.border = Borders.LIGHT


# ═══════════════════════════════════════════════════════════════════════════
# 8. Drawers (banner / section headers spanning multiple columns)
# ═══════════════════════════════════════════════════════════════════════════


def draw_banner(
    ws,
    row: int,
    end_col: int,
    text: str,
    bg: str = Palette.INK,
    fg: str = Palette.GOLD_LIGHT,
    size: float = Typography.SIZE_PAGE_TITLE - 2,
    height: float = 48,
) -> None:
    """
    Draw a wide banner spanning A..end_col on the given row.

    تعمل merge_cells + تطبّق font + fill + alignment + border + row height.
    """
    last_col = get_column_letter(end_col)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = get_font(size=size, bold=True, color=fg)
    cell.fill = get_fill(bg)
    cell.alignment = get_alignment(h="center")
    cell.border = Borders.BANNER_BOTTOM
    ws.row_dimensions[row].height = height


def draw_section(
    ws,
    row: int,
    end_col: int,
    text: str,
    bg: str = Palette.NAVY,
    height: float = 32,
) -> None:
    """Draw a section-header bar across A..end_col."""
    last_col = get_column_letter(end_col)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = get_font(
        size=Typography.SIZE_SECTION_HEADER, bold=True, color=Palette.WHITE,
    )
    cell.fill = get_fill(bg)
    cell.alignment = get_alignment(h="center")
    ws.row_dimensions[row].height = height


# ═══════════════════════════════════════════════════════════════════════════
# 9. Sheet-level defaults
# ═══════════════════════════════════════════════════════════════════════════


def apply_sheet_defaults(ws, freeze: bool = True) -> None:
    """
    Apply default visual settings to a worksheet:
      - RTL view
      - Optional freeze panes at A2 (header always visible)
    """
    ws.sheet_view.rightToLeft = True
    if freeze and ws.title not in {"شهادة"}:
        ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════════════
# 10. Diagnostic CLI
# ═══════════════════════════════════════════════════════════════════════════


def _print_theme_summary() -> None:
    """Quick visual sanity check from CLI."""
    print("\n🎨 EXPERT_SMART Midnight Gold Theme")
    print("─" * 60)
    print("\n  Palette (Navy spectrum):")
    for name in ["INK", "NAVY_DEEP", "NAVY", "NAVY_LIGHT", "INDIGO"]:
        print(f"    {name:<12} #{getattr(Palette, name)}")
    print("\n  Palette (Gold spectrum):")
    for name in ["GOLD", "GOLD_LIGHT", "GOLD_PALE"]:
        print(f"    {name:<12} #{getattr(Palette, name)}")
    print("\n  Palette (Status):")
    for name in ["EMERALD", "CORAL", "PURPLE", "TEAL"]:
        print(f"    {name:<12} #{getattr(Palette, name)}")
    print("\n  Typography:")
    print(f"    PRIMARY     {Typography.PRIMARY}")
    print(f"    FALLBACK    {Typography.FALLBACK}")
    print(f"    NUMERIC     {Typography.NUMERIC}")
    print("\n  Number Formats:")
    for name in ["CURRENCY", "PERCENT", "PERCENT_SIGNED", "RATIO", "DATE"]:
        print(f"    {name:<18} {getattr(NumFormat, name)}")
    print("\n  Exported style functions:")
    funcs = [
        "style_page_title", "style_section_header", "style_table_header",
        "style_label", "style_value", "style_input_value", "style_body",
        "draw_banner", "draw_section", "apply_sheet_defaults",
        "get_font", "get_fill", "get_alignment",
    ]
    for f in funcs:
        print(f"    • {f}")
    print()


if __name__ == "__main__":
    import io as _io
    import sys as _sys
    if hasattr(_sys.stdout, "buffer"):
        _sys.stdout = _io.TextIOWrapper(
            _sys.stdout.buffer, encoding="utf-8", errors="replace"
        )
    _print_theme_summary()
