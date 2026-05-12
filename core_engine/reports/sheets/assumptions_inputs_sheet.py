#!/usr/bin/env python3
"""
assumptions_inputs_sheet.py — Builder for شيت "الافتراضات والمدخلات"
(Assumptions & Inputs, Classic Office Blue theme).
"""
from __future__ import annotations

from typing import Any, Optional

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    BuilderPalette as _BP,
    NumFormat as _NF,
    Palette as _Palette,
    get_fill as _gf,
)

# ── Module-level style constants (Classic Office Blue) ────────────────────────
_FILL_HEADER      = _gf(_BP.HEADER)
_FILL_INPUT_SECT  = _gf(_BP.SECTION_DARK)
_FILL_INPUT_CELL  = _gf(_BP.INPUT_CELL)
_FILL_CALC_CELL   = _gf(_BP.CALC_CELL)
_FILL_FINAL_VALUE = _gf(_BP.SUCCESS_LIGHT)
_FONT_MUTED       = Font(italic=True, color=_BP.MUTED)
_FONT_FINAL_VALUE = Font(bold=True, size=12, color=_BP.SUCCESS_DARK)
_ALIGN_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER_THIN      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_BORDER_MEDIUM    = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)
_FMT_CURRENCY = _NF.CURRENCY_2DP
_FMT_PCT      = _NF.PERCENT_DETAILED


def apply_assumptions_inputs_sheet(
    ws: Worksheet,
    result: Optional[Any] = None,
    report_date: str = "",
) -> None:
    """Build شيت 'الافتراضات والمدخلات' (Assumptions & Inputs)."""
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22

    md = result.metadata if result else {}
    fv = float(result.primary_value) if result and result.primary_value else 0
    w  = result.weights_applied if result else {}

    # Banner
    ws.merge_cells("A1:D1")
    t           = ws["A1"]
    t.value     = "الافتراضات والمدخلات — Assumptions & Inputs"
    t.fill      = _FILL_HEADER
    t.font      = Font(bold=True, color=_Palette.WHITE, size=14)
    t.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    ws["A2"].value     = f"تاريخ التقرير: {report_date}"
    ws["A2"].font      = _FONT_MUTED
    ws["A2"].alignment = Alignment(horizontal="center")

    r = 4

    def _sect(label: str, en: str = "") -> None:
        nonlocal r
        ws.merge_cells(f"A{r}:D{r}")
        c           = ws.cell(row=r, column=1)
        c.value     = f"  {label}" + (f"  — {en}" if en else "")
        c.fill      = _FILL_INPUT_SECT
        c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

    def _inp(label: str, value: Any, note: str = "", is_calc: bool = False,
             fmt: str = "") -> None:
        nonlocal r
        fill = _FILL_CALC_CELL if is_calc else _FILL_INPUT_CELL
        lc           = ws.cell(row=r, column=2)
        lc.value     = label
        lc.font      = Font(bold=True, size=10)
        lc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        lc.fill      = fill
        lc.border    = _BORDER_THIN

        vc           = ws.cell(row=r, column=3)
        vc.value     = value if value is not None else "—"
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill      = fill
        vc.border    = _BORDER_THIN
        if fmt:
            vc.number_format = fmt

        if note:
            nc           = ws.cell(row=r, column=4)
            nc.value     = note
            nc.font      = Font(italic=True, color=_BP.NOTE, size=9)
            nc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            nc.fill      = fill
            nc.border    = _BORDER_THIN

        ws.row_dimensions[r].height = 18
        r += 1

    # ── Section 1: بيانات التقرير ─────────────────────────────────────────────
    _sect("بيانات التقرير", "Report Data")
    _inp("رقم التقرير",     md.get("report_id") or "—")
    _inp("تاريخ التقييم",  md.get("valuation_date") or report_date)
    _inp("تاريخ المعاينة", md.get("inspection_date") or "—")
    _inp("غرض التقييم",    result.primary_purpose if result else "—")
    _inp("أساس القيمة",    "القيمة السوقية")
    r += 1

    # ── Section 2: بيانات العميل والمقيم ─────────────────────────────────────
    _sect("بيانات العميل والمقيم", "Client & Appraiser")
    _inp("اسم العميل",    md.get("client_name") or md.get("borrower_name") or "—")
    _inp("اسم المقيم",   md.get("appraiser_name") or md.get("reviewer_name") or "—")
    _inp("رقم الترخيص",  md.get("license_no") or "—")
    _inp("الجهة المعينة", md.get("instructed_by") or "—")
    r += 1

    # ── Section 3: بيانات العقار ──────────────────────────────────────────────
    _sect("بيانات العقار", "Property Data")
    _inp("نوع الأصل",         result.asset_type if result else "—")
    _inp("الموقع / العنوان",  md.get("location") or md.get("address") or "—")
    _inp("المساحة (م²)",      md.get("area") or md.get("floor_area_m2") or "—")
    _inp("سنة الإنشاء",      md.get("year_built") or "—")
    _inp("الحالة",            md.get("condition") or "—")
    r += 1

    # ── Section 4: بيانات السوق ───────────────────────────────────────────────
    _sect("بيانات السوق", "Market Data")
    _inp("متوسط سعر السوق (EGP/م²)", md.get("market_avg_price_sqm") or "—")
    _inp("معدل الرسملة",  md.get("cap_rate") or "—")
    _inp("معدل الشاغر",  md.get("vacancy_rate") or "—")
    _inp("عدد المقارنات", md.get("comparables_count") or
         len(md.get("comparables") or []) or "—")
    r += 1

    # ── Section 5: افتراضات التقييم ───────────────────────────────────────────
    _sect("افتراضات التقييم", "Valuation Assumptions")
    comp = float(md.get("comparable") or 0)
    cost = float(md.get("cost")       or 0)
    inc  = float(md.get("income")     or 0)

    _inp("القيمة — المقارنة البيعية (EGP)", comp, is_calc=True, fmt=_FMT_CURRENCY)
    _inp("الوزن — المقارنة",               w.get("comparable", 0), fmt=_FMT_PCT)
    _inp("القيمة — طريقة التكلفة (EGP)",   cost, is_calc=True, fmt=_FMT_CURRENCY)
    _inp("الوزن — التكلفة",                w.get("cost", 0), fmt=_FMT_PCT)
    _inp("القيمة — رأسمالة الدخل (EGP)",   inc,  is_calc=True, fmt=_FMT_CURRENCY)
    _inp("الوزن — الدخل",                  w.get("income", 0), fmt=_FMT_PCT)
    r += 1

    # Final value banner
    ws.merge_cells(f"B{r}:D{r}")
    fv_cell           = ws.cell(row=r, column=2)
    fv_cell.value     = f"القيمة السوقية النهائية (EGP):  {fv:,.0f}"
    fv_cell.font      = _FONT_FINAL_VALUE
    fv_cell.fill      = _FILL_FINAL_VALUE
    fv_cell.alignment = Alignment(horizontal="center", vertical="center")
    fv_cell.border    = _BORDER_MEDIUM
    ws.row_dimensions[r].height = 24
    r += 2

    # ── Section 6: حدود ومحددات الاستخدام ────────────────────────────────────
    _sect("حدود ومحددات الاستخدام", "Limiting Conditions")
    for cond in (
        "يُعدّ هذا التقرير سارياً فقط بالغرض المُبيَّن أعلاه.",
        "لا يجوز الاستشهاد بجزء منه دون الرجوع إلى التقرير الكامل.",
        "القيمة محددة وفق أحوال السوق في تاريخ التقييم.",
        "لم يُراعَ في التقدير أي تكاليف بيع أو ضرائب.",
    ):
        ws.merge_cells(f"B{r}:D{r}")
        c           = ws.cell(row=r, column=2)
        c.value     = f"• {cond}"
        c.font      = Font(size=9)
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 16
        r += 1
    r += 1

    # ── Section 7: إعدادات التقرير ────────────────────────────────────────────
    _sect("إعدادات التقرير", "Report Settings")
    _inp("نمط التقرير",      "Legacy — أساسي")
    _inp("مستوى الثقة",     result.confidence if result else "—")
    _inp("المعايير المطبقة", "EGVS / IFRS 13")
    r += 2

    # ── Legend ────────────────────────────────────────────────────────────────
    ws.cell(row=r, column=2).value     = "دليل الألوان:"
    ws.cell(row=r, column=2).font      = Font(bold=True, size=9)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    r += 1
    for lbl, clr in (
        ("خلية إدخال يدوي",     _BP.INPUT_CELL),
        ("خلية محسوبة / ناتجة", _BP.CALC_CELL),
        ("القيمة النهائية",     _BP.SUCCESS_LIGHT),
    ):
        ws.cell(row=r, column=2).value     = lbl
        ws.cell(row=r, column=2).font      = Font(size=9)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
        ic        = ws.cell(row=r, column=3)
        ic.fill   = _gf(clr)
        ic.border = _BORDER_THIN
        r += 1

    ws.freeze_panes = "B3"
