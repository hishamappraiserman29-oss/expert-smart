#!/usr/bin/env python3
"""
certificate_sheet.py — Builder for شيت "شهادة" (Appraiser Certification).

ملاحظة: freeze_panes لا تُطبَّق على هذا الشيت
(apply_sheet_defaults تتجاهل freeze للشيت المسمى "شهادة").
"""
from __future__ import annotations

import io
import sys
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    NumFormat,
    Palette,
    Typography,
    apply_sheet_defaults,
    draw_banner,
    draw_section,
    get_alignment,
    get_fill,
    get_font,
)

_COL_A    = 1
_COL_B    = 2
_NUM_COLS = 2

_CERTIFICATION_TEXT = (
    "أشهد أنا الموقع أدناه بأن البيانات الواردة فى هذا التقرير صحيحة وصادقة "
    "وفق أفضل ما لديَّ من معرفة ومعلومات، وأن الاستنتاجات المُبيَّنة فيه تمثل "
    "رأيي المهني المستقل والمحايد، وقد توصلتُ إليها وفق المعايير المهنية المعتمدة.\n\n"
    "أُقرّ بأنه لا توجد لديَّ أى مصلحة حالية أو مستقبلية فى العقار موضوع التقرير، "
    "ولم أتلقَّ مقابلاً مرتبطاً بنتيجة التقييم.\n\n"
    "تم إعداد هذا التقرير وفقاً للمعايير المصرية للتقييم (EgVS) "
    "والمعيار الدولى للتقارير المالية (IFRS 13)، "
    "ومعايير التقييم الدولية (IVS 2022)."
)

_EGVS_REFS = (
    ("EgVS 1.0",   "تعريف القيمة السوقية"),
    ("EgVS 2.0",   "الافتراضات والقيود"),
    ("EgVS 3.0",   "أساليب التقييم الثلاثة"),
    ("IFRS 13",    "قياس القيمة العادلة"),
    ("IVS 105",    "معايير التقييم الدولية — الأساليب"),
    ("Basel III",  "متطلبات ضمانات الاقتراض"),
)

_DEFAULTS: dict[str, Any] = {
    "appraiser_name": "هشام محمد محمد المهدى",
    "license_id":     29,
    "authority":      "الهيئة العامة للرقابة المالية",
    "email":          "appraiserman29@gmail.com",
    "valuation_date": "2026/04/12",
    "primary_value":  4_899_960,
    "property_type":  "شقة سكنية — التجمع الخامس",
    "report_id":      "VAL-20260412-1153",
}


def apply_certificate_sheet(
    ws: Worksheet, data: dict | None = None
) -> dict[str, tuple[int, int]]:
    """
    Build شيت الشهادة.

    ملاحظة: يجب أن يكون ws.title == "شهادة" قبل الاستدعاء
    حتى تعمل apply_sheet_defaults بصورة صحيحة (بدون freeze).
    """
    d = {**_DEFAULTS, **(data or {})}
    apply_sheet_defaults(ws)
    locs: dict[str, tuple[int, int]] = {}

    draw_banner(
        ws, row=1, end_col=_NUM_COLS,
        text="شهادة المقيم — Appraiser Certification",
        bg=Palette.INK, fg=Palette.GOLD_LIGHT,
        size=Typography.SIZE_PAGE_TITLE - 2, height=46,
    )

    row = 3

    # ─── نص الشهادة ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:B{row + 5}")
    cert_cell = ws.cell(row=row, column=_COL_A, value=_CERTIFICATION_TEXT)
    cert_cell.font = get_font(size=11, color=Palette.INK)
    cert_cell.fill = get_fill(Palette.GREY_50)
    cert_cell.alignment = get_alignment(h="right", v="top", wrap=True)
    ws.row_dimensions[row].height = 120
    locs["cert_text"] = (row, _COL_A)
    row += 7

    # ─── بيانات التقييم ───────────────────────────────────────────────────
    draw_section(ws, row, _NUM_COLS, "بيانات التقييم", bg=Palette.NAVY_DEEP)
    row += 1

    def _kv_cert(key: str, label: str, fmt: str = "") -> None:
        nonlocal row
        lbl_cell = ws.cell(row=row, column=_COL_A, value=f"{label}:")
        lbl_cell.font = get_font(size=11, bold=True, color=Palette.INK)
        lbl_cell.alignment = get_alignment(h="right")

        val_cell = ws.cell(row=row, column=_COL_B, value=d.get(key, "—"))
        val_cell.font = get_font(size=11, color=Palette.NAVY)
        val_cell.alignment = get_alignment(h="right")
        if fmt:
            val_cell.number_format = fmt

        ws.row_dimensions[row].height = 26
        locs[key] = (row, _COL_B)
        row += 1

    _kv_cert("report_id",      "رقم التقرير")
    _kv_cert("property_type",  "وصف العقار")
    _kv_cert("primary_value",  "القيمة السوقية المُقدَّرة",
             fmt=NumFormat.CURRENCY)
    _kv_cert("valuation_date", "تاريخ التقييم")
    row += 1

    # ─── بيانات المُقيِّم ──────────────────────────────────────────────────
    draw_section(ws, row, _NUM_COLS, "بيانات المُقيِّم", bg=Palette.NAVY)
    row += 1
    _kv_cert("appraiser_name", "اسم المُقيِّم")
    _kv_cert("license_id",     "رقم القيد")
    _kv_cert("authority",      "الهيئة المرخصة")
    _kv_cert("email",          "البريد الإلكترونى")
    row += 1

    # ─── التوقيع ─────────────────────────────────────────────────────────
    draw_section(ws, row, _NUM_COLS, "التوقيع — Signature", bg=Palette.GOLD)
    row += 1

    for sig_label in (
        "التوقيع:",
        "الختم الرسمى:",
    ):
        ws.cell(row=row, column=_COL_A, value=sig_label)
        cell = ws.cell(row=row, column=_COL_A)
        cell.font = get_font(size=12, bold=True, color=Palette.INK)
        cell.alignment = get_alignment(h="right")
        ws.cell(row=row, column=_COL_B,
                value="________________________________")
        ws.row_dimensions[row].height = 36
        row += 2

    # ─── المراجع التنظيمية ────────────────────────────────────────────────
    row += 1
    draw_section(ws, row, _NUM_COLS, "المراجع التنظيمية — Regulatory References",
                 bg=Palette.GREY_700)
    row += 1
    for ref, desc in _EGVS_REFS:
        ws.cell(row=row, column=_COL_A, value=ref)
        cell_r = ws.cell(row=row, column=_COL_A)
        cell_r.font = get_font(size=10, bold=True, color=Palette.NAVY)
        cell_r.alignment = get_alignment(h="right")

        ws.cell(row=row, column=_COL_B, value=desc)
        cell_d = ws.cell(row=row, column=_COL_B)
        cell_d.font = get_font(size=10, color=Palette.INK)
        cell_d.alignment = get_alignment(h="right")

        ws.row_dimensions[row].height = 20
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    return locs


def get_field_count() -> int:
    return len(_DEFAULTS)


def get_section_count() -> int:
    return 4  # cert text + valuation + appraiser + signature


def _print_summary() -> None:
    print("\n  شهادة — Certificate Sheet")
    print("  " + "-" * 40)
    print(f"  أقسام: {get_section_count()}")
    print(f"  حقول:  {get_field_count()}")
    print(f"  المراجع التنظيمية: {len(_EGVS_REFS)}")
    print()


if __name__ == "__main__":
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace"
        )
    _print_summary()
