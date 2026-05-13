#!/usr/bin/env python3
"""
certification_sheet.py — Builder for شيت "شهادة"
(Appraiser Certification, Classic Office Blue theme).

استخدام:
    from core_engine.reports.sheets.certification_sheet import apply_certification_sheet

    locs = apply_certification_sheet(ws, inputs_dict, profile_key="legacy")
    # locs: dict[str, tuple[int, int]] — خريطة field_key → (row, col)
"""
from __future__ import annotations

from typing import Any, Mapping

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    BuilderPalette as _BP,
    NumFormat as _NF,
    Palette as _Palette,
    get_fill as _gf,
)

# ── Module-level style constants (Classic Office Blue) ────────────────────────
_FILL_HEADER     = _gf(_BP.HEADER)
_FILL_INPUT_CELL = _gf(_BP.INPUT_CELL)
_FONT_MUTED      = Font(italic=True, color=_BP.MUTED)
_ALIGN_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER_THIN     = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

_CERT_TEXT_AR = (
    "أشهد أنا الموقع أدناه بأن البيانات الواردة في هذا التقرير صحيحة وصادقة "
    "وفق أفضل ما لديَّ من معرفة ومعلومات، وأن الاستنتاجات المُبيَّنة فيه تمثل "
    "رأيي المهني المستقل والمحايد، وقد توصلتُ إليها وفق المعايير المهنية المعتمدة.\n\n"
    "أُقرّ بأنه لا توجد لديَّ أي مصلحة حالية أو مستقبلية في العقار موضوع التقرير.\n\n"
    "تم إعداد هذا التقرير وفقاً للمعايير المصرية للتقييم (EgVS) "
    "والمعيار الدولي للتقارير المالية (IFRS 13)."
)


def apply_certification_sheet(
    ws: Worksheet,
    inputs: Mapping[str, Any],
    *,
    profile_key: str = "legacy",
) -> dict[str, tuple[int, int]]:
    """
    Build شيت شهادة المقيم.

    Args:
        ws: الشيت المراد تعديله (يجب أن يكون فارغاً).
        inputs: قاموس يحتوى بيانات المقيم والتقرير. المفاتيح المدعومة:
                appraiser_name / reviewer_name, license_no, report_date.
        profile_key: "legacy" فقط مدعوم حالياً (محفوظ للتوسع).

    Returns:
        dict[str, tuple[int, int]] — خريطة field_key → (row, col) لخلايا القيم.
    """
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 70

    # ── Unpack inputs ─────────────────────────────────────────────────────────
    appraiser_name = str(
        inputs.get("appraiser_name") or inputs.get("reviewer_name") or
        "________________________________"
    )
    license_no  = str(inputs.get("license_no") or "________________________________")
    report_date = str(inputs.get("report_date") or "")

    locs: dict[str, tuple[int, int]] = {}

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    b           = ws["A1"]
    b.value     = "شهادة المقيم — Appraiser Certification"
    b.fill      = _FILL_HEADER
    b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
    b.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:B2")
    ws["A2"].value     = f"تاريخ التقرير: {report_date}"
    ws["A2"].font      = _FONT_MUTED
    ws["A2"].alignment = Alignment(horizontal="center")

    r = 4

    # ── Certification text block ──────────────────────────────────────────────
    ws.merge_cells(f"B{r}:B{r + 5}")
    cc           = ws.cell(row=r, column=2)
    cc.value     = _CERT_TEXT_AR
    cc.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    cc.fill      = _FILL_INPUT_CELL
    cc.border    = _BORDER_THIN
    ws.row_dimensions[r].height = 120
    locs["cert_text"] = (r, 2)
    r += 7

    # ── Signature block ───────────────────────────────────────────────────────
    for field_key, lbl, val in (
        ("appraiser_name", "اسم المقيم:",    appraiser_name),
        ("license_no",     "رقم الترخيص:", license_no),
        ("report_date",    "التاريخ:",       report_date),
        ("signature",      "التوقيع:",       "________________________________"),
    ):
        ws.row_dimensions[r].height = 22
        c           = ws.cell(row=r, column=2)
        c.value     = f"{lbl}  {val}"
        c.font      = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="right", vertical="center")
        locs[field_key] = (r, 2)
        r += 2

    return locs
