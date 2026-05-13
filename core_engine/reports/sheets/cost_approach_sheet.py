#!/usr/bin/env python3
"""
cost_approach_sheet.py — Builder for شيت "طريقة التكلفة"
(Cost Approach, Classic Office Blue theme).

استخدام:
    from core_engine.reports.sheets.cost_approach_sheet import apply_cost_approach_sheet

    locs = apply_cost_approach_sheet(ws, inputs_dict, profile_key="legacy")
    # locs: dict[str, tuple[int, int]] — خريطة field_key → (row, col)
"""
from __future__ import annotations

from typing import Any, Mapping

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    BuilderPalette as _BP,
    NumFormat as _NF,
    Palette as _Palette,
    get_fill as _gf,
)

# ── Module-level style constants (Classic Office Blue) ────────────────────────
_FILL_HEADER      = _gf(_BP.HEADER)
_FILL_INPUT_SECT  = _gf(_BP.SECTION_DARK)
_FILL_INPUT_CELL  = _gf(_BP.INPUT_CELL)
_FILL_CALC_CELL   = _gf(_BP.CALC_CELL)
_FILL_FINAL_VALUE = _gf(_BP.SUCCESS_LIGHT)
_FILL_ROW_BAND    = _gf(_BP.ROW_BAND)
_FONT_MUTED       = Font(italic=True, color=_BP.MUTED)
_FONT_FINAL_VALUE = Font(bold=True, size=12, color=_BP.SUCCESS_DARK)
_ALIGN_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER_THIN      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_BORDER_MEDIUM    = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)
_FMT_CURRENCY = _NF.CURRENCY_2DP
_FMT_PCT      = _NF.PERCENT_DETAILED


def apply_cost_approach_sheet(
    ws: Worksheet,
    inputs: Mapping[str, Any],
    *,
    profile_key: str = "legacy",
) -> dict[str, tuple[int, int]]:
    """
    Build شيت طريقة التكلفة.

    Args:
        ws: الشيت المراد تعديله (يجب أن يكون فارغاً).
        inputs: قاموس يحتوى بيانات منهج التكلفة. المفاتيح المدعومة:
                area / floor_area_m2, land_value / land_price,
                construction_cost_sqm (default 8000),
                construction_cost / replacement_cost,
                depreciation_rate / depreciation,
                additional_items, report_date.
        profile_key: "legacy" فقط مدعوم حالياً (محفوظ للتوسع).

    Returns:
        dict[str, tuple[int, int]] — خريطة field_key → (row, col) لخلايا القيم.
    """
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 16

    # ── Unpack inputs ─────────────────────────────────────────────────────────
    area       = float(inputs.get("area") or inputs.get("floor_area_m2") or 0)
    land_val   = float(inputs.get("land_value") or inputs.get("land_price") or 0)
    cost_sqm   = float(inputs.get("construction_cost_sqm") or 8000)
    const_cost = float(
        inputs.get("construction_cost") or
        inputs.get("replacement_cost") or
        (area * cost_sqm)
    )
    depr_pct   = float(inputs.get("depreciation_rate") or inputs.get("depreciation") or 0)
    add_items  = float(inputs.get("additional_items") or 0)
    report_date = str(inputs.get("report_date") or "")

    depr_amt  = const_cost * depr_pct
    indicated = land_val + const_cost + add_items - depr_amt

    locs: dict[str, tuple[int, int]] = {}

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    b           = ws["A1"]
    b.value     = "طريقة التكلفة — Cost Approach"
    b.fill      = _FILL_HEADER
    b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
    b.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    ws["A2"].value     = f"تاريخ التقرير: {report_date}  |  المعيار: EGVS 3.2 / IVS 105"
    ws["A2"].font      = _FONT_MUTED
    ws["A2"].alignment = Alignment(horizontal="center")

    r = 4

    def _sect(label: str) -> None:
        nonlocal r
        ws.merge_cells(f"A{r}:D{r}")
        c           = ws.cell(row=r, column=1)
        c.value     = f"  {label}"
        c.fill      = _FILL_INPUT_SECT
        c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

    def _row(label: str, value: Any, fmt: str = _FMT_CURRENCY,
             is_calc: bool = False, alt: bool = False,
             key: str = "") -> None:
        nonlocal r
        fill         = _FILL_CALC_CELL if is_calc else (_FILL_ROW_BAND if alt else _FILL_INPUT_CELL)
        lc           = ws.cell(row=r, column=2)
        lc.value     = label
        lc.font      = Font(bold=True, size=10)
        lc.fill      = fill
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border    = _BORDER_THIN
        vc           = ws.cell(row=r, column=3)
        vc.value     = value
        vc.number_format = fmt
        vc.alignment = _ALIGN_CENTER
        vc.fill      = fill
        vc.border    = _BORDER_THIN
        if key:
            locs[key] = (r, 3)
        ws.row_dimensions[r].height = 18
        r += 1

    # ── Section 1: قيمة الأرض ────────────────────────────────────────────────
    _sect("قيمة الأرض")
    _row("مساحة الأرض (م²)",            area,     '#,##0.00 "م²"', key="area")
    _row("قيمة الأرض الإجمالية (EGP)",  land_val,  key="land_value")
    r += 1

    # ── Section 2: تكلفة الإنشاء والبناء ─────────────────────────────────────
    _sect("تكلفة الإنشاء والبناء")
    _row("المساحة المبنية (م²)",              area,      '#,##0.00 "م²"')
    _row("تكلفة البناء للمتر (EGP/م²)",      cost_sqm,  alt=True,  key="cost_sqm")
    _row("إجمالي تكلفة البناء (EGP)",        const_cost, is_calc=True, key="const_cost")
    _row("بنود إضافية (EGP)",               add_items,  alt=True,  key="add_items")
    r += 1

    # ── Section 3: الاستهلاك والتقادم ────────────────────────────────────────
    _sect("الاستهلاك والتقادم")
    _row("نسبة الاستهلاك",          depr_pct, _FMT_PCT,   key="depr_pct")
    _row("قيمة الاستهلاك (EGP)",    depr_amt, is_calc=True, key="depr_amt")
    r += 1

    # ── Section 4: القيمة الاستدلالية ────────────────────────────────────────
    _sect("القيمة الاستدلالية بطريقة التكلفة")
    _row("المباني بعد الاستهلاك (EGP)",
         const_cost + add_items - depr_amt, is_calc=True, key="net_building_value")

    ws.merge_cells(f"B{r}:D{r}")
    fc           = ws.cell(row=r, column=2)
    fc.value     = f"القيمة الاستدلالية الإجمالية (EGP):  {indicated:,.0f}"
    fc.font      = _FONT_FINAL_VALUE
    fc.fill      = _FILL_FINAL_VALUE
    fc.alignment = Alignment(horizontal="center", vertical="center")
    fc.border    = _BORDER_MEDIUM
    ws.row_dimensions[r].height = 24
    locs["indicated_value"] = (r, 2)

    ws.freeze_panes = "B3"
    return locs
