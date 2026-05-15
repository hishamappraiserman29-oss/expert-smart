#!/usr/bin/env python3
"""
cost_method.py — Builder for شيت "طريقة التكلفة" (Cost Approach).

أربعة أقسام:
  1. قيمة الأرض
  2. تكلفة الإنشاء والبناء
  3. التهالك والاستهلاك
  4. القيمة الاستدلالية بطريقة التكلفة
"""
from __future__ import annotations

import io
import sys
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    NumFormat,
    Palette,
    Typography,
    apply_sheet_defaults,
    draw_banner,
    draw_section,
    get_alignment,
    get_fill,
    get_font,
    style_label,
    style_value,
)

_COL_LABEL = 1
_COL_VALUE = 2
_COL_NOTE  = 3
_NUM_COLS  = 3

_SECTION_NAMES = [
    "قيمة الأرض",
    "تكلفة الإنشاء والبناء",
    "التهالك والاستهلاك",
    "القيمة الاستدلالية بطريقة التكلفة",
]

_DEFAULTS: dict[str, Any] = {
    # Land
    "land_area_m2":         150.0,
    "land_price_per_m2":    15_000,
    "land_value":           2_250_000,
    # Construction
    "built_area_m2":        150.0,
    "construction_cost_m2": 12_000,
    "gross_construction":   1_800_000,
    "contractor_profit":    0.18,
    "total_construction":   2_124_000,
    # Depreciation
    "actual_age":           16,
    "economic_life":        60,
    "dep_physical":         0.027,
    "dep_functional":       0.05,
    "dep_external":         0.025,
    "dep_total":            0.096,
    "dep_amount":           204_000,
    "net_building_value":   1_920_000,
    # Final
    "indicated_value":      4_170_000,
    "valuation_date":       "2026/04/12",
}


def _compute_defaults(d: dict) -> None:
    """Fill derived fields if not supplied."""
    lv  = d["land_area_m2"] * d["land_price_per_m2"]
    gc  = d["built_area_m2"] * d["construction_cost_m2"]
    tc  = gc * (1 + d["contractor_profit"])
    dep = min(d["dep_physical"] + d["dep_functional"] + d["dep_external"], 0.80)
    da  = tc * dep
    nb  = tc - da
    d.setdefault("land_value",         lv)
    d.setdefault("gross_construction", gc)
    d.setdefault("total_construction", tc)
    d.setdefault("dep_total",          dep)
    d.setdefault("dep_amount",         da)
    d.setdefault("net_building_value", nb)
    d.setdefault("indicated_value",    lv + nb)


def apply_cost_method(
    ws: Worksheet, data: dict | None = None
) -> dict[str, tuple[int, int]]:
    """Build شيت طريقة التكلفة. Returns {key: (row, col)}."""
    d = {**_DEFAULTS, **(data or {})}
    _compute_defaults(d)
    apply_sheet_defaults(ws)
    locs: dict[str, tuple[int, int]] = {}
    row = [1]

    draw_banner(
        ws, row=1, end_col=_NUM_COLS,
        text="طريقة التكلفة — Cost Approach  |  EGVS 3.2 / IVS 105",
        bg=Palette.INK, fg=Palette.GOLD_LIGHT,
        size=Typography.SIZE_PAGE_TITLE - 2, height=46,
    )
    row[0] = 3

    def _section(title: str, bg: str = Palette.NAVY) -> None:
        draw_section(ws, row[0], _NUM_COLS, title, bg=bg)
        row[0] += 1

    def _kv(key: str, label: str, fmt: str = NumFormat.CURRENCY,
            note: str = "", is_calc: bool = False) -> None:
        r = row[0]
        cell_l = ws.cell(row=r, column=_COL_LABEL, value=label)
        style_label(cell_l)

        cell_v = ws.cell(row=r, column=_COL_VALUE, value=d.get(key, "—"))
        if is_calc:
            cell_v.font = get_font(size=10, bold=True, color=Palette.EMERALD)
            cell_v.fill = get_fill(Palette.EMERALD_VLT)
            cell_v.alignment = get_alignment(h="center")
        else:
            style_value(cell_v)
        if fmt:
            cell_v.number_format = fmt

        if note:
            cell_n = ws.cell(row=r, column=_COL_NOTE, value=note)
            cell_n.font = get_font(size=9, color=Palette.GREY_500, italic=True)
            cell_n.alignment = get_alignment(h="right")

        ws.row_dimensions[r].height = 24
        locs[key] = (r, _COL_VALUE)
        row[0] += 1

    # ─── Section 1: قيمة الأرض ───────────────────────────────────────────
    _section("قيمة الأرض", Palette.EMERALD)
    _kv("land_area_m2",       "مساحة الأرض",
        fmt='#,##0.00" م²"',  note="المساحة الصافية للأرض")
    _kv("land_price_per_m2",  "سعر المتر المرجعى للأرض",
        fmt=NumFormat.CURRENCY, note="من تحليل المقارنات")
    _kv("land_value",         "قيمة الأرض الإجمالية",
        is_calc=True,         note="= مساحة × سعر المتر")
    row[0] += 1

    # ─── Section 2: تكلفة الإنشاء ────────────────────────────────────────
    _section("تكلفة الإنشاء والبناء", Palette.INDIGO)
    _kv("built_area_m2",      "المساحة المبنية",
        fmt='#,##0.00" م²"',  note="مساحة البناء الإجمالية")
    _kv("construction_cost_m2","تكلفة الإنشاء للمتر (EGP/م²)",
        fmt=NumFormat.CURRENCY, note="أسعار السوق الحالية")
    _kv("gross_construction",  "إجمالى تكلفة الإنشاء",
        is_calc=True,          note="= مساحة × تكلفة/م²")
    _kv("contractor_profit",   "ربح المقاول والمصروفات الإدارية",
        fmt=NumFormat.PERCENT,  note="15-20% معيار السوق")
    _kv("total_construction",  "إجمالى التكلفة مع الربح",
        is_calc=True,          note="= إجمالى × (1 + ربح)")
    row[0] += 1

    # ─── Section 3: التهالك ───────────────────────────────────────────────
    _section("التهالك والاستهلاك", Palette.CORAL)
    _kv("actual_age",    "العمر الفعلى للمبنى",
        fmt=NumFormat.YEAR + '" سنة"', note="منذ إنشاء المبنى")
    _kv("economic_life", "العمر الاقتصادى الافتراضى",
        fmt=NumFormat.YEAR + '" سنة"', note="60 سنة للخرسانى")
    _kv("dep_physical",  "التهالك المادى (Physical)",
        fmt=NumFormat.PERCENT,  note="تدهور مادى / عمر")
    _kv("dep_functional","التهالك الوظيفى (Functional)",
        fmt=NumFormat.PERCENT,  note="تقادم التصميم")
    _kv("dep_external",  "التهالك الخارجى (External)",
        fmt=NumFormat.PERCENT,  note="عوامل خارج العقار")
    _kv("dep_total",     "إجمالى التهالك (max 80%)",
        fmt=NumFormat.PERCENT,  is_calc=True)
    _kv("dep_amount",    "قيمة التهالك (EGP)",
        is_calc=True,           note="= إجمالى التكلفة × نسبة التهالك")
    _kv("net_building_value","قيمة المبانى بعد التهالك",
        is_calc=True,           note="= إجمالى التكلفة − التهالك")
    row[0] += 1

    # ─── Section 4: القيمة الاستدلالية ───────────────────────────────────
    _section("القيمة الاستدلالية بطريقة التكلفة", Palette.GOLD)
    _kv("land_value",        "قيمة الأرض",          fmt=NumFormat.CURRENCY)
    _kv("net_building_value","قيمة المبانى بعد التهالك", fmt=NumFormat.CURRENCY)

    r = row[0]
    cell_lbl = ws.cell(row=r, column=_COL_LABEL, value="القيمة الاستدلالية الإجمالية")
    cell_lbl.font = get_font(
        size=Typography.SIZE_SECTION_HEADER, bold=True, color=Palette.INK
    )
    cell_lbl.fill = get_fill(Palette.GOLD_PALE)
    cell_lbl.alignment = get_alignment(h="right")

    cell_val = ws.cell(row=r, column=_COL_VALUE, value=d["indicated_value"])
    cell_val.font = get_font(
        size=Typography.SIZE_KPI_VALUE, bold=True, color=Palette.NAVY
    )
    cell_val.fill = get_fill(Palette.GOLD_PALE)
    cell_val.alignment = get_alignment(h="center")
    cell_val.number_format = NumFormat.CURRENCY
    ws.row_dimensions[r].height = 40
    locs["indicated_value"] = (r, _COL_VALUE)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.freeze_panes = "A3"
    return locs


def get_field_count() -> int:
    return len([k for k in _DEFAULTS if k != "valuation_date"])


def get_section_count() -> int:
    return len(_SECTION_NAMES)


def _print_summary() -> None:
    print("\n  طريقة التكلفة — Cost Approach Sheet")
    print("  " + "-" * 40)
    print(f"  أقسام: {get_section_count()}")
    print(f"  حقول:  {get_field_count()}")
    print()
    for s in _SECTION_NAMES:
        print(f"  • {s}")
    print()


if __name__ == "__main__":
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace"
        )
    _print_summary()
