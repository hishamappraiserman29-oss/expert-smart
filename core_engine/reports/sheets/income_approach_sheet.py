#!/usr/bin/env python3
"""
income_approach_sheet.py — Builder for شيت "رأسمالة الدخل"
(Income Capitalization Approach, Classic Office Blue theme).

استخدام:
    from core_engine.reports.sheets.income_approach_sheet import apply_income_approach_sheet

    locs = apply_income_approach_sheet(ws, inputs_dict, profile_key="legacy")
    # locs: dict[str, tuple[int, int]] — خريطة field_key → (row, col)
"""
from __future__ import annotations

from typing import Any, Mapping

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    BuilderPalette as _BP,
    NumFormat as _NF,
    Palette as _Palette,
    get_fill as _gf,
)

# ── Module-level style constants (Classic Office Blue) ────────────────────────
_FILL_HEADER      = _gf(_BP.HEADER)
_FILL_INPUT_SECT  = _gf(_BP.SECTION_DARK)
_FILL_INPUT_CELL  = _gf(_BP.INPUT_CELL)
_FILL_CALC_CELL   = _gf(_BP.CALC_CELL)
_FILL_FINAL_VALUE = _gf(_BP.SUCCESS_LIGHT)
_FILL_ROW_BAND    = _gf(_BP.ROW_BAND)
_FONT_MUTED       = Font(italic=True, color=_BP.MUTED)
_FONT_FINAL_VALUE = Font(bold=True, size=12, color=_BP.SUCCESS_DARK)
_ALIGN_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER_THIN      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_BORDER_MEDIUM    = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)
_FMT_CURRENCY = _NF.CURRENCY_2DP
_FMT_PCT      = _NF.PERCENT_DETAILED


def apply_income_approach_sheet(
    ws: Worksheet,
    inputs: Mapping[str, Any],
    *,
    profile_key: str = "legacy",
) -> dict[str, tuple[int, int]]:
    """
    Build شيت رأسمالة الدخل.

    Args:
        ws: الشيت المراد تعديله (يجب أن يكون فارغاً).
        inputs: قاموس يحتوى بيانات منهج رأسمالة الدخل. المفاتيح المدعومة:
                area / floor_area_m2, rent_sqm / rent_per_sqm,
                vacancy_rate (default 0.05), management_rate (default 0.05),
                maintenance_rate (default 0.02), tax_rate (default 0.01),
                cap_rate / capitalization_rate (default 0.08),
                gross_income (optional override), report_date.
        profile_key: "legacy" فقط مدعوم حالياً (محفوظ للتوسع).

    Returns:
        dict[str, tuple[int, int]] — خريطة field_key → (row, col) لخلايا القيم.
    """
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 16

    # ── Unpack inputs ─────────────────────────────────────────────────────────
    area       = float(inputs.get("area") or inputs.get("floor_area_m2") or 0)
    rent_sqm   = float(inputs.get("rent_sqm") or inputs.get("rent_per_sqm") or 0)
    vacancy    = float(inputs.get("vacancy_rate") or 0.05)
    mgmt_pct   = float(inputs.get("management_rate") or 0.05)
    maint_pct  = float(inputs.get("maintenance_rate") or 0.02)
    tax_pct    = float(inputs.get("tax_rate") or 0.01)
    cap_rate   = float(inputs.get("cap_rate") or inputs.get("capitalization_rate") or 0.08)
    report_date = str(inputs.get("report_date") or "")

    gross_income = float(
        inputs.get("gross_income") or
        (area * rent_sqm * 12 if area and rent_sqm else 0)
    )
    vac_loss  = gross_income * vacancy
    egi       = gross_income - vac_loss
    total_exp = egi * (mgmt_pct + maint_pct + tax_pct)
    noi       = egi - total_exp
    indicated = noi / cap_rate if cap_rate else 0

    locs: dict[str, tuple[int, int]] = {}

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    b           = ws["A1"]
    b.value     = "رأسمالة الدخل — Income Capitalization"
    b.fill      = _FILL_HEADER
    b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
    b.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    ws["A2"].value     = f"تاريخ التقرير: {report_date}  |  المعيار: EGVS 3.3 / IVS 105"
    ws["A2"].font      = _FONT_MUTED
    ws["A2"].alignment = Alignment(horizontal="center")

    r = 4

    def _sect(label: str) -> None:
        nonlocal r
        ws.merge_cells(f"A{r}:D{r}")
        c           = ws.cell(row=r, column=1)
        c.value     = f"  {label}"
        c.fill      = _FILL_INPUT_SECT
        c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

    def _row(label: str, value: Any, fmt: str = _FMT_CURRENCY,
             is_calc: bool = False, alt: bool = False,
             key: str = "") -> None:
        nonlocal r
        fill         = _FILL_CALC_CELL if is_calc else (_FILL_ROW_BAND if alt else _FILL_INPUT_CELL)
        lc           = ws.cell(row=r, column=2)
        lc.value     = label
        lc.font      = Font(bold=True, size=10)
        lc.fill      = fill
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border    = _BORDER_THIN
        vc           = ws.cell(row=r, column=3)
        vc.value     = value
        vc.number_format = fmt
        vc.alignment = _ALIGN_CENTER
        vc.fill      = fill
        vc.border    = _BORDER_THIN
        if key:
            locs[key] = (r, 3)
        ws.row_dimensions[r].height = 18
        r += 1

    # ── Section 1: بيانات الدخل الإجمالي ─────────────────────────────────────
    _sect("بيانات الدخل الإجمالي")
    _row("المساحة القابلة للإيجار (م²)", area, '#,##0.00 "م²"', key="area")
    _row("معدل الإيجار (EGP/م²/شهر)", rent_sqm, alt=True, key="rent_sqm")
    _row("الدخل الإجمالي السنوي (EGP)", gross_income, is_calc=True, key="gross_income")
    r += 1

    # ── Section 2: الشاغر والخسائر ───────────────────────────────────────────
    _sect("الشاغر والخسائر")
    _row("نسبة الشاغر", vacancy, _FMT_PCT, key="vacancy_rate")
    _row("خسارة الشاغر (EGP)", vac_loss, is_calc=True, alt=True, key="vac_loss")
    _row("الدخل الفعلي الإجمالي — EGI (EGP)", egi, is_calc=True, key="egi")
    r += 1

    # ── Section 3: المصروفات والنتيجة ────────────────────────────────────────
    _sect("المصروفات التشغيلية والنتيجة")
    _row("إجمالي المصروفات (EGP)", total_exp, is_calc=True, key="total_exp")
    _row("صافي الدخل التشغيلي — NOI (EGP)", noi, is_calc=True, alt=True, key="noi")
    _row("معدل الرسملة", cap_rate, _FMT_PCT, key="cap_rate")

    # ── Final value ───────────────────────────────────────────────────────────
    ws.merge_cells(f"B{r}:D{r}")
    fc           = ws.cell(row=r, column=2)
    fc.value     = f"القيمة الاستدلالية (EGP):  {indicated:,.0f}"
    fc.font      = _FONT_FINAL_VALUE
    fc.fill      = _FILL_FINAL_VALUE
    fc.alignment = Alignment(horizontal="center", vertical="center")
    fc.border    = _BORDER_MEDIUM
    ws.row_dimensions[r].height = 24
    locs["indicated_value"] = (r, 2)

    ws.freeze_panes = "B3"
    return locs
