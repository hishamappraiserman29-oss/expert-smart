#!/usr/bin/env python3
"""
inputs_sheet.py — Builder for شيت "الافتراضات والمدخلات"
(Assumptions & Inputs, Classic Office Blue theme).

استخدام:
    from core_engine.reports.sheets.inputs_sheet import apply_inputs_sheet

    wb = Workbook()
    ws = wb.create_sheet("الافتراضات والمدخلات")
    locs = apply_inputs_sheet(ws, inputs_dict, profile_key="legacy")
    # locs: dict[str, tuple[int, int]] — خريطة field_key → (row, col)
"""
from __future__ import annotations

from typing import Any, Mapping

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    BuilderPalette as _BP,
    NumFormat as _NF,
    Palette as _Palette,
    get_fill as _gf,
)

# ── Module-level style constants (Classic Office Blue) ────────────────────────
_FILL_HEADER      = _gf(_BP.HEADER)
_FILL_INPUT_SECT  = _gf(_BP.SECTION_DARK)
_FILL_INPUT_CELL  = _gf(_BP.INPUT_CELL)
_FILL_CALC_CELL   = _gf(_BP.CALC_CELL)
_FILL_FINAL_VALUE = _gf(_BP.SUCCESS_LIGHT)
_FONT_MUTED       = Font(italic=True, color=_BP.MUTED)
_FONT_FINAL_VALUE = Font(bold=True, size=12, color=_BP.SUCCESS_DARK)
_ALIGN_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER_THIN      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_BORDER_MEDIUM    = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)
_FMT_CURRENCY = _NF.CURRENCY_2DP
_FMT_PCT      = _NF.PERCENT_DETAILED


def apply_inputs_sheet(
    ws: Worksheet,
    inputs: Mapping[str, Any],
    *,
    profile_key: str = "legacy",
) -> dict[str, tuple[int, int]]:
    """
    Build شيت الافتراضات والمدخلات.

    Args:
        ws: الشيت المراد تعديله (يجب أن يكون فارغاً).
        inputs: قاموس يحتوى بيانات التقرير والعقار والتقييم. المفاتيح المعترَف
                بها موثّقة فى المصدر أدناه — أى مفتاح غير موجود يُعامَل كـ "—".
        profile_key: "legacy" فقط مدعوم حالياً (محفوظ للتوسع مستقبلاً).

    Returns:
        dict[str, tuple[int, int]] — خريطة field_key → (row, col) لموقع خلية
        القيمة لكل حقل مُتتبَّع.
    """
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22

    # ── Flatten inputs ────────────────────────────────────────────────────────
    fv          = float(inputs.get("primary_value") or 0)
    report_date = str(inputs.get("report_date") or "")

    # Weights: support both nested {"weights": {"comparable": ...}} and flat keys
    _w          = inputs.get("weights") or {}
    w_comp      = float(_w.get("comparable", inputs.get("weights_comparable", 0)))
    w_cost      = float(_w.get("cost",       inputs.get("weights_cost", 0)))
    w_income    = float(_w.get("income",     inputs.get("weights_income", 0)))

    locs: dict[str, tuple[int, int]] = {}

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    t           = ws["A1"]
    t.value     = "الافتراضات والمدخلات — Assumptions & Inputs"
    t.fill      = _FILL_HEADER
    t.font      = Font(bold=True, color=_Palette.WHITE, size=14)
    t.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    ws["A2"].value     = f"تاريخ التقرير: {report_date}"
    ws["A2"].font      = _FONT_MUTED
    ws["A2"].alignment = Alignment(horizontal="center")

    r = 4

    def _sect(label: str, en: str = "") -> None:
        nonlocal r
        ws.merge_cells(f"A{r}:D{r}")
        c           = ws.cell(row=r, column=1)
        c.value     = f"  {label}" + (f"  — {en}" if en else "")
        c.fill      = _FILL_INPUT_SECT
        c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

    def _inp(label: str, value: Any, note: str = "", is_calc: bool = False,
             fmt: str = "", key: str = "") -> None:
        nonlocal r
        fill = _FILL_CALC_CELL if is_calc else _FILL_INPUT_CELL
        lc           = ws.cell(row=r, column=2)
        lc.value     = label
        lc.font      = Font(bold=True, size=10)
        lc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        lc.fill      = fill
        lc.border    = _BORDER_THIN

        vc           = ws.cell(row=r, column=3)
        vc.value     = value if value is not None else "—"
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill      = fill
        vc.border    = _BORDER_THIN
        if fmt:
            vc.number_format = fmt

        if note:
            nc           = ws.cell(row=r, column=4)
            nc.value     = note
            nc.font      = Font(italic=True, color=_BP.NOTE, size=9)
            nc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            nc.fill      = fill
            nc.border    = _BORDER_THIN

        if key:
            locs[key] = (r, 3)

        ws.row_dimensions[r].height = 18
        r += 1

    # ── Section 1: بيانات التقرير ─────────────────────────────────────────────
    _sect("بيانات التقرير", "Report Data")
    _inp("رقم التقرير",     inputs.get("report_id") or "—",          key="report_id")
    _inp("تاريخ التقييم",  inputs.get("valuation_date") or report_date, key="valuation_date")
    _inp("تاريخ المعاينة", inputs.get("inspection_date") or "—",     key="inspection_date")
    _inp("غرض التقييم",    inputs.get("primary_purpose", "—"),        key="primary_purpose")
    _inp("أساس القيمة",    "القيمة السوقية")
    r += 1

    # ── Section 2: بيانات العميل والمقيم ─────────────────────────────────────
    _sect("بيانات العميل والمقيم", "Client & Appraiser")
    _inp("اسم العميل",
         inputs.get("client_name") or inputs.get("borrower_name") or "—",
         key="client_name")
    _inp("اسم المقيم",
         inputs.get("appraiser_name") or inputs.get("reviewer_name") or "—",
         key="appraiser_name")
    _inp("رقم الترخيص",  inputs.get("license_no") or "—",   key="license_no")
    _inp("الجهة المعينة", inputs.get("instructed_by") or "—", key="instructed_by")
    r += 1

    # ── Section 3: بيانات العقار ──────────────────────────────────────────────
    _sect("بيانات العقار", "Property Data")
    _inp("نوع الأصل",
         inputs.get("asset_type", "—"),                               key="asset_type")
    _inp("الموقع / العنوان",
         inputs.get("location") or inputs.get("address") or "—",     key="location")
    _inp("المساحة (م²)",
         inputs.get("area") or inputs.get("floor_area_m2") or "—",   key="area")
    _inp("سنة الإنشاء",  inputs.get("year_built") or "—",            key="year_built")
    _inp("الحالة",       inputs.get("condition") or "—",             key="condition")
    r += 1

    # ── Section 4: بيانات السوق ───────────────────────────────────────────────
    _sect("بيانات السوق", "Market Data")
    _inp("متوسط سعر السوق (EGP/م²)",
         inputs.get("market_avg_price_sqm") or "—",                  key="market_avg_price_sqm")
    _inp("معدل الرسملة",  inputs.get("cap_rate") or "—",             key="cap_rate")
    _inp("معدل الشاغر",  inputs.get("vacancy_rate") or "—",          key="vacancy_rate")
    _inp("عدد المقارنات",
         inputs.get("comparables_count") or
         len(inputs.get("comparables") or []) or "—",                key="comparables_count")
    r += 1

    # ── Section 5: افتراضات التقييم ───────────────────────────────────────────
    _sect("افتراضات التقييم", "Valuation Assumptions")
    comp = float(inputs.get("comparable") or 0)
    cost = float(inputs.get("cost")       or 0)
    inc  = float(inputs.get("income")     or 0)

    _inp("القيمة — المقارنة البيعية (EGP)", comp, is_calc=True, fmt=_FMT_CURRENCY,
         key="comparable_value")
    _inp("الوزن — المقارنة",               w_comp, fmt=_FMT_PCT, key="weight_comparable")
    _inp("القيمة — طريقة التكلفة (EGP)",   cost, is_calc=True, fmt=_FMT_CURRENCY,
         key="cost_value")
    _inp("الوزن — التكلفة",                w_cost, fmt=_FMT_PCT, key="weight_cost")
    _inp("القيمة — رأسمالة الدخل (EGP)",   inc,  is_calc=True, fmt=_FMT_CURRENCY,
         key="income_value")
    _inp("الوزن — الدخل",                  w_income, fmt=_FMT_PCT, key="weight_income")
    r += 1

    # Final value banner
    ws.merge_cells(f"B{r}:D{r}")
    fv_cell           = ws.cell(row=r, column=2)
    fv_cell.value     = f"القيمة السوقية النهائية (EGP):  {fv:,.0f}"
    fv_cell.font      = _FONT_FINAL_VALUE
    fv_cell.fill      = _FILL_FINAL_VALUE
    fv_cell.alignment = Alignment(horizontal="center", vertical="center")
    fv_cell.border    = _BORDER_MEDIUM
    ws.row_dimensions[r].height = 24
    locs["primary_value"] = (r, 2)
    r += 2

    # ── Section 6: حدود ومحددات الاستخدام ────────────────────────────────────
    _sect("حدود ومحددات الاستخدام", "Limiting Conditions")
    for cond in (
        "يُعدّ هذا التقرير سارياً فقط بالغرض المُبيَّن أعلاه.",
        "لا يجوز الاستشهاد بجزء منه دون الرجوع إلى التقرير الكامل.",
        "القيمة محددة وفق أحوال السوق في تاريخ التقييم.",
        "لم يُراعَ في التقدير أي تكاليف بيع أو ضرائب.",
    ):
        ws.merge_cells(f"B{r}:D{r}")
        c           = ws.cell(row=r, column=2)
        c.value     = f"• {cond}"
        c.font      = Font(size=9)
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 16
        r += 1
    r += 1

    # ── Section 7: إعدادات التقرير ────────────────────────────────────────────
    _sect("إعدادات التقرير", "Report Settings")
    _inp("نمط التقرير",      "Legacy — أساسي")
    _inp("مستوى الثقة",     inputs.get("confidence", "—"), key="confidence")
    _inp("المعايير المطبقة", "EGVS / IFRS 13")
    r += 2

    # ── Legend ────────────────────────────────────────────────────────────────
    ws.cell(row=r, column=2).value     = "دليل الألوان:"
    ws.cell(row=r, column=2).font      = Font(bold=True, size=9)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    r += 1
    for lbl, clr in (
        ("خلية إدخال يدوي",     _BP.INPUT_CELL),
        ("خلية محسوبة / ناتجة", _BP.CALC_CELL),
        ("القيمة النهائية",     _BP.SUCCESS_LIGHT),
    ):
        ws.cell(row=r, column=2).value     = lbl
        ws.cell(row=r, column=2).font      = Font(size=9)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
        ic        = ws.cell(row=r, column=3)
        ic.fill   = _gf(clr)
        ic.border = _BORDER_THIN
        r += 1

    ws.freeze_panes = "B3"
    return locs
