#!/usr/bin/env python3
"""
main_report_sheet.py — Builder for شيت "التقرير"
(Executive Main Report, Classic Office Blue theme).
"""
from __future__ import annotations

from typing import Any, Optional

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    BuilderPalette as _BP,
    NumFormat as _NF,
    Palette as _Palette,
    get_fill as _gf,
)

# ── Module-level style constants (Classic Office Blue) ────────────────────────
_FILL_HEADER      = _gf(_BP.HEADER)
_FILL_SECTION     = _gf(_BP.SECTION_MID)
_FILL_INPUT_SECT  = _gf(_BP.SECTION_DARK)
_FILL_FINAL_VALUE = _gf(_BP.SUCCESS_LIGHT)
_FILL_ROW_BAND    = _gf(_BP.ROW_BAND)
_FILL_SUBHEAD     = _gf(_BP.SUBHEAD)
_FONT_BOLD        = Font(bold=True)
_FONT_FINAL_VALUE = Font(bold=True, size=12, color=_BP.SUCCESS_DARK)
_ALIGN_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER_THIN      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_BORDER_MEDIUM    = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)
_FMT_CURRENCY = _NF.CURRENCY_2DP
_FMT_PCT      = _NF.PERCENT_DETAILED


def apply_main_report_sheet(
    ws: Worksheet,
    result: Optional[Any] = None,
    report_date: str = "",
) -> None:
    """Build شيت 'التقرير' (Executive Main Report)."""
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    md   = result.metadata if result else {}
    fv   = float(result.primary_value) if result and result.primary_value else 0
    comp = float(md.get("comparable") or 0)
    cost = float(md.get("cost")       or 0)
    inc  = float(md.get("income")     or 0)
    w    = result.weights_applied if result else {}

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    t           = ws["A1"]
    t.value     = "تقرير التقييم العقاري"
    t.fill      = _FILL_HEADER
    t.font      = Font(bold=True, color=_Palette.WHITE, size=16)
    t.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    ws["A2"].value     = "Real Estate Valuation Report"
    ws["A2"].fill      = _FILL_SECTION
    ws["A2"].font      = Font(bold=True, color=_Palette.WHITE, size=12, italic=True)
    ws["A2"].alignment = _ALIGN_CENTER
    ws.row_dimensions[2].height = 22

    asset_type = result.asset_type if result else "—"
    confidence = result.confidence if result else "—"
    ws.merge_cells("A3:E3")
    ws["A3"].value     = (f"تاريخ التقرير: {report_date}  |  "
                          f"نوع الأصل: {asset_type}  |  الثقة: {confidence}")
    ws["A3"].fill      = _FILL_INPUT_SECT
    ws["A3"].font      = Font(italic=True, color=_Palette.WHITE, size=10)
    ws["A3"].alignment = _ALIGN_CENTER
    ws.row_dimensions[3].height = 18

    # ── Final Value card ──────────────────────────────────────────────────────
    ws.merge_cells("B5:E6")
    fv_cell           = ws["B5"]
    fv_cell.value     = f"القيمة السوقية النهائية\n{fv:,.0f}  جنيه مصري"
    fv_cell.fill      = _FILL_FINAL_VALUE
    fv_cell.font      = Font(bold=True, size=14, color=_BP.SUCCESS_DARK)
    fv_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fv_cell.border    = _BORDER_MEDIUM
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 28

    r = 8

    def _section_hd(label: str) -> None:
        nonlocal r
        ws.merge_cells(f"B{r}:E{r}")
        c           = ws.cell(row=r, column=2)
        c.value     = f"  {label}"
        c.fill      = _FILL_INPUT_SECT
        c.font      = Font(bold=True, color=_Palette.WHITE, size=11)
        c.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

    def _kv(label: str, value: Any, currency: bool = False,
            pct: bool = False, alt: bool = False) -> None:
        nonlocal r
        fill = _FILL_ROW_BAND if alt else None
        lc           = ws.cell(row=r, column=2)
        lc.value     = label
        lc.font      = Font(bold=True, size=10)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border    = _BORDER_THIN
        if fill:
            lc.fill = fill

        vc           = ws.cell(row=r, column=3)
        vc.value     = value if value is not None else "—"
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = _BORDER_THIN
        if fill:
            vc.fill = fill
        if currency:
            vc.number_format = _FMT_CURRENCY
        if pct:
            vc.number_format = _FMT_PCT
        ws.row_dimensions[r].height = 18
        r += 1

    # ── Property basics ───────────────────────────────────────────────────────
    _section_hd("بيانات العقار الأساسية")
    _kv("نوع الأصل",      asset_type)
    _kv("الموقع",         md.get("location") or md.get("address") or "—", alt=True)
    _kv("غرض التقييم",   result.primary_purpose if result else "—")
    _kv("تاريخ التقييم", md.get("valuation_date") or report_date, alt=True)
    _kv("مستوى الثقة",   confidence)
    r += 1

    # ── Three approaches table ────────────────────────────────────────────────
    _section_hd("نتائج أساليب التقييم الثلاثة")

    for col, hdr in enumerate(("الأسلوب", "القيمة (EGP)", "الوزن"), 2):
        hc           = ws.cell(row=r, column=col)
        hc.value     = hdr
        hc.fill      = _FILL_SUBHEAD
        hc.font      = Font(bold=True, size=10)
        hc.alignment = _ALIGN_CENTER
        hc.border    = _BORDER_THIN
        ws.row_dimensions[r].height = 18
    ws.cell(row=r, column=5).fill   = _FILL_SUBHEAD
    ws.cell(row=r, column=5).border = _BORDER_THIN
    r += 1

    for i, (lbl, val, wk) in enumerate((
        ("المقارنة البيعية", comp, "comparable"),
        ("طريقة التكلفة",   cost, "cost"),
        ("رأسمالة الدخل",   inc,  "income"),
    )):
        fill = _FILL_ROW_BAND if i % 2 else None
        nc           = ws.cell(row=r, column=2)
        nc.value     = lbl
        nc.font      = Font(bold=True, size=10)
        nc.alignment = Alignment(horizontal="right", vertical="center")
        nc.border    = _BORDER_THIN
        if fill:
            nc.fill = fill

        vc               = ws.cell(row=r, column=3)
        vc.value         = val
        vc.number_format = _FMT_CURRENCY
        vc.alignment     = _ALIGN_CENTER
        vc.border        = _BORDER_THIN
        if fill:
            vc.fill = fill

        wc               = ws.cell(row=r, column=4)
        wc.value         = w.get(wk, 0)
        wc.number_format = _FMT_PCT
        wc.alignment     = _ALIGN_CENTER
        wc.border        = _BORDER_THIN
        if fill:
            wc.fill = fill

        ws.row_dimensions[r].height = 18
        r += 1

    # Reconciled total
    rc           = ws.cell(row=r, column=2)
    rc.value     = "القيمة التوفيقية النهائية"
    rc.font      = _FONT_FINAL_VALUE
    rc.fill      = _FILL_FINAL_VALUE
    rc.alignment = Alignment(horizontal="right", vertical="center")
    rc.border    = _BORDER_MEDIUM

    vc               = ws.cell(row=r, column=3)
    vc.value         = fv
    vc.number_format = _FMT_CURRENCY
    vc.font          = _FONT_FINAL_VALUE
    vc.fill          = _FILL_FINAL_VALUE
    vc.alignment     = _ALIGN_CENTER
    vc.border        = _BORDER_MEDIUM
    ws.row_dimensions[r].height = 22
    r += 2

    # ── Simple bar chart ──────────────────────────────────────────────────────
    chart_data_row = r
    for col, hdr in enumerate(("الأسلوب", "القيمة"), 2):
        ws.cell(row=r, column=col).value = hdr
        ws.cell(row=r, column=col).font  = _FONT_BOLD
    r += 1
    for lbl, val in (
        ("المقارنة", comp),
        ("التكلفة",  cost),
        ("الدخل",    inc),
        ("النهائية", fv),
    ):
        ws.cell(row=r, column=2).value          = lbl
        ws.cell(row=r, column=3).value          = val
        ws.cell(row=r, column=3).number_format  = _FMT_CURRENCY
        r += 1
    chart_last_row = r - 1

    try:
        from openpyxl.chart import BarChart, Reference
        chart              = BarChart()
        chart.type         = "col"
        chart.title        = "مقارنة أساليب التقييم"
        chart.y_axis.title = "القيمة (EGP)"
        chart.style        = 10
        chart.width        = 18
        chart.height       = 12
        data = Reference(ws, min_col=3, min_row=chart_data_row,
                         max_row=chart_last_row)
        cats = Reference(ws, min_col=2, min_row=chart_data_row + 1,
                         max_row=chart_last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"B{r + 1}")
    except Exception:
        pass  # chart is optional

    ws.freeze_panes = "B4"
