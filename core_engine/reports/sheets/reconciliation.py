#!/usr/bin/env python3
"""
reconciliation.py — Builder for شيت "توفيق النتائج".

EGVS 3.0 / IVS 105 — Reconciliation of Value Indications.
"""
from __future__ import annotations

import io
import sys
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    NumFormat,
    Palette,
    Typography,
    apply_sheet_defaults,
    draw_banner,
    draw_section,
    get_alignment,
    get_fill,
    get_font,
    style_label,
    style_value,
)

_COL_METHOD   = 1
_COL_VALUE    = 2
_COL_WEIGHT   = 3
_COL_WEIGHTED = 4
_COL_NOTE     = 5
_NUM_COLS     = 5

_SECTION_NAMES = [
    "ملخص نتائج الأساليب",
    "القيمة السوقية النهائية",
    "ملاحظات التوفيق",
]

_DEFAULTS: dict[str, Any] = {
    "comparable_value": 4_989_300,
    "cost_value":       4_800_000,
    "income_value":     4_750_000,
    "w_comparable":     0.60,
    "w_cost":           0.20,
    "w_income":         0.20,
    "primary_value":    4_899_960,
    "price_per_m2":     32_666,
    "area_total":       150,
    "confidence":       "عالى",
    "valuation_date":   "2026/04/12",
}

_RECONCILIATION_NOTES = (
    "• تم إعطاء الوزن الأكبر لأسلوب المقارنة البيعية لتوافر بيانات السوق وكفايتها.",
    "• تم دعم النتيجة بأسلوب التكلفة وأسلوب الدخل لتأكيد القيمة.",
    "• تتوافق نتائج الأساليب الثلاثة مع مؤشرات السوق الحالية.",
)


def apply_reconciliation(
    ws: Worksheet, data: dict | None = None
) -> dict[str, tuple[int, int]]:
    """Build شيت توفيق النتائج. Returns {key: (row, col)}."""
    d = {**_DEFAULTS, **(data or {})}
    apply_sheet_defaults(ws)
    locs: dict[str, tuple[int, int]] = {}
    row = [1]

    draw_banner(
        ws, row=1, end_col=_NUM_COLS,
        text="توفيق النتائج — Reconciliation of Value Indications  |  EGVS 3.0",
        bg=Palette.INK, fg=Palette.GOLD_LIGHT,
        size=Typography.SIZE_PAGE_TITLE - 2, height=46,
    )
    row[0] = 3

    draw_section(ws, row[0], _NUM_COLS, "ملخص نتائج الأساليب", bg=Palette.NAVY_DEEP)
    row[0] += 1

    # ─── Table header ─────────────────────────────────────────────────────
    for col, hdr in enumerate(
        ["أسلوب التقييم",
         "القيمة الاستدلالية (EGP)",
         "الوزن",
         "القيمة الموزونة (EGP)",
         "ملاحظة"], 1
    ):
        cell = ws.cell(row=row[0], column=col, value=hdr)
        cell.font = get_font(
            size=Typography.SIZE_TABLE_HEADER, bold=True, color=Palette.WHITE
        )
        cell.fill = get_fill(Palette.NAVY_DEEP)
        cell.alignment = get_alignment(h="center")
    ws.row_dimensions[row[0]].height = 28
    row[0] += 1

    # ─── Three-approach rows ──────────────────────────────────────────────
    approaches = [
        ("comparable", "أسلوب المقارنة البيعية", Palette.WHITE),
        ("cost",       "أسلوب التكلفة",          Palette.GREY_50),
        ("income",     "رأسمالة الدخل",           Palette.WHITE),
    ]

    for key_prefix, label, bg in approaches:
        r = row[0]
        val = d[f"{key_prefix}_value"]
        wt  = d[f"w_{key_prefix}"]
        weighted = val * wt

        cell_m = ws.cell(row=r, column=_COL_METHOD, value=label)
        cell_m.font = get_font(size=10, bold=True, color=Palette.INK)
        cell_m.fill = get_fill(bg)
        cell_m.alignment = get_alignment(h="right")

        cell_v = ws.cell(row=r, column=_COL_VALUE, value=val)
        cell_v.fill = get_fill(bg)
        cell_v.alignment = get_alignment(h="center")
        cell_v.number_format = NumFormat.CURRENCY
        cell_v.font = get_font(size=10, color=Palette.INK)

        cell_w = ws.cell(row=r, column=_COL_WEIGHT, value=wt)
        cell_w.fill = get_fill(bg)
        cell_w.alignment = get_alignment(h="center")
        cell_w.number_format = NumFormat.PERCENT
        cell_w.font = get_font(size=10, color=Palette.INK)

        cell_wv = ws.cell(row=r, column=_COL_WEIGHTED, value=weighted)
        cell_wv.fill = get_fill(bg)
        cell_wv.alignment = get_alignment(h="center")
        cell_wv.number_format = NumFormat.CURRENCY
        cell_wv.font = get_font(size=10, color=Palette.INK)

        ws.row_dimensions[r].height = 22
        locs[f"{key_prefix}_value"]  = (r, _COL_VALUE)
        locs[f"w_{key_prefix}"]      = (r, _COL_WEIGHT)
        locs[f"weighted_{key_prefix}"] = (r, _COL_WEIGHTED)
        row[0] += 1

    # Final sum row
    r = row[0]
    fv = d["primary_value"]
    for col, (val, fmt) in enumerate((
        ("القيمة التوفيقية النهائية", None),
        (fv, NumFormat.CURRENCY),
        ("—",  None),
        (fv, NumFormat.CURRENCY),
        ("",   None),
    ), 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = get_font(
            size=Typography.SIZE_SECTION_HEADER, bold=True, color=Palette.WHITE
        )
        cell.fill = get_fill(Palette.NAVY)
        cell.alignment = get_alignment(
            h="right" if col == 1 else "center"
        )
        if fmt:
            cell.number_format = fmt
    ws.row_dimensions[r].height = 28
    locs["primary_value"] = (r, _COL_VALUE)
    row[0] += 2

    # ─── Final value KPI ──────────────────────────────────────────────────
    draw_section(ws, row[0], _NUM_COLS, "القيمة السوقية النهائية", bg=Palette.GOLD)
    row[0] += 1

    r = row[0]
    cell_lbl = ws.cell(row=r, column=_COL_METHOD, value="القيمة السوقية النهائية (EGP)")
    cell_lbl.font = get_font(
        size=Typography.SIZE_SECTION_HEADER, bold=True, color=Palette.INK
    )
    cell_lbl.fill = get_fill(Palette.GOLD_PALE)
    cell_lbl.alignment = get_alignment(h="right")

    cell_val = ws.cell(row=r, column=_COL_VALUE, value=fv)
    cell_val.font = get_font(
        size=Typography.SIZE_KPI_VALUE, bold=True, color=Palette.NAVY
    )
    cell_val.fill = get_fill(Palette.GOLD_PALE)
    cell_val.alignment = get_alignment(h="center")
    cell_val.number_format = NumFormat.CURRENCY
    ws.row_dimensions[r].height = 42
    locs["final_value_kpi"] = (r, _COL_VALUE)
    row[0] += 1

    for key, label, fmt in (
        ("price_per_m2",  "سعر المتر المُشتق (EGP/م²)", NumFormat.CURRENCY),
        ("area_total",    "المساحة الإجمالية (م²)",      NumFormat.INTEGER),
        ("confidence",    "مستوى الثقة",                  ""),
    ):
        r = row[0]
        style_label(ws.cell(row=r, column=_COL_METHOD, value=label))
        cell_v = ws.cell(row=r, column=_COL_VALUE, value=d.get(key, "—"))
        style_value(cell_v)
        if fmt:
            cell_v.number_format = fmt
        ws.row_dimensions[r].height = 24
        locs[key] = (r, _COL_VALUE)
        row[0] += 1

    row[0] += 1

    # ─── Notes ────────────────────────────────────────────────────────────
    draw_section(ws, row[0], _NUM_COLS, "ملاحظات التوفيق", bg=Palette.GREY_700)
    row[0] += 1
    for note in _RECONCILIATION_NOTES:
        r = row[0]
        ws.cell(row=r, column=_COL_METHOD, value=note)
        cell = ws.cell(row=r, column=_COL_METHOD)
        cell.font = get_font(size=9, color=Palette.INK)
        cell.alignment = get_alignment(h="right", wrap=True)
        ws.row_dimensions[r].height = 20
        row[0] += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 20
    ws.freeze_panes = "A3"
    return locs


def get_field_count() -> int:
    return len(_DEFAULTS)


def get_section_count() -> int:
    return len(_SECTION_NAMES)


def _print_summary() -> None:
    print("\n  توفيق النتائج — Reconciliation Sheet")
    print("  " + "-" * 40)
    print(f"  أقسام: {get_section_count()}")
    print(f"  حقول:  {get_field_count()}")
    print()
    for s in _SECTION_NAMES:
        print(f"  • {s}")
    print()


if __name__ == "__main__":
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace"
        )
    _print_summary()
