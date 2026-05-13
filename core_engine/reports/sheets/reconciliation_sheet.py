#!/usr/bin/env python3
"""
reconciliation_sheet.py — Builder for شيت "توفيق النتائج"
(Reconciliation of Value Indications, Classic Office Blue theme).

استخدام:
    from core_engine.reports.sheets.reconciliation_sheet import apply_reconciliation_sheet

    locs = apply_reconciliation_sheet(ws, inputs_dict, profile_key="legacy")
    # locs: dict[str, tuple[int, int]] — خريطة field_key → (row, col)
"""
from __future__ import annotations

from typing import Any, Mapping

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    BuilderPalette as _BP,
    NumFormat as _NF,
    Palette as _Palette,
    get_fill as _gf,
)

# ── Module-level style constants (Classic Office Blue) ────────────────────────
_FILL_HEADER      = _gf(_BP.HEADER)
_FILL_INPUT_SECT  = _gf(_BP.SECTION_DARK)
_FILL_INPUT_CELL  = _gf(_BP.INPUT_CELL)
_FILL_FINAL_VALUE = _gf(_BP.SUCCESS_LIGHT)
_FILL_ROW_BAND    = _gf(_BP.ROW_BAND)
_FONT_MUTED       = Font(italic=True, color=_BP.MUTED)
_FONT_FINAL_VALUE = Font(bold=True, size=12, color=_BP.SUCCESS_DARK)
_ALIGN_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER_THIN      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_BORDER_MEDIUM    = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)
_FMT_CURRENCY = _NF.CURRENCY_2DP
_FMT_PCT      = _NF.PERCENT_DETAILED


def apply_reconciliation_sheet(
    ws: Worksheet,
    inputs: Mapping[str, Any],
    *,
    profile_key: str = "legacy",
) -> dict[str, tuple[int, int]]:
    """
    Build شيت توفيق النتائج.

    Args:
        ws: الشيت المراد تعديله (يجب أن يكون فارغاً).
        inputs: قاموس يحتوى نتائج أساليب التقييم. المفاتيح المدعومة:
                primary_value / final_value, comparable, cost, income,
                weights (nested dict) or weights_comparable/cost/income (flat),
                report_date.
        profile_key: "legacy" فقط مدعوم حالياً (محفوظ للتوسع).

    Returns:
        dict[str, tuple[int, int]] — خريطة field_key → (row, col) لخلايا القيم.
    """
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 24

    # ── Unpack inputs ─────────────────────────────────────────────────────────
    fv   = float(inputs.get("primary_value") or inputs.get("final_value") or 0)
    comp = float(inputs.get("comparable") or 0)
    cost = float(inputs.get("cost") or 0)
    inc  = float(inputs.get("income") or 0)
    report_date = str(inputs.get("report_date") or "")

    w: dict = inputs.get("weights") or {}
    if not w:
        w = {
            "comparable": float(inputs.get("weights_comparable") or 0),
            "cost":       float(inputs.get("weights_cost")       or 0),
            "income":     float(inputs.get("weights_income")     or 0),
        }

    locs: dict[str, tuple[int, int]] = {}

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    b           = ws["A1"]
    b.value     = "توفيق النتائج — Reconciliation of Value Indications"
    b.fill      = _FILL_HEADER
    b.font      = Font(bold=True, color=_Palette.WHITE, size=14)
    b.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    ws["A2"].value     = f"تاريخ التقرير: {report_date}  |  المعيار: EGVS 3.0 / IVS 105"
    ws["A2"].font      = _FONT_MUTED
    ws["A2"].alignment = Alignment(horizontal="center")

    r = 4

    # ── Table header row ──────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 22
    for ci, hdr in enumerate(
        ("أسلوب التقييم", "القيمة الاستدلالية (EGP)", "الوزن", "القيمة الموزونة (EGP)"), 2
    ):
        c           = ws.cell(row=r, column=ci)
        c.value     = hdr
        c.fill      = _FILL_INPUT_SECT
        c.font      = Font(bold=True, color=_Palette.WHITE, size=10)
        c.alignment = _ALIGN_CENTER
        c.border    = _BORDER_THIN
    r += 1

    # ── Three approach rows ───────────────────────────────────────────────────
    for i, (lbl, val, wt_key) in enumerate((
        ("أسلوب المقارنة البيعية", comp, "comparable"),
        ("أسلوب التكلفة",          cost, "cost"),
        ("رأسمالة الدخل",          inc,  "income"),
    )):
        wt   = w.get(wt_key, 0)
        fill = _FILL_ROW_BAND if i % 2 else _FILL_INPUT_CELL
        ws.row_dimensions[r].height = 18

        lc           = ws.cell(row=r, column=2)
        lc.value     = lbl
        lc.font      = Font(bold=True, size=10)
        lc.fill      = fill
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border    = _BORDER_THIN

        for ci, (v, fmt) in enumerate(
            ((val, _FMT_CURRENCY), (wt, _FMT_PCT), (val * wt, _FMT_CURRENCY)), 3
        ):
            c                = ws.cell(row=r, column=ci)
            c.value          = v
            c.number_format  = fmt
            c.alignment      = _ALIGN_CENTER
            c.fill           = fill
            c.border         = _BORDER_THIN

        locs[f"val_{wt_key}"]    = (r, 3)
        locs[f"weight_{wt_key}"] = (r, 4)
        r += 1

    # ── Reconciled total row ──────────────────────────────────────────────────
    ws.row_dimensions[r].height = 26
    for ci, (v, fmt) in enumerate((
        ("القيمة التوفيقية النهائية", None),
        (fv, _FMT_CURRENCY),
        ("—", None),
        (fv, _FMT_CURRENCY),
    ), 2):
        c           = ws.cell(row=r, column=ci)
        c.value     = v
        c.font      = _FONT_FINAL_VALUE
        c.fill      = _FILL_FINAL_VALUE
        c.alignment = _ALIGN_CENTER
        c.border    = _BORDER_MEDIUM
        if fmt:
            c.number_format = fmt
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="right", vertical="center")
    locs["final_value"] = (r, 3)
    r += 2

    # ── Notes section ─────────────────────────────────────────────────────────
    ws.merge_cells(f"B{r}:E{r}")
    nh           = ws.cell(row=r, column=2)
    nh.value     = "  ملاحظات التوفيق"
    nh.fill      = _FILL_INPUT_SECT
    nh.font      = Font(bold=True, color=_Palette.WHITE, size=11)
    nh.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    for note in (
        "• تم إعطاء الوزن الأكبر لأسلوب المقارنة البيعية لتوافر بيانات السوق.",
        "• تم دعم النتيجة بأسلوب التكلفة والدخل لتأكيد القيمة.",
        "• تتوافق نتائج الأساليب الثلاثة مع مؤشرات السوق الحالية.",
    ):
        ws.merge_cells(f"B{r}:E{r}")
        nc           = ws.cell(row=r, column=2)
        nc.value     = note
        nc.font      = Font(size=9)
        nc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 16
        r += 1

    ws.freeze_panes = "B3"
    return locs
