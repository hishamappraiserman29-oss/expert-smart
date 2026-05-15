#!/usr/bin/env python3
"""
rental_sheet.py — Builder for شيت "المقارنات الإيجارية".
"""
from __future__ import annotations

import io
import sys
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    NumFormat,
    Palette,
    Typography,
    apply_sheet_defaults,
    draw_banner,
    draw_section,
    get_alignment,
    get_fill,
    get_font,
    style_label,
    style_value,
)

_COL_LABEL = 1
_COL_SUBJ  = 2
_COL_C1    = 3
_COL_C2    = 4
_COL_C3    = 5
_NUM_COLS  = 5

_COMP_ROWS: tuple[tuple[str, str], ...] = (
    ("الموقع",             "location"),
    ("المساحة (م²)",       "area"),
    ("الطابق",             "floor"),
    ("العمر (سنة)",        "age"),
    ("الحالة",             "condition"),
    ("الإيجار (EGP/م²/سنة)", "rent_sqm"),
    ("نسبة الشاغر",        "vacancy_rate"),
    ("تاريخ الإيجار",      "lease_date"),
)

_DEFAULT_COMPS: list[dict[str, Any]] = [
    {
        "location": "التجمع الخامس", "area": 130, "floor": 2, "age": 10,
        "condition": "جيد جداً", "rent_sqm": 380, "vacancy_rate": 0.05,
        "lease_date": "2025/09/01",
    },
    {
        "location": "القاهرة الجديدة", "area": 120, "floor": 4, "age": 8,
        "condition": "ممتاز", "rent_sqm": 420, "vacancy_rate": 0.03,
        "lease_date": "2026/01/15",
    },
    {
        "location": "الرحاب", "area": 145, "floor": 1, "age": 14,
        "condition": "جيد", "rent_sqm": 350, "vacancy_rate": 0.07,
        "lease_date": "2025/11/01",
    },
]

_DEFAULT_SUBJECT: dict[str, Any] = {
    "location": "التجمع الخامس", "area": 150, "floor": 3, "age": 16,
    "condition": "جيد جداً", "rent_sqm": 0, "vacancy_rate": 0.05,
    "lease_date": "2026/04/12",
}

_DEFAULTS: dict[str, Any] = {
    "indicated_rent_sqm": 383,
    "total_annual_rent":  57_450,
    "cap_rate":           0.08,
    "income_value":       718_125,
    "grm":                12.5,
    "valuation_date":     "2026/04/12",
}


def apply_rental_sheet(
    ws: Worksheet, data: dict | None = None
) -> dict[str, tuple[int, int]]:
    """Build شيت المقارنات الإيجارية. Returns {key: (row, col)}."""
    d = {**_DEFAULTS, **(data or {})}
    comps = list((d.get("comparables") or _DEFAULT_COMPS))[:3]
    while len(comps) < 3:
        comps.append({})
    subject = d.get("subject") or _DEFAULT_SUBJECT

    apply_sheet_defaults(ws)
    locs: dict[str, tuple[int, int]] = {}
    row = [1]

    draw_banner(
        ws, row=1, end_col=_NUM_COLS,
        text="المقارنات الإيجارية — Rental Comparables",
        bg=Palette.INK, fg=Palette.GOLD_LIGHT,
        size=Typography.SIZE_PAGE_TITLE - 2, height=46,
    )
    row[0] = 3

    # Subtitle / standard reference
    ws.cell(row=row[0], column=1,
            value=f"تاريخ التقرير: {d['valuation_date']}  |  المعيار: IVS 105")
    cell_sub = ws.cell(row=row[0], column=1)
    cell_sub.font = get_font(size=9, color=Palette.GREY_500, italic=True)
    cell_sub.alignment = get_alignment(h="center")
    row[0] += 2

    # ─── Column headers ───────────────────────────────────────────────────
    for col, hdr in enumerate(
        ["البند", "الموضوع", "مقارن 1", "مقارن 2", "مقارن 3"], 1
    ):
        cell = ws.cell(row=row[0], column=col, value=hdr)
        cell.font = get_font(
            size=Typography.SIZE_TABLE_HEADER, bold=True, color=Palette.WHITE
        )
        cell.fill = get_fill(Palette.NAVY_DEEP)
        cell.alignment = get_alignment(h="center")
    ws.row_dimensions[row[0]].height = 26
    row[0] += 1

    # ─── Comparable rows ──────────────────────────────────────────────────
    for ri, (label, key) in enumerate(_COMP_ROWS):
        r = row[0]
        alt = ri % 2 == 1
        bg = Palette.GREY_50 if alt else Palette.WHITE

        cell_l = ws.cell(row=r, column=_COL_LABEL, value=label)
        style_label(cell_l)

        subj_val = subject.get(key, "—")
        cell_s = ws.cell(row=r, column=_COL_SUBJ, value=subj_val)
        cell_s.font = get_font(size=10, bold=True, color=Palette.NAVY)
        cell_s.fill = get_fill(Palette.GOLD_PALE)
        cell_s.alignment = get_alignment(h="center")

        for ci, comp in zip([_COL_C1, _COL_C2, _COL_C3], comps):
            cell = ws.cell(row=r, column=ci, value=comp.get(key, "—"))
            cell.fill = get_fill(bg)
            cell.alignment = get_alignment(h="center")
            cell.font = get_font(size=10, color=Palette.INK)
            if key == "vacancy_rate" and isinstance(comp.get(key), float):
                cell.number_format = NumFormat.PERCENT
            elif key == "rent_sqm" and isinstance(comp.get(key), (int, float)):
                cell.number_format = NumFormat.CURRENCY

        ws.row_dimensions[r].height = 22
        locs[f"comp_{key}"] = (r, _COL_SUBJ)
        row[0] += 1

    row[0] += 1

    # ─── الإيجار المستنتج ─────────────────────────────────────────────────
    draw_section(ws, row[0], _NUM_COLS, "الإيجار المستنتج والقيمة بالرسملة",
                 bg=Palette.EMERALD)
    row[0] += 1

    result_rows: list[tuple[str, str, str]] = [
        ("indicated_rent_sqm", "الإيجار السوقى المستنتج (EGP/م²/سنة)",
         NumFormat.CURRENCY),
        ("total_annual_rent",  "الإيجار الإجمالى السنوى (EGP)",
         NumFormat.CURRENCY),
        ("cap_rate",           "معدل الرسملة المُطبَّق", NumFormat.PERCENT),
        ("income_value",       "القيمة الاستدلالية بالرسملة (EGP)", NumFormat.CURRENCY),
        ("grm",                "GRM — مضاعف الإيجار الإجمالى", NumFormat.QUANTITY),
    ]
    for key, label, fmt in result_rows:
        r = row[0]
        style_label(ws.cell(row=r, column=_COL_LABEL, value=label))
        cell_v = ws.cell(row=r, column=_COL_SUBJ, value=d.get(key, "—"))
        cell_v.font = get_font(size=11, bold=True, color=Palette.NAVY)
        cell_v.fill = get_fill(Palette.EMERALD_VLT)
        cell_v.alignment = get_alignment(h="center")
        if fmt:
            cell_v.number_format = fmt
        ws.row_dimensions[r].height = 26
        locs[key] = (r, _COL_SUBJ)
        row[0] += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "B6"
    return locs


def get_field_count() -> int:
    return len(_COMP_ROWS) + 5  # comp rows + result rows


def get_section_count() -> int:
    return 2  # comparables + indicated value


def _print_summary() -> None:
    print("\n  المقارنات الإيجارية — Rental Comparables Sheet")
    print("  " + "-" * 40)
    print(f"  صفوف المقارنة:  {len(_COMP_ROWS)}")
    print(f"  صفوف النتائج:   5")
    print()


if __name__ == "__main__":
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace"
        )
    _print_summary()
