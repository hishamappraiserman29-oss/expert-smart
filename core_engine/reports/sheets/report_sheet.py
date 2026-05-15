#!/usr/bin/env python3
"""
report_sheet.py — Builder for شيت "التقرير" (Main Report Executive Summary).

استخدام:
    from core_engine.reports.sheets.report_sheet import apply_report_sheet
    locs = apply_report_sheet(ws, data={...})
"""
from __future__ import annotations

import io
import sys
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    NumFormat,
    Palette,
    Typography,
    apply_sheet_defaults,
    draw_banner,
    draw_section,
    get_alignment,
    get_fill,
    get_font,
    style_label,
    style_value,
)

_COL_LABEL = 1
_COL_VALUE = 2
_NUM_COLS = 2

_SECTION_NAMES = [
    "بيانات التقرير",
    "بيانات العقار",
    "نتائج أساليب التقييم",
    "الأوزان المُطبَّقة",
    "القيمة السوقية النهائية",
]

_DEFAULTS: dict[str, Any] = {
    "report_id":        "VAL-20260412-1153",
    "valuation_date":   "2026/04/12",
    "report_purpose":   "تقدير القيمة السوقية",
    "appraiser_name":   "هشام محمد محمد المهدى",
    "license_id":       29,
    "client_name":      "د/ عبد الرؤوف محمد عبد الباقى",
    "property_type":    "شقة سكنية",
    "region":           "التجمع الخامس — القاهرة الجديدة",
    "area_total":       150,
    "comparable_value": 4_989_300,
    "cost_value":       4_800_000,
    "income_value":     4_750_000,
    "w_comparable":     0.60,
    "w_cost":           0.20,
    "w_income":         0.20,
    "primary_value":    4_899_960,
    "price_per_m2":     32_666,
    "confidence":       "عالى",
}


def apply_report_sheet(
    ws: Worksheet, data: dict | None = None
) -> dict[str, tuple[int, int]]:
    """
    Build شيت التقرير كاملاً.

    Returns:
        { field_key: (row, col) }
    """
    d = {**_DEFAULTS, **(data or {})}
    apply_sheet_defaults(ws)
    locs: dict[str, tuple[int, int]] = {}
    row = [1]

    draw_banner(
        ws, row=1, end_col=_NUM_COLS,
        text="تقرير التقييم العقاري — Real Estate Valuation Report",
        bg=Palette.INK, fg=Palette.GOLD_LIGHT,
        size=Typography.SIZE_PAGE_TITLE, height=52,
    )
    row[0] = 3

    def _section(title: str, bg: str = Palette.NAVY) -> None:
        draw_section(ws, row[0], _NUM_COLS, title, bg=bg)
        row[0] += 1

    def _kv(key: str, label: str, fmt: str = "") -> None:
        cell_l = ws.cell(row=row[0], column=_COL_LABEL, value=label)
        style_label(cell_l)
        cell_v = ws.cell(row=row[0], column=_COL_VALUE, value=d.get(key, "—"))
        style_value(cell_v)
        if fmt:
            cell_v.number_format = fmt
        ws.row_dimensions[row[0]].height = 26
        locs[key] = (row[0], _COL_VALUE)
        row[0] += 1

    # ─── بيانات التقرير ──────────────────────────────────────────────────
    _section("بيانات التقرير", Palette.NAVY_DEEP)
    _kv("report_id",       "رقم التقرير")
    _kv("valuation_date",  "تاريخ التقييم")
    _kv("report_purpose",  "الغرض من التقييم")
    _kv("appraiser_name",  "اسم المُقيِّم")
    _kv("license_id",      "رقم القيد فى الهيئة")
    _kv("client_name",     "اسم العميل / المالك")
    row[0] += 1

    # ─── بيانات العقار ───────────────────────────────────────────────────
    _section("بيانات العقار", Palette.EMERALD)
    _kv("property_type",   "نوع العقار")
    _kv("region",          "الموقع")
    _kv("area_total",      "المساحة الإجمالية")
    row[0] += 1

    # ─── نتائج أساليب التقييم ────────────────────────────────────────────
    _section("نتائج أساليب التقييم", Palette.NAVY)
    _kv("comparable_value", "مقارنة البيوع",  fmt=NumFormat.CURRENCY)
    _kv("cost_value",        "طريقة التكلفة", fmt=NumFormat.CURRENCY)
    _kv("income_value",      "رأسمالة الدخل",  fmt=NumFormat.CURRENCY)
    row[0] += 1

    # ─── الأوزان ──────────────────────────────────────────────────────────
    _section("الأوزان المُطبَّقة", Palette.INDIGO)
    _kv("w_comparable",    "وزن مقارنة البيوع", fmt=NumFormat.PERCENT)
    _kv("w_cost",          "وزن التكلفة",        fmt=NumFormat.PERCENT)
    _kv("w_income",        "وزن الدخل",           fmt=NumFormat.PERCENT)
    row[0] += 1

    # ─── القيمة النهائية ─────────────────────────────────────────────────
    _section("القيمة السوقية النهائية", Palette.GOLD)

    cell_lbl = ws.cell(row=row[0], column=_COL_LABEL,
                       value="القيمة السوقية النهائية")
    cell_lbl.font = get_font(
        size=Typography.SIZE_SECTION_HEADER, bold=True, color=Palette.INK
    )
    cell_lbl.fill = get_fill(Palette.GOLD_PALE)
    cell_lbl.alignment = get_alignment(h="right")

    cell_val = ws.cell(row=row[0], column=_COL_VALUE, value=d["primary_value"])
    cell_val.font = get_font(
        size=Typography.SIZE_KPI_VALUE, bold=True, color=Palette.NAVY
    )
    cell_val.fill = get_fill(Palette.GOLD_PALE)
    cell_val.alignment = get_alignment(h="center")
    cell_val.number_format = NumFormat.CURRENCY
    ws.row_dimensions[row[0]].height = 42
    locs["primary_value"] = (row[0], _COL_VALUE)
    row[0] += 1

    _kv("price_per_m2",   "سعر المتر المُشتق", fmt=NumFormat.CURRENCY)
    _kv("confidence",     "مستوى الثقة")

    # ─── Bar chart (optional) ─────────────────────────────────────────────
    chart_top = row[0] + 1
    for i, (lbl, val) in enumerate((
        ("مقارنة البيوع",  d["comparable_value"]),
        ("طريقة التكلفة", d["cost_value"]),
        ("رأسمالة الدخل",  d["income_value"]),
        ("القيمة النهائية", d["primary_value"]),
    ), chart_top):
        ws.cell(row=i, column=_COL_LABEL).value = lbl
        ws.cell(row=i, column=_COL_VALUE).value = val
    try:
        from openpyxl.chart import BarChart, Reference
        chart = BarChart()
        chart.type = "col"
        chart.title = "مقارنة أساليب التقييم"
        chart.y_axis.title = "القيمة (EGP)"
        chart.style = 10
        chart.width = 16
        chart.height = 10
        data_ref = Reference(ws, min_col=_COL_VALUE, min_row=chart_top,
                             max_row=chart_top + 3)
        cats_ref = Reference(ws, min_col=_COL_LABEL, min_row=chart_top,
                             max_row=chart_top + 3)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{chart_top + 6}")
    except Exception:
        pass

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 26
    return locs


def get_field_count() -> int:
    return len(_DEFAULTS)


def get_section_count() -> int:
    return len(_SECTION_NAMES)


def _print_summary() -> None:
    print("\n  التقرير — Main Report Sheet")
    print("  " + "-" * 40)
    print(f"  أقسام: {get_section_count()}")
    print(f"  حقول:  {get_field_count()}")
    print()
    for s in _SECTION_NAMES:
        print(f"  • {s}")
    print()


if __name__ == "__main__":
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace"
        )
    _print_summary()
