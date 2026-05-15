#!/usr/bin/env python3
"""
sales_comparison.py — Builder for شيت "مقارنات البيوع".

يكتب مصفوفة الضبط الاحترافية بصيغ Excel حقيقية:
  =SUM(...)          ← إجمالى الضبط الصافى
  =price*(1+net_adj) ← السعر بعد الضبط
  =SUMPRODUCT(...)   ← السعر النهائي الموزون
"""
from __future__ import annotations

import io
import sys
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core_engine.reports.report_theme import (
    NumFormat,
    Palette,
    Typography,
    apply_sheet_defaults,
    draw_banner,
    draw_section,
    get_alignment,
    get_fill,
    get_font,
    style_input_value,
    style_label,
    style_value,
)

_COL_LABEL = 1
_COL_SUBJ  = 2
_COL_C1    = 3
_COL_C2    = 4
_COL_C3    = 5
_COL_NOTE  = 6
_NUM_COLS  = 6

_ADJ_LABELS: tuple[tuple[str, str], ...] = (
    ("adj_location",  "الموقع"),
    ("adj_area",      "المساحة"),
    ("adj_floor",     "الدور"),
    ("adj_age",       "العمر"),
    ("adj_condition", "التشطيب"),
    ("adj_view",      "الإطلالة"),
    ("adj_timing",    "التوقيت"),
    ("adj_facade",    "الواجهة"),
    ("adj_services",  "الخدمات"),
)

_DEFAULT_COMPS: list[dict[str, Any]] = [
    {
        "location": "القاهرة الجديدة", "area": 120, "price_per_sqm": 35_000,
        "adj_location": 0.05,  "adj_area": -0.02, "adj_floor":  0.00,
        "adj_age":       0.03,  "adj_condition": 0.00, "adj_view": -0.01,
        "adj_timing":    0.02,  "adj_facade": 0.00, "adj_services": 0.01,
        "weight": 0.40,
    },
    {
        "location": "مدينة نصر", "area": 115, "price_per_sqm": 32_000,
        "adj_location": -0.03, "adj_area":  0.01, "adj_floor":  0.00,
        "adj_age":      -0.02, "adj_condition": 0.05, "adj_view":  0.00,
        "adj_timing":    0.02, "adj_facade": 0.00, "adj_services": -0.01,
        "weight": 0.35,
    },
    {
        "location": "الرحاب", "area": 130, "price_per_sqm": 34_000,
        "adj_location":  0.00, "adj_area": -0.03, "adj_floor":  0.02,
        "adj_age":        0.00, "adj_condition": 0.00, "adj_view":  0.03,
        "adj_timing":     0.01, "adj_facade": 0.00, "adj_services": 0.00,
        "weight": 0.25,
    },
]

_DEFAULT_SUBJECT: dict[str, Any] = {
    "location": "التجمع الخامس", "area": 150, "price_per_sqm": 0,
}


def apply_sales_comparison(
    ws: Worksheet, data: dict | None = None
) -> dict[str, tuple[int, int]]:
    """
    Build شيت مقارنات البيوع بصيغ Excel حقيقية.

    data keys:
        comparables: list[dict]  (up to 3 comps)
        subject:     dict        (الموضوع)
    """
    d = data or {}
    comps = list((d.get("comparables") or _DEFAULT_COMPS))[:3]
    while len(comps) < 3:
        comps.append({})
    subject = d.get("subject") or _DEFAULT_SUBJECT

    apply_sheet_defaults(ws)
    locs: dict[str, tuple[int, int]] = {}
    row = [1]

    draw_banner(
        ws, row=1, end_col=_NUM_COLS,
        text="مصفوفة الضبط الاحترافية — مقارنات البيوع",
        bg=Palette.INK, fg=Palette.GOLD_LIGHT,
        size=Typography.SIZE_PAGE_TITLE - 2, height=46,
    )
    row[0] = 3
    draw_section(ws, row[0], _NUM_COLS, "مصفوفة مقارنة البيوع", bg=Palette.NAVY_DEEP)
    row[0] += 1

    # ─── Column headers ───────────────────────────────────────────────────
    for col, hdr in enumerate(
        ["البيان", "الموضوع", "مقارن 1", "مقارن 2", "مقارن 3", "ملاحظة"], 1
    ):
        cell = ws.cell(row=row[0], column=col, value=hdr)
        cell.font = get_font(
            size=Typography.SIZE_TABLE_HEADER, bold=True, color=Palette.WHITE
        )
        cell.fill = get_fill(Palette.NAVY_DEEP)
        cell.alignment = get_alignment(h="center")
    ws.row_dimensions[row[0]].height = 28
    row[0] += 1

    def _info_row(label: str, key: str, subj_val: Any,
                  fmt: str = "") -> int:
        """Write one info row; return the row index."""
        r = row[0]
        style_label(ws.cell(row=r, column=_COL_LABEL, value=label))
        style_value(ws.cell(row=r, column=_COL_SUBJ, value=subj_val))
        for ci, comp in zip([_COL_C1, _COL_C2, _COL_C3], comps):
            v = comp.get(key, "—")
            cell = ws.cell(row=r, column=ci, value=v)
            style_value(cell)
            if fmt and isinstance(v, (int, float)):
                cell.number_format = fmt
        ws.row_dimensions[r].height = 22
        locs[f"info_{key}"] = (r, _COL_SUBJ)
        row[0] += 1
        return r

    _info_row("الموقع / العنوان",    "location",      subject.get("location", "—"))
    _info_row("المساحة (م²)",        "area",          subject.get("area", "—"))
    price_info_row = _info_row(
        "سعر المتر (EGP/م²)", "price_per_sqm", subject.get("price_per_sqm", "—"),
        fmt=NumFormat.CURRENCY,
    )
    row[0] += 1

    # ─── Adjustments ──────────────────────────────────────────────────────
    draw_section(
        ws, row[0], _NUM_COLS, "التسويات النسبية — Adjustments",
        bg=Palette.INDIGO,
    )
    row[0] += 1

    adj_start_row = row[0]
    for adj_key, adj_label in _ADJ_LABELS:
        r = row[0]
        style_label(ws.cell(row=r, column=_COL_LABEL, value=adj_label))
        style_value(ws.cell(row=r, column=_COL_SUBJ, value="—"))
        for ci, comp in zip([_COL_C1, _COL_C2, _COL_C3], comps):
            cell = ws.cell(row=r, column=ci, value=comp.get(adj_key, 0))
            style_input_value(cell)
            cell.number_format = NumFormat.PERCENT_SIGNED
        ws.row_dimensions[r].height = 22
        locs[adj_key] = (r, _COL_C1)
        row[0] += 1
    adj_end_row = row[0] - 1

    row[0] += 1
    draw_section(ws, row[0], _NUM_COLS, "النتائج — Results", bg=Palette.NAVY)
    row[0] += 1

    # ─── Net adjustment (SUM formula) ─────────────────────────────────────
    net_adj_row = row[0]
    style_label(
        ws.cell(row=net_adj_row, column=_COL_LABEL, value="إجمالى الضبط الصافى")
    )
    style_value(ws.cell(row=net_adj_row, column=_COL_SUBJ, value="—"))
    for ci in [_COL_C1, _COL_C2, _COL_C3]:
        col = get_column_letter(ci)
        cell = ws.cell(
            row=net_adj_row, column=ci,
            value=f"=SUM({col}{adj_start_row}:{col}{adj_end_row})",
        )
        cell.font = get_font(size=10, bold=True, color=Palette.NAVY)
        cell.fill = get_fill(Palette.GOLD_PALE)
        cell.alignment = get_alignment(h="center")
        cell.number_format = NumFormat.PERCENT_SIGNED
    ws.row_dimensions[net_adj_row].height = 24
    locs["net_adj"] = (net_adj_row, _COL_C1)
    row[0] += 1

    # ─── Adjusted price (multiply formula) ────────────────────────────────
    adj_price_row = row[0]
    style_label(
        ws.cell(row=adj_price_row, column=_COL_LABEL,
                value="السعر بعد الضبط (EGP/م²)")
    )
    style_value(ws.cell(row=adj_price_row, column=_COL_SUBJ, value="—"))
    for ci in [_COL_C1, _COL_C2, _COL_C3]:
        col = get_column_letter(ci)
        cell = ws.cell(
            row=adj_price_row, column=ci,
            value=f"={col}{price_info_row}*(1+{col}{net_adj_row})",
        )
        cell.font = get_font(size=10, bold=True, color=Palette.NAVY)
        cell.fill = get_fill(Palette.EMERALD_LT)
        cell.alignment = get_alignment(h="center")
        cell.number_format = NumFormat.CURRENCY
    ws.row_dimensions[adj_price_row].height = 24
    locs["adj_price"] = (adj_price_row, _COL_C1)
    row[0] += 1

    # ─── Weight ───────────────────────────────────────────────────────────
    weight_row = row[0]
    style_label(ws.cell(row=weight_row, column=_COL_LABEL, value="وزن المقارن"))
    style_value(ws.cell(row=weight_row, column=_COL_SUBJ, value="—"))
    for ci, comp in zip([_COL_C1, _COL_C2, _COL_C3], comps):
        cell = ws.cell(row=weight_row, column=ci, value=comp.get("weight", 0))
        style_input_value(cell)
        cell.number_format = NumFormat.PERCENT
    ws.row_dimensions[weight_row].height = 22
    locs["weight"] = (weight_row, _COL_C1)
    row[0] += 1

    # ─── SUMPRODUCT final weighted price ──────────────────────────────────
    final_row = row[0]
    c1, c2, c3 = (get_column_letter(c) for c in [_COL_C1, _COL_C2, _COL_C3])
    sumproduct_formula = (
        f"=SUMPRODUCT({c1}{adj_price_row}:{c3}{adj_price_row},"
        f"{c1}{weight_row}:{c3}{weight_row})"
        f"/SUM({c1}{weight_row}:{c3}{weight_row})"
    )

    cell_lbl = ws.cell(row=final_row, column=_COL_LABEL,
                       value="السعر النهائي الموزون (EGP/م²)")
    cell_lbl.font = get_font(size=11, bold=True, color=Palette.INK)
    cell_lbl.fill = get_fill(Palette.GOLD_PALE)
    cell_lbl.alignment = get_alignment(h="right")

    cell_v = ws.cell(row=final_row, column=_COL_SUBJ, value=sumproduct_formula)
    ws.merge_cells(
        f"{get_column_letter(_COL_SUBJ)}{final_row}"
        f":{get_column_letter(_COL_C3)}{final_row}"
    )
    cell_v.font = get_font(size=Typography.SIZE_KPI_VALUE, bold=True, color=Palette.NAVY)
    cell_v.fill = get_fill(Palette.GOLD_PALE)
    cell_v.alignment = get_alignment(h="center")
    cell_v.number_format = NumFormat.CURRENCY
    ws.row_dimensions[final_row].height = 40
    locs["final_weighted_price"] = (final_row, _COL_SUBJ)
    row[0] += 2

    # ─── Legend ───────────────────────────────────────────────────────────
    draw_section(ws, row[0], _NUM_COLS, "مفتاح الرموز — Legend", bg=Palette.GREY_700)
    row[0] += 1
    for bg, text in (
        (Palette.EMERALD_LT, "ضبط موجب (+): يرفع قيمة المقارن"),
        (Palette.CORAL_LT,   "ضبط سالب (-): يخفض قيمة المقارن"),
        (Palette.GREY_100,   "بدون ضبط (0): تطابق مع الموضوع"),
    ):
        r = row[0]
        cell = ws.cell(row=r, column=_COL_LABEL, value=text)
        cell.fill = get_fill(bg)
        cell.alignment = get_alignment(h="right")
        cell.font = get_font(size=10, color=Palette.INK)
        ws.row_dimensions[r].height = 20
        row[0] += 1

    row[0] += 1
    draw_section(
        ws, row[0], _NUM_COLS, "الأسلوب المنهجى — Methodology",
        bg=Palette.GREY_700,
    )
    row[0] += 1
    for step in (
        "1. اختيار المقارنات المناسبة من نفس المنطقة والفئة",
        "2. تطبيق التسويات النسبية على كل عامل بشكل مستقل",
        "3. حساب إجمالى الضبط الصافى كنسبة من سعر المقارن",
        "4. السعر المُعدَّل = سعر المقارن × (1 + إجمالى الضبط)",
        "5. تطبيق الأوزان حسب جودة التطابق والموثوقية",
        "6. السعر الموزون = SUMPRODUCT / SUM(أوزان)",
    ):
        r = row[0]
        ws.cell(row=r, column=_COL_LABEL, value=step)
        cell = ws.cell(row=r, column=_COL_LABEL)
        cell.alignment = get_alignment(h="right")
        cell.font = get_font(size=9, color=Palette.INK)
        ws.row_dimensions[r].height = 18
        row[0] += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 24
    return locs


def get_field_count() -> int:
    return 3 + len(_ADJ_LABELS) + 4  # info rows + adj rows + result rows


def get_section_count() -> int:
    return 4  # header + adjustments + results + legend+methodology


def _print_summary() -> None:
    print("\n  مقارنات البيوع — Sales Comparison Sheet")
    print("  " + "-" * 40)
    print(f"  صفوف المعلومات:    3 (موقع، مساحة، سعر)")
    print(f"  عوامل التسوية:    {len(_ADJ_LABELS)}")
    print(f"  صفوف النتائج:     4 (صافى، مُعدَّل، وزن، موزون)")
    print()


if __name__ == "__main__":
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace"
        )
    _print_summary()
