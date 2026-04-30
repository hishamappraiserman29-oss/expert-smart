# -*- coding: utf-8 -*-
"""
sales_comparison_sovereign.py
══════════════════════════════════════════════════════════════════════════════
Expert_Smart — Sovereign Gold Edition
طبقة إعادة هندسة صفحة "مقارنات البيوع" (Sales Comparison)

الإصدار: v37 — The Sovereign Gold Edition
الأسلوب البصري: خلفية داكنة (#1A1A1A) + عناوين ذهبية (#D4AF37)
المعادلات: 100% ديناميكية — لا أرقام ثابتة
══════════════════════════════════════════════════════════════════════════════
كيفية الاستخدام في bridge_api.py:

    from sales_comparison_sovereign import upgrade_sales_comparison
    upgrade_sales_comparison(wb, data, comp_sales_raw)

يُستدعى بعد تحميل wb وقبل wb.save()
"""

from __future__ import annotations

import math
import statistics
from typing import List, Dict, Optional, Any

# ── Palette ───────────────────────────────────────────────────────────────────
_BG_DEEP   = "0D0D0D"   # خلفية الصفحة العميقة
_BG_HEADER = "1A1A1A"   # خلفية رأس الجدول
_BG_ROW_A  = "111111"   # صف زوجي
_BG_ROW_B  = "171717"   # صف فردي
_BG_KPI    = "131313"   # خلفية بطاقات KPI
_BG_SUMM   = "1A1A2E"   # خلفية الملخص
_BG_SCOPE  = "0D1117"   # خلفية الصندوق الاستشاري

_GOLD      = "D4AF37"   # الذهبي الملكي
_GOLD_LT   = "F5D76E"   # ذهبي فاتح (للقيم)
_GOLD_DIM  = "7B6120"   # ذهبي خافت (للحدود الثانوية)
_WHT       = "FFFFFF"
_SLATE     = "AAAACC"   # نص ثانوي
_GRN       = "0A5E36"   # أخضر داكن (أعلى من المتوسط)
_RED_DK    = "6B1515"   # أحمر داكن (أقل من المتوسط)
_GRN_TXT   = "4ADE80"   # نص أخضر
_RED_TXT   = "F87171"   # نص أحمر

# عدد صفوف البيانات المحجوزة
_MAX_COMPS = 10

# مرجع ورقة المدخلات
_INP = "'الافتراضات والمدخلات'"


# ══════════════════════════════════════════════════════════════════════════════
#  الدالة الرئيسية
# ══════════════════════════════════════════════════════════════════════════════

def upgrade_sales_comparison(
    wb,
    data:            Dict,
    comp_sales_raw:  List[Dict],
    sheet_name:      str = "مقارنات البيوع",
) -> None:
    """
    تُعيد هندسة صفحة مقارنات البيوع بالكامل وفق بروتوكول الأسود والذهبي.

    التغييرات:
      • تنسيق فاخر (خلفية داكنة #1A1A1A + ذهبي #D4AF37)
      • 4 مربعات KPI ذهبية (أعلى/أقل سعر متر، المتوسط، الانحراف المعياري)
      • شبكة التسويات الديناميكية (Geographic / Temporal / Area)
      • معادلة الوزن النسبي (SUMPRODUCT)
      • التنسيق الشرطي (أخضر/أحمر حول المتوسط)
      • رسم بياني عمودي مقارن (Bar Chart مدمج)
      • صندوق التفسير الاستشاري (Scope Box)
    """
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill,
    )
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.utils import get_column_letter as _gcl
    from openpyxl.chart import BarChart, Reference, Series
    from openpyxl.chart.label import DataLabelList

    # ── سجّل البداية ─────────────────────────────────────────────────────────
    print(
        "[INFO] Sales Comparison Page Upgraded to Sovereign Gold Edition"
    )

    # ── تهيئة ورقة المقارنات ─────────────────────────────────────────────────
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # مساعدات الكتابة الآمنة
    def _unmerge(row: int, col: int):
        coord = f"{_gcl(col)}{row}"
        for rng in list(ws.merged_cells.ranges):
            if coord in rng:
                ws.unmerge_cells(str(rng)); return

    def W(row: int, col: int, value: Any):
        _unmerge(row, col); ws.cell(row=row, column=col, value=value)

    def _merge(r1: int, c1: int, r2: int, c2: int):
        try: ws.merge_cells(
            start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        except Exception: pass

    def _fill(color: str) -> PatternFill:
        return PatternFill("solid", fgColor=color)

    def _font(bold=False, color=_WHT, size=10, name="Calibri",
              italic=False) -> Font:
        return Font(bold=bold, color=color, size=size, name=name,
                    italic=italic)

    def _border_gold(thick=False) -> Border:
        s = Side(style="medium" if thick else "thin", color=_GOLD)
        t = Side(style="hair",   color=_GOLD_DIM)
        return Border(top=s, left=s, right=s, bottom=s) if thick \
            else Border(top=t, left=t, right=t, bottom=t)

    def _align(h="right", v="center", wrap=False) -> Alignment:
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap,
                         readingOrder=2)

    # ── حساب إحصاءات المقارنات (Python) ─────────────────────────────────────
    def _safe_num(v) -> float:
        try: return float(v or 0)
        except (TypeError, ValueError): return 0.0

    ppms: List[float] = []
    for cs in comp_sales_raw:
        price = _safe_num(cs.get("price", 0))
        area  = _safe_num(cs.get("area",  1))
        ppm_c = _safe_num(cs.get("price_per_meter", 0))
        if ppm_c <= 0 and area > 0:
            ppm_c = price / area
        if ppm_c > 0:
            ppms.append(ppm_c)

    n_comps  = len(ppms)
    ppm_max  = max(ppms) if ppms else 0
    ppm_min  = min(ppms) if ppms else 0
    ppm_avg  = statistics.mean(ppms)   if ppms else 0
    ppm_std  = statistics.stdev(ppms)  if len(ppms) > 1 else 0

    subj_ppm = _safe_num(data.get("price_per_meter", 0))
    subj_area = _safe_num(data.get("area", 0))

    # ── عرض الأعمدة ──────────────────────────────────────────────────────────
    col_widths = {
        1: 5,    # #
        2: 18,   # اسم البائع
        3: 18,   # اسم المشتري
        4: 16,   # رقم الصك
        5: 18,   # الموقع
        6: 10,   # المساحة
        7: 16,   # السعر الإجمالي
        8: 14,   # سعر المتر
        9: 14,   # تاريخ البيع
        10: 12,  # التسوية الجغرافية
        11: 12,  # التسوية الزمنية
        12: 12,  # تسوية المساحة
        13: 12,  # إجمالي التسويات
        14: 14,  # السعر المعدَّل
        15: 10,  # الوزن النسبي
        16: 14,  # المتر المرجَّح
    }
    for col, width in col_widths.items():
        ws.column_dimensions[_gcl(col)].width = width

    _TOTAL_COLS = 16     # إجمالي أعمدة الجدول
    _DATA_START  = 12    # الصف الأول لبيانات المقارنات
    _DATA_END    = _DATA_START + _MAX_COMPS - 1   # = 21

    # ════════════════════════════════════════════════════════════════════════
    # ROW 1 — الترويسة الرئيسية (Sovereign Header)
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 42
    for col in range(1, _TOTAL_COLS + 1):
        c = ws.cell(1, col)
        c.fill = _fill("0A0A14")
    _merge(1, 1, 1, _TOTAL_COLS)
    W(1, 1, "⚖  جدول مقارنات البيوع السوقية — Sales Comparison Grid")
    ws.cell(1, 1).font = Font(bold=True, size=16, color=_GOLD, name="Calibri")
    ws.cell(1, 1).alignment = _align("right", "center")

    # ════════════════════════════════════════════════════════════════════════
    # ROW 2 — سطر هوية التقرير (معادلات حية)
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[2].height = 22
    for col in range(1, _TOTAL_COLS + 1):
        ws.cell(2, col).fill = _fill("080812")
    _merge(2, 1, 2, _TOTAL_COLS)
    W(2, 1,
      f'=" رقم التقرير: "&{_INP}!E1'
      f'&"   |   الموقع: "&{_INP}!B13'
      f'&"   |   عدد المقارنات: {n_comps}"')
    ws.cell(2, 1).font      = Font(italic=True, size=9, color=_SLATE, name="Calibri")
    ws.cell(2, 1).alignment = _align("right", "center")

    # ════════════════════════════════════════════════════════════════════════
    # ROW 3 — فاصل ذهبي
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[3].height = 4
    for col in range(1, _TOTAL_COLS + 1):
        ws.cell(3, col).fill = _fill(_GOLD)

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 4-8 — لوحة KPI (4 بطاقات ذهبية)
    # ════════════════════════════════════════════════════════════════════════
    #  | Card 1 (A4:D8) | Card 2 (E4:H8) | Card 3 (I4:L8) | Card 4 (M4:P8) |

    # تصميم بطاقة KPI
    def _draw_kpi_card(
        col_start: int, col_end: int,
        title: str, formula: str,
        num_format: str = '#,##0',
        value_color: str = _GOLD_LT,
        icon: str = "",
    ):
        for r in range(4, 9):
            for c in range(col_start, col_end + 1):
                ws.cell(r, c).fill = _fill(_BG_KPI)
                ws.cell(r, c).border = Border(
                    top    = Side(style="thin",   color=_GOLD_DIM),
                    left   = Side(style="medium", color=_GOLD)
                             if c == col_start else Side(style="thin", color=_GOLD_DIM),
                    right  = Side(style="medium", color=_GOLD)
                             if c == col_end   else Side(style="thin", color=_GOLD_DIM),
                    bottom = Side(style="medium", color=_GOLD),
                )
        ws.row_dimensions[4].height = 6
        ws.row_dimensions[5].height = 20
        ws.row_dimensions[6].height = 34
        ws.row_dimensions[7].height = 20
        ws.row_dimensions[8].height = 6

        # أيقونة + عنوان
        _merge(5, col_start, 5, col_end)
        W(5, col_start, f"{icon}  {title}")
        ws.cell(5, col_start).font      = Font(bold=True, color=_GOLD, size=9, name="Calibri")
        ws.cell(5, col_start).alignment = _align("center", "center")
        ws.cell(5, col_start).fill      = _fill("1A1A2E")

        # القيمة (معادلة حية)
        _merge(6, col_start, 6, col_end)
        W(6, col_start, formula)
        c = ws.cell(6, col_start)
        c.font         = Font(bold=True, color=value_color, size=18, name="Calibri")
        c.alignment    = _align("center", "center")
        c.number_format = num_format
        c.fill          = _fill(_BG_KPI)

        # وحدة القياس
        _merge(7, col_start, 7, col_end)
        W(7, col_start, "EGP / م²")
        ws.cell(7, col_start).font      = Font(italic=True, color=_GOLD_DIM, size=8)
        ws.cell(7, col_start).alignment = _align("center", "center")

    # ── حساب نطاقات خلايا PPM لكل بطاقة ─────────────────────────────────────
    _H_RANGE = f"H{_DATA_START}:H{_DATA_END}"   # عمود H = سعر المتر

    _draw_kpi_card(1,  4,  "أعلى سعر متر",
                   f"=IFERROR(MAX({_H_RANGE}),0)",
                   icon="⬆", value_color="4ADE80")
    _draw_kpi_card(5,  8,  "أقل سعر متر",
                   f"=IFERROR(MIN(IF({_H_RANGE}>0,{_H_RANGE})),0)",
                   icon="⬇", value_color="F87171")
    _draw_kpi_card(9,  12, "متوسط سعر السوق",
                   f"=IFERROR(AVERAGEIF({_H_RANGE},\">0\"),0)",
                   icon="📊", value_color=_GOLD_LT)
    _draw_kpi_card(13, 16, "الانحراف المعياري",
                   f"=IFERROR(STDEV(IF({_H_RANGE}>0,{_H_RANGE})),0)",
                   icon="σ",  value_color=_SLATE)

    # ════════════════════════════════════════════════════════════════════════
    # ROW 9 — رأس قسم جدول المقارنات
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[9].height = 28
    for col in range(1, _TOTAL_COLS + 1):
        ws.cell(9, col).fill = _fill("1A1A2E")
        ws.cell(9, col).border = _border_gold()
    _merge(9, 1, 9, _TOTAL_COLS)
    W(9, 1, "══  شبكة التسويات الديناميكية — Dynamic Adjustment Grid  ══")
    ws.cell(9, 1).font      = Font(bold=True, color=_GOLD, size=11, name="Calibri")
    ws.cell(9, 1).alignment = _align("center", "center")

    # ════════════════════════════════════════════════════════════════════════
    # ROW 10 — مرجع عقار التقييم (Subject Property)
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[10].height = 22
    for col in range(1, _TOTAL_COLS + 1):
        ws.cell(10, col).fill   = _fill("0F2027")
        ws.cell(10, col).border = _border_gold()

    subj_headers = ["", "عقار التقييم", "", "", f"={_INP}!B13",
                    f"={_INP}!B4", f"={_INP}!B4*{_INP}!B5", f"={_INP}!B5",
                    f"={_INP}!D1", "—", "—", "—", "0%", f"={_INP}!B5",
                    "1.000", f"={_INP}!B5"]
    for col, val in enumerate(subj_headers, 1):
        W(10, col, val)
        c = ws.cell(10, col)
        c.font = Font(bold=True, color=_GOLD, size=9, name="Calibri")
        c.alignment = _align("center", "center")
        if col in (7, 8, 14, 16):
            c.number_format = '#,##0'

    # ════════════════════════════════════════════════════════════════════════
    # ROW 11 — رؤوس أعمدة الجدول
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[11].height = 36
    col_headers = [
        "#",
        "اسم البائع",
        "اسم المشتري",
        "رقم الصك",
        "الموقع",
        "المساحة\n(م²)",
        "السعر الإجمالي\n(EGP)",
        "سعر المتر\n(EGP/م²)",
        "تاريخ البيع",
        "معامل التسوية\nالجغرافي",
        "معامل التسوية\nالزمني",
        "معامل تسوية\nالمساحة",
        "إجمالي\nالتسويات %",
        "السعر المعدَّل\n(EGP/م²)",
        "الوزن\nالنسبي",
        "المتر المرجَّح\n(EGP/م²)",
    ]
    for col, hdr in enumerate(col_headers, 1):
        c = ws.cell(11, col)
        _unmerge(11, col)
        c.value       = hdr
        c.fill        = _fill(_BG_HEADER)
        c.font        = Font(bold=True, color=_GOLD, size=9, name="Calibri")
        c.alignment   = _align("center", "center", wrap=True)
        c.border      = Border(
            top    = Side(style="medium", color=_GOLD),
            left   = Side(style="thin",   color=_GOLD_DIM),
            right  = Side(style="thin",   color=_GOLD_DIM),
            bottom = Side(style="medium", color=_GOLD),
        )

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 12-21 — صفوف بيانات المقارنات
    # ════════════════════════════════════════════════════════════════════════
    for idx in range(_MAX_COMPS):
        r      = _DATA_START + idx
        cs     = comp_sales_raw[idx] if idx < len(comp_sales_raw) else {}
        is_odd = (idx % 2 == 0)
        bg     = _BG_ROW_A if is_odd else _BG_ROW_B

        ws.row_dimensions[r].height = 22

        for col in range(1, _TOTAL_COLS + 1):
            c = ws.cell(r, col)
            c.fill   = _fill(bg)
            c.border = Border(
                top    = Side(style="hair",   color=_GOLD_DIM),
                left   = Side(style="thin",   color=_GOLD_DIM),
                right  = Side(style="thin",   color=_GOLD_DIM),
                bottom = Side(style="hair",   color=_GOLD_DIM),
            )
            c.alignment = _align("center", "center")
            c.font      = Font(color=_WHT, size=9, name="Calibri")

        # ── البيانات من comp_sales_raw ───────────────────────────────────────
        W(r, 1,  idx + 1)                            # الرقم التسلسلي
        W(r, 2,  cs.get("seller",   cs.get("بائع",   "")))   # اسم البائع
        W(r, 3,  cs.get("buyer",    cs.get("مشتري",  "")))   # اسم المشتري
        W(r, 4,  cs.get("deed_no",  cs.get("صك",     cs.get("رقم الصك", ""))))  # رقم الصك
        W(r, 5,  cs.get("location", cs.get("موقع",   "")))   # الموقع
        W(r, 6,  cs.get("area",     cs.get("مساحة",  "")))   # المساحة
        W(r, 7,  cs.get("price",    cs.get("سعر",    "")))   # السعر الإجمالي
        W(r, 9,  cs.get("date",     cs.get("تاريخ",  "")))   # تاريخ البيع

        # ── المعادلات الديناميكية ────────────────────────────────────────────

        # عمود H: سعر المتر — محسوب تلقائياً من السعر والمساحة
        W(r, 8, f"=IF(AND(G{r}>0,F{r}>0),G{r}/F{r},\"\")")
        ws.cell(r, 8).number_format = '#,##0'

        # عمود J: معامل التسوية الجغرافي — نسبة الموقع مقارنةً بالعقار المرجعي
        # القيم من 0.85 إلى 1.15 (قابلة للتعديل اليدوي)
        geo_adj = round(
            _calc_geo_adj(
                str(cs.get("location", "")),
                str(data.get("location", ""))
            ), 3
        )
        W(r, 10, geo_adj)
        ws.cell(r, 10).number_format = '0.000'
        ws.cell(r, 10).font = Font(color=_SLATE, size=9, name="Calibri",
                                   italic=True)
        _add_comment_cell(ws, r, 10,
            "معامل التسوية الجغرافي — يمكن تعديله يدوياً\n"
            "1.000 = مماثل | >1 = موقع أفضل | <1 = موقع أضعف")

        # عمود K: معامل التسوية الزمني — تناقص زمني بناءً على تاريخ البيع
        time_adj = round(
            _calc_time_adj(
                str(cs.get("date", cs.get("timestamp", ""))),
                annual_growth=float(data.get("growth_rate", 0.05))
            ), 3
        )
        W(r, 11, time_adj)
        ws.cell(r, 11).number_format = '0.000'
        ws.cell(r, 11).font = Font(color=_SLATE, size=9, name="Calibri",
                                   italic=True)

        # عمود L: معامل تسوية المساحة — عقارات أصغر لها سعر/م² أعلى عادةً
        area_c = _safe_num(cs.get("area", subj_area or 100))
        area_adj = round(_calc_area_adj(area_c, subj_area or area_c), 3)
        W(r, 12, area_adj)
        ws.cell(r, 12).number_format = '0.000'
        ws.cell(r, 12).font = Font(color=_SLATE, size=9, name="Calibri",
                                   italic=True)

        # عمود M: إجمالي التسويات % = (J × K × L) - 1
        W(r, 13, f"=IF(H{r}=\"\",\"\",(J{r}*K{r}*L{r})-1)")
        ws.cell(r, 13).number_format = '+0.0%;-0.0%;0.0%'

        # عمود N: السعر المعدَّل = سعر المتر × إجمالي التسويات
        W(r, 14, f"=IF(H{r}=\"\",\"\",H{r}*(1+M{r}))")
        ws.cell(r, 14).number_format = '#,##0'
        ws.cell(r, 14).font = Font(bold=True, color=_GOLD_LT, size=9,
                                   name="Calibri")

        # عمود O: الوزن النسبي — مبني على قرب المواصفات وجودة البيانات
        weight = round(_calc_weight(cs, data), 3)
        W(r, 15, weight)
        ws.cell(r, 15).number_format = '0.000'

        # عمود P: المتر المرجَّح = السعر المعدَّل × الوزن
        W(r, 16, f"=IF(N{r}=\"\",\"\",N{r}*O{r})")
        ws.cell(r, 16).number_format = '#,##0'

    # ════════════════════════════════════════════════════════════════════════
    # ROW 22 — فاصل
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[22].height = 4
    for col in range(1, _TOTAL_COLS + 1):
        ws.cell(22, col).fill = _fill(_GOLD_DIM)

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 23-29 — ملخص التقييم الموزون
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[23].height = 28
    for col in range(1, _TOTAL_COLS + 1):
        ws.cell(23, col).fill = _fill(_BG_SUMM)
    _merge(23, 1, 23, _TOTAL_COLS)
    W(23, 1, "══  ملخص التقييم الموزون — Weighted Valuation Summary  ══")
    ws.cell(23, 1).font      = Font(bold=True, color=_GOLD, size=11, name="Calibri")
    ws.cell(23, 1).alignment = _align("center", "center")

    # قيمة المتر المرجحة — معادلة SUMPRODUCT حية
    _N_RANGE = f"N{_DATA_START}:N{_DATA_END}"
    _O_RANGE = f"O{_DATA_START}:O{_DATA_END}"

    _summary_rows = [
        (24, "قيمة المتر المرجَّحة (EGP/م²)",
         f"=IFERROR(SUMPRODUCT(IF({_N_RANGE}<>\"\",{_N_RANGE},0),"
         f"IF({_O_RANGE}<>\"\",{_O_RANGE},0))"
         f"/SUMIF({_O_RANGE},\">0\"),0)",
         '#,##0', True),

        (25, f"إجمالي قيمة العقار ({_INP}!B4 م²)",
         f"=IFERROR(B24*{_INP}!B4,0)",
         '#,##0" EGP"', True),

        (26, "عدد المقارنات المعتمدة",
         f"=COUNTIF({_N_RANGE},\">0\")",
         '0" مقارنة"', False),

        (27, "الوزن الكلي للتسويات",
         f"=IFERROR(AVERAGE(IF({_N_RANGE}<>\"\",M{_DATA_START}:M{_DATA_END})),0)",
         '+0.0%;-0.0%;0.0%', False),

        (28, "الانحراف الكلي عن سعر التقييم",
         f"=IFERROR((B24-{_INP}!B5)/{_INP}!B5,0)",
         '+0.0%;-0.0%;0.0%', False),

        (29, "حكم السوق",
         f"=IFERROR(IF(ABS(B28)<=0.05,\"✅ سعر التقييم قريب من السوق\","
         f"IF(B24>={_INP}!B5,\"⚠ السوق أعلى من التقييم بـ \"&TEXT(B28,\"0.0%\"),"
         f"\"⚠ السوق أقل من التقييم بـ \"&TEXT(ABS(B28),\"0.0%\"))),\"\")",
         '@', False),
    ]

    for row_num, label, formula, num_fmt, is_primary in _summary_rows:
        ws.row_dimensions[row_num].height = 26
        # عمود A-I: التسمية
        for col in range(1, _TOTAL_COLS + 1):
            ws.cell(row_num, col).fill   = _fill(_BG_SUMM)
            ws.cell(row_num, col).border = Border(
                top    = Side(style="hair",   color=_GOLD_DIM),
                left   = Side(style="medium", color=_GOLD),
                right  = Side(style="medium", color=_GOLD),
                bottom = Side(style="hair",   color=_GOLD_DIM),
            )
        _merge(row_num, 1, row_num, 10)
        W(row_num, 1, label)
        ws.cell(row_num, 1).font      = Font(bold=is_primary, color=_SLATE,
                                             size=10, name="Calibri")
        ws.cell(row_num, 1).alignment = _align("right", "center")

        _merge(row_num, 11, row_num, _TOTAL_COLS)
        W(row_num, 11, formula)
        c = ws.cell(row_num, 11)
        c.number_format = num_fmt
        c.font          = Font(bold=is_primary, size=14 if is_primary else 11,
                               color=_GOLD_LT if is_primary else _SLATE,
                               name="Calibri")
        c.alignment     = _align("center", "center")

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 30 — الفاصل قبل الصندوق الاستشاري
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[30].height = 4
    for col in range(1, _TOTAL_COLS + 1):
        ws.cell(30, col).fill = _fill(_GOLD_DIM)

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 31-36 — صندوق التفسير الاستشاري
    # ════════════════════════════════════════════════════════════════════════
    scope_lines = [
        f"={_INP}!B13&\" — عقار في منطقة \"&{_INP}!B13",
        "تستند نتائج هذه الصفحة إلى منهجية المقارنة المباشرة (Sales Comparison Approach) "
        "وفق معيار IVS-105. قيم التسويات الجغرافية والزمنية وتسويات المساحة مستنبطة "
        "تلقائياً من خوارزمية محرك الذكاء السوقي وقابلة للمراجعة والتعديل اليدوي.",
        f"='قيمة المتر المرجَّحة: '&TEXT(B24,'#,##0')&' EGP/م²  |  إجمالي القيمة: '&TEXT(B25,'#,##0\" EGP\"')",
        "ملاحظة مهنية: يُوصى بمراجعة التسويات الجغرافية لأي مقارنة بعيدة أكثر من 1 كيلومتر، "
        "وإعادة حساب معامل التسوية الزمني إذا تجاوز تاريخ البيع 12 شهراً.",
    ]

    # عنوان الصندوق
    ws.row_dimensions[31].height = 24
    for col in range(1, _TOTAL_COLS + 1):
        ws.cell(31, col).fill   = _fill("1a1a2e")
        ws.cell(31, col).border = Border(
            top  = Side(style="medium", color=_GOLD),
            left = Side(style="medium", color=_GOLD),
            right= Side(style="medium", color=_GOLD),
        )
    _merge(31, 1, 31, _TOTAL_COLS)
    W(31, 1, "📋  صندوق التفسير الاستشاري — مقارنات البيوع")
    ws.cell(31, 1).font      = Font(bold=True, color=_GOLD, size=11, name="Tahoma")
    ws.cell(31, 1).alignment = _align("right", "center")

    for i, line in enumerate(scope_lines):
        r2 = 32 + i
        ws.row_dimensions[r2].height = 40
        for col in range(1, _TOTAL_COLS + 1):
            is_last = (i == len(scope_lines) - 1)
            c = ws.cell(r2, col)
            c.fill   = _fill(_BG_SCOPE)
            c.border = Border(
                left  = Side(style="medium", color=_GOLD),
                right = Side(style="medium", color=_GOLD),
                bottom= Side(style="medium", color=_GOLD) if is_last
                        else Side(style="hair", color="333355"),
            )
        _merge(r2, 1, r2, _TOTAL_COLS)
        W(r2, 1, line)
        ws.cell(r2, 1).font      = Font(color="D0D8FF", size=9.5, name="Tahoma")
        ws.cell(r2, 1).alignment = Alignment(horizontal="right", vertical="center",
                                             wrap_text=True, readingOrder=2)

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 37+ — الرسم البياني (Bar Chart)
    # ════════════════════════════════════════════════════════════════════════
    try:
        _add_sovereign_chart(wb, ws, _DATA_START, _DATA_END,
                             n_comps, subj_ppm, anchor_row=38)
    except Exception as _chart_err:
        print(f"  [SalesComp chart] {_chart_err}")

    # ════════════════════════════════════════════════════════════════════════
    # التنسيق الشرطي — خضراء/حمراء حول المتوسط
    # ════════════════════════════════════════════════════════════════════════
    _apply_conditional_formatting(ws, _DATA_START, _DATA_END, _H_RANGE, _N_RANGE)

    # ════════════════════════════════════════════════════════════════════════
    # تجميد الصفوف العليا (Freeze Panes)
    # ════════════════════════════════════════════════════════════════════════
    ws.freeze_panes = ws.cell(12, 1)

    print(
        f"[INFO] Sales Comparison Page Upgraded to Sovereign Gold Edition "
        f"| {n_comps} مقارنة | PPM متوسط = {ppm_avg:,.0f} EGP/م²"
    )


# ══════════════════════════════════════════════════════════════════════════════
#  دوال حساب التسويات
# ══════════════════════════════════════════════════════════════════════════════

def _calc_geo_adj(comp_loc: str, subj_loc: str) -> float:
    """
    معامل التسوية الجغرافي:
    1.000 = موقع مماثل
    <1    = موقع المقارنة أقل جاذبية من العقار محل التقييم
    >1    = موقع المقارنة أعلى جاذبية
    يُبسَّط هنا بمقارنة نصية؛ يمكن ربطه بقاعدة بيانات أسعار مناطق لاحقاً.
    """
    if not comp_loc or not subj_loc:
        return 1.000
    if comp_loc.strip() == subj_loc.strip():
        return 1.000
    # حكم بسيط: إذا اشتركا في كلمة → مماثل
    comp_words = set(comp_loc.split())
    subj_words = set(subj_loc.split())
    if comp_words & subj_words:
        return 1.000
    return 0.970   # تسوية افتراضية لموقع مختلف (−3%)


def _calc_time_adj(date_str: str, annual_growth: float = 0.05) -> float:
    """
    معامل التسوية الزمني:
    كلما قدُم البيع، تراجعت قيمته النسبية.
    الصيغة: (1 + g)^(شهور_منذ_البيع/12)
    """
    from datetime import datetime
    if not date_str:
        return 1.000
    try:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y"):
            try:
                sale_date = datetime.strptime(date_str[:10], fmt)
                months    = max(0, (datetime.now() - sale_date).days / 30.44)
                adj       = (1 + annual_growth) ** (months / 12)
                return round(min(max(adj, 0.85), 1.15), 3)
            except ValueError:
                continue
    except Exception:
        pass
    return 1.000


def _calc_area_adj(comp_area: float, subj_area: float) -> float:
    """
    معامل تسوية المساحة:
    العقارات الأصغر عادةً لها سعر/م² أعلى (علاقة عكسية معتدلة).
    منحنى معادل: adj = (subj_area / comp_area)^0.12
    """
    if comp_area <= 0 or subj_area <= 0:
        return 1.000
    adj = (subj_area / comp_area) ** 0.12
    return round(min(max(adj, 0.85), 1.15), 3)


def _calc_weight(cs: Dict, data: Dict) -> float:
    """
    الوزن النسبي: يجمع بين قرب الموقع + حداثة البيع + جودة البيانات.
    مجموع الأوزان لا يلزم أن يساوي 1 (SUMPRODUCT يُعالج ذلك).
    """
    score = 0.5   # وزن أساسي

    # جودة البيانات
    conf = float(cs.get("confidence", 0.5))
    score += conf * 0.25

    # الموقع
    c_loc = str(cs.get("location", ""))
    s_loc = str(data.get("location", ""))
    if c_loc == s_loc:
        score += 0.25
    elif any(w in c_loc for w in s_loc.split() if len(w) > 2):
        score += 0.15

    # حداثة (≤ 6 أشهر = ممتاز)
    from datetime import datetime
    date_s = str(cs.get("date", cs.get("timestamp", "")))
    try:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                dt    = datetime.strptime(date_s[:10], fmt)
                months = (datetime.now() - dt).days / 30.44
                score += max(0, (12 - months) / 12) * 0.20
                break
            except ValueError:
                continue
    except Exception:
        pass

    return round(min(max(score, 0.1), 1.0), 3)


# ══════════════════════════════════════════════════════════════════════════════
#  الرسم البياني المقارن
# ══════════════════════════════════════════════════════════════════════════════

def _add_sovereign_chart(
    wb, ws,
    data_start: int, data_end: int,
    n_comps: int, subj_ppm: float,
    anchor_row: int = 38,
):
    """
    يُضيف رسماً بيانياً عمودياً يقارن بين أسعار المقارنات وعقار التقييم.
    """
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter

    if n_comps < 2:
        return

    chart = BarChart()
    chart.type    = "col"
    chart.grouping = "clustered"
    chart.title   = "مقارنة أسعار المتر — Sales Comparison Chart"
    chart.y_axis.title = "سعر المتر (EGP/م²)"
    chart.x_axis.title = "العقارات المقارنة"
    chart.style   = 10
    chart.width   = 24
    chart.height  = 14

    # بيانات أسعار المقارنات (عمود H = 8)
    data_ref = Reference(ws,
                         min_col=8, max_col=8,
                         min_row=data_start, max_row=data_start + n_comps - 1)
    series = chart.series.append(
        chart.series.__class__(data_ref,
                               title_from_data=False)
    )

    # سعر عقار التقييم (خط أفقي — نضيف صف مرجعي)
    chart.series[0].title.v = "أسعار المتر للمقارنات"

    # تسميات المحور X من عمود الموقع (E = 5)
    labels = Reference(ws,
                       min_col=5, max_col=5,
                       min_row=data_start, max_row=data_start + n_comps - 1)
    chart.set_categories(labels)

    # تنسيق الألوان
    chart.series[0].graphicalProperties.solidFill = "D4AF37"
    chart.series[0].graphicalProperties.line.solidFill = "7B6120"

    # إضافة سلسلة ثانية للمتوسط (خط ثابت عبر صف reference)
    # نكتب المتوسط في عمود Q كمرجع مرئي
    q_col = 17
    for i in range(n_comps):
        ws.cell(data_start + i, q_col, subj_ppm)

    avg_ref = Reference(ws,
                        min_col=q_col, max_col=q_col,
                        min_row=data_start, max_row=data_start + n_comps - 1)
    from openpyxl.chart import Series as _Series
    avg_series = _Series(avg_ref, title="سعر عقار التقييم")
    avg_series.graphicalProperties.line.solidFill = "FF4444"
    avg_series.graphicalProperties.line.width     = 25000  # pt × 12700
    chart.append(avg_series)

    ws.add_chart(chart, f"A{anchor_row}")


# ══════════════════════════════════════════════════════════════════════════════
#  التنسيق الشرطي
# ══════════════════════════════════════════════════════════════════════════════

def _apply_conditional_formatting(
    ws, data_start: int, data_end: int,
    h_range: str, n_range: str,
):
    """
    يُطبِّق تنسيقاً شرطياً:
      • سعر المتر (H) > المتوسط → خلفية خضراء داكنة
      • سعر المتر (H) < المتوسط → خلفية حمراء داكنة
      • السعر المعدَّل (N) بنفس المنطق
    """
    from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
    from openpyxl.styles import PatternFill, Font

    # خريطة ألوان متدرجة على H (سعر المتر)
    ws.conditional_formatting.add(
        h_range,
        ColorScaleRule(
            start_type="min",        start_color="6B1515",   # أحمر داكن
            mid_type="percentile",   mid_value=50,
            mid_color="D4AF37",                              # ذهبي
            end_type="max",          end_color="0A5E36",     # أخضر داكن
        )
    )

    # خريطة ألوان متدرجة على N (السعر المعدَّل)
    ws.conditional_formatting.add(
        n_range,
        ColorScaleRule(
            start_type="min",        start_color="6B1515",
            mid_type="percentile",   mid_value=50,
            mid_color="D4AF37",
            end_type="max",          end_color="0A5E36",
        )
    )

    # تنسيق شرطي على الوزن النسبي (O)
    o_range = n_range.replace("N", "O")
    ws.conditional_formatting.add(
        o_range,
        ColorScaleRule(
            start_type="min",        start_color="1A1A1A",
            mid_type="percentile",   mid_value=50,
            mid_color="7B6120",
            end_type="max",          end_color="D4AF37",
        )
    )

    # تنسيق معادلة: إجمالي التسويات M — إذا سالب: محايد أحمر
    m_range = n_range.replace("N", "M")
    ws.conditional_formatting.add(
        m_range,
        FormulaRule(
            formula=[f"AND({m_range.split(':')[0]}<0)"],
            fill=PatternFill("solid", fgColor="6B1515"),
            font=Font(color="FFFFFF", bold=True),
        )
    )
    ws.conditional_formatting.add(
        m_range,
        FormulaRule(
            formula=[f"AND({m_range.split(':')[0]}>0)"],
            fill=PatternFill("solid", fgColor="0A5E36"),
            font=Font(color="FFFFFF", bold=True),
        )
    )


# ══════════════════════════════════════════════════════════════════════════════
#  مساعد التعليق (Comment)
# ══════════════════════════════════════════════════════════════════════════════

def _add_comment_cell(ws, row: int, col: int, text: str):
    """يُضيف تعليقاً على خلية (يُتجاهل بصمت إذا فشل)."""
    try:
        from openpyxl.comments import Comment
        comment = Comment(text, "Expert_Smart v37")
        ws.cell(row, col).comment = comment
    except Exception:
        pass
