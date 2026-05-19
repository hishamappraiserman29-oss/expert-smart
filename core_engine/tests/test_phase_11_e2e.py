"""
test_phase_11_e2e.py — Phase 11: Portfolio Analysis E2E Tests

16 tests covering:
  1.  PortfolioScenario dataclass + to_dict roundtrip
  2.  create_standard_scenarios() produces exactly 3 scenarios
  3.  run_scenarios() returns 3 ScenarioResult objects
  4.  Pessimistic scenario reduces NOI vs base
  5.  Optimistic scenario increases portfolio value vs base
  6.  Base scenario value_change_pct == 0
  7.  All IRR estimates are positive
  8.  get_performance_summary() structure (required keys present)
  9.  Pessimistic value < base value < optimistic value
  10. value_at_risk_pct is between 0 and 1
  11. POST /api/valuation/portfolio/performance — 200 OK, standard scenarios
  12. POST /api/valuation/portfolio/performance — custom scenario passthrough
  13. sheet_portfolio_performance creates "Portfolio Performance" sheet
  14. Missing optional VaR keys do not crash sheet builder
  15. Empty scenarios shows "No scenario data available"
  16. Baseline block cells contain expected portfolio name
"""
from __future__ import annotations

import sys
from pathlib import Path

_CORE = Path(__file__).resolve().parents[1]   # …/core_engine/
_ROOT = _CORE.parent
for _p in (str(_CORE), str(_ROOT)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import os
import tempfile
os.environ.setdefault("JWT_SECRET", "test-secret-e2e-bundle")
os.chdir(str(_CORE))

from bridge_api import app
from auth.tokens import generate_token as _gen_token
_AUTH_HDR = {"Authorization": f"Bearer {_gen_token('test-user-e2e')}"}
from openpyxl import load_workbook
from reports.excel_builder import ExcelReportBuilder
from adapters.portfolio import PortfolioBuilder
from adapters.portfolio_performance import (
    PortfolioPerformanceAnalyzer,
    PortfolioScenario,
    ScenarioResult,
)

# ── Shared fixture ────────────────────────────────────────────────────────────

def _build_portfolio() -> PortfolioBuilder:
    pb = PortfolioBuilder("Test Portfolio")
    pb.add_property("P1", "Downtown Office", "commercial", 5_000_000, "high",    400_000, 500_000)
    pb.add_property("P2", "Maadi Apartment",  "residential", 2_000_000, "medium", 150_000, 180_000)
    pb.add_property("P3", "New Cairo Land",   "land",        1_500_000, "low",     60_000,  80_000)
    return pb


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_portfolio_scenario_dataclass():
    """PortfolioScenario stores params and to_dict() is complete."""
    sc = PortfolioScenario("stress", noi_shock=0.75, value_shock=0.80, cap_rate_shift=0.015)
    d  = sc.to_dict()
    assert d["label"]          == "stress"
    assert d["noi_shock"]      == 0.75
    assert d["value_shock"]    == 0.80
    assert d["cap_rate_shift"] == 0.015
    print("[PASS] test_portfolio_scenario_dataclass")


def test_standard_scenarios_created():
    """create_standard_scenarios() produces exactly 3 presets."""
    pb  = _build_portfolio()
    ppa = PortfolioPerformanceAnalyzer(pb)
    ppa.create_standard_scenarios()
    labels = [sc.label for sc in ppa.scenarios]
    assert len(ppa.scenarios) == 3
    assert "pessimistic" in labels
    assert "base"        in labels
    assert "optimistic"  in labels
    print("[PASS] test_standard_scenarios_created")


def test_run_scenarios_returns_three_results():
    """run_scenarios() returns one ScenarioResult per scenario."""
    pb  = _build_portfolio()
    ppa = PortfolioPerformanceAnalyzer(pb)
    ppa.create_standard_scenarios()
    results = ppa.run_scenarios()
    assert len(results) == 3
    for r in results:
        assert isinstance(r, ScenarioResult)
    print("[PASS] test_run_scenarios_returns_three_results")


def test_pessimistic_reduces_noi():
    """Pessimistic stressed NOI < base stressed NOI."""
    pb  = _build_portfolio()
    ppa = PortfolioPerformanceAnalyzer(pb)
    ppa.create_standard_scenarios()
    results = {r.scenario_label: r for r in ppa.run_scenarios()}
    assert results["pessimistic"].stressed_total_noi < results["base"].stressed_total_noi
    print(f"[PASS] test_pessimistic_reduces_noi  — "
          f"pessimistic NOI {results['pessimistic'].stressed_total_noi:,.0f} "
          f"< base {results['base'].stressed_total_noi:,.0f}")


def test_optimistic_increases_value():
    """Optimistic stressed value > base stressed value."""
    pb  = _build_portfolio()
    ppa = PortfolioPerformanceAnalyzer(pb)
    ppa.create_standard_scenarios()
    results = {r.scenario_label: r for r in ppa.run_scenarios()}
    assert results["optimistic"].stressed_portfolio_value > results["base"].stressed_portfolio_value
    print(f"[PASS] test_optimistic_increases_value  — "
          f"optimistic {results['optimistic'].stressed_portfolio_value:,.0f} "
          f"> base {results['base'].stressed_portfolio_value:,.0f}")


def test_base_value_change_pct_is_zero():
    """Base scenario value_change_pct == 0 (no shock applied)."""
    pb  = _build_portfolio()
    ppa = PortfolioPerformanceAnalyzer(pb)
    ppa.create_standard_scenarios()
    results = {r.scenario_label: r for r in ppa.run_scenarios()}
    assert results["base"].value_change_pct == 0.0
    assert results["base"].noi_change_pct   == 0.0
    print("[PASS] test_base_value_change_pct_is_zero")


def test_irr_estimates_positive():
    """Every scenario produces a positive IRR estimate."""
    pb  = _build_portfolio()
    ppa = PortfolioPerformanceAnalyzer(pb)
    ppa.create_standard_scenarios()
    for r in ppa.run_scenarios():
        assert r.irr_estimate > 0, f"{r.scenario_label} IRR {r.irr_estimate} not positive"
    print("[PASS] test_irr_estimates_positive")


def test_performance_summary_structure():
    """get_performance_summary() contains all required top-level keys."""
    pb  = _build_portfolio()
    ppa = PortfolioPerformanceAnalyzer(pb)
    ppa.create_standard_scenarios()
    s = ppa.get_performance_summary()
    required = {
        "portfolio_name", "scenario_count", "base_portfolio_value", "base_total_noi",
        "base_cap_rate", "diversification_score", "scenarios",
        "min_stressed_value", "max_stressed_value",
        "min_stressed_noi",   "max_stressed_noi",
        "value_at_risk_pct",
    }
    missing = required - set(s.keys())
    assert not missing, f"Missing keys: {missing}"
    assert s["scenario_count"] == 3
    print(f"[PASS] test_performance_summary_structure  — {len(s)} keys present")


def test_scenario_value_ordering():
    """Pessimistic portfolio value < base < optimistic."""
    pb  = _build_portfolio()
    ppa = PortfolioPerformanceAnalyzer(pb)
    ppa.create_standard_scenarios()
    results = {r.scenario_label: r for r in ppa.run_scenarios()}
    pess = results["pessimistic"].stressed_portfolio_value
    base = results["base"].stressed_portfolio_value
    opti = results["optimistic"].stressed_portfolio_value
    assert pess < base < opti, f"Ordering violated: {pess:.0f} < {base:.0f} < {opti:.0f}"
    print(f"[PASS] test_scenario_value_ordering  — {pess:,.0f} < {base:,.0f} < {opti:,.0f}")


def test_value_at_risk_pct_bounded():
    """value_at_risk_pct is in (0, 1) for a portfolio with a pessimistic shock."""
    pb  = _build_portfolio()
    ppa = PortfolioPerformanceAnalyzer(pb)
    ppa.create_standard_scenarios()
    s = ppa.get_performance_summary()
    var = s["value_at_risk_pct"]
    assert 0.0 < var < 1.0, f"VaR {var} out of (0,1)"
    print(f"[PASS] test_value_at_risk_pct_bounded  — VaR = {var:.2%}")


_SAMPLE_PROPERTIES = [
    {"property_id": "A1", "property_name": "Tower A", "property_type": "commercial",
     "valuation_value": 10_000_000, "valuation_confidence": "high",
     "annual_noi": 700_000, "annual_gross_income": 900_000},
    {"property_id": "A2", "property_name": "Villa B", "property_type": "residential",
     "valuation_value": 3_000_000, "valuation_confidence": "medium",
     "annual_noi": 180_000, "annual_gross_income": 220_000},
]


def test_api_performance_standard_scenarios():
    """POST /api/valuation/portfolio/performance returns 200 with 3 standard scenarios."""
    with app.test_client() as client:
        resp = client.post("/api/valuation/portfolio/performance", json={
            "portfolio_name": "API Test Portfolio",
            "properties": _SAMPLE_PROPERTIES,
            "use_standard_scenarios": True,
        }, headers=_AUTH_HDR)
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["status"] == "success"
        pa = data["performance_analysis"]
        assert pa["scenario_count"] == 3
        assert pa["value_at_risk_pct"] > 0
    print(f"[PASS] test_api_performance_standard_scenarios  — "
          f"{pa['scenario_count']} scenarios, VaR {pa['value_at_risk_pct']:.2%}")


def test_api_performance_custom_scenario():
    """POST /api/valuation/portfolio/performance accepts a custom scenario list."""
    with app.test_client() as client:
        resp = client.post("/api/valuation/portfolio/performance", json={
            "portfolio_name": "Custom Scenario Portfolio",
            "properties": _SAMPLE_PROPERTIES,
            "use_standard_scenarios": False,
            "scenarios": [
                {"label": "stress_test", "noi_shock": 0.70, "value_shock": 0.75, "cap_rate_shift": 0.02},
            ],
        }, headers=_AUTH_HDR)
        assert resp.status_code == 200
        data = resp.get_json()
        pa = data["performance_analysis"]
        assert pa["scenario_count"] == 1
        assert pa["scenarios"][0]["scenario_label"] == "stress_test"
    print(f"[PASS] test_api_performance_custom_scenario  — "
          f"1 custom scenario, value {pa['scenarios'][0]['stressed_portfolio_value']:,.0f}")


# ── Sheet tests (Task 11.3) ───────────────────────────────────────────────────

def _make_perf_summary(with_scenarios: bool = True) -> dict:
    """Build a minimal performance summary dict matching get_performance_summary() shape."""
    base = {
        "portfolio_name":        "Test Portfolio",
        "scenario_count":        0,
        "base_portfolio_value":  8_500_000.0,
        "base_total_noi":        610_000.0,
        "base_cap_rate":         0.0718,
        "diversification_score": 0.59,
        "scenarios":             [],
    }
    if with_scenarios:
        base["scenarios"] = [
            {
                "scenario_label":           "pessimistic",
                "stressed_portfolio_value": 7_225_000.0,
                "stressed_total_noi":       488_000.0,
                "stressed_cap_rate":        0.0685,
                "stressed_noi_margin":      0.62,
                "value_change_pct":         -0.15,
                "noi_change_pct":           -0.20,
                "irr_estimate":             0.1185,
                "diversification_score":    0.59,
            },
        ]
        base["scenario_count"]      = 1
        base["min_stressed_value"]  = 7_225_000.0
        base["max_stressed_value"]  = 7_225_000.0
        base["min_stressed_noi"]    = 488_000.0
        base["max_stressed_noi"]    = 488_000.0
        base["value_at_risk_pct"]   = 0.15
    return base


def test_sheet_portfolio_performance_created():
    """sheet_portfolio_performance() creates a 'Portfolio Performance' sheet."""
    perf = _make_perf_summary(with_scenarios=True)
    b = ExcelReportBuilder()
    b.sheet_portfolio_performance(perf)
    sheet_names = [ws.title for ws in b.workbook.worksheets]
    assert "Portfolio Performance" in sheet_names, f"Sheet not found in: {sheet_names}"
    print(f"[PASS] test_sheet_portfolio_performance_created  — sheets: {sheet_names}")


def test_sheet_missing_optional_keys_no_crash():
    """Missing conditional VaR keys (no scenarios run) must not raise."""
    perf = _make_perf_summary(with_scenarios=False)
    # Intentionally omit all conditional keys
    for k in ("min_stressed_value", "max_stressed_value",
               "min_stressed_noi", "max_stressed_noi", "value_at_risk_pct"):
        perf.pop(k, None)
    b = ExcelReportBuilder()
    try:
        b.sheet_portfolio_performance(perf)
        print("[PASS] test_sheet_missing_optional_keys_no_crash")
    except Exception as exc:
        raise AssertionError(f"Crashed on missing keys: {exc}") from exc


def test_sheet_empty_scenarios_message():
    """Empty scenarios list renders 'No scenario data available' on the sheet."""
    perf = _make_perf_summary(with_scenarios=False)
    b = ExcelReportBuilder()
    b.sheet_portfolio_performance(perf)
    ws = b.workbook["Portfolio Performance"]
    found = any(
        ws.cell(row=r, column=1).value == "No scenario data available"
        for r in range(1, ws.max_row + 1)
    )
    assert found, "Expected 'No scenario data available' cell not found"
    print("[PASS] test_sheet_empty_scenarios_message")


def test_sheet_baseline_block_values():
    """Baseline block contains the portfolio name from the input dict."""
    perf = _make_perf_summary(with_scenarios=True)
    b = ExcelReportBuilder()
    b.sheet_portfolio_performance(perf)
    ws = b.workbook["Portfolio Performance"]
    all_values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert "Test Portfolio" in all_values, \
        f"Portfolio name not found in column B values: {all_values}"
    print("[PASS] test_sheet_baseline_block_values")


# ── Runner ────────────────────────────────────────────────────────────────────

def test_all_phase_11_tests():
    print("\n=== Phase 11 E2E Tests ===")
    tests = [
        test_portfolio_scenario_dataclass,
        test_standard_scenarios_created,
        test_run_scenarios_returns_three_results,
        test_pessimistic_reduces_noi,
        test_optimistic_increases_value,
        test_base_value_change_pct_is_zero,
        test_irr_estimates_positive,
        test_performance_summary_structure,
        test_scenario_value_ordering,
        test_value_at_risk_pct_bounded,
        test_api_performance_standard_scenarios,
        test_api_performance_custom_scenario,
        test_sheet_portfolio_performance_created,
        test_sheet_missing_optional_keys_no_crash,
        test_sheet_empty_scenarios_message,
        test_sheet_baseline_block_values,
    ]
    passed = 0
    for fn in tests:
        try:
            fn()
            passed += 1
        except Exception as exc:
            print(f"[FAIL] {fn.__name__} — {exc}")
    print(f"\n{passed}/{len(tests)} tests passed.")
    if passed == len(tests):
        print("Phase 11 E2E COMPLETE")


if __name__ == "__main__":
    test_all_phase_11_tests()
