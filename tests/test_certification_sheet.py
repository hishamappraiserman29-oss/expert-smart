#!/usr/bin/env python3
"""Smoke tests for certification_sheet.apply_certification_sheet."""

import io
import sys
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parent.parent
_CORE = _ROOT / "core_engine"
for _p in (str(_ROOT), str(_CORE)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from openpyxl import Workbook

from core_engine.reports.sheets.certification_sheet import apply_certification_sheet


# ── Shared fixtures ───────────────────────────────────────────────────────────

_FULL: dict = {
    "appraiser_name": "هشام محمد المهدى",
    "license_no":     "29",
    "report_date":    "2026/05/12",
}

_MINIMAL: dict = {}


def _build(inputs: dict) -> object:
    wb = Workbook()
    ws = wb.active
    ws.title = "شهادة"
    return ws, apply_certification_sheet(ws, inputs)


def _all_values(ws) -> list[str]:
    return [str(c.value or "") for row in ws.iter_rows() for c in row]


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_apply_creates_banner_at_row_1():
    """A1 must contain the sheet banner text."""
    ws, _ = _build(_FULL)
    assert ws["A1"].value is not None
    assert "شهادة المقيم" in str(ws["A1"].value)
    print("✓ test_apply_creates_banner_at_row_1")


def test_apply_returns_field_locations():
    """Return value must be a non-empty dict of (row, col) tuples."""
    _, locs = _build(_FULL)
    assert isinstance(locs, dict)
    assert len(locs) > 0
    for key, val in locs.items():
        assert isinstance(val, tuple) and len(val) == 2
    print(f"✓ test_apply_returns_field_locations ({len(locs)} fields tracked)")


def test_apply_writes_appraiser_name():
    """Appraiser name must appear in the sheet."""
    ws, _ = _build(_FULL)
    all_vals = _all_values(ws)
    assert any("هشام محمد المهدى" in v for v in all_vals), "appraiser_name not written"
    print("✓ test_apply_writes_appraiser_name")


def test_apply_writes_certification_text():
    """Arabic certification text must be present (keyword 'أشهد')."""
    ws, _ = _build(_FULL)
    all_vals = _all_values(ws)
    assert any("أشهد" in v for v in all_vals), "cert text missing"
    print("✓ test_apply_writes_certification_text")


def test_apply_applies_rtl():
    """Sheet must be set to right-to-left reading direction."""
    ws, _ = _build(_FULL)
    assert ws.sheet_view.rightToLeft is True
    print("✓ test_apply_applies_rtl")


def test_apply_no_freeze_panes():
    """Certification sheet must NOT set freeze_panes."""
    ws, _ = _build(_FULL)
    assert ws.freeze_panes is None or ws.freeze_panes == "A1"
    print("✓ test_apply_no_freeze_panes")


def test_apply_handles_empty_inputs():
    """Empty inputs dict must not raise and must still return a dict."""
    ws, locs = _build(_MINIMAL)
    assert isinstance(locs, dict)
    assert ws["A1"].value is not None
    print("✓ test_apply_handles_empty_inputs")


def test_apply_tracked_keys_present():
    """appraiser_name, license_no, report_date, signature must be tracked."""
    _, locs = _build(_FULL)
    for key in ("cert_text", "appraiser_name", "license_no", "report_date", "signature"):
        assert key in locs, f"Missing tracked key: {key}"
    print("✓ test_apply_tracked_keys_present (5 keys)")


def test_apply_reviewer_name_alias():
    """reviewer_name alias must work when appraiser_name is absent."""
    inputs = {"reviewer_name": "Test Reviewer", "report_date": "2026/05/12"}
    ws, _ = _build(inputs)
    all_vals = _all_values(ws)
    assert any("Test Reviewer" in v for v in all_vals), "reviewer_name alias not used"
    print("✓ test_apply_reviewer_name_alias")


if __name__ == "__main__":
    test_apply_creates_banner_at_row_1()
    test_apply_returns_field_locations()
    test_apply_writes_appraiser_name()
    test_apply_writes_certification_text()
    test_apply_applies_rtl()
    test_apply_no_freeze_panes()
    test_apply_handles_empty_inputs()
    test_apply_tracked_keys_present()
    test_apply_reviewer_name_alias()
    print("\n✅ كل الاختبارات (9) نجحت")
