#!/usr/bin/env python3
"""Smoke tests for cost_approach_sheet.apply_cost_approach_sheet."""

import io
import sys
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parent.parent
_CORE = _ROOT / "core_engine"
for _p in (str(_ROOT), str(_CORE)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from openpyxl import Workbook

from core_engine.reports.sheets.cost_approach_sheet import apply_cost_approach_sheet


# ── Shared fixtures ───────────────────────────────────────────────────────────

_FULL: dict = {
    "area":                   150.0,
    "land_value":             1_500_000,
    "construction_cost_sqm":  8_000,
    "depreciation_rate":      0.10,
    "additional_items":       50_000,
    "report_date":            "2026/05/12",
}

_MINIMAL: dict = {}


def _build(inputs: dict) -> object:
    wb = Workbook()
    ws = wb.active
    ws.title = "طريقة التكلفة"
    return ws, apply_cost_approach_sheet(ws, inputs)


def _all_values(ws) -> list[str]:
    return [str(c.value or "") for row in ws.iter_rows() for c in row]


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_apply_creates_banner_at_row_1():
    """A1 must contain the sheet banner text."""
    ws, _ = _build(_FULL)
    assert ws["A1"].value is not None
    assert "طريقة التكلفة" in str(ws["A1"].value)
    print("✓ test_apply_creates_banner_at_row_1")


def test_apply_returns_field_locations():
    """Return value must be a non-empty dict of (row, col) tuples."""
    _, locs = _build(_FULL)
    assert isinstance(locs, dict)
    assert len(locs) > 0
    for key, val in locs.items():
        assert isinstance(val, tuple) and len(val) == 2
        assert isinstance(val[0], int) and isinstance(val[1], int)
    print(f"✓ test_apply_returns_field_locations ({len(locs)} fields tracked)")


def test_apply_writes_all_section_headers():
    """All 4 section headings must appear somewhere in the sheet."""
    ws, _ = _build(_FULL)
    all_vals = _all_values(ws)
    for heading in (
        "قيمة الأرض",
        "تكلفة الإنشاء والبناء",
        "الاستهلاك والتقادم",
        "القيمة الاستدلالية بطريقة التكلفة",
    ):
        assert any(heading in v for v in all_vals), f"Missing section: {heading}"
    print("✓ test_apply_writes_all_section_headers (4 sections)")


def test_apply_applies_rtl():
    """Sheet must be set to right-to-left reading direction."""
    ws, _ = _build(_FULL)
    assert ws.sheet_view.rightToLeft is True
    print("✓ test_apply_applies_rtl")


def test_apply_freeze_panes_set():
    """Freeze panes must be B3 for the cost approach sheet."""
    ws, _ = _build(_FULL)
    assert ws.freeze_panes == "B3"
    print("✓ test_apply_freeze_panes_set")


def test_apply_handles_empty_inputs():
    """Empty inputs dict must not raise and must still return a dict."""
    ws, locs = _build(_MINIMAL)
    assert isinstance(locs, dict)
    assert ws["A1"].value is not None
    print("✓ test_apply_handles_empty_inputs")


def test_apply_indicated_value_in_locations():
    """'indicated_value' key must be present in the returned location map."""
    _, locs = _build(_FULL)
    assert "indicated_value" in locs, "indicated_value not tracked in locs"
    row, col = locs["indicated_value"]
    assert row > 1
    print(f"✓ test_apply_indicated_value_in_locations (row={row}, col={col})")


def test_apply_tracked_keys_present():
    """All expected tracked keys must appear in returned locs."""
    _, locs = _build(_FULL)
    for key in ("area", "land_value", "cost_sqm", "const_cost",
                "add_items", "depr_pct", "depr_amt", "net_building_value",
                "indicated_value"):
        assert key in locs, f"Missing tracked key: {key}"
    print("✓ test_apply_tracked_keys_present (9 keys)")


def test_apply_alternate_input_keys():
    """floor_area_m2 / land_price / replacement_cost aliases must work."""
    inputs = {
        "floor_area_m2":    200.0,
        "land_price":       2_000_000,
        "replacement_cost": 1_600_000,
        "depreciation":     0.05,
    }
    ws, locs = _build(inputs)
    assert isinstance(locs, dict)
    assert ws["A1"].value is not None
    print("✓ test_apply_alternate_input_keys")


if __name__ == "__main__":
    test_apply_creates_banner_at_row_1()
    test_apply_returns_field_locations()
    test_apply_writes_all_section_headers()
    test_apply_applies_rtl()
    test_apply_freeze_panes_set()
    test_apply_handles_empty_inputs()
    test_apply_indicated_value_in_locations()
    test_apply_tracked_keys_present()
    test_apply_alternate_input_keys()
    print("\n✅ كل الاختبارات (9) نجحت")
