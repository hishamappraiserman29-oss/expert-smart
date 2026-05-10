"""
Quick self-test for ExcelTemplateRenderer.

Runs without touching any existing reports or templates.
Creates a minimal in-memory template, renders it, reloads and verifies.

Run from project root:
    python _test_excel_template_renderer.py
"""

import sys
import tempfile
from pathlib import Path

# Allow imports from core_engine/
sys.path.insert(0, str(Path(__file__).parent / "core_engine"))

from openpyxl import Workbook, load_workbook

from reports.excel_template_renderer import ExcelTemplateRenderer, build_from_template


def _make_template(path: str) -> None:
    """Create a minimal .xlsx template with placeholders and a TABLE anchor."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Scalar placeholders — mixed (pure + embedded)
    ws["A1"] = "Report Date: {{report_date}}"
    ws["A2"] = "Purpose: {{valuation_purpose}}"
    ws["A3"] = "{{total_portfolio_value}}"          # pure numeric placeholder
    ws["A3"].number_format = "#,##0.00"             # should survive
    ws["A4"] = "Reviewer: {{reviewer_name}}"
    ws["A5"] = "COD: {{cod}} — PRD: {{prd}}"        # two placeholders, same cell

    # Unrecognised key — should be left as-is
    ws["B1"] = "{{UNKNOWN_KEY}}"

    # Formula — must NOT be touched
    ws["B2"] = "=SUM(A3:A3)"

    # TABLE anchor
    ws["A7"] = "{{TABLE:properties_results}}"

    wb.save(path)


def test_renderer() -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        template_path = Path(tmpdir) / "template.xlsx"
        output_path   = Path(tmpdir) / "output.xlsx"

        _make_template(str(template_path))

        context = {
            "report_date":           "2026-05-10",
            "valuation_purpose":     "Mortgage Lending",
            "total_portfolio_value": 125_000_000,
            "avg_price_per_meter":   18_500,
            "median_ratio":          0.982,
            "cod":                   4.5,
            "prd":                   1.02,
            "reviewer_name":         "Hisham Elmahdy",
        }

        tables = {
            "properties_results": [
                {"Property ID": "P001", "Address": "Cairo",      "Value (EGP)": 5_000_000},
                {"Property ID": "P002", "Address": "Alexandria", "Value (EGP)": 3_200_000},
            ],
        }

        result = build_from_template(
            template_path=template_path,
            output_path=output_path,
            context=context,
            tables=tables,
        )

        assert result is not None, "build_from_template returned None — template not found?"
        assert result.exists(), "Output file was not written to disk"

        # ── Reload and verify ────────────────────────────────────────────────
        wb2 = load_workbook(str(output_path), data_only=True)
        ws2 = wb2["Report"]

        # Embedded placeholder
        assert ws2["A1"].value == "Report Date: 2026-05-10", (
            f"A1 mismatch: {ws2['A1'].value!r}"
        )
        assert ws2["A2"].value == "Purpose: Mortgage Lending", (
            f"A2 mismatch: {ws2['A2'].value!r}"
        )

        # Pure numeric placeholder — raw value, not string
        assert ws2["A3"].value == 125_000_000, (
            f"A3 mismatch (expected int 125000000): {ws2['A3'].value!r}"
        )

        # Reviewer
        assert ws2["A4"].value == "Reviewer: Hisham Elmahdy", (
            f"A4 mismatch: {ws2['A4'].value!r}"
        )

        # Two placeholders in one cell
        assert ws2["A5"].value == "COD: 4.5 — PRD: 1.02", (
            f"A5 mismatch: {ws2['A5'].value!r}"
        )

        # Unknown key must be preserved unchanged
        assert ws2["B1"].value == "{{UNKNOWN_KEY}}", (
            f"B1 mismatch — unknown key should not be touched: {ws2['B1'].value!r}"
        )

        # TABLE anchor row (row 7) becomes header row
        assert ws2["A7"].value == "Property ID", (
            f"A7 (table header) mismatch: {ws2['A7'].value!r}"
        )
        assert ws2["B7"].value == "Address", (
            f"B7 (table header) mismatch: {ws2['B7'].value!r}"
        )
        assert ws2["C7"].value == "Value (EGP)", (
            f"C7 (table header) mismatch: {ws2['C7'].value!r}"
        )

        # Data rows inserted at rows 8 and 9
        assert ws2["A8"].value == "P001", f"A8 mismatch: {ws2['A8'].value!r}"
        assert ws2["B8"].value == "Cairo", f"B8 mismatch: {ws2['B8'].value!r}"
        assert ws2["C8"].value == 5_000_000, f"C8 mismatch: {ws2['C8'].value!r}"
        assert ws2["A9"].value == "P002", f"A9 mismatch: {ws2['A9'].value!r}"

        wb2.close()

        print("PASS  Placeholder replacement  -- passed")
        print("PASS  Pure numeric placeholder -- passed (raw int, not string)")
        print("PASS  Multi-placeholder cell   -- passed")
        print("PASS  Unknown key preserved    -- passed")
        print("PASS  Table header row         -- passed")
        print("PASS  Table data rows          -- passed")
        print("PASS  Output file exists       -- passed")
        print()
        print("All assertions passed.")


def test_fallback_no_template() -> None:
    """Renderer must return None (not raise) when no template is given."""
    renderer = ExcelTemplateRenderer(template_path=None, context={})
    result = renderer.build("/tmp/should_not_exist.xlsx")
    assert result is None, f"Expected None, got {result!r}"

    renderer2 = ExcelTemplateRenderer(
        template_path="/nonexistent/path/template.xlsx", context={}
    )
    result2 = renderer2.build("/tmp/should_not_exist_2.xlsx")
    assert result2 is None, f"Expected None, got {result2!r}"

    print("PASS  Fallback (no template)    -- passed")


if __name__ == "__main__":
    print("=== ExcelTemplateRenderer self-test ===\n")
    test_fallback_no_template()
    test_renderer()
    print("\nAll tests passed. Safe to commit.")
