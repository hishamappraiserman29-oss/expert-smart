#!/usr/bin/env python3
"""Smoke tests for inputs_sheet.apply_inputs_sheet."""

import io
import sys
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parent.parent
_CORE = _ROOT / "core_engine"
for _p in (str(_ROOT), str(_CORE)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from openpyxl import Workbook

from core_engine.reports.sheets.inputs_sheet import apply_inputs_sheet


# ── Shared minimal inputs dict ────────────────────────────────────────────────
_MINIMAL: dict = {}

_FULL: dict = {
    "report_id":         "VAL-TEST-001",
    "valuation_date":    "2026/05/12",
    "inspection_date":   "2026/05/01",
    "primary_purpose":   "تمويل بنكى",
    "client_name":       "Test Client",
    "appraiser_name":    "Test Appraiser",
    "license_no":        "29",
    "instructed_by":     "Test Bank",
    "asset_type":        "شقة سكنية",
    "location":          "التجمع الخامس",
    "area":              150,
    "year_built":        2010,
    "condition":         "ممتاز",
    "market_avg_price_sqm": 33000,
    "cap_rate":          0.08,
    "vacancy_rate":      0.05,
    "comparable":        4_900_000,
    "cost":              4_800_000,
    "income":            5_000_000,
    "primary_value":     4_900_000,
    "confidence":        "High",
    "report_date":       "2026/05/12",
    "weights": {"comparable": 0.6, "cost": 0.2, "income": 0.2},
}


def _build(inputs: dict) -> object:
    wb = Workbook()
    ws = wb.active
    ws.title = "الافتراضات والمدخلات"
    return ws, apply_inputs_sheet(ws, inputs)


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_apply_creates_banner_at_row_1():
    """A1 must contain the sheet banner text."""
    ws, _ = _build(_FULL)
    assert ws["A1"].value is not None
    assert "الافتراضات والمدخلات" in str(ws["A1"].value)
    print("✓ test_apply_creates_banner_at_row_1")


def test_apply_returns_field_locations():
    """Return value must be a non-empty dict of (row, col) tuples."""
    _, locs = _build(_FULL)
    assert isinstance(locs, dict)
    assert len(locs) > 0
    for key, val in locs.items():
        assert isinstance(val, tuple) and len(val) == 2
        assert isinstance(val[0], int) and isinstance(val[1], int)
    print(f"✓ test_apply_returns_field_locations ({len(locs)} fields tracked)")


def test_apply_writes_all_section_headers():
    """All 7 section headings must be present somewhere in the sheet."""
    ws, _ = _build(_FULL)
    all_vals = [str(c.value or "") for row in ws.iter_rows() for c in row]
    required = [
        "بيانات التقرير",
        "بيانات العميل والمقيم",
        "بيانات العقار",
        "بيانات السوق",
        "افتراضات التقييم",
        "حدود ومحددات الاستخدام",
        "إعدادات التقرير",
    ]
    for heading in required:
        assert any(heading in v for v in all_vals), f"Missing section: {heading}"
    print("✓ test_apply_writes_all_section_headers (7 sections)")


def test_apply_applies_rtl_view():
    """Sheet must be set to right-to-left reading direction."""
    ws, _ = _build(_FULL)
    assert ws.sheet_view.rightToLeft is True
    print("✓ test_apply_applies_rtl_view")


def test_apply_handles_empty_inputs():
    """Empty inputs dict must not raise and must still return a dict."""
    ws, locs = _build(_MINIMAL)
    assert isinstance(locs, dict)
    # Banner must still render
    assert ws["A1"].value is not None
    print("✓ test_apply_handles_empty_inputs")


def test_apply_primary_value_in_locations():
    """'primary_value' key must be in the returned location map."""
    _, locs = _build(_FULL)
    assert "primary_value" in locs, "primary_value not tracked in locs"
    row, col = locs["primary_value"]
    assert row > 1 and col >= 2
    print(f"✓ test_apply_primary_value_in_locations (row={row}, col={col})")


def test_apply_field_value_written_correctly():
    """The value written to the cell must match the inputs dict."""
    ws, locs = _build(_FULL)
    assert "report_id" in locs
    row, col = locs["report_id"]
    assert ws.cell(row=row, column=col).value == "VAL-TEST-001"
    print("✓ test_apply_field_value_written_correctly")


def test_apply_freeze_panes_set():
    """Freeze panes must be set to B3 for the inputs sheet."""
    ws, _ = _build(_FULL)
    assert ws.freeze_panes == "B3"
    print("✓ test_apply_freeze_panes_set")


def test_apply_weights_flat_keys():
    """Flat weights_comparable/cost/income keys work as fallback."""
    inputs = {
        "primary_value": 1_000_000,
        "comparable": 1_000_000,
        "weights_comparable": 0.5,
        "weights_cost":       0.3,
        "weights_income":     0.2,
    }
    ws, locs = _build(inputs)
    # Should not raise; weight fields must be present
    assert "weight_comparable" in locs
    print("✓ test_apply_weights_flat_keys")


if __name__ == "__main__":
    test_apply_creates_banner_at_row_1()
    test_apply_returns_field_locations()
    test_apply_writes_all_section_headers()
    test_apply_applies_rtl_view()
    test_apply_handles_empty_inputs()
    test_apply_primary_value_in_locations()
    test_apply_field_value_written_correctly()
    test_apply_freeze_panes_set()
    test_apply_weights_flat_keys()
    print("\n✅ كل الاختبارات (9) نجحت")
