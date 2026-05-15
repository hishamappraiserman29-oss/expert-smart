#!/usr/bin/env python3
"""Smoke tests for main_report_sheet.apply_main_report_sheet."""

import io
import sys
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parent.parent
_CORE = _ROOT / "core_engine"
for _p in (str(_ROOT), str(_CORE)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from openpyxl import Workbook

from core_engine.reports.sheets.main_report_sheet import apply_main_report_sheet


# ── Shared fixtures ───────────────────────────────────────────────────────────

_APPRAISER = {
    "appraiser_name": "هشام محمد المهدى",
    "license_id":     29,
}

_PROPERTY_INFO = {
    "asset_type":      "شقة سكنية",
    "location":        "التجمع الخامس، القاهرة",
    "primary_purpose": "تمويل بنكى",
    "valuation_date":  "2026/05/12",
    "confidence":      "High",
    "report_date":     "2026/05/12",
}

_VALUATION_RESULTS = {
    "primary_value": 4_900_000,
    "comparable":    4_900_000,
    "cost":          4_800_000,
    "income":        5_000_000,
    "weights":       {"comparable": 0.6, "cost": 0.2, "income": 0.2},
}


def _build(profile_key: str = "legacy", appraiser: dict | None = None) -> object:
    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير"
    apply_main_report_sheet(
        ws,
        appraiser=_APPRAISER if appraiser is None else appraiser,
        property_info=_PROPERTY_INFO,
        valuation_results=_VALUATION_RESULTS,
        profile_key=profile_key,
    )
    return ws


def _all_values(ws) -> list[str]:
    return [str(c.value or "") for row in ws.iter_rows() for c in row]


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_apply_writes_property_info():
    """Property type and location must appear in the sheet."""
    ws  = _build()
    vals = _all_values(ws)
    assert any("شقة سكنية" in v for v in vals), "asset_type not written"
    assert any("تقرير التقييم العقاري" in v for v in vals), "title header missing"
    print("✓ test_apply_writes_property_info")


def test_apply_writes_appraiser_data():
    """Appraiser name must appear when a non-empty appraiser dict is supplied."""
    ws   = _build(appraiser={"appraiser_name": "Test Appraiser Name"})
    vals = _all_values(ws)
    assert any("Test Appraiser Name" in v for v in vals), "appraiser name not written"
    print("✓ test_apply_writes_appraiser_data")


def test_apply_skips_appraiser_when_empty():
    """No appraiser row when appraiser dict has no name key."""
    ws   = _build(appraiser={})
    vals = _all_values(ws)
    assert not any("المُقيِّم:" in v for v in vals), "unexpected appraiser row"
    print("✓ test_apply_skips_appraiser_when_empty")


def test_apply_kpi_dashboard_only_for_pro_profile():
    """'مؤشرات الأداء الرئيسية' section only appears for professional/detailed."""
    ws_pro     = _build(profile_key="professional")
    ws_detail  = _build(profile_key="detailed")
    for ws, label in ((ws_pro, "professional"), (ws_detail, "detailed")):
        vals = _all_values(ws)
        assert any("مؤشرات الأداء" in v for v in vals), \
            f"KPI section missing for profile_key='{label}'"
    print("✓ test_apply_kpi_dashboard_only_for_pro_profile")


def test_apply_skips_kpi_for_legacy():
    """'مؤشرات الأداء الرئيسية' section must NOT appear for legacy."""
    ws   = _build(profile_key="legacy")
    vals = _all_values(ws)
    assert not any("مؤشرات الأداء الرئيسية" in v for v in vals), \
        "KPI section must be absent for legacy profile"
    print("✓ test_apply_skips_kpi_for_legacy")


def test_apply_applies_rtl():
    """Sheet must be set to right-to-left reading direction."""
    ws = _build()
    assert ws.sheet_view.rightToLeft is True
    print("✓ test_apply_applies_rtl")


def test_apply_has_three_approaches_table():
    """All three approach labels must appear in the sheet."""
    ws   = _build()
    vals = _all_values(ws)
    for label in ("المقارنة البيعية", "طريقة التكلفة", "رأسمالة الدخل"):
        assert any(label in v for v in vals), f"Approach row missing: {label}"
    print("✓ test_apply_has_three_approaches_table")


def test_apply_has_final_value_card():
    """Final value card text must be present in the sheet."""
    ws   = _build()
    vals = _all_values(ws)
    assert any("القيمة السوقية النهائية" in v for v in vals)
    print("✓ test_apply_has_final_value_card")


def test_apply_freeze_panes_set():
    """Freeze panes must be B4 for the main report sheet."""
    ws = _build()
    assert ws.freeze_panes == "B4"
    print("✓ test_apply_freeze_panes_set")


def test_apply_handles_empty_dicts():
    """All three mapping params as empty dicts must not raise."""
    wb = Workbook()
    ws = wb.active
    apply_main_report_sheet(
        ws,
        appraiser={},
        property_info={},
        valuation_results={},
        profile_key="legacy",
    )
    assert ws["A1"].value is not None
    print("✓ test_apply_handles_empty_dicts")


if __name__ == "__main__":
    test_apply_writes_property_info()
    test_apply_writes_appraiser_data()
    test_apply_skips_appraiser_when_empty()
    test_apply_kpi_dashboard_only_for_pro_profile()
    test_apply_skips_kpi_for_legacy()
    test_apply_applies_rtl()
    test_apply_has_three_approaches_table()
    test_apply_has_final_value_card()
    test_apply_freeze_panes_set()
    test_apply_handles_empty_dicts()
    print("\n✅ كل الاختبارات (10) نجحت")
