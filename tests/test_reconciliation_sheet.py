#!/usr/bin/env python3
"""Smoke tests for reconciliation_sheet.apply_reconciliation_sheet."""

import io
import sys
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parent.parent
_CORE = _ROOT / "core_engine"
for _p in (str(_ROOT), str(_CORE)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from openpyxl import Workbook

from core_engine.reports.sheets.reconciliation_sheet import apply_reconciliation_sheet


# ── Shared fixtures ───────────────────────────────────────────────────────────

_FULL: dict = {
    "primary_value": 4_900_000,
    "comparable":    4_900_000,
    "cost":          4_800_000,
    "income":        5_000_000,
    "weights":       {"comparable": 0.6, "cost": 0.2, "income": 0.2},
    "report_date":   "2026/05/12",
}

_MINIMAL: dict = {}


def _build(inputs: dict) -> object:
    wb = Workbook()
    ws = wb.active
    ws.title = "توفيق النتائج"
    return ws, apply_reconciliation_sheet(ws, inputs)


def _all_values(ws) -> list[str]:
    return [str(c.value or "") for row in ws.iter_rows() for c in row]


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_apply_creates_banner_at_row_1():
    """A1 must contain the sheet banner text."""
    ws, _ = _build(_FULL)
    assert ws["A1"].value is not None
    assert "توفيق النتائج" in str(ws["A1"].value)
    print("✓ test_apply_creates_banner_at_row_1")


def test_apply_returns_field_locations():
    """Return value must be a non-empty dict of (row, col) tuples."""
    _, locs = _build(_FULL)
    assert isinstance(locs, dict)
    assert len(locs) > 0
    for key, val in locs.items():
        assert isinstance(val, tuple) and len(val) == 2
    print(f"✓ test_apply_returns_field_locations ({len(locs)} fields tracked)")


def test_apply_writes_three_approach_labels():
    """All three approach labels must appear in the sheet."""
    ws, _ = _build(_FULL)
    all_vals = _all_values(ws)
    for label in ("أسلوب المقارنة البيعية", "أسلوب التكلفة", "رأسمالة الدخل"):
        assert any(label in v for v in all_vals), f"Missing row: {label}"
    print("✓ test_apply_writes_three_approach_labels")


def test_apply_writes_notes_section():
    """Notes section header 'ملاحظات التوفيق' must be present."""
    ws, _ = _build(_FULL)
    all_vals = _all_values(ws)
    assert any("ملاحظات التوفيق" in v for v in all_vals)
    print("✓ test_apply_writes_notes_section")


def test_apply_applies_rtl():
    """Sheet must be set to right-to-left reading direction."""
    ws, _ = _build(_FULL)
    assert ws.sheet_view.rightToLeft is True
    print("✓ test_apply_applies_rtl")


def test_apply_freeze_panes_set():
    """Freeze panes must be B3 for the reconciliation sheet."""
    ws, _ = _build(_FULL)
    assert ws.freeze_panes == "B3"
    print("✓ test_apply_freeze_panes_set")


def test_apply_handles_empty_inputs():
    """Empty inputs dict must not raise and must still return a dict."""
    ws, locs = _build(_MINIMAL)
    assert isinstance(locs, dict)
    assert ws["A1"].value is not None
    print("✓ test_apply_handles_empty_inputs")


def test_apply_final_value_in_locations():
    """'final_value' key must be present in the returned location map."""
    _, locs = _build(_FULL)
    assert "final_value" in locs, "final_value not tracked in locs"
    row, col = locs["final_value"]
    assert row > 4
    print(f"✓ test_apply_final_value_in_locations (row={row}, col={col})")


def test_apply_weight_keys_in_locations():
    """weight_comparable, weight_cost, weight_income must all be tracked."""
    _, locs = _build(_FULL)
    for key in ("weight_comparable", "weight_cost", "weight_income"):
        assert key in locs, f"Missing tracked key: {key}"
    print("✓ test_apply_weight_keys_in_locations")


def test_apply_flat_weight_keys():
    """Flat weights_comparable/cost/income keys work as fallback."""
    inputs = {
        "primary_value":      4_900_000,
        "comparable":         4_900_000,
        "weights_comparable": 0.6,
        "weights_cost":       0.2,
        "weights_income":     0.2,
    }
    ws, locs = _build(inputs)
    assert isinstance(locs, dict)
    assert ws["A1"].value is not None
    print("✓ test_apply_flat_weight_keys")


if __name__ == "__main__":
    test_apply_creates_banner_at_row_1()
    test_apply_returns_field_locations()
    test_apply_writes_three_approach_labels()
    test_apply_writes_notes_section()
    test_apply_applies_rtl()
    test_apply_freeze_panes_set()
    test_apply_handles_empty_inputs()
    test_apply_final_value_in_locations()
    test_apply_weight_keys_in_locations()
    test_apply_flat_weight_keys()
    print("\n✅ كل الاختبارات (10) نجحت")
