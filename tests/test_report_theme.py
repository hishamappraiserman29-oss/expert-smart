#!/usr/bin/env python3
"""Smoke tests لـ report_theme."""

import io
import sys
from pathlib import Path

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_ROOT = Path(__file__).resolve().parent.parent
_CORE = _ROOT / "core_engine"
for _p in (str(_ROOT), str(_CORE)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill

from core_engine.reports.report_theme import (
    BuilderPalette,
    Borders,
    NumFormat,
    Palette,
    Typography,
    apply_sheet_defaults,
    draw_banner,
    draw_section,
    get_alignment,
    get_fill,
    get_font,
    style_body,
    style_input_value,
    style_label,
    style_page_title,
    style_section_header,
    style_table_header,
    style_value,
)


def test_palette_has_required_colors():
    """كل ألوان الـ Midnight Gold الأساسية موجودة."""
    required = [
        "INK", "NAVY_DEEP", "NAVY", "INDIGO",
        "GOLD", "GOLD_LIGHT", "GOLD_PALE",
        "EMERALD", "CORAL", "PURPLE",
        "WHITE", "GREY_300",
    ]
    for name in required:
        assert hasattr(Palette, name), f"Missing color: {name}"
        val = getattr(Palette, name)
        assert isinstance(val, str)
        assert len(val) == 6, f"{name} should be 6-char hex, got {val}"
        int(val, 16)
    print("✓ test_palette_has_required_colors")


def test_palette_hex_no_hash():
    """ألوان الـ palette بدون علامة #."""
    assert not Palette.INK.startswith("#")
    assert not Palette.GOLD.startswith("#")
    print("✓ test_palette_hex_no_hash")


def test_typography_constants():
    assert Typography.PRIMARY == "Cairo"
    assert Typography.FALLBACK == "Calibri"
    assert Typography.SIZE_PAGE_TITLE == 22
    assert Typography.SIZE_LABEL == 10.5
    print("✓ test_typography_constants")


def test_numformat_codes():
    assert "#,##0" in NumFormat.CURRENCY
    assert "%" in NumFormat.PERCENT
    assert "0.0%" == NumFormat.PERCENT
    assert "EGP" in NumFormat.CURRENCY_EGP
    print("✓ test_numformat_codes")


def test_get_font_returns_font():
    f = get_font(size=11, bold=True, color=Palette.INK)
    assert isinstance(f, Font)
    assert f.bold is True
    assert f.name == "Cairo"
    assert f.size == 11
    print("✓ test_get_font_returns_font")


def test_get_fill_returns_patternfill():
    p = get_fill(Palette.GOLD)
    assert isinstance(p, PatternFill)
    assert p.patternType == "solid"
    print("✓ test_get_fill_returns_patternfill")


def test_get_alignment_rtl():
    a = get_alignment(h="right")
    assert isinstance(a, Alignment)
    assert a.horizontal == "right"
    assert a.readingOrder == 2  # RTL
    print("✓ test_get_alignment_rtl")


def test_borders_presets():
    assert isinstance(Borders.LIGHT, Border)
    assert isinstance(Borders.GOLD, Border)
    assert isinstance(Borders.HEAD, Border)
    assert isinstance(Borders.BANNER_BOTTOM, Border)
    print("✓ test_borders_presets")


def test_style_page_title_applies_correctly():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "اختبار"
    style_page_title(ws["A1"])
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.color.rgb == "00" + Palette.GOLD_LIGHT
    assert ws["A1"].fill.start_color.rgb == "00" + Palette.INK
    print("✓ test_style_page_title_applies_correctly")


def test_style_input_value_uses_blue():
    """قيم الإدخال يجب تكون بخط أزرق (Wall Street convention)."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 100
    style_input_value(ws["A1"])
    assert ws["A1"].font.color.rgb == "00" + Palette.INPUT_BLUE
    assert ws["A1"].font.bold is True
    print("✓ test_style_input_value_uses_blue")


def test_draw_banner_merges_cells():
    wb = Workbook()
    ws = wb.active
    draw_banner(ws, 1, 8, "عنوان تجريبى")
    assert ws["A1"].value == "عنوان تجريبى"
    merged_ranges = [str(r) for r in ws.merged_cells.ranges]
    assert "A1:H1" in merged_ranges
    assert ws.row_dimensions[1].height == 48
    print("✓ test_draw_banner_merges_cells")


def test_draw_section_creates_header():
    wb = Workbook()
    ws = wb.active
    draw_section(ws, 3, 5, "قسم")
    assert ws["A3"].value == "قسم"
    merged_ranges = [str(r) for r in ws.merged_cells.ranges]
    assert "A3:E3" in merged_ranges
    print("✓ test_draw_section_creates_header")


def test_apply_sheet_defaults_sets_rtl():
    wb = Workbook()
    ws = wb.active
    ws.title = "تجريبى"
    apply_sheet_defaults(ws)
    assert ws.sheet_view.rightToLeft is True
    assert ws.freeze_panes == "A2"
    print("✓ test_apply_sheet_defaults_sets_rtl")


def test_apply_sheet_defaults_no_freeze_for_certificate():
    """شيت 'شهادة' لا يجب أن يحتوى freeze panes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "شهادة"
    apply_sheet_defaults(ws)
    assert ws.sheet_view.rightToLeft is True
    assert ws.freeze_panes is None
    print("✓ test_apply_sheet_defaults_no_freeze_for_certificate")


def test_apply_sheet_defaults_freeze_optional():
    """freeze=False يلغى الـ freeze panes."""
    wb = Workbook()
    ws = wb.active
    apply_sheet_defaults(ws, freeze=False)
    assert ws.freeze_panes is None
    print("✓ test_apply_sheet_defaults_freeze_optional")


def test_builder_palette_has_required_tokens():
    """BuilderPalette يحتوى على كل الـ 22 hex token المطلوبة."""
    required = [
        "HEADER", "SECTION_MID", "SECTION_DARK", "SUBHEAD",
        "INPUT_CELL", "CALC_CELL", "SUCCESS_LIGHT", "ROW_BAND",
        "PORT_COL", "PORTFOLIO", "CB_GREEN",
        "ERROR", "WARNING", "ADJ_NEG", "ADJ_ZERO", "ADJ_GOLD", "ADJ_EMERALD",
        "SUCCESS_DARK", "MUTED", "NOTE", "GOLD_DARK", "AMBER_DARK",
    ]
    for name in required:
        assert hasattr(BuilderPalette, name), f"Missing token: {name}"
    print("✓ test_builder_palette_has_required_tokens")


def test_builder_palette_hex_valid():
    """كل tokens في BuilderPalette هى hex صالحة من 6 أحرف."""
    tokens = [
        "HEADER", "SECTION_MID", "SECTION_DARK", "SUBHEAD",
        "INPUT_CELL", "CALC_CELL", "SUCCESS_LIGHT", "ROW_BAND",
        "PORT_COL", "PORTFOLIO", "CB_GREEN",
        "ERROR", "WARNING", "ADJ_NEG", "ADJ_ZERO", "ADJ_GOLD", "ADJ_EMERALD",
        "SUCCESS_DARK", "MUTED", "NOTE", "GOLD_DARK", "AMBER_DARK",
    ]
    for name in tokens:
        val = getattr(BuilderPalette, name)
        assert isinstance(val, str), f"{name} not a str"
        assert len(val) == 6, f"{name} should be 6-char hex, got {val!r}"
        int(val, 16)  # raises if not valid hex
    print("✓ test_builder_palette_hex_valid")


def test_builder_palette_no_hash():
    """BuilderPalette tokens بدون علامة #."""
    assert not BuilderPalette.HEADER.startswith("#")
    assert not BuilderPalette.SUCCESS_LIGHT.startswith("#")
    print("✓ test_builder_palette_no_hash")


def test_numformat_currency_2dp():
    """NumFormat.CURRENCY_2DP يطابق '#,##0.00' المستخدمة فى ExcelReportBuilder."""
    assert NumFormat.CURRENCY_2DP == "#,##0.00"
    assert "0.00" in NumFormat.CURRENCY_2DP
    print("✓ test_numformat_currency_2dp")


def test_builder_palette_get_fill_integration():
    """get_fill(BuilderPalette.X) يُنتج PatternFill صحيح."""
    from openpyxl.styles import PatternFill
    fill = get_fill(BuilderPalette.HEADER)
    assert isinstance(fill, PatternFill)
    assert fill.patternType == "solid"
    assert fill.fgColor.rgb.endswith(BuilderPalette.HEADER)
    print("✓ test_builder_palette_get_fill_integration")


if __name__ == "__main__":
    test_palette_has_required_colors()
    test_palette_hex_no_hash()
    test_typography_constants()
    test_numformat_codes()
    test_get_font_returns_font()
    test_get_fill_returns_patternfill()
    test_get_alignment_rtl()
    test_borders_presets()
    test_style_page_title_applies_correctly()
    test_style_input_value_uses_blue()
    test_draw_banner_merges_cells()
    test_draw_section_creates_header()
    test_apply_sheet_defaults_sets_rtl()
    test_apply_sheet_defaults_no_freeze_for_certificate()
    test_apply_sheet_defaults_freeze_optional()
    test_builder_palette_has_required_tokens()
    test_builder_palette_hex_valid()
    test_builder_palette_no_hash()
    test_numformat_currency_2dp()
    test_builder_palette_get_fill_integration()
    print("\n✅ كل الاختبارات (20) نجحت")
