#!/usr/bin/env python3
"""
tests/test_sheets_bundle.py — Integration tests for the 6 sheet modules.

Covers:
  report_sheet       → apply_report_sheet
  sales_comparison   → apply_sales_comparison
  rental_sheet       → apply_rental_sheet
  cost_method        → apply_cost_method
  reconciliation     → apply_reconciliation
  certificate_sheet  → apply_certificate_sheet

Run:
  $env:PYTHONPATH = (Get-Location).Path + ';' + (Get-Location).Path + '/core_engine'
  python tests/test_sheets_bundle.py
"""
from __future__ import annotations

import io
import sys
import os

# UTF-8 fix for Arabic + emoji on Windows cp1252
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(
        sys.stdout.buffer, encoding="utf-8", errors="replace"
    )

# Make sure both the project root and core_engine/ are importable
_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
_CORE = os.path.join(_ROOT, "core_engine")
for _p in (_ROOT, _CORE):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import unittest
from openpyxl import Workbook


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _ws(title: str):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# 1. report_sheet
# ═══════════════════════════════════════════════════════════════════════════════
from reports.sheets.report_sheet import (
    apply_report_sheet,
    get_field_count as rs_field_count,
    get_section_count as rs_section_count,
)


class TestReportSheet(unittest.TestCase):

    def setUp(self):
        self.ws = _ws("التقرير")
        self.locs = apply_report_sheet(self.ws)

    def test_returns_dict(self):
        self.assertIsInstance(self.locs, dict)

    def test_locs_not_empty(self):
        self.assertGreater(len(self.locs), 0)

    def test_required_keys_present(self):
        for key in ("primary_value", "property_type", "valuation_date"):
            self.assertIn(key, self.locs, f"Missing key: {key}")

    def test_locs_are_row_col_tuples(self):
        for k, v in self.locs.items():
            self.assertIsInstance(v, tuple, f"locs[{k!r}] is not a tuple")
            self.assertEqual(len(v), 2, f"locs[{k!r}] should be (row, col)")
            self.assertGreater(v[0], 0)
            self.assertGreater(v[1], 0)

    def test_rtl(self):
        self.assertTrue(self.ws.sheet_view.rightToLeft)

    def test_banner_row1_has_value(self):
        self.assertIsNotNone(self.ws.cell(row=1, column=1).value)

    def test_field_count(self):
        self.assertGreater(rs_field_count(), 0)

    def test_section_count(self):
        self.assertGreaterEqual(rs_section_count(), 3)

    def test_default_data_passthrough(self):
        ws2 = _ws("التقرير")
        locs2 = apply_report_sheet(ws2, data={"primary_value": 9_000_000})
        self.assertIn("primary_value", locs2)

    def test_no_crash_with_empty_data(self):
        ws2 = _ws("التقرير")
        apply_report_sheet(ws2, data={})


# ═══════════════════════════════════════════════════════════════════════════════
# 2. sales_comparison
# ═══════════════════════════════════════════════════════════════════════════════
from reports.sheets.sales_comparison import (
    apply_sales_comparison,
    get_field_count as sc_field_count,
    get_section_count as sc_section_count,
    _ADJ_LABELS,
    _DEFAULT_COMPS,
)


class TestSalesComparison(unittest.TestCase):

    def setUp(self):
        self.ws = _ws("مقارنات البيوع")
        self.locs = apply_sales_comparison(self.ws)

    def test_returns_dict(self):
        self.assertIsInstance(self.locs, dict)

    def test_adj_labels_present_in_locs(self):
        for adj_key, _ in _ADJ_LABELS:
            self.assertIn(adj_key, self.locs, f"Missing adj key: {adj_key}")

    def test_result_keys_present(self):
        for key in ("net_adj", "adj_price", "weight", "final_weighted_price"):
            self.assertIn(key, self.locs, f"Missing result key: {key}")

    def test_final_weighted_price_has_formula(self):
        row, col = self.locs["final_weighted_price"]
        val = self.ws.cell(row=row, column=col).value
        self.assertIsInstance(val, str)
        self.assertIn("SUMPRODUCT", val.upper())

    def test_net_adj_has_sum_formula(self):
        row, col = self.locs["net_adj"]
        val = self.ws.cell(row=row, column=col).value
        self.assertIsInstance(val, str)
        self.assertIn("SUM", val.upper())

    def test_adj_price_has_formula(self):
        row, col = self.locs["adj_price"]
        val = self.ws.cell(row=row, column=col).value
        self.assertIsInstance(val, str)
        self.assertTrue(val.startswith("="))

    def test_rtl(self):
        self.assertTrue(self.ws.sheet_view.rightToLeft)

    def test_nine_adjustment_factors(self):
        self.assertEqual(len(_ADJ_LABELS), 9)

    def test_field_count(self):
        self.assertGreater(sc_field_count(), 0)

    def test_section_count(self):
        self.assertGreaterEqual(sc_section_count(), 3)

    def test_custom_comps_accepted(self):
        ws2 = _ws("مقارنات البيوع")
        comps = [
            {"location": "X", "area": 100, "price_per_sqm": 30_000,
             "adj_location": 0.0, "adj_area": 0.0, "adj_floor": 0.0,
             "adj_age": 0.0, "adj_condition": 0.0, "adj_view": 0.0,
             "adj_timing": 0.0, "adj_facade": 0.0, "adj_services": 0.0,
             "weight": 1.0},
        ]
        apply_sales_comparison(ws2, data={"comparables": comps})

    def test_default_comps_count(self):
        self.assertEqual(len(_DEFAULT_COMPS), 3)


# ═══════════════════════════════════════════════════════════════════════════════
# 3. rental_sheet
# ═══════════════════════════════════════════════════════════════════════════════
from reports.sheets.rental_sheet import (
    apply_rental_sheet,
    get_field_count as rn_field_count,
    get_section_count as rn_section_count,
    _COMP_ROWS,
)


class TestRentalSheet(unittest.TestCase):

    def setUp(self):
        self.ws = _ws("المقارنات الإيجارية")
        self.locs = apply_rental_sheet(self.ws)

    def test_returns_dict(self):
        self.assertIsInstance(self.locs, dict)

    def test_comp_rows_eight(self):
        self.assertEqual(len(_COMP_ROWS), 8)

    def test_result_keys_present(self):
        for key in ("indicated_rent_sqm", "total_annual_rent",
                    "cap_rate", "income_value", "grm"):
            self.assertIn(key, self.locs, f"Missing key: {key}")

    def test_comp_row_keys_in_locs(self):
        for _, field_key in _COMP_ROWS:
            expected = f"comp_{field_key}"
            self.assertIn(expected, self.locs, f"Missing comp key: {expected}")

    def test_rtl(self):
        self.assertTrue(self.ws.sheet_view.rightToLeft)

    def test_freeze_panes_set(self):
        self.assertIsNotNone(self.ws.freeze_panes)

    def test_field_count(self):
        self.assertGreater(rn_field_count(), 0)

    def test_section_count(self):
        self.assertGreaterEqual(rn_section_count(), 2)

    def test_custom_data_accepted(self):
        ws2 = _ws("المقارنات الإيجارية")
        apply_rental_sheet(ws2, data={"cap_rate": 0.09, "income_value": 800_000})


# ═══════════════════════════════════════════════════════════════════════════════
# 4. cost_method
# ═══════════════════════════════════════════════════════════════════════════════
from reports.sheets.cost_method import (
    apply_cost_method,
    get_field_count as cm_field_count,
    get_section_count as cm_section_count,
    _DEFAULTS as cm_defaults,
    _SECTION_NAMES as cm_sections,
    _compute_defaults,
)


class TestCostMethod(unittest.TestCase):

    def setUp(self):
        self.ws = _ws("طريقة التكلفة")
        self.locs = apply_cost_method(self.ws)

    def test_returns_dict(self):
        self.assertIsInstance(self.locs, dict)

    def test_four_sections(self):
        self.assertEqual(len(cm_sections), 4)

    def test_indicated_value_key(self):
        self.assertIn("indicated_value", self.locs)

    def test_land_and_building_keys(self):
        for key in ("land_value", "net_building_value", "total_construction",
                    "dep_total", "dep_amount"):
            self.assertIn(key, self.locs, f"Missing key: {key}")

    def test_rtl(self):
        self.assertTrue(self.ws.sheet_view.rightToLeft)

    def test_freeze_panes(self):
        self.assertIsNotNone(self.ws.freeze_panes)

    def test_compute_defaults_land_value(self):
        d = {
            "land_area_m2": 100.0, "land_price_per_m2": 10_000,
            "built_area_m2": 100.0, "construction_cost_m2": 8_000,
            "contractor_profit": 0.15,
            "dep_physical": 0.05, "dep_functional": 0.03, "dep_external": 0.02,
        }
        _compute_defaults(d)
        self.assertAlmostEqual(d["land_value"], 1_000_000)
        self.assertAlmostEqual(d["gross_construction"], 800_000)

    def test_compute_defaults_dep_cap(self):
        d = {
            "land_area_m2": 100.0, "land_price_per_m2": 10_000,
            "built_area_m2": 100.0, "construction_cost_m2": 8_000,
            "contractor_profit": 0.0,
            "dep_physical": 0.50, "dep_functional": 0.40, "dep_external": 0.30,
        }
        _compute_defaults(d)
        self.assertLessEqual(d["dep_total"], 0.80)

    def test_field_count(self):
        self.assertGreater(cm_field_count(), 0)

    def test_section_count(self):
        self.assertEqual(cm_section_count(), 4)

    def test_no_crash_custom_data(self):
        ws2 = _ws("طريقة التكلفة")
        apply_cost_method(ws2, data={"land_area_m2": 200.0, "land_price_per_m2": 20_000})


# ═══════════════════════════════════════════════════════════════════════════════
# 5. reconciliation
# ═══════════════════════════════════════════════════════════════════════════════
from reports.sheets.reconciliation import (
    apply_reconciliation,
    get_field_count as rec_field_count,
    get_section_count as rec_section_count,
    _SECTION_NAMES as rec_sections,
    _DEFAULTS as rec_defaults,
)


class TestReconciliation(unittest.TestCase):

    def setUp(self):
        self.ws = _ws("توفيق النتائج")
        self.locs = apply_reconciliation(self.ws)

    def test_returns_dict(self):
        self.assertIsInstance(self.locs, dict)

    def test_three_approach_keys(self):
        for prefix in ("comparable", "cost", "income"):
            self.assertIn(f"{prefix}_value", self.locs,
                          f"Missing key: {prefix}_value")
            self.assertIn(f"w_{prefix}", self.locs,
                          f"Missing key: w_{prefix}")

    def test_primary_value_key(self):
        self.assertIn("primary_value", self.locs)

    def test_final_value_kpi_key(self):
        self.assertIn("final_value_kpi", self.locs)

    def test_additional_kpi_keys(self):
        for key in ("price_per_m2", "area_total", "confidence"):
            self.assertIn(key, self.locs, f"Missing key: {key}")

    def test_rtl(self):
        self.assertTrue(self.ws.sheet_view.rightToLeft)

    def test_freeze_panes(self):
        self.assertIsNotNone(self.ws.freeze_panes)

    def test_weights_sum_to_one(self):
        w = (rec_defaults["w_comparable"] + rec_defaults["w_cost"]
             + rec_defaults["w_income"])
        self.assertAlmostEqual(w, 1.0, places=6)

    def test_field_count(self):
        self.assertEqual(rec_field_count(), len(rec_defaults))

    def test_three_section_names(self):
        self.assertEqual(rec_section_count(), len(rec_sections))

    def test_custom_data_accepted(self):
        ws2 = _ws("توفيق النتائج")
        apply_reconciliation(ws2, data={"primary_value": 5_000_000})


# ═══════════════════════════════════════════════════════════════════════════════
# 6. certificate_sheet
# ═══════════════════════════════════════════════════════════════════════════════
from reports.sheets.certificate_sheet import (
    apply_certificate_sheet,
    get_field_count as cert_field_count,
    get_section_count as cert_section_count,
    _EGVS_REFS,
    _DEFAULTS as cert_defaults,
)


class TestCertificateSheet(unittest.TestCase):

    def setUp(self):
        # Must be titled "شهادة" for freeze exemption
        self.ws = _ws("شهادة")
        self.locs = apply_certificate_sheet(self.ws)

    def test_returns_dict(self):
        self.assertIsInstance(self.locs, dict)

    def test_appraiser_keys_present(self):
        for key in ("appraiser_name", "license_id", "authority", "email"):
            self.assertIn(key, self.locs, f"Missing key: {key}")

    def test_valuation_keys_present(self):
        for key in ("report_id", "property_type", "primary_value",
                    "valuation_date"):
            self.assertIn(key, self.locs, f"Missing key: {key}")

    def test_cert_text_key_present(self):
        self.assertIn("cert_text", self.locs)

    def test_rtl(self):
        self.assertTrue(self.ws.sheet_view.rightToLeft)

    def test_no_freeze_panes_for_shahada(self):
        # Certificate sheet must NOT have freeze panes
        self.assertIsNone(self.ws.freeze_panes)

    def test_six_egvs_refs(self):
        self.assertEqual(len(_EGVS_REFS), 6)

    def test_egvs_refs_have_two_items(self):
        for ref in _EGVS_REFS:
            self.assertEqual(len(ref), 2)

    def test_field_count(self):
        self.assertEqual(cert_field_count(), len(cert_defaults))

    def test_four_sections(self):
        self.assertEqual(cert_section_count(), 4)

    def test_custom_data_overrides(self):
        ws2 = _ws("شهادة")
        locs2 = apply_certificate_sheet(
            ws2, data={"appraiser_name": "تست مقيم", "license_id": 99}
        )
        row, col = locs2["appraiser_name"]
        self.assertEqual(ws2.cell(row=row, column=col).value, "تست مقيم")

    def test_banner_row1_has_value(self):
        self.assertIsNotNone(self.ws.cell(row=1, column=1).value)


# ═══════════════════════════════════════════════════════════════════════════════
# Runner
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    loader = unittest.TestLoader()
    suite  = unittest.TestSuite()
    for cls in (
        TestReportSheet,
        TestSalesComparison,
        TestRentalSheet,
        TestCostMethod,
        TestReconciliation,
        TestCertificateSheet,
    ):
        suite.addTests(loader.loadTestsFromTestCase(cls))

    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    total  = result.testsRun
    passed = total - len(result.failures) - len(result.errors)
    print(f"\n{'='*55}")
    print(f"  Sheet bundle: {passed}/{total} passed")
    print(f"{'='*55}")
    sys.exit(0 if result.wasSuccessful() else 1)
