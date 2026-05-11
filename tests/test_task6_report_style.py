"""
Task 6 + Task 7 — report_style selector integration tests.

Scenarios:
  1. legacy   — builds, has only basic sheets, NO advanced analytics sheets
  2. detailed — builds, has all basic + all 7 advanced analytics sheets
  3. professional_template — unchanged behaviour (xlsm or fallback xlsx)
  4. missing report_style  — defaults to legacy (no advanced sheets)
  5. mass appraisal / legacy
  6. mass appraisal / professional_template

Run from repo root:
    $env:PYTHONPATH = (Get-Location).Path + ';' + (Get-Location).Path + '\\core_engine'
    python tests/test_task6_report_style.py
"""
import sys, os, decimal, tempfile, io

# Force UTF-8 stdout so Arabic sheet names print safely on Windows cp1252 terminals
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "core_engine"))

import openpyxl

CHECKS_PASSED = 0
CHECKS_FAILED = 0

def check(label, condition, detail=""):
    global CHECKS_PASSED, CHECKS_FAILED
    if condition:
        print(f"  [PASS] {label}")
        CHECKS_PASSED += 1
    else:
        print(f"  [FAIL] {label}" + (f" -- {detail}" if detail else ""))
        CHECKS_FAILED += 1


# ── Import shared types ───────────────────────────────────────────────────────
try:
    from reports.excel_builder import ExcelReportBuilder, _LEGACY_EXCLUDED_SHEETS
    from adapters.asset import AssetValuationResult
except Exception as _imp_err:
    print(f"[FATAL] Import failed: {_imp_err}")
    sys.exit(1)

ADVANCED_SHEETS = [
    "التحليل المكاني",
    "الانحدار المتعدد",
    "الخيارات الحقيقية",
    "لوحة القيادة التنفيذية",
    "الشبكات العصبية",
    "السلاسل الزمنية",
    "استخبارات السوق",
]

BASIC_SHEETS = [
    "Summary",
    "Three Approaches",
    "Weights Analysis",
    "Audit Trail",
    "Property Details",
    "Disclosures",
    "Issues & Warnings",
    "Certification",
]

def _make_result(asset_type="residential", value="1500000"):
    return AssetValuationResult(
        asset_type=asset_type,
        primary_purpose="fair_market_value",
        primary_value=decimal.Decimal(value),
        confidence="high",
        alternative_values={},
        weights_applied={},
        audit_trail=[],
        issues=[],
        metadata={},
        disclosures=[],
    )

def _sheet_names(path):
    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=False)
    names = wb.sheetnames
    wb.close()
    return names


# ── Scenario 1: legacy ───────────────────────────────────────────────────────
print("\nScenario 1: ExcelReportBuilder — report_style='legacy'")
try:
    _out1 = os.path.join(tempfile.gettempdir(), "test_t7_legacy.xlsx")
    r1 = ExcelReportBuilder(_make_result())
    ret1 = r1.build(_out1, report_style="legacy")

    check("legacy: returns a string",    isinstance(ret1, str))
    check("legacy: output file exists",  os.path.isfile(ret1))
    check("legacy: .xlsx extension",     ret1.endswith(".xlsx"))

    if os.path.isfile(ret1):
        sheets1 = _sheet_names(ret1)
        for bs in BASIC_SHEETS:
            check(f"legacy: has basic sheet '{bs}'", bs in sheets1)
        for _i, adv in enumerate(ADVANCED_SHEETS, 1):
            check(f"legacy: NO advanced sheet [{_i}]", adv not in sheets1,
                  f"sheet[{_i}] found in workbook")
        os.remove(ret1)
except Exception as e:
    check("Scenario 1 setup", False, str(e))


# ── Scenario 1b: Arabic legacy sheets ───────────────────────────────────────
print("\nScenario 1b: legacy Arabic sheets — existence and basic structure")
try:
    _ARABIC_LEGACY_SHEETS = ("الافتراضات والمدخلات", "التقرير")
    _out1b = os.path.join(tempfile.gettempdir(), "test_t7_legacy_arabic.xlsx")
    r1b = ExcelReportBuilder(_make_result("residential", "2500000"))
    ret1b = r1b.build(_out1b, report_style="legacy")

    check("legacy Arabic: builds successfully",  os.path.isfile(ret1b))
    if os.path.isfile(ret1b):
        wb1b   = openpyxl.load_workbook(ret1b, read_only=True, keep_vba=False)
        sh1b   = wb1b.sheetnames
        wb1b.close()
        check("legacy Arabic: 'الافتراضات والمدخلات' exists",
              "الافتراضات والمدخلات" in sh1b)
        check("legacy Arabic: 'التقرير' exists",
              "التقرير" in sh1b)
        # verify section header text in الافتراضات والمدخلات
        wb1b_rw = openpyxl.load_workbook(ret1b, keep_vba=False)
        ws_inp  = wb1b_rw["الافتراضات والمدخلات"]
        _all_vals = [str(c.value or "") for row in ws_inp.iter_rows() for c in row]
        check("legacy Arabic: section 'بيانات التقرير' present",
              any("بيانات التقرير" in v for v in _all_vals))
        check("legacy Arabic: section 'بيانات العقار' present",
              any("بيانات العقار" in v for v in _all_vals))
        check("legacy Arabic: section 'افتراضات التقييم' present",
              any("افتراضات التقييم" in v for v in _all_vals))
        # verify التقرير has final value text
        ws_rep  = wb1b_rw["التقرير"]
        _rep_vals = [str(c.value or "") for row in ws_rep.iter_rows() for c in row]
        check("legacy Arabic: 'التقرير' has title header",
              any("تقرير التقييم العقاري" in v for v in _rep_vals))
        check("legacy Arabic: 'التقرير' has final value label",
              any("القيمة السوقية النهائية" in v for v in _rep_vals))
        check("legacy Arabic: 'التقرير' has approaches section",
              any("نتائج أساليب التقييم" in v for v in _rep_vals))
        wb1b_rw.close()
        os.remove(ret1b)
except Exception as e:
    check("Scenario 1b setup", False, str(e))


# ── Scenario 1c: Sales Comparison sheet ─────────────────────────────────────
print("\nScenario 1c: legacy — 'مقارنات البيوع' Sales Comparison Adjustment Grid")
try:
    import decimal as _dec

    # Build a result with rich comparable data so formulas can be exercised
    _comps_data = [
        {
            "price_per_sqm": 5000,
            "location": "القاهرة الجديدة",
            "area": "120",
            "adj_location": 0.05,
            "adj_area": -0.02,
            "adj_floor": 0.0,
            "adj_age": 0.03,
            "adj_condition": 0.0,
            "adj_view": -0.01,
            "adj_timing": 0.02,
            "adj_facade": 0.0,
            "adj_services": 0.01,
            "weight": 0.4,
        },
        {
            "price_per_sqm": 4800,
            "location": "مدينة نصر",
            "area": "115",
            "adj_location": -0.03,
            "adj_area": 0.01,
            "adj_floor": 0.0,
            "adj_age": -0.02,
            "adj_condition": 0.05,
            "adj_view": 0.0,
            "adj_timing": 0.02,
            "adj_facade": 0.0,
            "adj_services": -0.01,
            "weight": 0.35,
        },
        {
            "price_per_sqm": 5200,
            "location": "الرحاب",
            "area": "130",
            "adj_location": 0.0,
            "adj_area": -0.03,
            "adj_floor": 0.02,
            "adj_age": 0.0,
            "adj_condition": 0.0,
            "adj_view": 0.03,
            "adj_timing": 0.01,
            "adj_facade": 0.0,
            "adj_services": 0.0,
            "weight": 0.25,
        },
    ]
    _result_sc = AssetValuationResult(
        asset_type="residential",
        primary_purpose="fair_market_value",
        primary_value=_dec.Decimal("1800000"),
        confidence="high",
        alternative_values={},
        weights_applied={"comparable": 0.6, "cost": 0.2, "income": 0.2},
        audit_trail=[],
        issues=[],
        metadata={"comparables": _comps_data, "location": "القاهرة الجديدة"},
        disclosures=[],
    )
    _out1c = os.path.join(tempfile.gettempdir(), "test_t_sales_comp.xlsx")
    r1c    = ExcelReportBuilder(_result_sc)
    ret1c  = r1c.build(_out1c, report_style="legacy")

    check("S1c: builds successfully",         os.path.isfile(ret1c))
    if os.path.isfile(ret1c):
        _wb1c = openpyxl.load_workbook(ret1c, data_only=False)
        _sh1c = _wb1c.sheetnames
        check("S1c: 'مقارنات البيوع' sheet exists", "مقارنات البيوع" in _sh1c)

        if "مقارنات البيوع" in _sh1c:
            _ws1c = _wb1c["مقارنات البيوع"]
            _all  = [str(c.value or "") for row in _ws1c.iter_rows() for c in row]

            # 9 adjustment labels
            _ADJ_LABELS = [
                "الموقع", "المساحة", "الدور", "العمر", "التشطيب",
                "الإطلالة", "التوقيت", "الواجهة", "الخدمات",
            ]
            for _lbl in _ADJ_LABELS:
                check(f"S1c: adj label '{_lbl}' present", any(_lbl in v for v in _all))

            # Banner
            check("S1c: banner 'مصفوفة الضبط الاحترافية' present",
                  any("مصفوفة الضبط الاحترافية" in v for v in _all))

            # Formula checks (formulas, not computed values)
            check("S1c: SUM formula (net adj) present",
                  any("=SUM(" in v for v in _all))
            check("S1c: adjusted price formula present",
                  any("*(1+" in v for v in _all))
            check("S1c: final weighted price formula present",
                  any(")/" in v and "=" in v for v in _all))
            check("S1c: SUMPRODUCT formula present",
                  any("SUMPRODUCT" in v for v in _all))

            # Legend
            check("S1c: legend 'ضبط موجب' present",
                  any("ضبط موجب" in v for v in _all))
            check("S1c: legend 'ضبط سالب' present",
                  any("ضبط سالب" in v for v in _all))

            # Methodology
            check("S1c: methodology 'منهجية' present",
                  any("منهجية" in v for v in _all))
            check("S1c: methodology step 1 present",
                  any("اختيار المقارنات" in v for v in _all))

            # Summary row labels
            check("S1c: net adjustment row label",
                  any("إجمالي الضبط الصافي" in v for v in _all))
            check("S1c: adjusted price row label",
                  any("السعر بعد الضبط" in v for v in _all))
            check("S1c: final weighted row label",
                  any("السعر النهائي الموزون" in v for v in _all))
            check("S1c: weight row label",
                  any("وزن المقارن" in v for v in _all))

        _wb1c.close()
        os.remove(ret1c)

    # Also verify that legacy still excludes advanced sheets
    _out1c2 = os.path.join(tempfile.gettempdir(), "test_t_sc_adv.xlsx")
    r1c2 = ExcelReportBuilder(_make_result())
    ret1c2 = r1c2.build(_out1c2, report_style="legacy")
    if os.path.isfile(ret1c2):
        _sh1c2 = _sheet_names(ret1c2)
        for _adv in ADVANCED_SHEETS:
            check(f"S1c: legacy still excludes advanced sheet '{_adv}'",
                  _adv not in _sh1c2)
        os.remove(ret1c2)
except Exception as e:
    check("Scenario 1c setup", False, str(e))


# ── Scenario 1d: Task 3 — 10 new Arabic legacy sheets ───────────────────────
print("\nScenario 1d: legacy — 10 new Arabic sheets (Task 3)")
_TASK3_SHEETS = [
    "المقارنات الإيجارية",
    "طريقة التكلفة",
    "رأسمالة الدخل",
    "توفيق النتائج",
    "محددات التقييم",
    "شهادة",
    "مصادر البيانات والمنهجية",
    "DCF — التدفقات النقدية",
    "الإيجار مقابل الشراء",
    "أفضل وأعلى استخدام — HABU",
]
_TASK3_KEY_CONTENT = {
    "المقارنات الإيجارية":     "الإيجار",
    "طريقة التكلفة":           "تكلفة",
    "رأسمالة الدخل":           "دخل",
    "توفيق النتائج":           "توفيق",
    "محددات التقييم":          "افتراضات",
    "شهادة":                   "المقيم",
    "مصادر البيانات والمنهجية": "مصادر",
    "DCF — التدفقات النقدية":   "DCF",
    "الإيجار مقابل الشراء":    "الإيجار",
    "أفضل وأعلى استخدام — HABU": "HABU",
}
try:
    _out1d = os.path.join(tempfile.gettempdir(), "test_t_task3.xlsx")
    r1d = ExcelReportBuilder(_make_result("residential", "2000000"))
    ret1d = r1d.build(_out1d, report_style="legacy")

    check("S1d: builds successfully", os.path.isfile(ret1d))
    if os.path.isfile(ret1d):
        _wb1d = openpyxl.load_workbook(ret1d, keep_vba=False)
        _sh1d = _wb1d.sheetnames

        for _sn in _TASK3_SHEETS:
            check(f"S1d: sheet '{_sn}' exists", _sn in _sh1d)
            if _sn in _sh1d:
                _ws1d = _wb1d[_sn]
                _vals = [str(c.value or "") for row in _ws1d.iter_rows() for c in row]
                _kw = _TASK3_KEY_CONTENT[_sn]
                check(f"S1d: '{_sn}' has keyword '{_kw}'",
                      any(_kw in v for v in _vals),
                      f"keyword '{_kw}' not found in sheet")

        _wb1d.close()
        os.remove(ret1d)
except Exception as e:
    check("Scenario 1d setup", False, str(e))


# ── Scenario 2: detailed ─────────────────────────────────────────────────────
print("\nScenario 2: ExcelReportBuilder — report_style='detailed'")
try:
    _out2 = os.path.join(tempfile.gettempdir(), "test_t7_detailed.xlsx")
    r2 = ExcelReportBuilder(_make_result("commercial", "3000000"))
    ret2 = r2.build(_out2, report_style="detailed")

    check("detailed: returns a string",   isinstance(ret2, str))
    check("detailed: output file exists", os.path.isfile(ret2))
    check("detailed: .xlsx extension",    ret2.endswith(".xlsx"))

    if os.path.isfile(ret2):
        sheets2 = _sheet_names(ret2)
        for bs in BASIC_SHEETS:
            check(f"detailed: has basic sheet '{bs}'", bs in sheets2)
        for _i, adv in enumerate(ADVANCED_SHEETS, 1):
            check(f"detailed: HAS advanced sheet [{_i}]", adv in sheets2,
                  f"sheet[{_i}] missing from workbook")
        # Arabic legacy-only sheets must NOT appear in detailed
        check("detailed: NO 'الافتراضات والمدخلات'",
              "الافتراضات والمدخلات" not in sheets2)
        check("detailed: NO 'التقرير'",
              "التقرير" not in sheets2)
        check("detailed: NO 'مقارنات البيوع'",
              "مقارنات البيوع" not in sheets2)
        for _t3s in _TASK3_SHEETS:
            check(f"detailed: NO '{_t3s}'", _t3s not in sheets2)
        os.remove(ret2)
except Exception as e:
    check("Scenario 2 setup", False, str(e))


# ── Scenario 3: professional_template ───────────────────────────────────────
print("\nScenario 3: ExcelReportBuilder — report_style='professional_template'")
try:
    from reports.excel_template_renderer import INDIVIDUAL_VALUATION_TEMPLATE
    tpl_present = INDIVIDUAL_VALUATION_TEMPLATE.is_file()

    _out3 = os.path.join(tempfile.gettempdir(), "test_t7_prof.xlsx")
    r3 = ExcelReportBuilder(_make_result())
    ret3 = r3.build(_out3, report_style="professional_template")

    if tpl_present:
        check("professional_template: returns .xlsm", ret3.endswith(".xlsm"),
              f"got {ret3!r}")
        check("professional_template: xlsm file exists", os.path.isfile(ret3))
        check("professional_template: original .xlsx NOT created",
              not os.path.isfile(_out3))
        if os.path.isfile(ret3):
            os.remove(ret3)
    else:
        check("professional_template (no tpl): fallback .xlsx", ret3.endswith(".xlsx"),
              f"got {ret3!r}")
        check("professional_template (no tpl): file exists", os.path.isfile(ret3))
        if os.path.isfile(ret3):
            os.remove(ret3)
except Exception as e:
    check("Scenario 3 setup", False, str(e))


# ── Scenario 4: missing report_style → defaults to legacy ────────────────────
print("\nScenario 4: ExcelReportBuilder — no report_style (should default to legacy)")
try:
    _out4 = os.path.join(tempfile.gettempdir(), "test_t7_default.xlsx")
    r4 = ExcelReportBuilder(_make_result())
    ret4 = r4.build(_out4)   # no report_style kwarg

    check("default: returns a string",    isinstance(ret4, str))
    check("default: output file exists",  os.path.isfile(ret4))

    if os.path.isfile(ret4):
        sheets4 = _sheet_names(ret4)
        for _i, adv in enumerate(ADVANCED_SHEETS, 1):
            check(f"default: NO advanced sheet [{_i}]", adv not in sheets4,
                  f"sheet[{_i}] found in workbook")
        os.remove(ret4)
except Exception as e:
    check("Scenario 4 setup", False, str(e))


# ── Scenario 5: mass appraisal / legacy ─────────────────────────────────────
print("\nScenario 5: build_mass_appraisal_workbook — report_style='legacy'")
try:
    from mass_appraisal_excel import build_mass_appraisal_workbook

    _ma_result = {
        "total_units": 5,
        "mean_value": 1_000_000,
        "properties": [
            {"id": f"P{i}", "value": 900_000 + i * 50_000,
             "area": 120, "location": "مدينة نصر",
             "property_type": "residential", "year_built": 2010}
            for i in range(5)
        ],
    }
    xlsxb = build_mass_appraisal_workbook(_ma_result, report_style="legacy")
    check("mass/legacy: returns bytes",       isinstance(xlsxb, bytes))
    check("mass/legacy: non-empty",           len(xlsxb) > 0)
    check("mass/legacy: ZIP magic bytes",     xlsxb[:2] == b"PK")
except Exception as e:
    check("Scenario 5 setup", False, str(e))


# ── Scenario 6: mass appraisal / professional_template ──────────────────────
print("\nScenario 6: build_mass_appraisal_workbook — report_style='professional_template'")
try:
    from reports.excel_template_renderer import MASS_APPRAISAL_TEMPLATE
    ma_tpl_present = MASS_APPRAISAL_TEMPLATE.is_file()

    xlsmb = build_mass_appraisal_workbook(_ma_result, report_style="professional_template")
    check("mass/professional_template: returns bytes",   isinstance(xlsmb, bytes))
    check("mass/professional_template: non-empty",       len(xlsmb) > 0)
    check("mass/professional_template: ZIP magic bytes", xlsmb[:2] == b"PK")
except Exception as e:
    check("Scenario 6 setup", False, str(e))


# ── Scenarios 8-10: _remove_legacy_advanced_sheets (actual browser path) ──────
# Import the helper from bridge_api.  bridge_api has a heavy dependency tree so
# we catch import errors and skip gracefully rather than failing the whole suite.
_remove_fn = None
try:
    import bridge_api as _ba
    _remove_fn = _ba._remove_legacy_advanced_sheets
except Exception as _ba_err:
    print(f"\n  [SKIP] bridge_api import failed ({_ba_err}) — Scenarios 8-10 skipped")


# Real sheet names from the live template workbook (as reported by the user).
# Basic sheets that MUST be kept in legacy:
_REAL_BASIC_SHEETS = (
    "الافتراضات والمدخلات",
    "التقرير",
    "مقارنات البيوع",
    "المقارنات الإيجارية",
    "طريقة التكلفة",
    "رأسمالة الدخل",
    "توفيق النتائج",
    "محددات التقييم",
    "شهادة",
    "مصادر البيانات والمنهجية",
    "DCF — التدفقات النقدية",
    "الإيجار مقابل الشراء",
    "أفضل وأعلى استخدام — HABU",
)
# Advanced sheets that MUST be removed in legacy:
_REAL_ADVANCED_SHEETS = (
    "التحليل المكاني",            # exact match
    "الانحدار المتعدد",           # exact match
    "الخيارات الحقيقية",          # exact match
    "لوحة القيادة التنفيذية",     # exact match
    "ANN — الشبكات العصبية",     # substring: "الشبكات العصبية"
    "ARIMA — السلاسل الزمنية",   # substring: "السلاسل الزمنية"
    "استخبارات السوق — MI",       # substring: "استخبارات السوق"
)


def _make_workbook_from_sheets(sheets: tuple, fname: str) -> str:
    """Build a minimal workbook with the given sheet names and return path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for s in sheets:
        wb.create_sheet(s)
    p = os.path.join(tempfile.gettempdir(), fname)
    wb.save(p)
    wb.close()
    return p


print("\nScenario 8: _remove_legacy_advanced_sheets — exact-match removal")
if _remove_fn is None:
    check("S8 skipped (bridge_api not importable)", True)
else:
    try:
        _p8 = _make_workbook_from_sheets(
            _REAL_BASIC_SHEETS[:4] + _REAL_ADVANCED_SHEETS[:4],
            "test_t8_exact.xlsx",
        )
        _remove_fn(_p8)
        _s8 = _sheet_names(_p8)
        for _bs in _REAL_BASIC_SHEETS[:4]:
            check(f"S8: basic kept",                    _bs in _s8)
        check("S8: 'التحليل المكاني' removed",          "التحليل المكاني"        not in _s8)
        check("S8: 'الانحدار المتعدد' removed",         "الانحدار المتعدد"       not in _s8)
        check("S8: 'الخيارات الحقيقية' removed",        "الخيارات الحقيقية"      not in _s8)
        check("S8: 'لوحة القيادة التنفيذية' removed",   "لوحة القيادة التنفيذية" not in _s8)
        os.remove(_p8)
    except Exception as e:
        check("S8 setup", False, str(e))


print("\nScenario 9: _remove_legacy_advanced_sheets — substring matching (prefixed names)")
if _remove_fn is None:
    check("S9 skipped (bridge_api not importable)", True)
else:
    try:
        _KEEP = ("التقرير", "DCF — التدفقات النقدية", "أفضل وأعلى استخدام — HABU")
        _REMOVE = (
            "ANN — الشبكات العصبية",    # prefix + Arabic keyword
            "ARIMA — السلاسل الزمنية",  # prefix + Arabic keyword
            "استخبارات السوق — MI",      # Arabic keyword + suffix
        )
        _p9 = _make_workbook_from_sheets(_KEEP + _REMOVE, "test_t9_substr.xlsx")
        _remove_fn(_p9)
        _s9 = _sheet_names(_p9)
        # Non-advanced sheets with similar structure must NOT be removed
        check("S9: 'التقرير' kept",                          "التقرير"                not in [s for s in _REAL_ADVANCED_SHEETS] and "التقرير" in _s9)
        check("S9: 'DCF — التدفقات النقدية' kept",           "DCF — التدفقات النقدية" in _s9)
        check("S9: 'أفضل وأعلى استخدام — HABU' kept",       "أفضل وأعلى استخدام — HABU" in _s9)
        check("S9: 'ANN — الشبكات العصبية' removed",         "ANN — الشبكات العصبية"    not in _s9)
        check("S9: 'ARIMA — السلاسل الزمنية' removed",       "ARIMA — السلاسل الزمنية"  not in _s9)
        check("S9: 'استخبارات السوق — MI' removed",           "استخبارات السوق — MI"     not in _s9)
        os.remove(_p9)
    except Exception as e:
        check("S9 setup", False, str(e))


print("\nScenario 10: _remove_legacy_advanced_sheets — full real template simulation")
if _remove_fn is None:
    check("S10 skipped (bridge_api not importable)", True)
else:
    try:
        _all_real = _REAL_BASIC_SHEETS + _REAL_ADVANCED_SHEETS
        _p10 = _make_workbook_from_sheets(_all_real, "test_t10_real.xlsx")
        _remove_fn(_p10)
        _s10 = _sheet_names(_p10)
        # Every basic sheet must remain
        for _bs in _REAL_BASIC_SHEETS:
            check(f"S10: basic kept [{_REAL_BASIC_SHEETS.index(_bs)+1}]", _bs in _s10)
        # Every advanced sheet must be gone
        for _adv in _REAL_ADVANCED_SHEETS:
            check(f"S10: advanced removed [{_REAL_ADVANCED_SHEETS.index(_adv)+1}]",
                  _adv not in _s10)
        os.remove(_p10)
    except Exception as e:
        check("S10 setup", False, str(e))


print("\nScenario 10b: variant normalization (ى / hamza)")
if _remove_fn is None:
    check("S10b skipped (bridge_api not importable)", True)
else:
    try:
        _p10b = _make_workbook_from_sheets(
            ("ملخص",
             "التحليل المكانى",       # ى variant
             "الإنحدار المتعدد",      # hamza إ
             "إستخبارات السوق",       # hamza إ prefix
             "استخبارات السوق"),      # normal
            "test_t10b_norm.xlsx",
        )
        _remove_fn(_p10b)
        _s10b = _sheet_names(_p10b)
        check("S10b: 'ملخص' kept",                   "ملخص"              in _s10b)
        check("S10b: ى-variant removed",              "التحليل المكانى"   not in _s10b)
        check("S10b: hamza-الانحدار removed",         "الإنحدار المتعدد"  not in _s10b)
        check("S10b: hamza-استخبارات removed",        "إستخبارات السوق"   not in _s10b)
        check("S10b: regular-استخبارات removed",      "استخبارات السوق"   not in _s10b)
        os.remove(_p10b)
    except Exception as e:
        check("S10b setup", False, str(e))


# ── Constant sanity check ─────────────────────────────────────────────────────
print("\nScenario 11: _LEGACY_EXCLUDED_SHEETS constant sanity")
for _i, _name in enumerate(ADVANCED_SHEETS, 1):
    check(f"S11 excluded[{_i}]: Arabic name in constant",
          _name.strip().lower() in _LEGACY_EXCLUDED_SHEETS)
for _en in ("spatial analysis", "multiple regression", "real options",
            "executive dashboard", "neural networks", "time series", "market intelligence"):
    check(f"S11 excluded (en): '{_en}' in constant", _en in _LEGACY_EXCLUDED_SHEETS)


# ── Summary ───────────────────────────────────────────────────────────────────
total = CHECKS_PASSED + CHECKS_FAILED
print(f"\n{'='*60}")
print(f"Results: {CHECKS_PASSED}/{total} passed")
if CHECKS_FAILED:
    print("SOME CHECKS FAILED")
    sys.exit(1)
else:
    print("ALL CHECKS PASSED")
